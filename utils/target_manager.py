"""
مدیریت تارگت‌ها (هدف‌های فروش)
"""

import os
import json
import random
import string
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from kivy.logger import Logger as logger
from utils.storage import get_data_path
from utils.jalali_date import (
    get_today_jalali,
    get_current_time,
    convert_to_gregorian,
    validate_jalali_date,
    to_jalali
)
from constants import TARGET_TYPES, TARGET_STATUSES

TARGETS_FILE = 'targets.json'


def _get_targets_path() -> str:
    """دریافت مسیر فایل تارگت‌ها"""
    return os.path.join(get_data_path(), TARGETS_FILE)


def _load_targets() -> List[Dict]:
    """بارگذاری همه تارگت‌ها"""
    try:
        path = _get_targets_path()
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
                return []
        return []
    except Exception as e:
        logger.error(f"خطا در بارگذاری تارگت‌ها: {e}")
        return []


def _save_targets(targets: List[Dict]) -> bool:
    """ذخیره تارگت‌ها"""
    try:
        path = _get_targets_path()
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(targets, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"خطا در ذخیره تارگت‌ها: {e}")
        return False


def _generate_target_id() -> str:
    """تولید شناسه یکتا برای تارگت"""
    chars = string.ascii_uppercase + string.digits
    random_part = ''.join(random.choices(chars, k=4))
    return f"TG{random_part}"


# ============================================================
# ✅ توابع محاسبه تاریخ با پشتیبانی از دوره‌ها
# ============================================================

# utils/target_manager.py

# ============================================================
# ✅ توابع محاسبه تاریخ با پشتیبانی از دوره‌ها (مثل ریزتارگت)
# ============================================================

def _add_days(date_str: str, days: int) -> str:
    """افزودن تعداد روز به تاریخ شمسی"""
    try:
        gregorian = convert_to_gregorian(date_str)
        if not gregorian or gregorian == date_str:
            return date_str
        dt = datetime.strptime(gregorian, '%Y-%m-%d')
        new_dt = dt + timedelta(days=days)
        return to_jalali(new_dt.year, new_dt.month, new_dt.day)
    except Exception as e:
        logger.error(f"خطا در add_days: {e}")
        return date_str


def _add_months(date_str: str, months: int) -> str:
    """
    افزودن تعداد ماه به تاریخ شمسی و برگرداندن آخرین روز ماه
    
    مثال‌ها:
    - _add_months('1405/05/01', 1) -> '1405/05/31'
    - _add_months('1405/05/15', 1) -> '1405/05/31'
    - _add_months('1405/12/01', 1) -> '1405/12/29' (اگر سال کبیسه نباشه)
    - _add_months('1405/12/01', 2) -> '1406/01/31'
    """
    try:
        parts = date_str.split('/')
        if len(parts) != 3:
            return date_str
        
        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
        
        # محاسبه ماه و سال جدید
        new_month = month + months
        new_year = year
        
        # تنظیم سال اگر ماه از 12 بیشتر شد
        while new_month > 12:
            new_month -= 12
            new_year += 1
        
        # تنظیم سال اگر ماه به صفر یا کمتر رسید
        while new_month < 1:
            new_month += 12
            new_year -= 1
        
        # ✅ محاسبه آخرین روز **ماه شروع** (نه ماه جدید!)
        # اینجا باید از month استفاده کنیم، چون میخوایم آخرین روز 
        # همان ماهی که تاریخ شروع در اون هست رو برگردونیم
        
        if 1 <= month <= 6:  # ماه‌های ۳۱ روزه
            max_day = 31
        elif 7 <= month <= 11:  # ماه‌های ۳۰ روزه
            max_day = 30
        else:  # اسفند (ماه ۱۲)
            # سال جدید رو برای محاسبه کبیسه در نظر بگیر
            max_day = 29 if new_year % 4 == 0 else 28
        
        # ✅ برگرداندن آخرین روز ماه شروع با سال جدید
        return f"{new_year:04d}/{month:02d}/{max_day:02d}"
        
    except Exception as e:
        logger.error(f"خطا در add_months: {e}")
        return date_str


def _add_years(date_str: str, years: int) -> str:
    """افزودن تعداد سال به تاریخ شمسی (مثل ریزتارگت)"""
    try:
        parts = date_str.split('/')
        if len(parts) != 3:
            return date_str
        
        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
        new_year = year + years
        
        # محاسبه حداکثر روز ماه
        if 1 <= month <= 6:
            max_day = 31
        elif 7 <= month <= 11:
            max_day = 30
        else:  # اسفند
            max_day = 29 if new_year % 4 == 0 else 28
        
        if day > max_day:
            day = max_day
        
        return f"{new_year:04d}/{month:02d}/{day:02d}"
        
    except Exception as e:
        logger.error(f"خطا در add_years: {e}")
        return date_str


def _calculate_end_date(start_date: str, period_type: str, duration: int) -> str:
    """
    محاسبه تاریخ پایان بر اساس دوره و مدت (مثل ریزتارگت)
    """
    try:
        logger.info(f"_calculate_end_date: start={start_date}, period={period_type}, duration={duration}")
        
        if period_type == 'daily':
            result = _add_days(start_date, duration)
        elif period_type == 'weekly':
            result = _add_days(start_date, duration * 7)
        elif period_type == 'monthly':
            result = _add_months(start_date, duration)
        elif period_type == 'seasonal':
            result = _add_months(start_date, duration * 3)
        elif period_type == 'yearly':
            result = _add_years(start_date, duration)
        else:
            logger.warning(f"نوع دوره نامشخص: {period_type} - استفاده از پیش‌فرض روزانه")
            result = _add_days(start_date, duration)
        
        logger.info(f"_calculate_end_date: نتیجه={result}")
        return result
        
    except Exception as e:
        logger.error(f"خطا در _calculate_end_date: {e}")
        return start_date


# ============================================================
# تابع اصلی ایجاد تارگت
# ============================================================

def create_target(
    agent_name: str,
    target_type: str,
    target_value: int,
    period_type: str,
    duration: int,
    start_date: str,
    description: str = '',
    created_by: str = 'supervisor'
) -> Tuple[bool, str, Optional[Dict]]:
    """
    ایجاد تارگت جدید با محاسبه تاریخ پایان بر اساس دوره
    
    Args:
        agent_name: نام عامل (بازاریاب)
        target_type: نوع تارگت (تعدادی, مبلغی, تعداد فاکتور, ...)
        target_value: مقدار هدف
        period_type: نوع دوره (daily, weekly, monthly, seasonal, yearly)
        duration: تعداد دوره‌ها
        start_date: تاریخ شروع (فرمت: 1405/01/31)
        description: توضیحات (اختیاری)
        created_by: نام ایجادکننده (سوپروایزر)
    
    Returns:
        Tuple[bool, str, Optional[Dict]]: (موفقیت, پیام, دیتای تارگت)
    """
    try:
        # ============================================================
        # اعتبارسنجی فیلدها
        # ============================================================
        if not agent_name:
            return False, 'نام عامل الزامی است', None

        if not target_type:
            return False, 'نوع تارگت الزامی است', None

        if target_value <= 0:
            return False, 'میزان هدف باید بزرگتر از صفر باشد', None

        if not period_type:
            return False, 'نوع دوره الزامی است', None

        if duration <= 0:
            return False, 'مدت تارگت باید بزرگتر از صفر باشد', None

        if not start_date:
            return False, 'تاریخ شروع الزامی است', None

        if not validate_jalali_date(start_date):
            return False, 'فرمت تاریخ نامعتبر است (مثال: 1405/01/31)', None

        # ============================================================
        # ✅ محاسبه تاریخ پایان بر اساس دوره
        # ============================================================
        end_date = _calculate_end_date(start_date, period_type, duration)
        
        logger.info(f"ایجاد تارگت: {agent_name} - {target_type} - {start_date} تا {end_date}")

        # ============================================================
        # تولید شناسه یکتا
        # ============================================================
        target_id = _generate_target_id()

        # ============================================================
        # ایجاد دیکشنری تارگت
        # ============================================================
        target = {
            'target_id': target_id,
            'agent_name': agent_name,
            'target_type': target_type,
            'target_value': target_value,
            'period_type': period_type,
            'duration': duration,
            'description': description or '',
            'start_date': start_date,
            'end_date': end_date,  # ✅ تاریخ پایان محاسبه شده
            'status': 'در انتظار',
            'is_active': True,
            'is_locked': False,
            'achieved_value': 0,
            'created_at': datetime.now().isoformat(),
            'created_by': created_by or 'supervisor',
            'finalized_at': ''
        }

        # ============================================================
        # بررسی تکراری بودن
        # ============================================================
        is_allowed, dup_msg, existing = check_duplicate_target(agent_name, target_type, period_type)
        if not is_allowed:
            return False, dup_msg, None

        # ============================================================
        # ذخیره در فایل
        # ============================================================
        targets = _load_targets()
        targets.append(target)

        if _save_targets(targets):
            logger.info(f"✅ تارگت جدید ایجاد شد: {target_id} - تاریخ پایان: {end_date} - توسط: {created_by}")
            return True, f'تارگت با شناسه {target_id} ثبت شد', target
        else:
            return False, 'خطا در ذخیره تارگت', None

    except Exception as e:
        logger.error(f"خطا در ایجاد تارگت: {e}")
        import traceback
        traceback.print_exc()
        return False, f'خطا: {str(e)}', None


# ============================================================
# توابع دریافت و فیلتر
# ============================================================

def get_all_targets() -> List[Dict]:
    """دریافت همه تارگت‌ها"""
    return _load_targets()


def get_targets_by_agent(agent_name: str) -> List[Dict]:
    """دریافت تارگت‌های یک عامل"""
    targets = _load_targets()
    return [t for t in targets if t.get('agent_name') == agent_name]


def get_targets_by_status(status: str) -> List[Dict]:
    """دریافت تارگت‌ها بر اساس وضعیت"""
    targets = _load_targets()
    return [t for t in targets if t.get('status') == status]


def get_targets_by_type(target_type: str) -> List[Dict]:
    """دریافت تارگت‌ها بر اساس نوع"""
    targets = _load_targets()
    return [t for t in targets if t.get('target_type') == target_type]


def get_targets_filtered(
    agent_name: str = None,
    target_type: str = None,
    status: str = None,
    period_type: str = None
) -> List[Dict]:
    """دریافت تارگت‌ها با فیلترهای دلخواه"""
    targets = _load_targets()
    result = targets

    if agent_name:
        result = [t for t in result if t.get('agent_name') == agent_name]

    if target_type:
        result = [t for t in result if t.get('target_type') == target_type]

    if status:
        result = [t for t in result if t.get('status') == status]

    if period_type:
        result = [t for t in result if t.get('period_type') == period_type]

    result.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return result


# ============================================================
# توابع برای تب تحقق تارگت
# ============================================================

def get_active_targets_by_agent(agent_name: str, start_date: str = None, end_date: str = None) -> List[Dict]:
    """
    دریافت تارگت‌های فعال یا در انتظار یک عامل در بازه زمانی مشخص
    """
    try:
        if not agent_name:
            return []
        
        targets = _load_targets()
        result = []
        
        for target in targets:
            if target.get('agent_name') != agent_name:
                continue
            
            status = target.get('status', '')
            if status not in ['فعال', 'در انتظار']:
                continue
            
            target_start = target.get('start_date', '')
            if start_date and target_start < start_date:
                continue
            if end_date and target_start > end_date:
                continue
            
            result.append(target)
        
        result.sort(key=lambda x: x.get('start_date', ''))
        return result
        
    except Exception as e:
        logger.error(f"خطا در دریافت تارگت‌های فعال: {e}")
        return []


def finalize_targets(target_ids: List[str], achieved_values: Dict[str, int], finalized_by: str = None) -> Tuple[bool, str]:
    """
    نهایی‌سازی تارگت‌های انتخاب شده
    
    Args:
        target_ids: لیست شناسه‌های تارگت
        achieved_values: دیکشنری {target_id: achieved_value}
        finalized_by: نام کاربری که نهایی‌سازی را انجام داده (اختیاری)
    """
    try:
        targets = _load_targets()
        updated = 0
        
        for i, target in enumerate(targets):
            target_id = target.get('target_id')
            if target_id in target_ids:
                status = target.get('status', '')
                if status not in ['فعال', 'در انتظار']:
                    continue
                
                achieved = achieved_values.get(target_id, 0)
                targets[i]['achieved_value'] = achieved
                targets[i]['status'] = 'تکمیل شده'
                targets[i]['finalized_at'] = datetime.now().isoformat()
                if finalized_by:
                    targets[i]['finalized_by'] = finalized_by  # ✅ اضافه شد
                updated += 1
        
        if updated == 0:
            return False, 'هیچ تارگت قابل نهایی‌سازی یافت نشد'
        
        if _save_targets(targets):
            logger.info(f"{updated} تارگت نهایی‌سازی شد")
            return True, f'{updated} تارگت با موفقیت نهایی‌سازی شد'
        else:
            return False, 'خطا در ذخیره تارگت‌ها'
        
    except Exception as e:
        logger.error(f"خطا در نهایی‌سازی تارگت‌ها: {e}")
        return False, f'خطا: {str(e)}'


def read_excel_summary(filepath: str) -> Dict[str, int]:
    """
    خواندن داده‌های خلاصه از فایل اکسل
    """
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        if 'خلاصه آمار' not in wb.sheetnames:
            return {}
        
        ws = wb['خلاصه آمار']
        summary_data = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row and len(row) >= 2:
                key = str(row[0]).strip()
                value = row[1]
                if value is not None:
                    try:
                        if isinstance(value, str):
                            value = value.replace(',', '').replace(' ', '')
                        summary_data[key] = int(value)
                    except (ValueError, TypeError):
                        summary_data[key] = 0
        
        wb.close()
        return summary_data
        
    except Exception as e:
        logger.error(f"خطا در خواندن فایل اکسل: {e}")
        return {}


def can_edit_target(target: Dict) -> bool:
    """
    بررسی اینکه آیا تارگت قابل ویرایش است یا نه
    """
    try:
        status = target.get('status', '')
        if status != 'تکمیل شده':
            return True

        finalized_at = target.get('finalized_at', '')
        if not finalized_at:
            return True

        finalized_date = datetime.fromisoformat(finalized_at)
        now = datetime.now()
        days_diff = (now - finalized_date).days

        return days_diff <= 5

    except Exception as e:
        logger.error(f"خطا در بررسی ویرایش تارگت: {e}")
        return False


def update_target(target_id: str, updates: Dict) -> Tuple[bool, str]:
    """
    به‌روزرسانی تارگت
    """
    try:
        targets = _load_targets()

        for i, target in enumerate(targets):
            if target.get('target_id') == target_id:
                if not can_edit_target(target):
                    return False, 'این تارگت قابل ویرایش نیست'

                allowed_fields = [
                    'target_type', 'target_value', 'duration',
                    'start_date', 'description', 'is_active', 'status'
                ]

                for field, value in updates.items():
                    if field in allowed_fields:
                        if field == 'start_date':
                            if validate_jalali_date(value):
                                duration_val = updates.get('duration', target.get('duration', 0))
                                period_type = target.get('period_type', 'daily')
                                target['end_date'] = _calculate_end_date(value, period_type, duration_val)
                            else:
                                return False, 'تاریخ شروع نامعتبر است'
                        target[field] = value

                if 'duration' in updates and 'start_date' in target:
                    period_type = target.get('period_type', 'daily')
                    target['end_date'] = _calculate_end_date(
                        target['start_date'],
                        period_type,
                        updates['duration']
                    )

                if _save_targets(targets):
                    return True, 'تارگت با موفقیت به‌روزرسانی شد'
                else:
                    return False, 'خطا در ذخیره تارگت'

        return False, 'تارگت یافت نشد'

    except Exception as e:
        logger.error(f"خطا در به‌روزرسانی تارگت: {e}")
        return False, f'خطا: {str(e)}'


def delete_target(target_id: str) -> Tuple[bool, str]:
    """
    حذف تارگت
    """
    try:
        targets = _load_targets()

        for i, target in enumerate(targets):
            if target.get('target_id') == target_id:
                if not can_edit_target(target):
                    return False, 'این تارگت قابل حذف نیست'

                targets.pop(i)

                if _save_targets(targets):
                    return True, 'تارگت با موفقیت حذف شد'
                else:
                    return False, 'خطا در ذخیره تارگت'

        return False, 'تارگت یافت نشد'

    except Exception as e:
        logger.error(f"خطا در حذف تارگت: {e}")
        return False, f'خطا: {str(e)}'


def get_target_statistics() -> Dict:
    """
    دریافت آمار کلی تارگت‌ها
    """
    try:
        targets = _load_targets()
        return {
            'total': len(targets),
            'pending': len([t for t in targets if t.get('status') == 'در انتظار']),
            'active': len([t for t in targets if t.get('status') == 'فعال']),
            'completed': len([t for t in targets if t.get('status') == 'تکمیل شده']),
            'cancelled': len([t for t in targets if t.get('status') == 'لغو شده'])
        }
    except Exception as e:
        logger.error(f"خطا در دریافت آمار تارگت‌ها: {e}")
        return {'total': 0, 'pending': 0, 'active': 0, 'completed': 0, 'cancelled': 0}


def export_targets_to_excel(targets: List[Dict], filename: str = None) -> Tuple[bool, str, str]:
    """
    خروجی گرفتن از تارگت‌ها به صورت فایل Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from utils.storage import get_backup_path
        from utils.jalali_date import get_today_jalali
        
        if not targets:
            return False, 'هیچ تارگتی برای خروجی وجود ندارد', ''
        
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "تارگت‌ها"
        ws1.sheet_view.rightToLeft = True
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['شناسه', 'عامل', 'نوع تارگت', 'میزان هدف', 'دوره', 'مدت (روز)',
                   'تاریخ شروع', 'تاریخ پایان', 'وضعیت', 'مقدار محقق شده', 'توضیحات', 'ایجاد شده توسط']
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        period_display_map = {
            'daily': 'روزانه',
            'weekly': 'هفتگی',
            'monthly': 'ماهانه',
            'seasonal': 'فصلی',
            'yearly': 'سالانه'
        }
        
        for row, target in enumerate(targets, 2):
            period_type = target.get('period_type', '')
            period_display = period_display_map.get(period_type, period_type)
            
            ws1.cell(row=row, column=1, value=target.get('target_id', ''))
            ws1.cell(row=row, column=2, value=target.get('agent_name', ''))
            ws1.cell(row=row, column=3, value=target.get('target_type', ''))
            ws1.cell(row=row, column=4, value=target.get('target_value', 0))
            ws1.cell(row=row, column=5, value=period_display)
            ws1.cell(row=row, column=6, value=target.get('duration', 0))
            ws1.cell(row=row, column=7, value=target.get('start_date', ''))
            ws1.cell(row=row, column=8, value=target.get('end_date', ''))
            ws1.cell(row=row, column=9, value=target.get('status', ''))
            ws1.cell(row=row, column=10, value=target.get('achieved_value', 0))
            ws1.cell(row=row, column=11, value=target.get('description', ''))
            ws1.cell(row=row, column=12, value=target.get('created_by', ''))
        
        column_widths = [14, 20, 16, 16, 14, 12, 14, 14, 14, 18, 30, 18]
        for i, width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = width
        
        for row in ws1.iter_rows(min_row=2, max_row=len(targets) + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        if not filename:
            today = get_today_jalali().replace('/', '-')
            timestamp = datetime.now().strftime('%H%M%S')
            filename = f'گزارش_تارگت_{today}_{timestamp}.xlsx'
        
        export_dir = get_backup_path()
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)
        
        wb.save(filepath)
        
        if os.path.exists(filepath):
            return True, f'فایل با موفقیت ذخیره شد:\n{filename}', filepath
        else:
            return False, 'فایل ساخته نشد', ''
        
    except ImportError:
        return False, 'ماژول openpyxl نصب نیست', ''
    except Exception as e:
        logger.error(f"خطا در خروجی اکسل: {e}")
        return False, f'خطا: {str(e)}', ''


def get_targets_filtered_advanced(
    agent_name: str = None,
    target_type: str = None,
    status: str = None,
    period_type: str = None,
    target_id: str = None,
    start_date: str = None,
    end_date: str = None
) -> List[Dict]:
    """
    دریافت تارگت‌ها با فیلترهای پیشرفته
    """
    try:
        targets = _load_targets()
        if not isinstance(targets, list):
            return []
        
        result = targets
        
        if agent_name:
            result = [t for t in result if isinstance(t, dict) and t.get('agent_name') == agent_name]
        
        if target_type:
            result = [t for t in result if isinstance(t, dict) and t.get('target_type') == target_type]
        
        if status:
            result = [t for t in result if isinstance(t, dict) and t.get('status') == status]
        
        if period_type:
            result = [t for t in result if isinstance(t, dict) and t.get('period_type') == period_type]
        
        if target_id:
            search_term = target_id.strip().upper()
            result = [t for t in result if isinstance(t, dict) and search_term in str(t.get('target_id', '')).upper()]
        
        if start_date:
            result = [t for t in result if isinstance(t, dict) and t.get('start_date', '') >= start_date]
        
        if end_date:
            result = [t for t in result if isinstance(t, dict) and t.get('start_date', '') <= end_date]
        
        result.sort(key=lambda x: x.get('created_at', '') if isinstance(x, dict) else '', reverse=True)
        
        return result
        
    except Exception as e:
        logger.error(f"خطا در فیلتر تارگت‌ها: {e}")
        return []


def check_duplicate_target(agent_name: str, target_type: str, period_type: str) -> Tuple[bool, str, Optional[Dict]]:
    """
    بررسی تکراری نبودن تارگت
    """
    all_targets = _load_targets()
    if not isinstance(all_targets, list):
        return True, '', None
    
    period_display_map = {
        'daily': 'روزانه',
        'weekly': 'هفتگی',
        'monthly': 'ماهانه',
        'seasonal': 'فصلی',
        'yearly': 'سالانه'
    }
    
    # ۱. بررسی تکراری بودن دقیق
    for t in all_targets:
        if not isinstance(t, dict):
            continue
        if (t.get('agent_name') == agent_name and 
            t.get('target_type') == target_type and 
            t.get('period_type') == period_type and
            t.get('status') in ['در انتظار', 'فعال']):
            
            period_display = period_display_map.get(period_type, period_type)
            return False, f'در دوره انتخابی، تارگت {target_type} {period_display} برای {agent_name} قبلاً تعریف شده است.', t
    
    # ۲. بررسی تکراری بودن با دوره متفاوت
    for t in all_targets:
        if not isinstance(t, dict):
            continue
        if (t.get('agent_name') == agent_name and 
            t.get('target_type') == target_type and 
            t.get('period_type') != period_type and
            t.get('status') in ['در انتظار', 'فعال']):
            
            existing_period = t.get('period_type', '')
            existing_display = period_display_map.get(existing_period, existing_period)
            new_display = period_display_map.get(period_type, period_type)
            
            return False, f'برای {agent_name} تارگت {target_type} {existing_display} ثبت شده است.آیا از ایجاد تارگت {target_type} {new_display} اطمینان دارید؟', t
    
    return True, '', None


# ============================================================
# تابع تست
# ============================================================

def test_target_manager():
    """تست توابع مدیریت تارگت"""
    print("\n" + "=" * 50)
    print("تست مدیریت تارگت‌ها")
    print("=" * 50)

    # ایجاد تارگت تست با دوره ماهانه
    success, msg, target = create_target(
        agent_name='حیدری ناصر',
        target_type='ریالی',
        target_value=30000000000,
        period_type='monthly',
        duration=1,
        start_date='1405/05/01',
        description='کف تارگت فروش ریالی',
        created_by='supervisor'
    )

    if success:
        print(f"✅ موفق: {msg}")
        print(f"   شناسه: {target['target_id']}")
        print(f"   تاریخ شروع: {target['start_date']}")
        print(f"   تاریخ پایان: {target['end_date']}")
        print(f"   نوع دوره: {target['period_type']}")
    else:
        print(f"❌ خطا: {msg}")

    print("=" * 50)


if __name__ == '__main__':
    test_target_manager()
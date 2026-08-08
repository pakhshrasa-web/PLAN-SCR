# utils/report_exporter.py
# ========== کلاس کمکی برای خروجی گزارشات با استفاده از توابع اصلی ==========

import os
import json
import traceback
from datetime import datetime
from kivy.clock import Clock
from kivy.metrics import dp, sp

from utils.storage import get_data_path, get_backup_path
from utils.jalali_date import get_today_jalali
from utils.attendance_manager import AttendanceManager
from utils.collection_manager import get_collections
from utils.delivery_manager import get_all_deliveries
from utils.detailed_target_manager import get_all_detailed_targets, export_to_excel as export_detailed_to_excel
from utils.target_manager import get_targets_filtered, export_targets_to_excel
from utils.supervisor_visits_manager import get_visits_filtered, export_visits_to_excel
from utils.excel_exporter_distributor import export_distributor_to_excel
from utils.file_manager import get_daily_logs


class ReportExporter:
    """
    کلاس کمکی برای تولید خروجی‌های اکسل
    از توابع اصلی موجود در اسکرین‌ها استفاده می‌کند
    """
    
    @staticmethod
    def export_attendance(user_id, from_date, to_date, user_name):
        """خروجی حضور و غیاب - با استفاده از AttendanceManager"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            records = AttendanceManager.load_attendance()
            
            filtered = []
            for r in records:
                if r.get('user_id') != user_id:
                    continue
                r_date = r.get('date', '')
                if from_date <= r_date <= to_date:
                    filtered.append(r)
            
            if not filtered:
                return None
            
            # ایجاد پوشه
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            filename = f'حضور_و_غیاب_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "حضور و غیاب"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['ردیف', 'تاریخ', 'ورود', 'خروج', 'کارکرد']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, record in enumerate(filtered, 1):
                check_in = record.get('check_in', '')
                check_out = record.get('check_out', '')
                diff = '-'
                if check_in and check_out:
                    try:
                        in_h, in_m = map(int, check_in.split(':'))
                        out_h, out_m = map(int, check_out.split(':'))
                        diff_min = (out_h - in_h) * 60 + (out_m - in_m)
                        diff_h = diff_min // 60
                        diff_m = diff_min % 60
                        diff = f'{diff_h:02d}:{diff_m:02d}'
                    except:
                        pass
                
                values = [idx, record.get('date', ''), check_in, check_out, diff]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی حضور و غیاب: {e}")
            return None
    
    @staticmethod
    def export_leave(user_id, from_date, to_date, user_name):
        """خروجی مرخصی"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            if not os.path.exists(file_path):
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_requests = json.load(f)
            
            filtered = []
            for r in all_requests:
                if r.get('user_id') != user_id:
                    continue
                r_date = r.get('created_at', '')
                if from_date <= r_date <= to_date:
                    filtered.append(r)
            
            if not filtered:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            filename = f'مرخصی_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "مرخصی"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['ردیف', 'نوع', 'مدت', 'شروع', 'پایان', 'وضعیت', 'تاریخ ثبت']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, req in enumerate(filtered, 1):
                values = [
                    idx,
                    req.get('leave_type', ''),
                    req.get('duration_display', ''),
                    req.get('start_date', ''),
                    req.get('end_date', ''),
                    req.get('status', ''),
                    req.get('created_at', '')
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی مرخصی: {e}")
            return None
    
    @staticmethod
    def export_mission(agent_name, from_date, to_date, user_name):
        """خروجی ماموریت - با استفاده از فایل do_missions.json"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            file_path = os.path.join(get_data_path(), 'do_missions.json')
            if not os.path.exists(file_path):
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_missions = json.load(f)
            
            filtered = []
            for m in all_missions:
                if m.get('agent_name', '') != agent_name:
                    continue
                m_date = m.get('start_date', '')
                if from_date <= m_date <= to_date:
                    filtered.append(m)
            
            if not filtered:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            filename = f'ماموریت_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ماموریت"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['ردیف', 'شناسه', 'نوع', 'روش', 'شروع', 'پایان', 'امتیاز', 'هدف', 'وضعیت']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, mission in enumerate(filtered, 1):
                values = [
                    idx,
                    mission.get('id', ''),
                    mission.get('type', ''),
                    mission.get('method', ''),
                    mission.get('start_date', ''),
                    mission.get('end_date', ''),
                    mission.get('score', 0),
                    mission.get('target', 0),
                    mission.get('status', '')
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی ماموریت: {e}")
            return None
    
    @staticmethod
    def export_daily_visits(agent_name, from_date, to_date, user_name):
        """خروجی ویزیت روزانه - با استفاده از daily_log.json"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            all_logs = get_daily_logs()
            
            filtered = []
            for date, logs in all_logs.items():
                if from_date <= date <= to_date:
                    if isinstance(logs, list):
                        for log in logs:
                            if isinstance(log, dict) and log.get('agent_name') == agent_name:
                                log['date'] = date
                                filtered.append(log)
            
            if not filtered:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            filename = f'ویزیت_روزانه_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ویزیت روزانه"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['ردیف', 'تاریخ', 'مسیر', 'مشتری', 'وضعیت ویزیت', 'وضعیت فروش', 'مبلغ فروش', 'واحد فروش', 'نحوه پرداخت']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, log in enumerate(filtered, 1):
                values = [
                    idx,
                    log.get('date', ''),
                    log.get('route', ''),
                    log.get('customer', ''),
                    log.get('visit_status', ''),
                    log.get('sales_status', ''),
                    log.get('sales_amount', 0),
                    log.get('units_sold', 0),
                    log.get('payment_method', '')
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی ویزیت روزانه: {e}")
            return None
    
    @staticmethod
    def export_collection(agent_name, from_date, to_date, user_name):
        """خروجی وصول - با استفاده از collection.json"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            all_collections = get_collections(agent_name=agent_name)
            
            filtered = []
            for c in all_collections:
                c_date = c.get('date', '')
                if from_date <= c_date <= to_date:
                    filtered.append(c)
            
            if not filtered:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            filename = f'وصول_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "وصول"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['ردیف', 'تاریخ', 'مشتری', 'مسیر', 'وضعیت', 'نوع پرداخت', 'مبلغ نقد', 'مبلغ چک', 'جمع کل']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, col in enumerate(filtered, 1):
                status = col.get('status', '')
                cash = col.get('net_cash', 0)
                check = col.get('total_check_amount', 0)
                total = cash + check
                
                if col.get('has_cash') and col.get('has_check'):
                    payment = 'نقد + چک'
                elif col.get('has_cash'):
                    payment = 'نقد'
                elif col.get('has_check'):
                    payment = 'چک'
                else:
                    payment = '-'
                
                values = [
                    idx,
                    col.get('date', ''),
                    col.get('customer', ''),
                    col.get('route', ''),
                    status,
                    payment,
                    cash,
                    check,
                    total
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی وصول: {e}")
            return None
    
    @staticmethod
    def export_detailed_targets(agent_name, from_date, to_date, user_name):
        """خروجی ریزتارگت‌ها - با استفاده از توابع اصلی"""
        try:
            all_targets = get_all_detailed_targets()
            
            filtered = []
            for t in all_targets:
                if t.get('agent_name', '') != agent_name:
                    continue
                t_date = t.get('start_date', '')
                if from_date <= t_date <= to_date:
                    filtered.append(t)
            
            if not filtered:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            # استفاده از تابع اصلی
            success, message, filepath = export_detailed_to_excel(filtered)
            if success:
                return os.path.basename(filepath)
            return None
            
        except Exception as e:
            print(f"خطا در خروجی ریزتارگت‌ها: {e}")
            return None
    
    @staticmethod
    def export_delivery(agent_name, from_date, to_date, user_name):
        """خروجی توزیع - با استفاده از تابع اصلی"""
        try:
            all_deliveries = get_all_deliveries()
            filtered_data = {}
            
            for date, deliveries in all_deliveries.items():
                if from_date <= date <= to_date:
                    # فیلتر بر اساس نام موزع
                    filtered = []
                    for d in deliveries:
                        if d.get('distributor_name') == agent_name:
                            filtered.append(d)
                    if filtered:
                        filtered_data[date] = filtered
            
            if not filtered_data:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            success, result = export_distributor_to_excel(filtered_data)
            if success:
                today = get_today_jalali().replace('/', '-')
                timestamp = datetime.now().strftime("%H%M%S")
                filename = f'توزیع_{user_name}_{today}_{timestamp}.xlsx'
                return filename
            return None
            
        except Exception as e:
            print(f"خطا در خروجی توزیع: {e}")
            return None
    
    @staticmethod
    def export_targets(agent_name, from_date, to_date, user_name):
        """خروجی تارگت‌ها - با استفاده از تابع اصلی"""
        try:
            all_targets = get_targets_filtered(agent_name=agent_name)
            
            filtered = []
            for t in all_targets:
                t_date = t.get('start_date', '')
                if from_date <= t_date <= to_date:
                    filtered.append(t)
            
            if not filtered:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            success, message, filepath = export_targets_to_excel(filtered)
            if success:
                return os.path.basename(filepath)
            return None
            
        except Exception as e:
            print(f"خطا در خروجی تارگت‌ها: {e}")
            return None
    
    @staticmethod
    def export_supervisor_visits(agent_name, from_date, to_date, user_name):
        """خروجی سرکشی‌ها - با استفاده از تابع اصلی"""
        try:
            all_visits = get_visits_filtered(agent_name=agent_name, start_date=from_date, end_date=to_date)
            
            if not all_visits:
                return None
            
            reports_dir = os.path.join(get_backup_path(), 'daily_reports', get_today_jalali().replace('/', '-'))
            os.makedirs(reports_dir, exist_ok=True)
            
            success, message, filepath = export_visits_to_excel(all_visits)
            if success:
                return os.path.basename(filepath)
            return None
            
        except Exception as e:
            print(f"خطا در خروجی سرکشی‌ها: {e}")
            return None


def get_report_exporter():
    """دریافت نمونه از ReportExporter"""
    return ReportExporter
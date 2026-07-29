"""
مدیریت ریزتارگت‌ها
"""

import os
import json
import random
import string
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from kivy.logger import Logger as logger
from utils.storage import get_data_path
from utils.jalali_date import (
    get_today_jalali,
    get_current_time,
    convert_to_gregorian,
    validate_jalali_date,
    to_jalali
)

DETAILED_TARGETS_FILE = 'detailed_targets.json'


def _get_path() -> str:
    """دریافت مسیر فایل ریزتارگت‌ها"""
    return os.path.join(get_data_path(), DETAILED_TARGETS_FILE)


def _load() -> List[Dict]:
    """بارگذاری همه ریزتارگت‌ها"""
    try:
        path = _get_path()
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
                return []
        return []
    except Exception as e:
        logger.error(f"خطا در بارگذاری ریزتارگت‌ها: {e}")
        return []


def _save(data: List[Dict]) -> bool:
    """ذخیره ریزتارگت‌ها"""
    try:
        path = _get_path()
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"خطا در ذخیره ریزتارگت‌ها: {e}")
        return False


def _generate_id() -> str:
    """تولید آیدی یکتا"""
    chars = string.ascii_uppercase + string.digits
    random_part = ''.join(random.choices(chars, k=4))
    return f"DTG{random_part}"


def _calculate_end_date(start_date: str, period: str, duration: int) -> str:
    """
    محاسبه تاریخ پایان بر اساس تقویم شمسی واقعی
    
    Args:
        start_date: تاریخ شروع (مثال: 1405/05/01)
        period: روزانه / ماهانه / فصلی / سالیانه
        duration: تعداد دوره
    
    Returns:
        تاریخ پایان شمسی - همیشه آخرین روز ماه/دوره
    """
    try:
        parts = start_date.split('/')
        if len(parts) != 3:
            return start_date
        
        jy = int(parts[0])
        jm = int(parts[1])
        jd = int(parts[2])
        
        def get_month_days(year, month):
            """تعداد روزهای واقعی هر ماه شمسی"""
            if 1 <= month <= 6:
                return 31
            elif 7 <= month <= 11:
                return 30
            else:  # اسفند
                leap_years = {1, 5, 9, 13, 17, 22, 26, 30}
                year_mod = year % 33
                return 30 if year_mod in leap_years else 29
        
        if period == 'روزانه':
            total_days = jd
            current_year, current_month = jy, jm
            
            for _ in range(duration - 1):
                total_days += 1
                if total_days > get_month_days(current_year, current_month):
                    total_days = 1
                    current_month += 1
                    if current_month > 12:
                        current_month = 1
                        current_year += 1
            
            return f"{current_year:04d}/{current_month:02d}/{total_days:02d}"
            
        elif period == 'ماهانه':
            total_months = jm + duration - 1
            current_year = jy
            
            while total_months > 12:
                total_months -= 12
                current_year += 1
            
            last_day = get_month_days(current_year, total_months)
            
            return f"{current_year:04d}/{total_months:02d}/{last_day:02d}"
            
        elif period == 'فصلی':
            total_months = jm + (duration * 3) - 1
            current_year = jy
            
            while total_months > 12:
                total_months -= 12
                current_year += 1
            
            last_day = get_month_days(current_year, total_months)
            
            return f"{current_year:04d}/{total_months:02d}/{last_day:02d}"
            
        elif period == 'سالیانه':
            current_year = jy + duration
            last_day = get_month_days(current_year, jm)
            
            return f"{current_year:04d}/{jm:02d}/{last_day:02d}"
            
        else:
            return start_date
        
    except Exception as e:
        logger.error(f"خطا در محاسبه تاریخ پایان: {e}")
        return start_date


def _calculate_daily_target(target_count: int, period: str, duration: int) -> int:
    """
    محاسبه تارگت روزانه
    
    روزانه: / duration
    ماهانه: / (duration × 26)
    فصلی:   / (duration × 78)  (3 ماه × 26)
    سالیانه: / (duration × 312) (12 ماه × 26)
    """
    if period == 'روزانه':
        divisor = duration
    elif period == 'ماهانه':
        divisor = duration * 26
    elif period == 'فصلی':
        divisor = duration * 78
    elif period == 'سالیانه':
        divisor = duration * 312
    else:
        divisor = duration
    
    if divisor <= 0:
        return target_count
    
    return max(1, round(target_count / divisor))


def create_detailed_target(
    agent_name: str,
    product_group: str,
    target_count: int,
    unit: str,
    period: str,
    duration: int,
    linked_target_id: str,
    start_date: str = '',
    created_by: str = 'supervisor'
) -> Tuple[bool, str, Optional[Dict]]:
    """ایجاد ریزتارگت جدید"""
    try:
        if not agent_name:
            return False, 'نام عامل الزامی است', None
        
        if not product_group:
            return False, 'گروه کالا الزامی است', None
        
        if target_count <= 0:
            return False, 'تعداد هدف باید بزرگتر از صفر باشد', None
        
        if not unit:
            return False, 'واحد تارگت الزامی است', None
        
        if not period:
            return False, 'دوره تارگت الزامی است', None
        
        if duration <= 0:
            return False, 'مدت باید بزرگتر از صفر باشد', None
        
        if not linked_target_id:
            return False, 'پیوند به تارگت اصلی الزامی است', None
        
        if not start_date:
            start_date = get_today_jalali()
        
        if not validate_jalali_date(start_date):
            return False, 'تاریخ شروع نامعتبر است', None
        
        end_date = _calculate_end_date(start_date, period, duration)
        daily_target = _calculate_daily_target(target_count, period, duration)
        target_id = _generate_id()
        
        target = {
            'id': target_id,
            'agent_name': agent_name,
            'product_group': product_group,
            'target_count': target_count,
            'unit': unit,
            'period': period,
            'duration': duration,
            'linked_target_id': linked_target_id,
            'start_date': start_date,
            'end_date': end_date,
            'daily_target': daily_target,
            'status': 'در انتظار',
            'achieved_value': 0,
            'created_at': datetime.now().isoformat(),
            'created_by': created_by
        }
        
        data = _load()
        data.append(target)
        
        if _save(data):
            logger.info(f"ریزتارگت جدید: {target_id}")
            return True, f'ریزتارگت با شناسه {target_id} ثبت شد', target
        else:
            return False, 'خطا در ذخیره', None
        
    except Exception as e:
        logger.error(f"خطا: {e}")
        return False, f'خطا: {str(e)}', None


def get_all_detailed_targets() -> List[Dict]:
    """دریافت همه ریزتارگت‌ها"""
    return _load()


def get_targets_by_agent(agent_name: str) -> List[Dict]:
    """دریافت ریزتارگت‌های یک عامل"""
    data = _load()
    return [t for t in data if isinstance(t, dict) and t.get('agent_name') == agent_name]


def get_targets_by_status(status: str) -> List[Dict]:
    """دریافت ریزتارگت‌ها بر اساس وضعیت"""
    data = _load()
    return [t for t in data if isinstance(t, dict) and t.get('status') == status]


def can_edit_target(target: Dict) -> bool:
    """
    بررسی قابلیت ویرایش
    
    شرایط:
    - وضعیت 'تکمیل شده' → قابل ویرایش نیست
    - وضعیت 'لغو شده' → قابل ویرایش نیست
    - حداکثر ۱۰ روز از تاریخ ایجاد گذشته باشد → قابل ویرایش نیست
    """
    try:
        if not isinstance(target, dict):
            return False
        
        # ✅ شرط اول: وضعیت تکمیل شده یا لغو شده = قفل
        status = target.get('status', '')
        if status in ['تکمیل شده', 'لغو شده']:
            return False
        
        # ✅ شرط دوم: حداکثر ۱۰ روز از ایجاد
        created_at = target.get('created_at', '')
        if not created_at:
            return True
        
        created_date = datetime.fromisoformat(created_at)
        now = datetime.now()
        days_diff = (now - created_date).days
        
        return days_diff <= 10
        
    except Exception as e:
        logger.error(f"خطا در بررسی ویرایش: {e}")
        return False


def update_detailed_target(target_id: str, updates: Dict) -> Tuple[bool, str]:
    """به‌روزرسانی ریزتارگت"""
    try:
        data = _load()
        
        for i, target in enumerate(data):
            if not isinstance(target, dict):
                continue
            
            if target.get('id') == target_id:
                if not can_edit_target(target):
                    return False, 'این ریزتارگت قابل ویرایش نیست (نهایی شده یا منقضی شده)'
                
                allowed = ['product_group', 'target_count', 'unit', 'period', 
                          'duration', 'linked_target_id', 'start_date']
                
                for field, value in updates.items():
                    if field in allowed:
                        target[field] = value
                
                target['daily_target'] = _calculate_daily_target(
                    target.get('target_count', 0),
                    target.get('period', ''),
                    target.get('duration', 1)
                )
                target['end_date'] = _calculate_end_date(
                    target.get('start_date', ''),
                    target.get('period', ''),
                    target.get('duration', 1)
                )
                
                if _save(data):
                    return True, 'به‌روزرسانی موفق'
                else:
                    return False, 'خطا در ذخیره'
        
        return False, 'ریزتارگت یافت نشد'
        
    except Exception as e:
        return False, f'خطا: {str(e)}'


def delete_detailed_target(target_id: str) -> Tuple[bool, str]:
    """حذف ریزتارگت"""
    try:
        data = _load()
        
        for i, target in enumerate(data):
            if not isinstance(target, dict):
                continue
            
            if target.get('id') == target_id:
                if not can_edit_target(target):
                    return False, 'این ریزتارگت قابل حذف نیست (نهایی شده یا منقضی شده)'
                
                data.pop(i)
                
                if _save(data):
                    return True, 'حذف موفق'
                else:
                    return False, 'خطا در ذخیره'
        
        return False, 'ریزتارگت یافت نشد'
        
    except Exception as e:
        return False, f'خطا: {str(e)}'
    
def get_detailed_targets_filtered(
    agent_name: str = None,
    product_group: str = None,
    status: str = None,
    period: str = None,
    target_id: str = None,
    linked_target_id: str = None,
    start_date: str = None,
    end_date: str = None
) -> List[Dict]:
    """
    دریافت ریزتارگت‌ها با فیلترهای دلخواه
    
    Args:
        agent_name: نام عامل (None = همه)
        product_group: گروه کالا (None = همه)
        status: وضعیت (None = همه)
        period: دوره (None = همه)
        target_id: جستجو در شناسه ریزتارگت (None = همه) - جستجوی شامل
        linked_target_id: جستجو در شناسه تارگت مادر (None = همه) - جستجوی شامل
        start_date: تاریخ شروع بازه (None = بدون فیلتر)
        end_date: تاریخ پایان بازه (None = بدون فیلتر)
    
    Returns:
        List[Dict]: لیست فیلتر شده
    """
    try:
        data = _load()
        if not isinstance(data, list):
            return []
        
        result = data
        
        # فیلتر عامل
        if agent_name:
            result = [t for t in result if isinstance(t, dict) and t.get('agent_name') == agent_name]
        
        # فیلتر گروه کالا
        if product_group:
            result = [t for t in result if isinstance(t, dict) and t.get('product_group') == product_group]
        
        # فیلتر وضعیت
        if status:
            result = [t for t in result if isinstance(t, dict) and t.get('status') == status]
        
        # فیلتر دوره
        if period:
            result = [t for t in result if isinstance(t, dict) and t.get('period') == period]
        
        # جستجوی شامل در شناسه ریزتارگت
        if target_id:
            search_term = target_id.strip().upper()
            result = [t for t in result if isinstance(t, dict) and search_term in str(t.get('id', '')).upper()]
        
        # جستجوی شامل در شناسه تارگت مادر
        if linked_target_id:
            search_term = linked_target_id.strip().upper()
            result = [t for t in result if isinstance(t, dict) and search_term in str(t.get('linked_target_id', '')).upper()]
        
        # فیلتر بازه زمانی (بر اساس تاریخ شروع)
        if start_date:
            result = [t for t in result if isinstance(t, dict) and t.get('start_date', '') >= start_date]
        
        if end_date:
            result = [t for t in result if isinstance(t, dict) and t.get('start_date', '') <= end_date]
        
        # مرتب‌سازی بر اساس تاریخ ایجاد (جدیدترین اول)
        result.sort(key=lambda x: x.get('created_at', '') if isinstance(x, dict) else '', reverse=True)
        
        return result
        
    except Exception as e:
        logger.error(f"خطا در فیلتر ریزتارگت‌ها: {e}")
        return []

def check_duplicate_detailed_target(agent_name: str, product_group: str, period: str, linked_target_id: str) -> Tuple[bool, str, Optional[Dict]]:
    """
    بررسی تکراری نبودن ریزتارگت
    
    Args:
        agent_name: نام عامل
        product_group: گروه کالا
        period: دوره (روزانه، ماهانه، فصلی، سالیانه)
        linked_target_id: شناسه تارگت مادر
    
    Returns:
        Tuple[bool, str, Optional[Dict]]: (مجاز است؟, پیام, ریزتارگت مشابه در صورت وجود)
    """
    all_targets = _load()
    if not isinstance(all_targets, list):
        return True, '', None
    
    # ۱. بررسی تکراری بودن دقیق (عامل + گروه کالا + دوره + تارگت مادر)
    for t in all_targets:
        if not isinstance(t, dict):
            continue
        if (t.get('agent_name') == agent_name and 
            t.get('product_group') == product_group and 
            t.get('period') == period and
            t.get('linked_target_id') == linked_target_id and
            t.get('status') in ['در انتظار', 'فعال']):
            
            return False, f'ریزتارگت "{product_group}" با دوره {period} برای {agent_name} قبلاً تعریف شده است.اطلاعات تکراری است.', t
    
    # ۲. بررسی تکراری بودن با دوره متفاوت (عامل + گروه کالا + تارگت مادر)
    for t in all_targets:
        if not isinstance(t, dict):
            continue
        if (t.get('agent_name') == agent_name and 
            t.get('product_group') == product_group and 
            t.get('period') != period and
            t.get('linked_target_id') == linked_target_id and
            t.get('status') in ['در انتظار', 'فعال']):
            
            existing_period = t.get('period', '')
            return False, f'برای {agent_name} ریزتارگت "{product_group}" با دوره {existing_period} ثبت شده است.آیا از ایجاد ریزتارگت با دوره {period} اطمینان دارید؟', t
    
    return True, '', None

def export_to_excel(targets: List[Dict], filename: str = None) -> Tuple[bool, str, str]:
    """خروجی اکسل ریزتارگت‌ها"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from utils.storage import get_backup_path
        
        if not targets:
            return False, 'هیچ ریزتارگتی وجود ندارد', ''
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ریزتارگت‌ها"
        
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = [
            'شناسه ریزتارگت',
            'نام عامل',
            'دوره تارگت',
            'تاریخ شروع',
            'تاریخ پایان',
            'تاریخ ایجاد',
            'نام گروه کالا',
            'تعداد تارگت',
            'واحد تارگت',
            'پیوند با تارگت مادر',
            'تارگت روزانه',
            'مقدار محقق شده',
            'وضعیت'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row, target in enumerate(targets, 2):
            if not isinstance(target, dict):
                continue
                
            created_at = target.get('created_at', '')
            created_date = ''
            if created_at:
                try:
                    dt = datetime.fromisoformat(created_at)
                    created_date = to_jalali(dt.year, dt.month, dt.day)
                except:
                    created_date = created_at[:10]
            
            values = [
                target.get('id', ''),
                target.get('agent_name', ''),
                target.get('period', ''),
                target.get('start_date', ''),
                target.get('end_date', ''),
                created_date,
                target.get('product_group', ''),
                target.get('target_count', 0),
                target.get('unit', ''),
                target.get('linked_target_id', ''),
                target.get('daily_target', 0),
                target.get('achieved_value', 0),
                target.get('status', '')
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        column_widths = [16, 18, 12, 14, 14, 14, 18, 14, 12, 20, 16, 16, 14]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        if not filename:
            today = get_today_jalali().replace('/', '-')
            timestamp = datetime.now().strftime('%H%M%S')
            filename = f'ریزتارگت_{today}_{timestamp}.xlsx'
        
        export_dir = get_backup_path()
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)
        
        wb.save(filepath)
        return True, f'فایل ذخیره شد:\n{filename}', filepath
        
    except ImportError:
        return False, 'ماژول openpyxl نصب نیست', ''
    except Exception as e:
        return False, f'خطا: {str(e)}', ''
"""
ساخت خروجی Excel از توزیع‌ها - نسخه مخصوص موزع
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.delivery_manager import get_all_deliveries
from utils.storage import get_backup_path


def safe_int(value, default=0):
    """تبدیل امن به عدد صحیح"""
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def safe_str(value, default=''):
    """تبدیل امن به رشته"""
    if value is None:
        return default
    return str(value)


def get_export_filename_distributor():
    """
    دریافت مسیر کامل فایل خروجی توزیع
    در اندروید به مسیر Download میرود
    """
    export_dir = get_backup_path()
    os.makedirs(export_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'گزارش_توزیع_{timestamp}.xlsx'
    return os.path.join(export_dir, filename)


def export_distributor_to_excel(data=None):
    """
    خروجی گرفتن از توزیع‌ها به فایل Excel
    
    Args:
        data: دیکشنری داده‌های فیلتر شده {date: [deliveries]}
              اگر None باشد، همه داده‌ها گرفته میشود
    
    Returns:
        (success, filepath) یا (success, error_message)
    """
    try:
        if data is None:
            all_deliveries = get_all_deliveries()
        else:
            all_deliveries = data
        
        if not all_deliveries:
            return False, "هیچ داده‌ای برای خروجی وجود ندارد"
        
        wb = Workbook()
        
        # ========== صفحه اول: گزارش توزیع‌ها ==========
        ws1 = wb.active
        ws1.title = "گزارش توزیع‌ها"
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            "ردیف", "تاریخ", "مسیر", "مشتری", "شماره فاکتور",
            "وضعیت توزیع", "تحویل کامل", "مبلغ فاکتور",
            "تعداد برگشتی", "مبلغ برگشتی", "علت برگشتی",
            "مبلغ نقدی", "مبلغ چکی", "مبلغ نسیه",
            "جمع دریافتی", "مانده", "درصد تخفیف",
            "مبلغ تخفیف", "سایر کسورات", "نوع تسویه",
            "توضیحات", "ساعت", "شناسه"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        total_amount = 0
        total_cash = 0
        total_check = 0
        total_credit = 0
        total_return_qty = 0
        total_return_amount = 0
        total_deliveries = 0
        total_full_deliveries = 0
        total_partial_deliveries = 0
        
        sorted_dates = sorted(all_deliveries.keys(), reverse=True)
        idx = 1
        
        for date in sorted_dates:
            delivery_list = all_deliveries[date]
            if not isinstance(delivery_list, list):
                continue
            
            for delivery in delivery_list:
                if not isinstance(delivery, dict):
                    continue
                
                delivery_status = delivery.get('delivery_status', '')
                full_delivery = delivery.get('full_delivery', True)
                invoice_amount = safe_int(delivery.get('invoice_amount', 0))
                cash_amount = safe_int(delivery.get('cash_amount', 0))
                check_amount = safe_int(delivery.get('check_amount', 0))
                credit_amount = safe_int(delivery.get('remaining_amount', 0))
                returned_qty = safe_int(delivery.get('returned_quantity', 0))
                returned_amount = safe_int(delivery.get('returned_amount', 0))
                discount_percent = safe_int(delivery.get('discount_percent', 0))
                discount_amount = safe_int(delivery.get('discount_amount', 0))
                other_deductions = safe_int(delivery.get('other_deductions_total', 0))
                total_received = safe_int(delivery.get('total_received', 0))
                
                timestamp = delivery.get('timestamp', '')
                time_part = ''
                if timestamp and ' ' in timestamp:
                    time_part = timestamp.split(' ')[1]
                
                ws1.cell(row=row, column=1, value=idx)
                ws1.cell(row=row, column=2, value=safe_str(date))
                ws1.cell(row=row, column=3, value=safe_str(delivery.get('route', '')))
                ws1.cell(row=row, column=4, value=safe_str(delivery.get('customer_name', '')))
                ws1.cell(row=row, column=5, value=safe_str(delivery.get('invoice_number', '')))
                ws1.cell(row=row, column=6, value=safe_str(delivery_status))
                ws1.cell(row=row, column=7, value="بله" if full_delivery else "خیر")
                ws1.cell(row=row, column=8, value=invoice_amount)
                ws1.cell(row=row, column=9, value=returned_qty)
                ws1.cell(row=row, column=10, value=returned_amount)
                ws1.cell(row=row, column=11, value=safe_str(delivery.get('return_reason', '')))
                ws1.cell(row=row, column=12, value=cash_amount)
                ws1.cell(row=row, column=13, value=check_amount)
                ws1.cell(row=row, column=14, value=credit_amount)
                ws1.cell(row=row, column=15, value=total_received)
                ws1.cell(row=row, column=16, value=credit_amount)
                ws1.cell(row=row, column=17, value=discount_percent)
                ws1.cell(row=row, column=18, value=discount_amount)
                ws1.cell(row=row, column=19, value=other_deductions)
                ws1.cell(row=row, column=20, value=safe_str(delivery.get('settlement_type', '')))
                ws1.cell(row=row, column=21, value=safe_str(delivery.get('description', '')))
                ws1.cell(row=row, column=22, value=time_part)
                ws1.cell(row=row, column=23, value=safe_str(delivery.get('id', '')))
                
                if delivery_status == 'موفق':
                    total_deliveries += 1
                    total_amount += invoice_amount
                    total_cash += cash_amount
                    total_check += check_amount
                    total_credit += credit_amount
                    total_return_qty += returned_qty
                    total_return_amount += returned_amount
                    
                    if full_delivery:
                        total_full_deliveries += 1
                    else:
                        total_partial_deliveries += 1
                
                row += 1
                idx += 1
        
        col_widths = [8, 14, 18, 22, 16, 16, 14, 16, 16, 16, 22, 16, 16, 16, 16, 16, 14, 16, 16, 16, 20, 14, 20]
        for col, width in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(col)].width = width
        
        # ========== صفحه دوم: خلاصه آمار ==========
        ws2 = wb.create_sheet("خلاصه آمار")
        
        summary_data = [
            ["شاخص", "مقدار"],
            ["کل مبلغ توزیع (ریال)", f"{total_amount:,}"],
            ["مبلغ نقدی (ریال)", f"{total_cash:,}"],
            ["مبلغ چکی (ریال)", f"{total_check:,}"],
            ["مبلغ نسیه (ریال)", f"{total_credit:,}"],
            ["تعداد کل توزیع‌ها", total_deliveries],
            ["تعداد تحویل کامل", total_full_deliveries],
            ["تعداد تحویل ناقص", total_partial_deliveries],
            ["تعداد برگشتی", total_return_qty],
            ["مبلغ برگشتی (ریال)", f"{total_return_amount:,}"],
            ["تعداد روزهای کاری", len(all_deliveries)],
            ["میانگین مبلغ هر توزیع", f"{total_amount // total_deliveries:,}" if total_deliveries > 0 else "۰"],
        ]
        
        for r, row_data in enumerate(summary_data, 1):
            for c, value in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=c, value=value)
                if r == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="28B463", end_color="28B463", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 25
        
        # ========== صفحه سوم: آمار روزانه ==========
        ws3 = wb.create_sheet("آمار روزانه")
        
        daily_headers = [
            "تاریخ", "تعداد توزیع", "تحویل کامل", "تحویل ناقص",
            "مبلغ کل", "مبلغ نقدی", "مبلغ چکی", "مبلغ نسیه",
            "تعداد برگشتی", "مبلغ برگشتی"
        ]
        
        for col, header in enumerate(daily_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for date in sorted_dates:
            delivery_list = all_deliveries[date]
            if not isinstance(delivery_list, list):
                continue
            
            daily_count = 0
            daily_full = 0
            daily_partial = 0
            daily_amount = 0
            daily_cash = 0
            daily_check = 0
            daily_credit = 0
            daily_return_qty = 0
            daily_return_amount = 0
            
            for delivery in delivery_list:
                if not isinstance(delivery, dict):
                    continue
                
                delivery_status = delivery.get('delivery_status', '')
                full_delivery = delivery.get('full_delivery', True)
                
                if delivery_status == 'موفق':
                    daily_count += 1
                    daily_amount += safe_int(delivery.get('invoice_amount', 0))
                    daily_cash += safe_int(delivery.get('cash_amount', 0))
                    daily_check += safe_int(delivery.get('check_amount', 0))
                    daily_credit += safe_int(delivery.get('remaining_amount', 0))
                    daily_return_qty += safe_int(delivery.get('returned_quantity', 0))
                    daily_return_amount += safe_int(delivery.get('returned_amount', 0))
                    
                    if full_delivery:
                        daily_full += 1
                    else:
                        daily_partial += 1
            
            ws3.cell(row=row, column=1, value=safe_str(date))
            ws3.cell(row=row, column=2, value=daily_count)
            ws3.cell(row=row, column=3, value=daily_full)
            ws3.cell(row=row, column=4, value=daily_partial)
            ws3.cell(row=row, column=5, value=daily_amount)
            ws3.cell(row=row, column=6, value=daily_cash)
            ws3.cell(row=row, column=7, value=daily_check)
            ws3.cell(row=row, column=8, value=daily_credit)
            ws3.cell(row=row, column=9, value=daily_return_qty)
            ws3.cell(row=row, column=10, value=daily_return_amount)
            row += 1
        
        for col in range(1, len(daily_headers) + 1):
            ws3.column_dimensions[get_column_letter(col)].width = 16
        
        # ========== ذخیره فایل ==========
        filepath = get_export_filename_distributor()
        wb.save(filepath)
        
        if os.path.exists(filepath):
            size = os.path.getsize(filepath)
            print(f"فایل اکسل توزیع با موفقیت ساخته شد: {filepath} ({size} bytes)")
            return True, filepath
        else:
            return False, "فایل ساخته نشد"
        
    except Exception as e:
        print(f"خطا در خروجی Excel توزیع: {e}")
        import traceback
        traceback.print_exc()
        return False, str(e)
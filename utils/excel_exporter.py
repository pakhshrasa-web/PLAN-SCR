"""
ساخت خروجی Excel از ویزیت‌ها - نسخه بهینه برای اندروید
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.file_manager import get_daily_logs
from utils.storage import get_backup_path


def safe_int(value, default=0):
    """تبدیل امن به عدد صحیح"""
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def safe_str(value, default=''):
    """تبدیل امن به رشته"""
    if value is None:
        return default
    return str(value)


def get_export_filename():
    """
    دریافت مسیر کامل فایل خروجی
    در اندروید به مسیر Download میرود
    """
    export_dir = get_backup_path()
    os.makedirs(export_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'گزارش_فروش_{timestamp}.xlsx'
    return os.path.join(export_dir, filename)


def export_to_excel(data=None):
    """
    خروجی گرفتن از ویزیت‌ها به فایل Excel
    
    Args:
        data: دیکشنری داده‌های فیلتر شده {date: [logs]}
              اگر None باشد، همه داده‌ها گرفته میشود
    
    Returns:
        (success, filepath) یا (success, error_message)
    """
    try:
        # ✅ اگر داده‌ای ارسال نشده، همه رو بگیر
        if data is None:
            all_logs = get_daily_logs()
        else:
            all_logs = data
        
        if not all_logs:
            return False, "هیچ داده‌ای برای خروجی وجود ندارد"
        
        wb = Workbook()
        
        # ========== صفحه اول: گزارش ویزیت‌ها ==========
        ws1 = wb.active
        ws1.title = "گزارش ویزیت‌ها"
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== هدرها با ستون‌های جدید برای دلایل ==========
        headers = ["ردیف", "تاریخ", "مسیر", "مشتری", "وضعیت ویزیت", 
                   "وضعیت فروش", "تعداد واحد", "مبلغ فروش", "نحوه تسویه",
                   "فروش نقدی", "فروش چکی", "مشتری جدید", "ساعت",
                   "علت ویزیت ناموفق", "علت فروش ناموفق"]  # ← ستون‌های جدید
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        total_sales = 0
        total_invoices = 0
        total_visits = 0
        total_units = 0
        total_successful_visits = 0
        total_failed_visits = 0
        total_successful_sales = 0
        total_failed_sales = 0
        total_cash = 0
        total_check = 0
        total_new_customers = 0
        
        sorted_dates = sorted(all_logs.keys(), reverse=True)
        idx = 1
        
        for date in sorted_dates:
            log_list = all_logs[date]
            if not isinstance(log_list, list):
                continue
            
            for log in log_list:
                if not isinstance(log, dict):
                    continue
                
                visit_status = log.get('visit_status', '')
                sales_status = log.get('sales_status', '')
                sales_amount = safe_int(log.get('sales_amount', 0))
                units_sold = safe_int(log.get('units_sold', 0))
                payment_method = log.get('payment_method', '')
                is_new_customer = log.get('is_new_customer', False)
                
                # ✅ دریافت دلایل
                fail_reason = log.get('fail_reason', '')
                fail_sales_reason = log.get('fail_sales_reason', '')
                
                cash_amount = 0
                check_amount = 0
                if sales_status == 'موفق' and sales_amount > 0:
                    if payment_method == 'نقد':
                        cash_amount = sales_amount
                        total_cash += sales_amount
                    elif payment_method == 'چک':
                        check_amount = sales_amount
                        total_check += sales_amount
                
                if is_new_customer:
                    total_new_customers += 1
                
                ws1.cell(row=row, column=1, value=idx)
                ws1.cell(row=row, column=2, value=safe_str(date))
                ws1.cell(row=row, column=3, value=safe_str(log.get('route', '')))
                ws1.cell(row=row, column=4, value=safe_str(log.get('customer', '')))
                ws1.cell(row=row, column=5, value=safe_str(visit_status))
                ws1.cell(row=row, column=6, value=safe_str(sales_status if sales_status else '---'))
                ws1.cell(row=row, column=7, value=units_sold if sales_status == 'موفق' else 0)
                ws1.cell(row=row, column=8, value=sales_amount if sales_status == 'موفق' else 0)
                ws1.cell(row=row, column=9, value=safe_str(payment_method if sales_status == 'موفق' else '---'))
                ws1.cell(row=row, column=10, value=cash_amount)
                ws1.cell(row=row, column=11, value=check_amount)
                ws1.cell(row=row, column=12, value="✅" if is_new_customer else "—")
                ws1.cell(row=row, column=13, value=safe_str(log.get('time', '')))
                # ✅ ستون‌های جدید برای دلایل
                ws1.cell(row=row, column=14, value=safe_str(fail_reason))
                ws1.cell(row=row, column=15, value=safe_str(fail_sales_reason))
                
                total_visits += 1
                if visit_status == 'موفق':
                    total_successful_visits += 1
                    if sales_status == 'موفق':
                        total_successful_sales += 1
                        total_invoices += 1
                        total_units += units_sold
                        total_sales += sales_amount
                    else:
                        total_failed_sales += 1
                else:
                    total_failed_visits += 1
                
                row += 1
                idx += 1
        
        # تنظیم عرض ستون‌ها با ستون‌های جدید
        col_widths = [8, 14, 18, 22, 14, 14, 14, 18, 14, 14, 14, 14, 14, 25, 25]
        for col, width in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(col)].width = width
        
        # ========== صفحه دوم: خلاصه آمار ==========
        ws2 = wb.create_sheet("خلاصه آمار")
        
        summary_data = [
            ["شاخص", "مقدار"],
            ["کل فروش (ریال)", f"{total_sales:,}"],
            ["فروش نقدی (ریال)", f"{total_cash:,}"],
            ["فروش چکی (ریال)", f"{total_check:,}"],
            ["تعداد کل فاکتورها", total_invoices],
            ["تعداد کل ویزیت‌ها", total_visits],
            ["تعداد ویزیت موفق", total_successful_visits],
            ["تعداد ویزیت ناموفق", total_failed_visits],
            ["تعداد فروش موفق", total_successful_sales],
            ["تعداد فروش ناموفق", total_failed_sales],
            ["تعداد کل واحد فروش", total_units],
            ["تعداد مشتری جدید", total_new_customers],
            ["تعداد روزهای کاری", len(all_logs)],
            ["میانگین مبلغ هر فاکتور", f"{total_sales // total_invoices:,}" if total_invoices > 0 else "۰"],
            ["میانگین فروش هر ویزیت موفق", f"{total_sales // total_successful_visits:,}" if total_successful_visits > 0 else "۰"],
        ]
        
        for r, row_data in enumerate(summary_data, 1):
            for c, value in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=c, value=value)
                if r == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="28B463", end_color="28B463", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 25
        
        # ========== صفحه سوم: آمار روزانه ==========
        ws3 = wb.create_sheet("آمار روزانه")
        
        daily_headers = ["تاریخ", "کل ویزیت", "ویزیت موفق", "ویزیت ناموفق", 
                        "فروش موفق", "فروش ناموفق", "تعداد واحد", "مبلغ فروش",
                        "فروش نقدی", "فروش چکی", "مشتری جدید"]
        
        for col, header in enumerate(daily_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for date in sorted_dates:
            log_list = all_logs[date]
            if not isinstance(log_list, list):
                continue
            
            daily_visits = 0
            daily_successful = 0
            daily_failed = 0
            daily_sales_success = 0
            daily_sales_failed = 0
            daily_units = 0
            daily_amount = 0
            daily_cash = 0
            daily_check = 0
            daily_new_customers = 0
            
            for log in log_list:
                if not isinstance(log, dict):
                    continue
                
                visit_status = log.get('visit_status', '')
                sales_status = log.get('sales_status', '')
                sales_amount = safe_int(log.get('sales_amount', 0))
                payment_method = log.get('payment_method', '')
                is_new_customer = log.get('is_new_customer', False)
                
                daily_visits += 1
                
                if is_new_customer:
                    daily_new_customers += 1
                
                if visit_status == 'موفق':
                    daily_successful += 1
                    if sales_status == 'موفق':
                        daily_sales_success += 1
                        daily_units += safe_int(log.get('units_sold', 0))
                        daily_amount += sales_amount
                        
                        if payment_method == 'نقد':
                            daily_cash += sales_amount
                        elif payment_method == 'چک':
                            daily_check += sales_amount
                    else:
                        daily_sales_failed += 1
                else:
                    daily_failed += 1
            
            ws3.cell(row=row, column=1, value=safe_str(date))
            ws3.cell(row=row, column=2, value=daily_visits)
            ws3.cell(row=row, column=3, value=daily_successful)
            ws3.cell(row=row, column=4, value=daily_failed)
            ws3.cell(row=row, column=5, value=daily_sales_success)
            ws3.cell(row=row, column=6, value=daily_sales_failed)
            ws3.cell(row=row, column=7, value=daily_units)
            ws3.cell(row=row, column=8, value=daily_amount)
            ws3.cell(row=row, column=9, value=daily_cash)
            ws3.cell(row=row, column=10, value=daily_check)
            ws3.cell(row=row, column=11, value=daily_new_customers)
            row += 1
        
        for col in range(1, len(daily_headers) + 1):
            ws3.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== ذخیره فایل ==========
        filepath = get_export_filename()
        wb.save(filepath)
        
        if os.path.exists(filepath):
            size = os.path.getsize(filepath)
            print(f"فایل اکسل با موفقیت ساخته شد: {filepath} ({size} bytes)")
            return True, filepath
        else:
            return False, "فایل ساخته نشد"
        
    except Exception as e:
        print(f"خطا در خروجی Excel: {e}")
        import traceback
        traceback.print_exc()
        return False, str(e)
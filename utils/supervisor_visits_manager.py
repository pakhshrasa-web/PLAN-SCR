# utils/supervisor_visits_manager.py
"""
مدیریت سرکشی‌های سوپروایزر (بررسی بازار)
"""

import os
import json
import random
import string
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from kivy.logger import Logger as logger
from utils.storage import get_data_path
from utils.jalali_date import get_today_jalali, get_current_time, validate_jalali_date

VISITS_FILE = 'supervisor_visits.json'


def _get_visits_path() -> str:
    """دریافت مسیر فایل سرکشی‌ها"""
    return os.path.join(get_data_path(), VISITS_FILE)


def _load_visits() -> List[Dict]:
    """بارگذاری همه سرکشی‌ها"""
    try:
        path = _get_visits_path()
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    if not data:
                        return []
                    return []
                if not isinstance(data, list):
                    return []
                return data
        return []
    except Exception as e:
        logger.error(f"خطا در بارگذاری سرکشی‌ها: {e}")
        return []


def _save_visits(visits: List[Dict]) -> bool:
    """ذخیره سرکشی‌ها"""
    try:
        path = _get_visits_path()
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(visits, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"خطا در ذخیره سرکشی‌ها: {e}")
        return False


def _generate_visit_id() -> str:
    """تولید آیدی یکتا برای سرکشی"""
    chars = string.ascii_uppercase + string.digits
    random_part = ''.join(random.choices(chars, k=4))
    return f"SV{random_part}"


def create_supervisor_visit(data: Dict) -> Tuple[bool, str, Optional[Dict]]:
    """
    ثبت سرکشی جدید با ذخیره نام کاربر
    
    Args:
        data: دیکشنری شامل تمام فیلدهای سرکشی
    
    Returns:
        Tuple[bool, str, Optional[Dict]]: (موفقیت, پیام, دیتای سرکشی)
    """
    try:
        # اعتبارسنجی فیلدهای اجباری
        required_fields = [
            'route', 'customer', 'visit_type', 'visit_reason',
            'customer_status', 'shelf_status', 'monthly_visits',
            'visit_sufficient', 'expected_purchase', 'inventory_status',
            'agent_behavior', 'distributor_behavior', 'customer_satisfaction',
            'target_achievement', 'need_followup'
        ]
        
        for field in required_fields:
            if not data.get(field):
                return False, f'فیلد "{field}" الزامی است', None
        
        # اعتبارسنجی تاریخ مراجعه بعدی
        need_followup = data.get('need_followup', '')
        next_visit_date = data.get('next_visit_date', '')
        
        if need_followup == 'بله':
            if not next_visit_date:
                return False, 'در صورت نیاز به پیگیری، تاریخ مراجعه بعدی الزامی است', None
            if not validate_jalali_date(next_visit_date):
                return False, 'فرمت تاریخ مراجعه بعدی نامعتبر است (مثال: 1405/01/31)', None
        else:
            data['next_visit_date'] = ''
        
        # تولید آیدی
        visit_id = _generate_visit_id()
        
        # ✅ دریافت نام کاربر از data (از supervisor_screen ارسال می‌شود)
        created_by = data.get('created_by', 'supervisor')
        agent_name = data.get('agent_name', created_by)
        
        # ایجاد سرکشی
        visit = {
            'id': visit_id,
            'date': get_today_jalali(),
            'time': get_current_time(),
            'route': data.get('route', ''),
            'customer': data.get('customer', ''),
            'visit_type': data.get('visit_type', ''),
            'visit_reason': data.get('visit_reason', ''),
            'supervisor_note': data.get('supervisor_note', ''),
            'customer_status': data.get('customer_status', ''),
            'shelf_status': data.get('shelf_status', ''),
            'monthly_visits': data.get('monthly_visits', ''),
            'visit_sufficient': data.get('visit_sufficient', ''),
            'expected_purchase': data.get('expected_purchase', ''),
            'inventory_status': data.get('inventory_status', ''),
            'agent_behavior': data.get('agent_behavior', ''),
            'distributor_behavior': data.get('distributor_behavior', ''),
            'customer_satisfaction': data.get('customer_satisfaction', ''),
            'customer_feedback': data.get('customer_feedback', ''),
            'target_achievement': data.get('target_achievement', ''),
            'supervisor_opinion': data.get('supervisor_opinion', ''),
            'need_followup': data.get('need_followup', ''),
            'next_visit_date': data.get('next_visit_date', ''),
            'created_by': created_by,  # ✅ ذخیره نام کاربر واقعی
            'agent_name': agent_name,  # ✅ برای هماهنگی
            'created_at': datetime.now().isoformat(),
            'reported_to_manager': False,
            'reported_date': ''
        }
        
        # بارگذاری سرکشی‌های موجود
        visits = _load_visits()
        visits.append(visit)
        
        # ذخیره
        if _save_visits(visits):
            logger.info(f"سرکشی جدید ثبت شد: {visit_id} توسط {created_by}")
            return True, f'سرکشی با شناسه {visit_id} ثبت شد', visit
        else:
            return False, 'خطا در ذخیره سرکشی', None
        
    except Exception as e:
        logger.error(f"خطا در ثبت سرکشی: {e}")
        return False, f'خطا: {str(e)}', None


def get_all_visits() -> List[Dict]:
    """دریافت همه سرکشی‌ها"""
    return _load_visits()


def get_visits_filtered(
    customer: str = None,
    start_date: str = None,
    end_date: str = None,
    created_by: str = None
) -> List[Dict]:
    """دریافت سرکشی‌ها با فیلتر"""
    visits = _load_visits()
    result = visits
    
    if customer:
        result = [v for v in result if v.get('customer') == customer]
    
    if start_date:
        result = [v for v in result if v.get('date') >= start_date]
    
    if end_date:
        result = [v for v in result if v.get('date') <= end_date]
    
    if created_by:
        result = [v for v in result if v.get('created_by') == created_by or v.get('agent_name') == created_by]
    
    # مرتب‌سازی بر اساس تاریخ (جدیدترین اول)
    result.sort(key=lambda x: x.get('date', ''), reverse=True)
    
    return result


def get_visits_by_customer(customer: str) -> List[Dict]:
    """دریافت سرکشی‌های یک مشتری"""
    visits = _load_visits()
    return [v for v in visits if v.get('customer') == customer]


def get_visits_by_route(route: str) -> List[Dict]:
    """دریافت سرکشی‌های یک مسیر"""
    visits = _load_visits()
    return [v for v in visits if v.get('route') == route]


def get_visits_by_creator(creator_name: str) -> List[Dict]:
    """دریافت سرکشی‌های یک سوپروایزر خاص"""
    visits = _load_visits()
    result = []
    for v in visits:
        v_creator = v.get('created_by', '') or v.get('agent_name', '')
        if creator_name in v_creator or v_creator in creator_name:
            result.append(v)
    return result


def get_visits_statistics() -> Dict:
    """دریافت آمار سرکشی‌ها"""
    try:
        visits = _load_visits()
        
        total = len(visits)
        routes = list(set(v.get('route', '') for v in visits))
        customers = list(set(v.get('customer', '') for v in visits))
        
        return {
            'total': total,
            'routes_count': len(routes),
            'customers_count': len(customers)
        }
        
    except Exception as e:
        logger.error(f"خطا در دریافت آمار سرکشی‌ها: {e}")
        return {
            'total': 0,
            'routes_count': 0,
            'customers_count': 0
        }


def export_visits_to_excel(visits: List[Dict], filename: str = None) -> Tuple[bool, str, str]:
    """
    خروجی گرفتن از سرکشی‌ها به صورت فایل Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from utils.storage import get_backup_path
        from utils.jalali_date import get_today_jalali
        
        if not visits:
            return False, 'هیچ سرکشی برای خروجی وجود ندارد', ''
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "بررسی بازار"
        ws.sheet_view.rightToLeft = True
        
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ✅ اضافه کردن ستون "ثبت شده توسط"
        headers = [
            'شناسه', 'تاریخ', 'ساعت', 'مسیر', 'مشتری',
            'نحوه سرکشی', 'علت سرکشی', 'وضعیت مشتری',
            'وضعیت حضور در شلف', 'تعداد سرکشی در ماه',
            'آیا سرکشی کافیست؟', 'خرید مورد انتظار',
            'وضعیت موجودی', 'برخورد بازاریاب',
            'برخورد موزع', 'رضایتمندی مشتری',
            'نظرات مشتری', 'تحقق هدف سرکشی',
            'نظریه سوپروایزر', 'نیاز به پیگیری',
            'تاریخ مراجعه بعدی', 'ثبت شده توسط'
        ]
        
        column_widths = [12, 12, 10, 18, 20, 12, 18, 14, 18, 16, 18, 20, 16, 16, 16, 16, 30, 16, 30, 14, 16, 18]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        for row, visit in enumerate(visits, 2):
            ws.cell(row=row, column=1, value=visit.get('id', ''))
            ws.cell(row=row, column=2, value=visit.get('date', ''))
            ws.cell(row=row, column=3, value=visit.get('time', ''))
            ws.cell(row=row, column=4, value=visit.get('route', ''))
            ws.cell(row=row, column=5, value=visit.get('customer', ''))
            ws.cell(row=row, column=6, value=visit.get('visit_type', ''))
            ws.cell(row=row, column=7, value=visit.get('visit_reason', ''))
            ws.cell(row=row, column=8, value=visit.get('customer_status', ''))
            ws.cell(row=row, column=9, value=visit.get('shelf_status', ''))
            ws.cell(row=row, column=10, value=visit.get('monthly_visits', ''))
            ws.cell(row=row, column=11, value=visit.get('visit_sufficient', ''))
            ws.cell(row=row, column=12, value=visit.get('expected_purchase', ''))
            ws.cell(row=row, column=13, value=visit.get('inventory_status', ''))
            ws.cell(row=row, column=14, value=visit.get('agent_behavior', ''))
            ws.cell(row=row, column=15, value=visit.get('distributor_behavior', ''))
            ws.cell(row=row, column=16, value=visit.get('customer_satisfaction', ''))
            ws.cell(row=row, column=17, value=visit.get('customer_feedback', ''))
            ws.cell(row=row, column=18, value=visit.get('target_achievement', ''))
            ws.cell(row=row, column=19, value=visit.get('supervisor_opinion', ''))
            ws.cell(row=row, column=20, value=visit.get('need_followup', ''))
            ws.cell(row=row, column=21, value=visit.get('next_visit_date', ''))
            ws.cell(row=row, column=22, value=visit.get('created_by', 'supervisor'))
        
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        for row in ws.iter_rows(min_row=2, max_row=len(visits) + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        if not filename:
            today = get_today_jalali().replace('/', '-')
            timestamp = datetime.now().strftime('%H%M%S')
            filename = f'بررسی_بازار_{today}_{timestamp}.xlsx'
        
        export_dir = get_backup_path()
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)
        
        wb.save(filepath)
        
        if os.path.exists(filepath):
            size = os.path.getsize(filepath)
            print(f"فایل اکسل بررسی بازار ذخیره شد: {filepath} ({size} bytes)")
            return True, f'فایل با موفقیت ذخیره شد:\n{filename}', filepath
        else:
            return False, 'فایل ساخته نشد', ''
        
    except ImportError:
        return False, 'ماژول openpyxl نصب نیست.\nلطفاً با دستور زیر نصب کنید:\npip install openpyxl', ''
    except Exception as e:
        print(f"خطا در خروجی اکسل بررسی بازار: {e}")
        import traceback
        traceback.print_exc()
        return False, f'خطا در ایجاد فایل اکسل:\n{str(e)}', ''
    

def mark_visit_as_reported(visit_id: str, reported_by: str = None) -> Tuple[bool, str]:
    """علامت‌گذاری یک سرکشی به عنوان گزارش شده به مدیر
    
    Args:
        visit_id: شناسه سرکشی
        reported_by: نام کاربری که گزارش را ارسال کرده (اختیاری)
    
    Returns:
        (success, message)
    """
    try:
        import json
        from utils.storage import get_data_path
        from utils.jalali_date import get_today_jalali
        
        filepath = os.path.join(get_data_path(), 'supervisor_visits.json')
        
        if not os.path.exists(filepath):
            return False, "فایل سرکشی‌ها یافت نشد"
        
        with open(filepath, 'r', encoding='utf-8') as f:
            visits = json.load(f)
        
        if not isinstance(visits, list):
            return False, "فرمت داده‌ها نامعتبر است"
        
        found = False
        for visit in visits:
            if visit.get('id') == visit_id:
                visit['reported_to_manager'] = True
                visit['reported_date'] = get_today_jalali()
                if reported_by:
                    visit['reported_by'] = reported_by
                found = True
                break
        
        if not found:
            return False, "سرکشی مورد نظر یافت نشد"
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(visits, f, ensure_ascii=False, indent=2)
        
        if reported_by:
            return True, f"گزارش با موفقیت به مدیر ارسال شد (توسط: {reported_by})"
        else:
            return True, "گزارش با موفقیت به مدیر ارسال شد"
    
    except Exception as e:
        return False, f"خطا در علامت‌گذاری: {str(e)}"
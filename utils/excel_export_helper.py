# utils/excel_export_helper.py
# ========== توابع کمکی برای خروجی‌های اکسل و تصویر ==========

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import arabic_reshaper
from bidi.algorithm import get_display

from utils.jalali_date import get_today_jalali
from utils.storage import get_backup_path


class ExcelExportHelper:
    """کلاس کمکی برای خروجی‌های اکسل و تصویر"""
    
    @staticmethod
    def export_checks_excel(filtered_data, show_message_func):
        """خروجی اکسل گزارش چک با جزئیات کامل"""
        try:
            # جمع‌آوری تمام چک‌ها از داده‌های فیلتر شده
            all_checks = []
            for collection in filtered_data:
                if collection.get('has_check') and collection.get('checks'):
                    for check in collection.get('checks', []):
                        all_checks.append({
                            'collection_id': collection.get('id', ''),
                            'collection_date': collection.get('date', ''),
                            'agent_name': collection.get('agent_name', ''),
                            'customer': collection.get('customer', ''),
                            'route': collection.get('route', ''),
                            'check_number': check.get('check_number', ''),
                            'due_date': check.get('due_date', ''),
                            'amount': check.get('amount', 0),
                            'sayadi_id': check.get('sayadi_id', ''),
                            'sayadi_status': check.get('sayadi_status', 'ثبت نشده'),
                            'description': collection.get('description', '')
                        })
            
            if not all_checks:
                show_message_func('خطا', 'هیچ چکی برای خروجی وجود ندارد')
                return False
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش چک"
            ws.right_to_left = True
            
            # استایل‌ها
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4A2C6A", end_color="4A2C6A", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # رنگ‌بندی وضعیت ثبت در سامانه صیادی
            sayadi_colors = {
                'ثبت شده': PatternFill(start_color="235347", end_color="235347", fill_type="solid"),
                'ثبت نشده': PatternFill(start_color="78281F", end_color="78281F", fill_type="solid"),
                'در انتظار ثبت': PatternFill(start_color="7D6B2C", end_color="7D6B2C", fill_type="solid")
            }
            default_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            
            # هدرها
            headers = [
                'ردیف', 'شناسه وصول', 'تاریخ وصول', 'عامل', 'مشتری', 'مسیر',
                'شماره چک', 'تاریخ سررسید', 'مبلغ چک (ریال)', 'شناسه صیادی', 'وضعیت صیادی', 'توضیحات'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # پر کردن داده‌ها
            for row_idx, check in enumerate(all_checks, 2):
                sayadi_status = check.get('sayadi_status', 'ثبت نشده')
                row_fill = sayadi_colors.get(sayadi_status, default_fill)
                
                values = [
                    row_idx - 1,
                    check.get('collection_id', ''),
                    check.get('collection_date', ''),
                    check.get('agent_name', ''),
                    check.get('customer', ''),
                    check.get('route', ''),
                    check.get('check_number', ''),
                    check.get('due_date', ''),
                    check.get('amount', 0),
                    check.get('sayadi_id', ''),
                    sayadi_status,
                    check.get('description', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.font = Font(color="FFFFFF")
            
            # تنظیم عرض ستون‌ها
            column_widths = [6, 14, 12, 16, 22, 14, 14, 14, 16, 18, 14, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # اضافه کردن خلاصه در پایین
            summary_row = len(all_checks) + 3
            
            # عنوان خلاصه
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
            summary_cell = ws.cell(row=summary_row, column=1, value='خلاصه گزارش چک')
            summary_cell.font = Font(bold=True, size=14, color="FFD700")
            summary_cell.alignment = Alignment(horizontal="center")
            
            # جمع کل مبلغ چک‌ها
            total_amount = sum(c['amount'] for c in all_checks)
            ws.cell(row=summary_row, column=9, value=f'جمع کل مبلغ: {total_amount:,.0f}')
            ws.cell(row=summary_row, column=9).font = Font(bold=True, size=12, color="FFD700")
            ws.cell(row=summary_row, column=9).alignment = Alignment(horizontal="center")
            
            # تعداد چک‌ها بر اساس وضعیت صیادی
            sayadi_counts = {}
            for check in all_checks:
                status = check.get('sayadi_status', 'ثبت نشده')
                sayadi_counts[status] = sayadi_counts.get(status, 0) + 1
            
            row_offset = summary_row + 2
            ws.cell(row=row_offset, column=1, value='توزیع وضعیت ثبت صیادی:')
            ws.cell(row=row_offset, column=1).font = Font(bold=True, size=11, color="FFFFFF")
            row_offset += 1
            
            for status, count in sayadi_counts.items():
                ws.cell(row=row_offset, column=1, value=f'  {status}:')
                ws.cell(row=row_offset, column=1).font = Font(size=11, color="FFFFFF")
                ws.cell(row=row_offset, column=2, value=count)
                ws.cell(row=row_offset, column=2).font = Font(bold=True, size=11, color="FFD700")
                row_offset += 1
            
            # تعداد کل چک‌ها
            ws.cell(row=row_offset, column=1, value=f'تعداد کل چک‌ها: {len(all_checks)}')
            ws.cell(row=row_offset, column=1).font = Font(bold=True, size=11, color="FFFFFF")
            ws.cell(row=row_offset, column=2, value=len(all_checks))
            ws.cell(row=row_offset, column=2).font = Font(bold=True, size=11, color="FFD700")
            
            today = get_today_jalali().replace('/', '-')
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            excel_filename = f'گزارش_چک_{today}_{datetime.now().strftime("%H%M%S")}.xlsx'
            excel_path = os.path.join(export_dir, excel_filename)
            wb.save(excel_path)
            
            show_message_func('موفق', f'فایل اکسل چک ذخیره شد:\n{excel_filename}\nتعداد چک‌ها: {len(all_checks)}')
            return True
            
        except ImportError:
            show_message_func('خطا', 'ماژول openpyxl نصب نیست')
            return False
        except Exception as e:
            show_message_func('خطا', f'خطا در خروجی اکسل چک: {str(e)}')
            return False
    
    @staticmethod
    def export_checks_image(filtered_data, show_message_func):
        """خروجی تصویری گزارش چک"""
        try:
            # جمع‌آوری تمام چک‌ها
            all_checks = []
            for collection in filtered_data:
                if collection.get('has_check') and collection.get('checks'):
                    for check in collection.get('checks', []):
                        all_checks.append({
                            'collection_date': collection.get('date', ''),
                            'agent_name': collection.get('agent_name', ''),
                            'customer': collection.get('customer', ''),
                            'route': collection.get('route', ''),
                            'check_number': check.get('check_number', ''),
                            'due_date': check.get('due_date', ''),
                            'amount': check.get('amount', 0),
                            'sayadi_id': check.get('sayadi_id', ''),
                            'sayadi_status': check.get('sayadi_status', 'ثبت نشده'),
                        })
            
            if not all_checks:
                show_message_func('خطا', 'هیچ چکی برای خروجی وجود ندارد')
                return False
            
            font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Vazir.ttf')
            if not os.path.exists(font_path):
                font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Amiri-Regular.ttf')
            
            try:
                font_title = ImageFont.truetype(font_path, 20)
                font_header = ImageFont.truetype(font_path, 12)
                font_row = ImageFont.truetype(font_path, 10)
            except:
                font_title = ImageFont.load_default()
                font_header = ImageFont.load_default()
                font_row = ImageFont.load_default()
            
            def fix_text(text):
                if not text:
                    return ''
                try:
                    reshaped = arabic_reshaper.reshape(str(text))
                    return get_display(reshaped)
                except:
                    return str(text)
            
            headers = ['ردیف', 'تاریخ وصول', 'عامل', 'مشتری', 'مسیر', 'شماره چک', 'سررسید', 'مبلغ', 'وضعیت صیادی']
            col_widths = [30, 65, 80, 100, 70, 70, 65, 70, 70]
            
            row_height = 28
            header_height = 35
            title_height = 45
            padding = 10
            
            table_width = sum(col_widths) + (len(col_widths) + 1) * 2
            table_height = title_height + header_height + len(all_checks) * row_height + padding * 3
            
            img = Image.new('RGB', (table_width + 20, table_height + 20), color=(18, 18, 24))
            draw = ImageDraw.Draw(img)
            
            title_text = fix_text(f'گزارش چک‌ها - {get_today_jalali()}')
            draw.text((table_width // 2, padding), title_text, fill=(255, 215, 0), font=font_title, anchor='ma')
            
            y = title_height + padding
            x_start = 10
            
            for col_idx, (header, width) in enumerate(zip(headers, col_widths)):
                x = x_start + sum(col_widths[:col_idx]) + col_idx * 2
                draw.rectangle([x, y, x + width, y + header_height], fill=(74, 44, 106))
                draw.rectangle([x, y, x + width, y + header_height], outline=(50, 50, 60))
                draw.text((x + width // 2, y + header_height // 2), fix_text(header), fill=(255, 255, 255), font=font_header, anchor='mm')
            
            for row_idx, check in enumerate(all_checks):
                y = title_height + header_height + padding * 2 + row_idx * row_height
                
                sayadi_status = check.get('sayadi_status', 'ثبت نشده')
                if sayadi_status == 'ثبت شده':
                    bg_color = (25, 45, 35)
                elif sayadi_status == 'ثبت نشده':
                    bg_color = (60, 25, 20)
                else:
                    bg_color = (60, 50, 25)
                
                row_values = [
                    str(row_idx + 1),
                    check.get('collection_date', ''),
                    check.get('agent_name', ''),
                    check.get('customer', ''),
                    check.get('route', ''),
                    check.get('check_number', ''),
                    check.get('due_date', ''),
                    f"{check.get('amount', 0):,.0f}",
                    sayadi_status,
                ]
                
                colors = [
                    (200, 200, 200), (255, 255, 255), (200, 200, 200), (255, 255, 255),
                    (200, 200, 200), (255, 255, 255), (255, 255, 255), (255, 215, 0),
                    (100, 255, 100) if sayadi_status == 'ثبت شده' else (255, 100, 100) if sayadi_status == 'ثبت نشده' else (255, 200, 100)
                ]
                
                for col_idx, (value, width) in enumerate(zip(row_values, col_widths)):
                    x = x_start + sum(col_widths[:col_idx]) + col_idx * 2
                    draw.rectangle([x, y, x + width, y + row_height], fill=bg_color)
                    draw.rectangle([x, y, x + width, y + row_height], outline=(40, 40, 50))
                    draw.text((x + width // 2, y + row_height // 2), fix_text(value), fill=colors[col_idx], font=font_row, anchor='mm')
            
            today = get_today_jalali().replace('/', '-')
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            image_filename = f'گزارش_چک_{today}_{datetime.now().strftime("%H%M%S")}.png'
            image_path = os.path.join(export_dir, image_filename)
            img.save(image_path, quality=95)
            
            show_message_func('موفق', f'تصویر گزارش چک ذخیره شد:\n{image_filename}\nتعداد چک‌ها: {len(all_checks)}')
            return True
            
        except ImportError:
            show_message_func('خطا', 'ماژول Pillow یا bidi نصب نیست')
            return False
        except Exception as e:
            show_message_func('خطا', f'خطا در خروجی تصویری چک: {str(e)}')
            return False
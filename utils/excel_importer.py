"""
وارد کردن اطلاعات از فایل Excel - نسخه بهینه‌سازی شده
"""

import os
import openpyxl
from kivy.logger import Logger as logger
from utils.file_manager import add_route, add_customer, get_routes, get_customers


class ExcelImportError(Exception):
    """خطاهای مربوط به وارد کردن اکسل"""
    pass


def _get_sheet_names(wb):
    """دریافت نام تمام شیت‌ها"""
    return wb.sheetnames


def _validate_headers(ws, expected_headers):
    """بررسی وجود ستون‌های مورد نیاز"""
    if ws.max_row < 1:
        raise ExcelImportError("فایل خالی است")
    
    # خواندن سطر اول (عنوان ستون‌ها)
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(str(cell_value).strip().lower() if cell_value else '')
    
    # بررسی وجود همه ستون‌های مورد نیاز
    missing = [h for h in expected_headers if h not in headers]
    if missing:
        raise ExcelImportError(f"ستون‌های زیر در فایل وجود ندارند: {', '.join(missing)}")
    
    return headers


def _safe_int(value, default=0):
    """تبدیل امن به عدد صحیح"""
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def load_excel_file(filepath):
    """
    بارگذاری فایل اکسل با مدیریت خطا
    
    Returns:
        (workbook, worksheet, sheet_name)
    
    Raises:
        ExcelImportError: در صورت بروز خطا
    """
    if not filepath:
        raise ExcelImportError("مسیر فایل نامعتبر است")
    
    if not os.path.exists(filepath):
        raise ExcelImportError(f"فایل وجود ندارد: {filepath}")
    
    # بررسی حجم فایل
    file_size = os.path.getsize(filepath)
    logger.info(f"📏 حجم فایل: {file_size} bytes")
    
    if file_size == 0:
        raise ExcelImportError("فایل خالی است")
    
    # بررسی پسوند فایل
    if not filepath.lower().endswith(('.xlsx', '.xls')):
        raise ExcelImportError("فایل باید با فرمت اکسل (.xlsx یا .xls) باشد")
    
    try:
        logger.info("📂 در حال بارگذاری فایل اکسل...")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        if not wb.sheetnames:
            raise ExcelImportError("فایل هیچ شیتی ندارد")
        
        ws = wb.active
        sheet_name = ws.title
        
        logger.info(f"✅ فایل بارگذاری شد: {sheet_name} - {ws.max_row} سطر")
        return wb, ws, sheet_name
        
    except openpyxl.utils.exceptions.InvalidFileException:
        raise ExcelImportError("فایل معتبر اکسل نیست")
    except Exception as e:
        raise ExcelImportError(f"خطا در خواندن فایل: {str(e)}")


def import_routes_from_excel(filepath):
    """
    وارد کردن مسیرها از فایل Excel
    فرمت فایل: ستون اول = name
    """
    logger.info(f"🔍 import_routes_from_excel: filepath={filepath}")
    
    wb = None
    try:
        # ✅ بارگذاری فایل
        wb, ws, sheet_name = load_excel_file(filepath)
        
        # ✅ اعتبارسنجی ستون‌ها
        expected_headers = ['name']
        headers = _validate_headers(ws, expected_headers)
        
        # ✅ دریافت مسیرهای موجود (با Set برای سرعت بالا)
        existing_routes = {r.get('name', '') for r in get_routes()}
        logger.info(f"📋 {len(existing_routes)} مسیر موجود در سیستم")
        
        imported_count = 0
        duplicate_count = 0
        error_count = 0
        
        # ✅ خواندن سطرها
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    error_count += 1
                    continue
                    
                name = str(row[0]).strip()
                if not name:
                    error_count += 1
                    continue
                
                logger.debug(f"📝 سطر {idx}: نام مسیر = '{name}'")
                
                if name not in existing_routes:
                    add_route({'name': name})
                    imported_count += 1
                    existing_routes.add(name)
                    logger.info(f"✅ مسیر '{name}' اضافه شد")
                else:
                    duplicate_count += 1
                    logger.debug(f"⚠️ مسیر '{name}' تکراری است")
                    
            except Exception as e:
                error_count += 1
                logger.warning(f"❌ خطا در سطر {idx}: {e}")
        
        logger.info(f"📊 نتیجه: {imported_count} جدید، {duplicate_count} تکراری، {error_count} خطا")
        
        msg = f"{imported_count} مسیر جدید وارد شد."
        if duplicate_count > 0:
            msg += f" {duplicate_count} مسیر تکراری نادیده گرفته شد."
        if error_count > 0:
            msg += f" {error_count} سطر خطا داشت."
        
        return True, msg
        
    except ExcelImportError as e:
        logger.error(f"❌ خطا: {e}")
        return False, str(e)
    except Exception as e:
        logger.error(f"❌ خطای غیرمنتظره: {e}")
        import traceback
        traceback.print_exc()
        return False, f"خطا: {str(e)}"
    finally:
        if wb:
            try:
                wb.close()
                logger.debug("📂 فایل اکسل بسته شد")
            except:
                pass


def import_customers_from_excel(filepath):
    """
    وارد کردن مشتریان از فایل Excel
    فرمت فایل: name, store_name, route_name, mobile, address
    """
    logger.info(f"🔍 import_customers_from_excel: filepath={filepath}")
    
    wb = None
    try:
        # ✅ بارگذاری فایل
        wb, ws, sheet_name = load_excel_file(filepath)
        
        # ✅ اعتبارسنجی ستون‌ها
        expected_headers = ['name', 'store_name', 'route_name', 'mobile', 'address']
        headers = _validate_headers(ws, expected_headers)
        
        # ✅ دریافت مسیرهای موجود (با Set برای سرعت بالا)
        routes = get_routes()
        route_names = {r.get('name', '') for r in routes}
        logger.info(f"📋 {len(route_names)} مسیر موجود در سیستم")
        
        # ✅ دریافت مشتریان موجود (با Set برای سرعت بالا)
        existing_customers = {c.get('name', '') for c in get_customers()}
        logger.info(f"📋 {len(existing_customers)} مشتری موجود در سیستم")
        
        imported_count = 0
        duplicate_count = 0
        error_count = 0
        new_routes_added = 0
        
        # ✅ خواندن سطرها
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row:
                    error_count += 1
                    continue
                
                # خواندن مقادیر با مدیریت None
                name = str(row[0]).strip() if row[0] else ''
                store_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                route_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                mobile = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                address = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                
                if not name:
                    error_count += 1
                    logger.warning(f"⚠️ سطر {idx}: نام مشتری خالی است")
                    continue
                
                logger.debug(f"📝 سطر {idx}: نام='{name}', مسیر='{route_name}', موبایل='{mobile}'")
                
                # ✅ بررسی تکراری بودن با Set
                if name in existing_customers:
                    duplicate_count += 1
                    logger.debug(f"⚠️ مشتری '{name}' تکراری است")
                    continue
                
                # ✅ بررسی وجود مسیر (با Set)
                if route_name and route_name not in route_names:
                    route_names.add(route_name)
                    add_route({'name': route_name})
                    new_routes_added += 1
                    logger.info(f"✅ مسیر جدید '{route_name}' اضافه شد")
                
                # ✅ افزودن مشتری
                customer = {
                    'name': name,
                    'store_name': store_name,
                    'route_name': route_name,
                    'mobile': mobile,
                    'address': address
                }
                add_customer(customer)
                imported_count += 1
                existing_customers.add(name)
                logger.info(f"✅ مشتری '{name}' اضافه شد")
                    
            except Exception as e:
                error_count += 1
                logger.warning(f"❌ خطا در سطر {idx}: {e}")
        
        logger.info(f"📊 نتیجه: {imported_count} جدید، {duplicate_count} تکراری، {error_count} خطا، {new_routes_added} مسیر جدید")
        
        msg = f"{imported_count} مشتری جدید وارد شد."
        if duplicate_count > 0:
            msg += f" {duplicate_count} مشتری تکراری نادیده گرفته شد."
        if error_count > 0:
            msg += f" {error_count} سطر خطا داشت."
        if new_routes_added > 0:
            msg += f" {new_routes_added} مسیر جدید ایجاد شد."
        
        return True, msg
        
    except ExcelImportError as e:
        logger.error(f"❌ خطا: {e}")
        return False, str(e)
    except Exception as e:
        logger.error(f"❌ خطای غیرمنتظره: {e}")
        import traceback
        traceback.print_exc()
        return False, f"خطا: {str(e)}"
    finally:
        if wb:
            try:
                wb.close()
                logger.debug("📂 فایل اکسل بسته شد")
            except:
                pass


def get_excel_template_routes():
    """ایجاد فایل نمونه برای مسیرها"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Routes"
    ws['A1'] = "name"
    ws['A2'] = "نمونه: مسیر شمال"
    ws['A3'] = "نمونه: مسیر جنوب"
    ws['A4'] = "نمونه: مسیر شرق"
    return wb


def get_excel_template_customers():
    """ایجاد فایل نمونه برای مشتریان"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws['A1'] = "name"
    ws['B1'] = "store_name"
    ws['C1'] = "route_name"
    ws['D1'] = "mobile"
    ws['E1'] = "address"
    
    ws['A2'] = "علی محمدی"
    ws['B2'] = "فروشگاه رضا"
    ws['C2'] = "مسیر شمال"
    ws['D2'] = "09121234567"
    ws['E2'] = "خیابان ولیعصر"
    
    ws['A3'] = "محمد کریمی"
    ws['B3'] = "فروشگاه سینا"
    ws['C3'] = "مسیر جنوب"
    ws['D3'] = "09129876543"
    ws['E3'] = "خیابان انقلاب"
    
    return wb


def save_excel_template(data_type='customers'):
    """
    ذخیره فایل نمونه در پوشه import
    data_type: 'customers' یا 'routes'
    """
    try:
        from utils.storage import get_import_path
        
        import_path = get_import_path()
        
        if data_type == 'routes':
            wb = get_excel_template_routes()
            filename = 'template_routes.xlsx'
        else:
            wb = get_excel_template_customers()
            filename = 'template_customers.xlsx'
        
        filepath = os.path.join(import_path, filename)
        wb.save(filepath)
        
        logger.info(f"✅ فایل نمونه ذخیره شد: {filepath}")
        return True, filepath
        
    except Exception as e:
        logger.error(f"❌ خطا در ذخیره فایل نمونه: {e}")
        return False, str(e)


def import_from_excel(filepath, data_type='customers'):
    """
    تابع عمومی برای وارد کردن از اکسل
    data_type: 'customers' یا 'routes'
    """
    logger.info(f"🔍 import_from_excel: filepath={filepath}, data_type={data_type}")
    
    if data_type == 'routes':
        return import_routes_from_excel(filepath)
    elif data_type == 'customers':
        return import_customers_from_excel(filepath)
    else:
        return False, "نوع داده نامعتبر است"

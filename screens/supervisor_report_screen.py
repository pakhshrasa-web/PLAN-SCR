# screens/supervisor_report_screen.py
# ========== صفحه گزارش سوپروایزر ==========

import traceback
import os
from datetime import datetime
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.graphics import Color, Rectangle, RoundedRectangle
from kivy.clock import Clock
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox

from utils.rtl_widgets import (
    PersianButton, PersianComboBox, RTLLabel,
    PersianPopup, RTLTextInput
)
from utils.file_manager import get_agents, get_settings, get_product_groups, get_customers, get_routes
from utils.jalali_date import get_today_jalali
from utils.target_manager import get_all_targets, get_targets_filtered, export_targets_to_excel
from utils.detailed_target_manager import (
    get_all_detailed_targets,
    get_detailed_targets_filtered,
    export_to_excel as export_detailed_to_excel
)
from constants import TARGET_TYPES, TARGET_STATUSES, PERIOD_DISPLAY
from error_handler import ErrorPopup


# ============================================================
# کلاس کمکی Progress Bar
# ============================================================
class DTProgressBar(BoxLayout): 
    """Progress bar فارسی - بدون کتابخانه اضافی"""

    def __init__(self, percent=0, height=dp(20), **kwargs):
        super().__init__(**kwargs)
        self.size_hint_y = None
        self.height = height
        self.percent = max(0, min(100, percent))

        # رنگ بر اساس درصد
        if self.percent >= 100:
            self.fill_color = (0.2, 0.8, 0.2, 1)
        elif self.percent >= 50:
            self.fill_color = (1, 0.7, 0, 1)
        else:
            self.fill_color = (0.8, 0.3, 0.3, 1)

        self.bind(pos=self._update, size=self._update)

        with self.canvas:
            # پس‌زمینه
            Color(0.1, 0.1, 0.1, 1)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(4)])

            # نوار پیشرفت
            Color(*self.fill_color)
            self.fill_rect = RoundedRectangle(
                pos=(self.pos[0], self.pos[1]),
                size=(0, self.size[1]),
                radius=[dp(4)]
            )

            # متن درصد
            Color(1, 1, 1, 1)
            self.percent_label = Label(
                text=f'{self.percent:.0f}%',
                font_size=sp(11),
                bold=True,
                halign='center',
                valign='middle',
                color=(1, 1, 1, 1)
            )
            self.percent_label.bind(texture_size=self._update_label_size)

        self._update()

    def _update(self, *args):
        self.bg_rect.pos = self.pos
        self.bg_rect.size = self.size

        fill_width = self.width * (self.percent / 100)
        self.fill_rect.pos = self.pos
        self.fill_rect.size = (fill_width, self.height)

        self.percent_label.pos = (self.center_x - dp(20), self.center_y - dp(8))

    def _update_label_size(self, instance, value):
        pass

    def set_percent(self, percent):
        self.percent = max(0, min(100, percent))
        if self.percent >= 100:
            self.fill_color = (0.2, 0.8, 0.2, 1)
        elif self.percent >= 50:
            self.fill_color = (1, 0.7, 0, 1)
        else:
            self.fill_color = (0.8, 0.3, 0.3, 1)

        # بروزرسانی رنگ fill_rect
        self.canvas.remove(self.fill_rect)
        with self.canvas:
            Color(*self.fill_color)
            self.fill_rect = RoundedRectangle(
                pos=(self.pos[0], self.pos[1]),
                size=(self.width * (self.percent / 100), self.height),
                radius=[dp(4)]
            )

        self.percent_label.text = f'{self.percent:.0f}%'
        self._update()


# ============================================================
# صفحه اصلی گزارش سوپروایزر
# ============================================================
class SupervisorReportScreen(Screen):
    """صفحه گزارش‌های سوپروایزر"""

    def __init__(self, **kwargs):
        try:
            super().__init__(**kwargs)
            with self.canvas.before:
                Color(0.08, 0.08, 0.08, 1)
                self.bg_rect = Rectangle(pos=self.pos, size=self.size)
                self.bind(pos=self._update_bg, size=self._update_bg)

            self.current_tab = 0
            self.tab_buttons = []

            # فیلترهای تب ۱
            self.filter_agent = None
            self.filter_period = None
            self.filter_type = None
            self.filter_status = None
            self.current_filtered_targets = []

            # فیلترهای تب ۲
            self.dt_filter_agent = None
            self.dt_filter_product = None
            self.dt_filter_status = None
            self.dt_filter_start = None
            self.dt_filter_end = None
            self.current_filtered_dt = []

            self.build_ui()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ساخت SupervisorReportScreen: {e}", error_details)
            raise

    def _update_bg(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size

    def build_ui(self):
        try:
            layout = BoxLayout(orientation='vertical', padding=[dp(5), dp(5), dp(5), dp(5)])

            # ========== تب‌ها ==========
            tabs_layout = BoxLayout(
                size_hint_y=None,
                height=dp(40),
                spacing=dp(2)
            )

            tab_names = [
                ('گزارش تارگت‌ها', 0),
                ('گزارش ریزتارگت‌ها', 1),
                ('گزارش بازاری', 2),
                ('آمار و ارزیابی', 3)
            ]

            for name, tab_id in tab_names:
                btn = PersianButton(
                    text=name,
                    background_color=(0.3, 0.5, 0.8, 0.6),
                    size_hint_y=None,
                    height=dp(36),
                    color=(1, 1, 1, 1),
                    font_size=sp(13)
                )
                btn.bind(on_press=lambda x, tid=tab_id: self.switch_tab(tid))
                tabs_layout.add_widget(btn)
                self.tab_buttons.append(btn)

            layout.add_widget(tabs_layout)

            self.content_area = BoxLayout(orientation='vertical')
            layout.add_widget(self.content_area)

            # ========== دکمه بازگشت ==========
            back_btn = PersianButton(
                text='بازگشت',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(36),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            back_btn.bind(on_press=self.go_back)
            layout.add_widget(back_btn)

            self.add_widget(layout)
            Clock.schedule_once(lambda dt: self.switch_tab(0), 0.1)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ساخت UI: {e}", error_details)
            raise

    def switch_tab(self, tab_id):
        try:
            self.current_tab = tab_id

            for i, btn in enumerate(self.tab_buttons):
                btn.background_color = (0.3, 0.5, 0.8, 1) if i == tab_id else (0.3, 0.5, 0.8, 0.6)

            self.content_area.clear_widgets()

            if tab_id == 0:
                self.show_targets_report_tab()
            elif tab_id == 1:
                self.show_detailed_targets_report_tab()
            elif tab_id == 2:
                self.show_market_report_tab()
            elif tab_id == 3:
                self.show_stats_evaluation_tab()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در تغییر تب: {e}", error_details)

    # ============================================================
    # تب ۱: گزارش تارگت‌ها (اصلی)
    # ============================================================

    def show_targets_report_tab(self):
        """نمایش تب گزارش تارگت‌های اصلی با فیلتر و کارت"""
        try:
            main_scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint=(1, 1),
                scroll_type=['bars', 'content'],
                bar_width=dp(8)
            )

            main_content = BoxLayout(
                orientation='vertical',
                spacing=dp(8),
                size_hint_y=None,
                padding=dp(10)
            )
            main_content.bind(minimum_height=main_content.setter('height'))

            # ========== عنوان ==========
            main_content.add_widget(RTLLabel(
                text='📊 گزارش تارگت‌های اصلی',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(22),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))

            # ========== فیلترها ==========
            filter_box = self._build_target_filter_box()
            main_content.add_widget(filter_box)

            # ========== کارت‌های آماری ==========
            self.targets_stats_box = BoxLayout(
                size_hint_y=None,
                height=dp(85),
                spacing=dp(6),
                padding=dp(5)
            )
            main_content.add_widget(self.targets_stats_box)

            # ========== لیست کارت‌های تارگت ==========
            self.targets_list_container = BoxLayout(
                orientation='vertical',
                spacing=dp(6),
                size_hint_y=None,
                padding=dp(5)
            )
            self.targets_list_container.bind(minimum_height=self.targets_list_container.setter('height'))
            main_content.add_widget(self.targets_list_container)

            main_scroll.add_widget(main_content)
            self.content_area.add_widget(main_scroll)

            # بارگذاری اولیه
            self._apply_targets_filter(None, initial_load=True)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب گزارش تارگت‌ها: {e}", error_details)

    def _build_target_filter_box(self):
        """ساخت باکس فیلتر تب تارگت‌های اصلی"""
        filter_box = BoxLayout(
            orientation='vertical',
            size_hint_y=None,
            height=dp(200),
            spacing=dp(6),
            padding=dp(8)
        )
        with filter_box.canvas.before:
            Color(0.12, 0.12, 0.12, 1)
            RoundedRectangle(pos=filter_box.pos, size=filter_box.size, radius=[dp(8)])

        filter_box.add_widget(RTLLabel(
            text='فیلترها:',
            size_hint_y=None, height=dp(25), font_size=sp(14),
            color=(1, 0.8, 0.2, 1), bold=True
        ))

        # ردیف ۱: عامل + دوره
        row1 = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(8))
        agents = get_agents()
        agent_names = ['همه'] + [a.get('name', '') for a in agents] if agents else ['همه']
        self.filter_agent = PersianComboBox(text='همه', values=agent_names, height=dp(55))
        self.filter_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.filter_agent.main_btn.color = (1, 1, 1, 1)
        self.filter_agent.main_btn.font_size = sp(16)
        self.filter_agent.size_hint_x = 0.5
        row1.add_widget(RTLLabel(
            text='عامل:', size_hint_x=0.12, size_hint_y=None, height=dp(55),
            font_size=sp(14), color=(1, 1, 1, 1)
        ))
        row1.add_widget(self.filter_agent)
        periods = ['همه'] + list(PERIOD_DISPLAY)
        self.filter_period = PersianComboBox(text='همه', values=periods, height=dp(55))
        self.filter_period.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.filter_period.main_btn.color = (1, 1, 1, 1)
        self.filter_period.main_btn.font_size = sp(16)
        self.filter_period.size_hint_x = 0.38
        row1.add_widget(self.filter_period)
        filter_box.add_widget(row1)

        # ردیف ۲: نوع + وضعیت
        row2 = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(8))
        types = ['همه'] + TARGET_TYPES
        self.filter_type = PersianComboBox(text='همه', values=types, height=dp(55))
        self.filter_type.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.filter_type.main_btn.color = (1, 1, 1, 1)
        self.filter_type.main_btn.font_size = sp(16)
        self.filter_type.size_hint_x = 0.5
        row2.add_widget(RTLLabel(
            text='نوع:', size_hint_x=0.12, size_hint_y=None, height=dp(55),
            font_size=sp(14), color=(1, 1, 1, 1)
        ))
        row2.add_widget(self.filter_type)
        statuses = ['همه'] + TARGET_STATUSES
        self.filter_status = PersianComboBox(text='همه', values=statuses, height=dp(55))
        self.filter_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.filter_status.main_btn.color = (1, 1, 1, 1)
        self.filter_status.main_btn.font_size = sp(16)
        self.filter_status.size_hint_x = 0.38
        row2.add_widget(self.filter_status)
        filter_box.add_widget(row2)

        # ردیف ۳: دکمه‌ها
        btn_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
        apply_btn = PersianButton(
            text='اعمال فیلتر', background_color=(0.2, 0.6, 0.2, 1),
            size_hint_x=0.35, size_hint_y=None, height=dp(38),
            color=(1, 1, 1, 1), font_size=sp(14)
        )
        apply_btn.bind(on_press=self._apply_targets_filter)
        btn_row.add_widget(apply_btn)
        excel_btn = PersianButton(
            text='📥 خروجی اکسل', background_color=(0.2, 0.7, 0.4, 1),
            size_hint_x=0.35, size_hint_y=None, height=dp(38),
            color=(1, 1, 1, 1), font_size=sp(14)
        )
        excel_btn.bind(on_press=self._export_targets_report)
        btn_row.add_widget(excel_btn)
        clear_btn = PersianButton(
            text='پاک کردن', background_color=(0.8, 0.4, 0.1, 1),
            size_hint_x=0.30, size_hint_y=None, height=dp(38),
            color=(1, 1, 1, 1), font_size=sp(14)
        )
        clear_btn.bind(on_press=self._clear_targets_filter)
        btn_row.add_widget(clear_btn)
        filter_box.add_widget(btn_row)

        return filter_box

    def _apply_targets_filter(self, instance, initial_load=False):
        """اعمال فیلتر و نمایش کارت‌های تارگت اصلی"""
        try:
            if not hasattr(self, 'filter_agent') or not self.filter_agent:
                return

            agent = self.filter_agent.text if self.filter_agent.text != 'همه' else None
            target_type = self.filter_type.text if hasattr(self, 'filter_type') and self.filter_type.text != 'همه' else None
            status = self.filter_status.text if hasattr(self, 'filter_status') and self.filter_status.text != 'همه' else None
            period = self.filter_period.text if hasattr(self, 'filter_period') and self.filter_period.text != 'همه' else None

            from constants import PERIOD_MAPPING
            period_type = PERIOD_MAPPING.get(period, None) if period else None

            if agent or target_type or status:
                filtered = get_targets_filtered(agent_name=agent, target_type=target_type, status=status)
            else:
                filtered = get_all_targets()

            if not isinstance(filtered, list):
                filtered = []

            if period_type:
                filtered = [t for t in filtered if isinstance(t, dict) and t.get('period_type') == period_type]

            self.current_filtered_targets = filtered
            self.targets_list_container.clear_widgets()
            self._update_targets_stats(filtered)

            if not filtered:
                self.targets_list_container.add_widget(RTLLabel(
                    text='هیچ تارگتی با این فیلترها یافت نشد',
                    size_hint_y=None, height=dp(50),
                    font_size=sp(18), color=(0.5, 0.5, 0.5, 1)
                ))
            else:
                for target in filtered:
                    if not isinstance(target, dict):
                        continue
                    card = self._build_target_card(target)
                    self.targets_list_container.add_widget(card)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال فیلتر: {e}", error_details)

    def _clear_targets_filter(self, instance):
        """پاک کردن فیلترهای تب ۱"""
        if hasattr(self, 'filter_agent') and self.filter_agent:
            self.filter_agent.text = 'همه'
        if hasattr(self, 'filter_type') and self.filter_type:
            self.filter_type.text = 'همه'
        if hasattr(self, 'filter_status') and self.filter_status:
            self.filter_status.text = 'همه'
        if hasattr(self, 'filter_period') and self.filter_period:
            self.filter_period.text = 'همه'
        self._apply_targets_filter(None)

    def _update_targets_stats(self, filtered_targets):
        """بروزرسانی کارت‌های آماری تب ۱"""
        if not hasattr(self, 'targets_stats_box'):
            return
        self.targets_stats_box.clear_widgets()

        total = len(filtered_targets)
        completed = sum(1 for t in filtered_targets if isinstance(t, dict) and t.get('status') == 'تکمیل شده')
        active = sum(1 for t in filtered_targets if isinstance(t, dict) and t.get('status') == 'فعال')

        total_percent = 0
        count = 0
        for t in filtered_targets:
            if isinstance(t, dict):
                tv = t.get('target_value', 1)
                av = t.get('achieved_value', 0)
                if tv > 0:
                    total_percent += (av / tv) * 100
                    count += 1
        avg = total_percent / count if count > 0 else 0

        cards = [
            ('کل تارگت‌ها', f'{total}', (0.2, 0.5, 0.8, 1)),
            ('تکمیل شده', f'{completed}', (0.2, 0.7, 0.2, 1)),
            ('فعال', f'{active}', (0.3, 0.6, 1, 1)),
            ('میانگین تحقق', f'{avg:.0f}%', (0.2, 0.8, 0.2, 1) if avg >= 50 else (1, 0.7, 0, 1))
        ]

        for title, value, color in cards:
            card = BoxLayout(orientation='vertical', size_hint_x=0.25, size_hint_y=None, height=dp(80), padding=dp(4), spacing=dp(2))
            with card.canvas.before:
                Color(*color)
                RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(6)])
            card.add_widget(RTLLabel(text=title, size_hint_y=None, height=dp(22), font_size=sp(11), color=(1, 1, 1, 1)))
            card.add_widget(RTLLabel(text=value, size_hint_y=None, height=dp(35), font_size=sp(24), bold=True, color=(1, 1, 1, 1)))
            self.targets_stats_box.add_widget(card)

    def _build_target_card(self, target):
        """ساخت کارت برای یک تارگت اصلی"""
        status = target.get('status', '')
        target_id = target.get('target_id', '')
        agent_name = target.get('agent_name', '')
        target_type = target.get('target_type', '')
        target_value = target.get('target_value', 0)
        achieved = target.get('achieved_value', 0)
        start_date = target.get('start_date', '')
        end_date = target.get('end_date', '')
        percent = (achieved / target_value * 100) if target_value > 0 else 0

        if status == 'تکمیل شده':
            card_bg, status_color = (0.1, 0.3, 0.1, 1), (0.2, 0.8, 0.2, 1)
        elif status == 'فعال':
            card_bg, status_color = (0.1, 0.2, 0.35, 1), (0.3, 0.6, 1, 1)
        elif status == 'در انتظار':
            card_bg, status_color = (0.25, 0.15, 0.05, 1), (1, 0.8, 0.2, 1)
        elif status == 'لغو شده':
            card_bg, status_color = (0.3, 0.1, 0.1, 1), (0.8, 0.2, 0.2, 1)
        else:
            card_bg, status_color = (0.15, 0.15, 0.15, 1), (0.5, 0.5, 0.5, 1)

        percent_color = (0.2, 0.8, 0.2, 1) if percent >= 100 else (1, 0.7, 0, 1) if percent >= 50 else (0.8, 0.3, 0.3, 1)

        card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(110), spacing=dp(4), padding=[dp(10), dp(8), dp(10), dp(8)])
        with card.canvas.before:
            Color(*card_bg)
            RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(8)])

        row1 = BoxLayout(size_hint_y=None, height=dp(28))
        row1.add_widget(RTLLabel(text=f'{target_id} | {agent_name}', size_hint_x=0.65, size_hint_y=None, height=dp(28), font_size=sp(15), color=(1, 1, 1, 1), bold=True))
        row1.add_widget(RTLLabel(text=status, size_hint_x=0.35, size_hint_y=None, height=dp(28), font_size=sp(14), color=status_color, bold=True, halign='right'))
        card.add_widget(row1)

        row2 = BoxLayout(size_hint_y=None, height=dp(30))
        row2.add_widget(RTLLabel(text=f'{target_type}', size_hint_x=0.25, size_hint_y=None, height=dp(30), font_size=sp(14), color=(1, 1, 1, 1)))
        row2.add_widget(RTLLabel(text=f'هدف: {target_value:,}', size_hint_x=0.35, size_hint_y=None, height=dp(30), font_size=sp(14), color=(0.8, 0.8, 0.8, 1)))
        row2.add_widget(RTLLabel(text=f'تحقق: {achieved:,}', size_hint_x=0.40, size_hint_y=None, height=dp(30), font_size=sp(14), color=(0.2, 0.8, 0.2, 1)))
        card.add_widget(row2)

        row3 = BoxLayout(size_hint_y=None, height=dp(28))
        row3.add_widget(RTLLabel(text=f'{percent:.0f}%', size_hint_x=0.3, size_hint_y=None, height=dp(28), font_size=sp(18), bold=True, color=percent_color))
        row3.add_widget(RTLLabel(text=f'{start_date} تا {end_date}', size_hint_x=0.7, size_hint_y=None, height=dp(28), font_size=sp(12), color=(0.6, 0.6, 0.6, 1), halign='right'))
        card.add_widget(row3)

        detail_btn = PersianButton(
            text='📋 مشاهده جزئیات', size_hint_y=None, height=dp(22),
            background_color=(0.2, 0.5, 0.8, 0.7), color=(1, 1, 1, 1), font_size=sp(12)
        )
        detail_btn.bind(on_press=lambda x, t=target: self._show_target_detail(t))
        card.add_widget(detail_btn)

        return card

    def _show_target_detail(self, target):
        """نمایش دیالوگ جزئیات تارگت اصلی"""
        try:
            if not isinstance(target, dict):
                return

            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v), size=lambda i, v: setattr(content_rect, 'size', v))

            status = target.get('status', '')
            status_colors = {'تکمیل شده': (0.2, 0.8, 0.2, 1), 'فعال': (0.3, 0.6, 1, 1), 'در انتظار': (1, 0.8, 0.2, 1), 'لغو شده': (0.8, 0.2, 0.2, 1)}
            st_color = status_colors.get(status, (0.5, 0.5, 0.5, 1))

            content.add_widget(RTLLabel(
                text=f'جزئیات تارگت {target.get("target_id", "")}',
                size_hint_y=None, height=dp(40), font_size=sp(22), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.7)
            detail_grid = GridLayout(cols=1, spacing=dp(6), size_hint_y=None, padding=dp(5))
            detail_grid.bind(minimum_height=detail_grid.setter('height'))

            tv = target.get('target_value', 0)
            av = target.get('achieved_value', 0)
            pct = (av / tv * 100) if tv > 0 else 0

            fields = [
                ('عامل', target.get('agent_name', ''), (1, 1, 1, 1)),
                ('نوع تارگت', target.get('target_type', ''), (1, 1, 1, 1)),
                ('میزان هدف', f'{tv:,}', (1, 1, 1, 1)),
                ('میزان تحقق', f'{av:,}', (1, 1, 1, 1)),
                ('درصد پیشرفت', f'{pct:.1f}%', (0.2, 0.8, 0.2, 1) if pct >= 100 else (1, 0.7, 0, 1)),
                ('وضعیت', status, st_color),
                ('دوره', target.get('period_type', ''), (1, 1, 1, 1)),
                ('مدت', str(target.get('duration', '')), (1, 1, 1, 1)),
                ('تاریخ شروع', target.get('start_date', ''), (1, 1, 1, 1)),
                ('تاریخ پایان', target.get('end_date', ''), (1, 1, 1, 1)),
                ('توضیحات', target.get('description', '---'), (1, 1, 1, 1))
            ]

            for label, value, color in fields:
                row = BoxLayout(size_hint_y=None, height=dp(32), spacing=dp(4))
                row.add_widget(RTLLabel(text=f'{label}:', size_hint_x=0.35, size_hint_y=None, height=dp(30), font_size=sp(15), color=(0.4, 0.7, 1, 1), halign='right'))
                row.add_widget(RTLLabel(text=value, size_hint_x=0.65, size_hint_y=None, height=dp(30), font_size=sp(15), color=color))
                detail_grid.add_widget(row)

            # ریزتارگت‌های مرتبط
            linked_id = target.get('target_id', '')
            all_dt = get_all_detailed_targets()
            linked = [dt for dt in all_dt if isinstance(dt, dict) and dt.get('linked_target_id') == linked_id]

            if linked:
                detail_grid.add_widget(RTLLabel(
                    text='── ریزتارگت‌های مرتبط ──',
                    size_hint_y=None, height=dp(30), font_size=sp(14), bold=True, color=(1, 0.5, 0, 1)
                ))
                for dt in linked:
                    ds = dt.get('status', '')
                    da = dt.get('achieved_value', 0)
                    dtv = dt.get('target_count', 0)
                    dpct = (da / dtv * 100) if dtv > 0 else 0
                    dc = (0.2, 0.8, 0.2, 1) if ds == 'تکمیل شده' else (0.3, 0.6, 1, 1) if ds == 'فعال' else (1, 0.8, 0.2, 1)
                    drow = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(4))
                    drow.add_widget(RTLLabel(text=f'{dt.get("id", "")} | {dt.get("product_group", "")}', size_hint_x=0.45, size_hint_y=None, height=dp(26), font_size=sp(12), color=(1, 1, 1, 1)))
                    drow.add_widget(RTLLabel(text=f'{da:,}/{dtv:,} ({dpct:.0f}%)', size_hint_x=0.35, size_hint_y=None, height=dp(26), font_size=sp(12), color=dc))
                    drow.add_widget(RTLLabel(text=ds, size_hint_x=0.20, size_hint_y=None, height=dp(26), font_size=sp(11), color=dc, halign='right'))
                    detail_grid.add_widget(drow)

            scroll.add_widget(detail_grid)
            content.add_widget(scroll)

            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            content.add_widget(close_btn)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.92, 0.85), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش جزئیات: {e}", error_details)

    def _export_targets_report(self, instance):
        """خروجی اکسل از تارگت‌های فیلترشده تب ۱"""
        try:
            if not self.current_filtered_targets:
                self.show_message('خطا', 'هیچ تارگتی برای خروجی وجود ندارد')
                return
            success, message, filepath = export_targets_to_excel(self.current_filtered_targets)
            if success:
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خروجی اکسل: {e}", error_details)

    # ============================================================
    # تب ۲: گزارش ریزتارگت‌ها
    # ============================================================

    def show_detailed_targets_report_tab(self):
        """نمایش تب گزارش ریزتارگت‌ها"""
        try:
            main_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint=(1, 1), scroll_type=['bars', 'content'], bar_width=dp(8)
            )

            main_content = BoxLayout(
                orientation='vertical', spacing=dp(8),
                size_hint_y=None, padding=dp(10)
            )
            main_content.bind(minimum_height=main_content.setter('height'))

            # عنوان
            main_content.add_widget(RTLLabel(
                text='📋 گزارش ریزتارگت‌ها',
                size_hint_y=None, height=dp(45),
                font_size=sp(22), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            # فیلترها
            filter_box = self._build_dt_filter_box()
            main_content.add_widget(filter_box)

            # کارت‌های آماری
            self.dt_stats_box = BoxLayout(
                size_hint_y=None, height=dp(85),
                spacing=dp(6), padding=dp(5)
            )
            main_content.add_widget(self.dt_stats_box)

            # لیست کارت‌ها
            self.dt_list_container = BoxLayout(
                orientation='vertical', spacing=dp(6),
                size_hint_y=None, padding=dp(5)
            )
            self.dt_list_container.bind(minimum_height=self.dt_list_container.setter('height'))
            main_content.add_widget(self.dt_list_container)

            main_scroll.add_widget(main_content)
            self.content_area.add_widget(main_scroll)

            self._apply_dt_filter(None, initial_load=True)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب ریزتارگت‌ها: {e}", error_details)

    def _build_dt_filter_box(self):
        """ساخت باکس فیلتر تب ریزتارگت‌ها"""
        filter_box = BoxLayout(
            orientation='vertical', size_hint_y=None,
            height=dp(230), spacing=dp(6), padding=dp(8)
        )
        with filter_box.canvas.before:
            Color(0.12, 0.12, 0.12, 1)
            RoundedRectangle(pos=filter_box.pos, size=filter_box.size, radius=[dp(8)])

        filter_box.add_widget(RTLLabel(
            text='فیلترها:', size_hint_y=None, height=dp(25),
            font_size=sp(14), color=(1, 0.8, 0.2, 1), bold=True
        ))

        # ردیف ۱: عامل + گروه کالا
        row1 = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(8))
        agents = get_agents()
        agent_names = ['همه'] + [a.get('name', '') for a in agents] if agents else ['همه']
        self.dt_filter_agent = PersianComboBox(text='همه', values=agent_names, height=dp(55))
        self.dt_filter_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.dt_filter_agent.main_btn.color = (1, 1, 1, 1)
        self.dt_filter_agent.main_btn.font_size = sp(16)
        self.dt_filter_agent.size_hint_x = 0.60
        row1.add_widget(RTLLabel(
            text='عامل', size_hint_x=0.10, size_hint_y=None, height=dp(55),
            font_size=sp(10), color=(1, 1, 1, 1)
        ))
        row1.add_widget(self.dt_filter_agent)

        product_groups = get_product_groups()
        if not isinstance(product_groups, list):
            product_groups = []
        product_list = ['همه'] + product_groups if product_groups else ['همه']
        self.dt_filter_product = PersianComboBox(text='همه', values=product_list, height=dp(55))
        self.dt_filter_product.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.dt_filter_product.main_btn.color = (1, 1, 1, 1)
        self.dt_filter_product.main_btn.font_size = sp(16)
        self.dt_filter_product.size_hint_x = 0.60
        row1.add_widget(self.dt_filter_product)
        filter_box.add_widget(row1)

        # ردیف ۲: وضعیت + بازه زمانی
        row2 = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(8))
        statuses = ['همه', 'در انتظار', 'فعال', 'تکمیل شده', 'لغو شده']
        self.dt_filter_status = PersianComboBox(text='همه', values=statuses, height=dp(55))
        self.dt_filter_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.dt_filter_status.main_btn.color = (1, 1, 1, 1)
        self.dt_filter_status.main_btn.font_size = sp(16)
        self.dt_filter_status.size_hint_x = 0.50
        row2.add_widget(RTLLabel(
            text='وضعیت', size_hint_x=0.20, size_hint_y=None, height=dp(55),
            font_size=sp(10), color=(1, 1, 1, 1)
        ))
        row2.add_widget(self.dt_filter_status)

        today = get_today_jalali()
        self.dt_filter_start = RTLTextInput(
            text='', hint_text='از تاریخ', multiline=False,
            size_hint_x=0.50, size_hint_y=None, height=dp(55), font_size=sp(16)
        )
        self.dt_filter_start.bg_color = (0.15, 0.15, 0.15, 1)
        self.dt_filter_start.border_color = (0.3, 0.3, 0.3, 1)
        self.dt_filter_start.border_color_focus = (0.2, 0.5, 0.9, 1)
        self.dt_filter_start._hidden_input.foreground_color = (1, 1, 1, 1)
        row2.add_widget(self.dt_filter_start)

        self.dt_filter_end = RTLTextInput(
            text='', hint_text='تا تاریخ', multiline=False,
            size_hint_x=0.50, size_hint_y=None, height=dp(55), font_size=sp(16)
        )
        self.dt_filter_end.bg_color = (0.15, 0.15, 0.15, 1)
        self.dt_filter_end.border_color = (0.3, 0.3, 0.3, 1)
        self.dt_filter_end.border_color_focus = (0.2, 0.5, 0.9, 1)
        self.dt_filter_end._hidden_input.foreground_color = (1, 1, 1, 1)
        row2.add_widget(self.dt_filter_end)
        filter_box.add_widget(row2)

        # ردیف ۳: دکمه‌ها
        btn_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
        apply_btn = PersianButton(
            text='اعمال فیلتر', background_color=(0.2, 0.6, 0.2, 1),
            size_hint_x=0.35, size_hint_y=None, height=dp(38),
            color=(1, 1, 1, 1), font_size=sp(14)
        )
        apply_btn.bind(on_press=self._apply_dt_filter)
        btn_row.add_widget(apply_btn)
        excel_btn = PersianButton(
            text='خروجی اکسل', background_color=(0.2, 0.7, 0.4, 1),
            size_hint_x=0.35, size_hint_y=None, height=dp(38),
            color=(1, 1, 1, 1), font_size=sp(14)
        )
        excel_btn.bind(on_press=self._export_dt_report)
        btn_row.add_widget(excel_btn)
        clear_btn = PersianButton(
            text='پاک کردن', background_color=(0.8, 0.4, 0.1, 1),
            size_hint_x=0.30, size_hint_y=None, height=dp(38),
            color=(1, 1, 1, 1), font_size=sp(14)
        )
        clear_btn.bind(on_press=self._clear_dt_filter)
        btn_row.add_widget(clear_btn)
        filter_box.add_widget(btn_row)

        return filter_box

    def _apply_dt_filter(self, instance, initial_load=False):
        """اعمال فیلتر و نمایش کارت‌های ریزتارگت"""
        try:
            if not hasattr(self, 'dt_filter_agent') or not self.dt_filter_agent:
                return

            agent = self.dt_filter_agent.text if self.dt_filter_agent.text != 'همه' else None
            product = self.dt_filter_product.text if hasattr(self, 'dt_filter_product') and self.dt_filter_product.text != 'همه' else None
            status = self.dt_filter_status.text if hasattr(self, 'dt_filter_status') and self.dt_filter_status.text != 'همه' else None
            start = self.dt_filter_start.text.strip() if hasattr(self, 'dt_filter_start') and self.dt_filter_start.text.strip() else None
            end = self.dt_filter_end.text.strip() if hasattr(self, 'dt_filter_end') and self.dt_filter_end.text.strip() else None

            filtered = get_detailed_targets_filtered(
                agent_name=agent,
                product_group=product,
                status=status,
                start_date=start,
                end_date=end
            )

            if not isinstance(filtered, list):
                filtered = []

            self.current_filtered_dt = filtered
            self.dt_list_container.clear_widgets()
            self._update_dt_stats(filtered)

            if not filtered:
                self.dt_list_container.add_widget(RTLLabel(
                    text='هیچ ریزتارگتی با این فیلترها یافت نشد',
                    size_hint_y=None, height=dp(50),
                    font_size=sp(18), color=(0.5, 0.5, 0.5, 1)
                ))
            else:
                for dt in filtered:
                    if not isinstance(dt, dict):
                        continue
                    card = self._build_dt_card(dt)
                    self.dt_list_container.add_widget(card)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال فیلتر ریزتارگت: {e}", error_details)

    def _clear_dt_filter(self, instance):
        """پاک کردن فیلترهای تب ۲"""
        if hasattr(self, 'dt_filter_agent') and self.dt_filter_agent:
            self.dt_filter_agent.text = 'همه'
        if hasattr(self, 'dt_filter_product') and self.dt_filter_product:
            self.dt_filter_product.text = 'همه'
        if hasattr(self, 'dt_filter_status') and self.dt_filter_status:
            self.dt_filter_status.text = 'همه'
        if hasattr(self, 'dt_filter_start') and self.dt_filter_start:
            self.dt_filter_start.text = ''
        if hasattr(self, 'dt_filter_end') and self.dt_filter_end:
            self.dt_filter_end.text = ''
        self._apply_dt_filter(None)

    def _update_dt_stats(self, filtered):
        """بروزرسانی کارت‌های آماری تب ۲"""
        if not hasattr(self, 'dt_stats_box'):
            return
        self.dt_stats_box.clear_widgets()

        total = len(filtered)
        completed = sum(1 for t in filtered if isinstance(t, dict) and t.get('status') == 'تکمیل شده')
        active = sum(1 for t in filtered if isinstance(t, dict) and t.get('status') == 'فعال')

        total_pct = 0
        cnt = 0
        for t in filtered:
            if isinstance(t, dict):
                tv = t.get('target_count', 1)
                av = t.get('achieved_value', 0)
                if tv > 0:
                    total_pct += (av / tv) * 100
                    cnt += 1
        avg = total_pct / cnt if cnt > 0 else 0

        cards = [
            ('کل ریزتارگت‌ها', f'{total}', (0.2, 0.5, 0.8, 1)),
            ('تکمیل شده', f'{completed}', (0.2, 0.7, 0.2, 1)),
            ('فعال', f'{active}', (0.3, 0.6, 1, 1)),
            ('میانگین تحقق', f'{avg:.0f}%', (0.2, 0.8, 0.2, 1) if avg >= 50 else (1, 0.7, 0, 1))
        ]

        for title, value, color in cards:
            card = BoxLayout(orientation='vertical', size_hint_x=0.25, size_hint_y=None, height=dp(80), padding=dp(4), spacing=dp(2))
            with card.canvas.before:
                Color(*color)
                RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(6)])
            card.add_widget(RTLLabel(text=title, size_hint_y=None, height=dp(22), font_size=sp(11), color=(1, 1, 1, 1)))
            card.add_widget(RTLLabel(text=value, size_hint_y=None, height=dp(35), font_size=sp(24), bold=True, color=(1, 1, 1, 1)))
            self.dt_stats_box.add_widget(card)

    def _build_dt_card(self, target):
        """ساخت کارت برای یک ریزتارگت"""
        status = target.get('status', '')
        dt_id = target.get('id', '')
        agent = target.get('agent_name', '')
        product = target.get('product_group', '')
        target_count = target.get('target_count', 0)
        achieved = target.get('achieved_value', 0)
        daily = target.get('daily_target', 0)
        unit = target.get('unit', '')
        period = target.get('period', '')
        percent = (achieved / target_count * 100) if target_count > 0 else 0

        if status == 'تکمیل شده':
            card_bg, st_color = (0.1, 0.3, 0.1, 1), (0.2, 0.8, 0.2, 1)
        elif status == 'فعال':
            card_bg, st_color = (0.1, 0.2, 0.35, 1), (0.3, 0.6, 1, 1)
        elif status == 'در انتظار':
            card_bg, st_color = (0.25, 0.15, 0.05, 1), (1, 0.8, 0.2, 1)
        else:
            card_bg, st_color = (0.15, 0.15, 0.15, 1), (0.5, 0.5, 0.5, 1)

        card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(125), spacing=dp(3), padding=[dp(10), dp(6), dp(10), dp(6)])
        with card.canvas.before:
            Color(*card_bg)
            RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(8)])

        # ردیف ۱: شناسه | عامل | وضعیت
        row1 = BoxLayout(size_hint_y=None, height=dp(26))
        row1.add_widget(RTLLabel(text=f'{dt_id} | {agent}', size_hint_x=0.60, size_hint_y=None, height=dp(26), font_size=sp(14), color=(1, 1, 1, 1), bold=True))
        row1.add_widget(RTLLabel(text=status, size_hint_x=0.40, size_hint_y=None, height=dp(26), font_size=sp(13), color=st_color, bold=True, halign='right'))
        card.add_widget(row1)

        # ردیف ۲: گروه کالا | هدف | واحد | دوره
        row2 = BoxLayout(size_hint_y=None, height=dp(26))
        row2.add_widget(RTLLabel(text=product, size_hint_x=0.25, size_hint_y=None, height=dp(26), font_size=sp(13), color=(1, 1, 1, 1)))
        row2.add_widget(RTLLabel(text=f'هدف: {target_count:,}', size_hint_x=0.30, size_hint_y=None, height=dp(26), font_size=sp(13), color=(0.8, 0.8, 0.8, 1)))
        row2.add_widget(RTLLabel(text=unit, size_hint_x=0.15, size_hint_y=None, height=dp(26), font_size=sp(13), color=(0.8, 0.8, 0.8, 1)))
        row2.add_widget(RTLLabel(text=period, size_hint_x=0.30, size_hint_y=None, height=dp(26), font_size=sp(13), color=(0.8, 0.8, 0.8, 1), halign='right'))
        card.add_widget(row2)

        # ردیف ۳: درصد + تحقق + تارگت روزانه
        row3 = BoxLayout(size_hint_y=None, height=dp(26))
        pct_color = (0.2, 0.8, 0.2, 1) if percent >= 100 else (1, 0.7, 0, 1) if percent >= 50 else (0.8, 0.3, 0.3, 1)
        row3.add_widget(RTLLabel(text=f'{percent:.0f}%', size_hint_x=0.25, size_hint_y=None, height=dp(26), font_size=sp(16), bold=True, color=pct_color))
        row3.add_widget(RTLLabel(text=f'تحقق: {achieved:,}', size_hint_x=0.40, size_hint_y=None, height=dp(26), font_size=sp(13), color=(0.2, 0.8, 0.2, 1)))
        row3.add_widget(RTLLabel(text=f'روزانه: {daily:,}', size_hint_x=0.35, size_hint_y=None, height=dp(26), font_size=sp(13), color=(0.6, 0.6, 0.6, 1), halign='right'))
        card.add_widget(row3)

        # ردیف ۴: Progress Bar
        progress = DTProgressBar(percent=min(percent, 100), height=dp(18))
        card.add_widget(progress)

        # دکمه جزئیات
        detail_btn = PersianButton(
            text='📋 مشاهده جزئیات و تاریخچه', size_hint_y=None, height=dp(22),
            background_color=(0.2, 0.5, 0.8, 0.7), color=(1, 1, 1, 1), font_size=sp(11)
        )
        detail_btn.bind(on_press=lambda x, t=target: self._show_dt_detail(t))
        card.add_widget(detail_btn)

        return card

    def _show_dt_detail(self, target):
        """نمایش دیالوگ جزئیات ریزتارگت با تاریخچه روزانه"""
        try:
            if not isinstance(target, dict):
                return

            content = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(6))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v), size=lambda i, v: setattr(content_rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'جزئیات ریزتارگت {target.get("id", "")}',
                size_hint_y=None, height=dp(38),
                font_size=sp(20), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.55)
            detail_grid = GridLayout(cols=1, spacing=dp(4), size_hint_y=None, padding=dp(3))
            detail_grid.bind(minimum_height=detail_grid.setter('height'))

            # اطلاعات پایه
            tv = target.get('target_count', 0)
            av = target.get('achieved_value', 0)
            pct = (av / tv * 100) if tv > 0 else 0
            status = target.get('status', '')

            info_fields = [
                ('عامل', target.get('agent_name', '')),
                ('گروه کالا', target.get('product_group', '')),
                ('هدف کل', f'{tv:,} {target.get("unit", "")}'),
                ('تارگت روزانه', f'{target.get("daily_target", 0):,}'),
                ('تحقق کل', f'{av:,}'),
                ('درصد پیشرفت', f'{pct:.1f}%'),
                ('وضعیت', status),
                ('دوره', target.get('period', '')),
                ('مدت', str(target.get('duration', ''))),
                ('تاریخ شروع', target.get('start_date', '')),
                ('تاریخ پایان', target.get('end_date', '')),
                ('پیوند به تارگت', target.get('linked_target_id', ''))
            ]

            for label, value in info_fields:
                row = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(3))
                row.add_widget(RTLLabel(text=f'{label}:', size_hint_x=0.35, size_hint_y=None, height=dp(24), font_size=sp(13), color=(0.4, 0.7, 1, 1), halign='right'))
                row.add_widget(RTLLabel(text=value, size_hint_x=0.65, size_hint_y=None, height=dp(24), font_size=sp(13), color=(1, 1, 1, 1)))
                detail_grid.add_widget(row)

            # تاریخچه روزانه
            daily_achievements = target.get('daily_achievements', {})
            if daily_achievements and isinstance(daily_achievements, dict) and len(daily_achievements) > 0:
                detail_grid.add_widget(RTLLabel(
                    text='── تاریخچه تحقق روزانه ──',
                    size_hint_y=None, height=dp(28),
                    font_size=sp(14), bold=True, color=(1, 0.5, 0, 1)
                ))

                # هدر جدول
                hdr = BoxLayout(size_hint_y=None, height=dp(24), spacing=dp(3))
                hdr.add_widget(RTLLabel(text='تاریخ', size_hint_x=0.25, size_hint_y=None, height=dp(22), font_size=sp(11), bold=True, color=(0.4, 0.7, 1, 1)))
                hdr.add_widget(RTLLabel(text='تحقق', size_hint_x=0.20, size_hint_y=None, height=dp(22), font_size=sp(11), bold=True, color=(0.4, 0.7, 1, 1)))
                hdr.add_widget(RTLLabel(text='پیشرفت', size_hint_x=0.55, size_hint_y=None, height=dp(22), font_size=sp(11), bold=True, color=(0.4, 0.7, 1, 1)))
                detail_grid.add_widget(hdr)

                daily_target = target.get('daily_target', 1)
                sorted_dates = sorted(daily_achievements.keys(), reverse=True)

                for date in sorted_dates:
                    val = daily_achievements.get(date, 0)
                    day_pct = (val / daily_target * 100) if daily_target > 0 else 0

                    day_row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(3))
                    day_row.add_widget(RTLLabel(text=date, size_hint_x=0.25, size_hint_y=None, height=dp(28), font_size=sp(12), color=(1, 1, 1, 1)))

                    val_color = (0.2, 0.8, 0.2, 1) if day_pct >= 100 else (1, 0.7, 0, 1) if day_pct >= 50 else (0.8, 0.3, 0.3, 1)
                    day_row.add_widget(RTLLabel(text=f'{val:,}', size_hint_x=0.20, size_hint_y=None, height=dp(28), font_size=sp(12), color=val_color))

                    # Progress bar کوچک برای هر روز
                    day_progress = DTProgressBar(percent=min(day_pct, 100), height=dp(20), size_hint_x=0.55)
                    day_row.add_widget(day_progress)

                    detail_grid.add_widget(day_row)

                # جمع کل
                total_daily = sum(daily_achievements.values())
                sum_row = BoxLayout(size_hint_y=None, height=dp(24))
                sum_row.add_widget(RTLLabel(
                    text=f'جمع کل: {total_daily:,}', size_hint_x=1,
                    size_hint_y=None, height=dp(22),
                    font_size=sp(12), bold=True, color=(0.2, 0.8, 0.2, 1), halign='center'
                ))
                detail_grid.add_widget(sum_row)
            else:
                detail_grid.add_widget(RTLLabel(
                    text='هنوز تحقق روزانه‌ای ثبت نشده است',
                    size_hint_y=None, height=dp(30),
                    font_size=sp(13), color=(0.5, 0.5, 0.5, 1)
                ))

            scroll.add_widget(detail_grid)
            content.add_widget(scroll)

            # دکمه بستن
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(15)
            )
            content.add_widget(close_btn)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.92, 0.85), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش جزئیات: {e}", error_details)

    def _export_dt_report(self, instance):
        """خروجی اکسل از ریزتارگت‌های فیلترشده تب ۲"""
        try:
            if not self.current_filtered_dt:
                self.show_message('خطا', 'هیچ ریزتارگتی برای خروجی وجود ندارد')
                return
            success, message, filepath = export_detailed_to_excel(self.current_filtered_dt)
            if success:
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خروجی اکسل: {e}", error_details)

    # ============================================================
    # تب ۳: گزارش بازاری
    # ============================================================

    def show_market_report_tab(self):
        """نمایش تب گزارش بازاری - لیست ساده سرکشی‌ها"""
        try:
            self.market_filter_customer = None
            self.market_filter_route = None
            self.market_filter_start = None
            self.market_filter_end = None
            self.market_filter_unreported = False
            self.current_market_visits = []

            main_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint=(1, 1), scroll_type=['bars', 'content'], bar_width=dp(8)
            )

            main_content = BoxLayout(
                orientation='vertical', spacing=dp(8),
                size_hint_y=None, padding=dp(10)
            )
            main_content.bind(minimum_height=main_content.setter('height'))

            # عنوان
            main_content.add_widget(RTLLabel(
                text='گزارش بازاری',
                size_hint_y=None, height=dp(45),
                font_size=sp(22), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            # ========== فیلترها ==========
            filter_box = BoxLayout(
                orientation='vertical', size_hint_y=None,
                height=dp(250), spacing=dp(5), padding=dp(8)
            )
            with filter_box.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                RoundedRectangle(pos=filter_box.pos, size=filter_box.size, radius=[dp(8)])

            filter_box.add_widget(RTLLabel(
                text='فیلترها:', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 0.8, 0.2, 1), bold=True
            ))

            # ردیف ۱: مشتری + مسیر
            row1 = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(8))

            all_customers = get_customers()
            customer_names = ['همه'] + list(set(c.get('name', '') for c in all_customers if c.get('name')))
            customer_names.sort()
            self.market_filter_customer = PersianComboBox(
                text='همه', values=customer_names, height=dp(50)
            )
            self.market_filter_customer.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_filter_customer.main_btn.color = (1, 1, 1, 1)
            self.market_filter_customer.main_btn.font_size = sp(14)
            self.market_filter_customer.size_hint_x = 0.5
            row1.add_widget(RTLLabel(
                text='مشتری:', size_hint_x=0.12, size_hint_y=None, height=dp(50),
                font_size=sp(13), color=(1, 1, 1, 1)
            ))
            row1.add_widget(self.market_filter_customer)

            routes = get_routes()
            route_names = ['همه'] + list(set(r.get('name', '') for r in routes if r.get('name')))
            route_names.sort()
            self.market_filter_route = PersianComboBox(
                text='همه', values=route_names, height=dp(50)
            )
            self.market_filter_route.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_filter_route.main_btn.color = (1, 1, 1, 1)
            self.market_filter_route.main_btn.font_size = sp(14)
            self.market_filter_route.size_hint_x = 0.38
            row1.add_widget(self.market_filter_route)
            filter_box.add_widget(row1)

            # ردیف ۲: از تاریخ + تا تاریخ + دکمه فیلتر گزارش نشده
            row2 = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(8))
            self.market_filter_start = RTLTextInput(
                text='', hint_text='از تاریخ', multiline=False,
                size_hint_x=0.30, size_hint_y=None, height=dp(50), font_size=sp(16)
            )
            self.market_filter_start.bg_color = (0.15, 0.15, 0.15, 1)
            self.market_filter_start.border_color = (0.3, 0.3, 0.3, 1)
            self.market_filter_start.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.market_filter_start._hidden_input.foreground_color = (1, 1, 1, 1)
            row2.add_widget(self.market_filter_start)

            self.market_filter_end = RTLTextInput(
                text='', hint_text='تا تاریخ', multiline=False,
                size_hint_x=0.30, size_hint_y=None, height=dp(50), font_size=sp(16)
            )
            self.market_filter_end.bg_color = (0.15, 0.15, 0.15, 1)
            self.market_filter_end.border_color = (0.3, 0.3, 0.3, 1)
            self.market_filter_end.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.market_filter_end._hidden_input.foreground_color = (1, 1, 1, 1)
            row2.add_widget(self.market_filter_end)

            # دکمه Toggle "فقط گزارش نشده"
            self.market_unreported_btn = PersianButton(
                text='[ ] گزارش نشده',
                size_hint_x=0.40,
                size_hint_y=None,
                height=dp(38),
                background_color=(0.25, 0.25, 0.25, 1),
                color=(0.7, 0.7, 0.7, 1),
                font_size=sp(12)
            )
            self.market_unreported_btn.bind(on_press=self._toggle_unreported_filter)
            row2.add_widget(self.market_unreported_btn)
            filter_box.add_widget(row2)

            # ردیف ۳: دکمه‌ها
            btn_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
            apply_btn = PersianButton(
                text='اعمال فیلتر', background_color=(0.2, 0.6, 0.2, 1),
                size_hint_x=0.35, size_hint_y=None, height=dp(38),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            apply_btn.bind(on_press=self._apply_market_filter)
            btn_row.add_widget(apply_btn)
            excel_btn = PersianButton(
                text='خروجی اکسل', background_color=(0.2, 0.7, 0.4, 1),
                size_hint_x=0.35, size_hint_y=None, height=dp(38),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            excel_btn.bind(on_press=self._export_market_report)
            btn_row.add_widget(excel_btn)
            clear_btn = PersianButton(
                text='پاک کردن', background_color=(0.8, 0.4, 0.1, 1),
                size_hint_x=0.30, size_hint_y=None, height=dp(38),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            clear_btn.bind(on_press=self._clear_market_filter)
            btn_row.add_widget(clear_btn)
            filter_box.add_widget(btn_row)

            main_content.add_widget(filter_box)

            # ========== تعداد نتایج ==========
            self.market_count_label = RTLLabel(
                text='', size_hint_y=None, height=dp(25),
                font_size=sp(13), color=(0.6, 0.6, 0.6, 1)
            )
            main_content.add_widget(self.market_count_label)

            # ========== لیست سرکشی‌ها ==========
            self.market_list_container = BoxLayout(
                orientation='vertical', spacing=dp(3),
                size_hint_y=None, padding=dp(3)
            )
            self.market_list_container.bind(minimum_height=self.market_list_container.setter('height'))
            main_content.add_widget(self.market_list_container)

            main_scroll.add_widget(main_content)
            self.content_area.add_widget(main_scroll)

            self._apply_market_filter(None, initial_load=True)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب گزارش بازاری: {e}", error_details)

    def _toggle_unreported_filter(self, instance):
        """تغییر وضعیت فیلتر گزارش نشده"""
        self.market_filter_unreported = not self.market_filter_unreported
        if self.market_filter_unreported:
            instance.text = '[X] گزارش نشده'
            instance.background_color = (0.15, 0.5, 0.15, 1)
            instance.color = (0.2, 0.9, 0.2, 1)
        else:
            instance.text = '[ ] گزارش نشده'
            instance.background_color = (0.25, 0.25, 0.25, 1)
            instance.color = (0.7, 0.7, 0.7, 1)

    def _apply_market_filter(self, instance, initial_load=False):
        """اعمال فیلتر و نمایش لیست سرکشی‌ها"""
        try:
            from utils.supervisor_visits_manager import get_all_visits, get_visits_filtered

            customer = self.market_filter_customer.text if hasattr(self, 'market_filter_customer') and self.market_filter_customer and self.market_filter_customer.text != 'همه' else None
            route = self.market_filter_route.text if hasattr(self, 'market_filter_route') and self.market_filter_route and self.market_filter_route.text != 'همه' else None
            start = self.market_filter_start.text.strip() if hasattr(self, 'market_filter_start') and self.market_filter_start and self.market_filter_start.text.strip() else None
            end = self.market_filter_end.text.strip() if hasattr(self, 'market_filter_end') and self.market_filter_end and self.market_filter_end.text.strip() else None
            unreported_only = self.market_filter_unreported if hasattr(self, 'market_filter_unreported') else False

            if customer or start or end:
                visits = get_visits_filtered(customer=customer, start_date=start, end_date=end)
            else:
                visits = get_all_visits()

            if not isinstance(visits, list):
                visits = []

            if route:
                visits = [v for v in visits if isinstance(v, dict) and v.get('route', '') == route]

            if unreported_only:
                visits = [v for v in visits if isinstance(v, dict) and not v.get('reported_to_manager', False)]

            visits.sort(key=lambda x: x.get('date', '') + x.get('time', ''), reverse=True)

            self.current_market_visits = visits
            self.market_list_container.clear_widgets()
            self.market_count_label.text = f'تعداد: {len(visits)} سرکشی'

            if not visits:
                self.market_list_container.add_widget(RTLLabel(
                    text='هیچ سرکشی با این فیلترها یافت نشد',
                    size_hint_y=None, height=dp(50),
                    font_size=sp(18), color=(0.5, 0.5, 0.5, 1)
                ))
                return

            # هدر جدول
            header = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(3))
            headers = [('تاریخ', 0.16), ('مسیر', 0.22), ('مشتری', 0.22), ('علت', 0.25), ('گزارش', 0.15)]
            for text, size in headers:
                header.add_widget(RTLLabel(
                    text=text, size_hint_x=size, size_hint_y=None, height=dp(26),
                    font_size=sp(12), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                ))
            self.market_list_container.add_widget(header)

            for i, visit in enumerate(visits):
                if not isinstance(visit, dict):
                    continue

                reported = visit.get('reported_to_manager', False)
                flag_text = '[X]' if reported else '[ ]'
                flag_color = (0.2, 0.8, 0.2, 1) if reported else (0.4, 0.4, 0.4, 1)

                row = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(3))
                with row.canvas.before:
                    Color(0.12, 0.12, 0.12, 1) if i % 2 == 0 else Color(0.15, 0.15, 0.15, 1)
                    RoundedRectangle(pos=row.pos, size=row.size, radius=[dp(3)])

                row.add_widget(RTLLabel(
                    text=visit.get('date', ''), size_hint_x=0.16, size_hint_y=None, height=dp(33),
                    font_size=sp(12), color=(1, 1, 1, 1), halign='center'
                ))
                row.add_widget(RTLLabel(
                    text=visit.get('route', ''), size_hint_x=0.22, size_hint_y=None, height=dp(33),
                    font_size=sp(12), color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=visit.get('customer', ''), size_hint_x=0.22, size_hint_y=None, height=dp(33),
                    font_size=sp(12), color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=visit.get('visit_reason', ''), size_hint_x=0.25, size_hint_y=None, height=dp(33),
                    font_size=sp(11), color=(0.8, 0.8, 0.8, 1)
                ))
                row.add_widget(RTLLabel(
                    text=flag_text, size_hint_x=0.15, size_hint_y=None, height=dp(33),
                    font_size=sp(14), bold=True, color=flag_color, halign='center'
                ))

                row.bind(on_touch_down=lambda instance, touch, v=visit: self._on_market_row_click(instance, touch, v))
                self.market_list_container.add_widget(row)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال فیلتر بازار: {e}", error_details)

    def _on_market_row_click(self, instance, touch, visit):
        """مدیریت کلیک روی ردیف سرکشی"""
        if instance.collide_point(*touch.pos):
            self._show_market_visit_detail(visit)

    def _clear_market_filter(self, instance):
        """پاک کردن فیلترهای تب بازار"""
        if hasattr(self, 'market_filter_customer') and self.market_filter_customer:
            self.market_filter_customer.text = 'همه'
        if hasattr(self, 'market_filter_route') and self.market_filter_route:
            self.market_filter_route.text = 'همه'
        if hasattr(self, 'market_filter_start') and self.market_filter_start:
            self.market_filter_start.text = ''
        if hasattr(self, 'market_filter_end') and self.market_filter_end:
            self.market_filter_end.text = ''
        if hasattr(self, 'market_unreported_btn') and self.market_unreported_btn:
            self.market_unreported_btn.text = '[ ] گزارش نشده'
            self.market_unreported_btn.background_color = (0.25, 0.25, 0.25, 1)
            self.market_unreported_btn.color = (0.7, 0.7, 0.7, 1)
        self.market_filter_unreported = False
        self._apply_market_filter(None)

    def _export_market_report(self, instance):
        """خروجی اکسل از سرکشی‌های فیلترشده"""
        try:
            from utils.supervisor_visits_manager import export_visits_to_excel

            if not self.current_market_visits:
                self.show_message('خطا', 'هیچ سرکشی برای خروجی وجود ندارد')
                return

            success, message, filepath = export_visits_to_excel(self.current_market_visits)
            if success:
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خروجی اکسل: {e}", error_details)

    def _show_market_visit_detail(self, visit):
        """نمایش دیالوگ جزئیات کامل یک سرکشی"""
        try:
            if not isinstance(visit, dict):
                return

            content = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(6))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))

            reported = visit.get('reported_to_manager', False)
            title_suffix = ' - [X] گزارش شده' if reported else ' - [ ] گزارش نشده'
            title_color = (0.2, 0.8, 0.2, 1) if reported else (1, 0.8, 0.2, 1)

            content.add_widget(RTLLabel(
                text=f'جزئیات سرکشی{title_suffix}',
                size_hint_y=None, height=dp(38),
                font_size=sp(18), bold=True, color=title_color
            ))

            scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.65)
            detail_grid = GridLayout(cols=1, spacing=dp(3), size_hint_y=None, padding=dp(3))
            detail_grid.bind(minimum_height=detail_grid.setter('height'))

            fields = [
                ('تاریخ', visit.get('date', '')),
                ('ساعت', visit.get('time', '')),
                ('مسیر', visit.get('route', '')),
                ('مشتری', visit.get('customer', '')),
                ('نحوه سرکشی', visit.get('visit_type', '')),
                ('علت سرکشی', visit.get('visit_reason', '')),
                ('وضعیت مشتری', visit.get('customer_status', '')),
                ('وضعیت حضور در شلف', visit.get('shelf_status', '')),
                ('تعداد سرکشی در ماه', visit.get('monthly_visits', '')),
                ('آیا سرکشی کافیست', visit.get('visit_sufficient', '')),
                ('خرید مورد انتظار', visit.get('expected_purchase', '')),
                ('وضعیت موجودی', visit.get('inventory_status', '')),
                ('برخورد بازاریاب', visit.get('agent_behavior', '')),
                ('برخورد موزع', visit.get('distributor_behavior', '')),
                ('رضایتمندی مشتری', visit.get('customer_satisfaction', '')),
                ('تحقق هدف سرکشی', visit.get('target_achievement', '')),
                ('نیاز به پیگیری', visit.get('need_followup', '')),
                ('تاریخ مراجعه بعدی', visit.get('next_visit_date', '')),
                ('توضیحات سوپروایزر', visit.get('supervisor_note', '')),
                ('نظرات مشتری', visit.get('customer_feedback', '')),
                ('نظریه سوپروایزر', visit.get('supervisor_opinion', '')),
            ]

            for label, value in fields:
                if value:
                    row = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(3))
                    row.add_widget(RTLLabel(
                        text=f'{label}:', size_hint_x=0.35, size_hint_y=None, height=dp(24),
                        font_size=sp(12), color=(0.4, 0.7, 1, 1), halign='right'
                    ))
                    row.add_widget(RTLLabel(
                        text=str(value), size_hint_x=0.65, size_hint_y=None, height=dp(24),
                        font_size=sp(12), color=(1, 1, 1, 1)
                    ))
                    detail_grid.add_widget(row)

            scroll.add_widget(detail_grid)
            content.add_widget(scroll)

            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))

            if not reported:
                send_report_btn = PersianButton(
                    text='ارسال گزارش به مدیر',
                    background_color=(0.2, 0.7, 0.2, 1),
                    size_hint_x=0.5, size_hint_y=None, height=dp(40),
                    color=(1, 1, 1, 1), font_size=sp(14)
                )
                send_report_btn.bind(on_press=lambda x, v=visit: self._show_manager_report(v))
                btn_layout.add_widget(send_report_btn)
            else:
                already_sent = PersianButton(
                    text='گزارش ارسال شده',
                    background_color=(0.2, 0.6, 0.2, 1),
                    size_hint_x=0.5, size_hint_y=None, height=dp(40),
                    color=(1, 1, 1, 1), font_size=sp(14), disabled=True
                )
                btn_layout.add_widget(already_sent)

            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            btn_layout.add_widget(close_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.92, 0.88), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش جزئیات: {e}", error_details)

    def _show_manager_report(self, visit):
        """نمایش نامه اداری برای ارسال به مدیر با نام کاربر"""
        try:
            from kivy.app import App
            
            # ✅ دریافت نام کاربر جاری
            app = App.get_running_app()
            current_username = app.current_username if hasattr(app, 'current_username') else ''
            
            if not current_username:
                try:
                    from utils.user_manager import get_current_user
                    user = get_current_user()
                    if user:
                        current_username = user.get('username', '') or user.get('name', '')
                except:
                    pass
            
            if not current_username:
                current_username = 'supervisor'
            
            content = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                        size=lambda i, v: setattr(content_rect, 'size', v))

            content.add_widget(RTLLabel(
                text='گزارش به مدیر',
                size_hint_y=None, height=dp(40),
                font_size=sp(20), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            letter_text = self._build_manager_letter(visit)

            letter_input = RTLTextInput(
                text=letter_text,
                multiline=True,
                size_hint_y=0.65,
                font_size=sp(15),
                disabled=False
            )
            letter_input.bg_color = (0.15, 0.15, 0.15, 1)
            letter_input.border_color = (0.3, 0.3, 0.3, 1)
            letter_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(letter_input)

            content.add_widget(RTLLabel(
                text='متن قابل ویرایش و کپی می باشد',
                size_hint_y=None, height=dp(22),
                font_size=sp(11), color=(0.5, 0.5, 0.5, 1)
            ))

            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))

            mark_btn = PersianButton(
                text='علامت‌گذاری و خروجی اکسل',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.6, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(13)
            )
            btn_layout.add_widget(mark_btn)

            cancel_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.4, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.94, 0.88), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_mark(inst):
                from utils.supervisor_visits_manager import mark_visit_as_reported
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                from utils.storage import get_backup_path
                
                visit_id = visit.get('id', '')
                
                # ✅ علامت‌گذاری با نام کاربر
                success, message = mark_visit_as_reported(visit_id, reported_by=current_username)
                
                if success:
                    # ساخت فایل اکسل
                    try:
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "گزارش بازاری"
                        
                        # تنظیم راست‌چین
                        ws.sheet_view.rightToLeft = True
                        
                        # استایل‌ها
                        title_font = Font(bold=True, size=14)
                        header_font = Font(bold=True, size=11, color="FFFFFF")
                        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
                        text_font = Font(size=11)
                        bold_font = Font(bold=True, size=11)
                        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
                        thin_border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        
                        # عنوان
                        ws.merge_cells('A1:B1')
                        title_cell = ws.cell(row=1, column=1, value='بنام خدا')
                        title_cell.font = title_font
                        title_cell.alignment = center_align
                        
                        ws.merge_cells('A2:B2')
                        report_title = ws.cell(row=2, column=1, value=f'گزارش بررسی بازار - شماره: {visit_id}')
                        report_title.font = title_font
                        report_title.alignment = center_align
                        
                        # اطلاعات
                        row_num = 4
                        fields = [
                            ('تاریخ سرکشی', f"{visit.get('date', '')} - ساعت: {visit.get('time', '')}"),
                            ('مسیر', visit.get('route', '')),
                            ('مشتری', visit.get('customer', '')),
                            ('', ''),
                            ('نحوه سرکشی', visit.get('visit_type', '')),
                            ('علت سرکشی', visit.get('visit_reason', '')),
                            ('', ''),
                            ('وضعیت مشتری', visit.get('customer_status', '')),
                            ('وضعیت حضور در شلف', visit.get('shelf_status', '')),
                            ('تعداد سرکشی در ماه', visit.get('monthly_visits', '')),
                            ('کافی بودن سرکشی', visit.get('visit_sufficient', '')),
                            ('خرید مورد انتظار', visit.get('expected_purchase', '')),
                            ('وضعیت موجودی', visit.get('inventory_status', '')),
                            ('', ''),
                            ('برخورد بازاریاب', visit.get('agent_behavior', '')),
                            ('برخورد موزع', visit.get('distributor_behavior', '')),
                            ('رضایتمندی مشتری', visit.get('customer_satisfaction', '')),
                            ('تحقق هدف سرکشی', visit.get('target_achievement', '')),
                            ('', ''),
                            ('نیاز به پیگیری', visit.get('need_followup', '')),
                            ('تاریخ مراجعه بعدی', visit.get('next_visit_date', '---')),
                            ('', ''),
                            ('توضیحات سوپروایزر', visit.get('supervisor_note', '---')),
                            ('نظرات مشتری', visit.get('customer_feedback', '---')),
                            ('نظریه نهایی', visit.get('supervisor_opinion', '---')),
                        ]
                        
                        for label, value in fields:
                            if label == '':
                                row_num += 1
                                continue
                            if value:
                                # لیبل
                                cell_a = ws.cell(row=row_num, column=1, value=f'{label}:')
                                cell_a.font = bold_font
                                cell_a.alignment = right_align
                                cell_a.border = thin_border
                                # مقدار
                                cell_b = ws.cell(row=row_num, column=2, value=str(value))
                                cell_b.font = text_font
                                cell_b.alignment = right_align
                                cell_b.border = thin_border
                                row_num += 1
                        
                        # خط جداکننده
                        row_num += 1
                        ws.merge_cells(f'A{row_num}:B{row_num}')
                        sep_cell = ws.cell(row=row_num, column=1, value='─' * 50)
                        sep_cell.alignment = center_align
                        sep_cell.font = text_font
                        sep_cell.border = thin_border
                        
                        # تاریخ ثبت
                        row_num += 1
                        ws.merge_cells(f'A{row_num}:B{row_num}')
                        date_cell = ws.cell(row=row_num, column=1, value=f'تاریخ ثبت گزارش: {get_today_jalali()}')
                        date_cell.font = text_font
                        date_cell.alignment = center_align
                        date_cell.border = thin_border
                        
                        # نام ثبت کننده
                        row_num += 1
                        ws.merge_cells(f'A{row_num}:B{row_num}')
                        by_cell = ws.cell(row=row_num, column=1, value=f'ثبت شده توسط: {current_username}')
                        by_cell.font = text_font
                        by_cell.alignment = center_align
                        by_cell.border = thin_border
                        
                        # عرض ستون‌ها
                        ws.column_dimensions['A'].width = 25
                        ws.column_dimensions['B'].width = 50
                        
                        # ارتفاع ردیف‌ها
                        ws.row_dimensions[1].height = 30
                        ws.row_dimensions[2].height = 30
                        for r in range(4, row_num + 1):
                            ws.row_dimensions[r].height = 24
                        
                        # ذخیره
                        export_dir = get_backup_path()
                        os.makedirs(export_dir, exist_ok=True)
                        today = get_today_jalali().replace('/', '-')
                        filename = f'گزارش_بازاری_{visit_id}_{today}.xlsx'
                        filepath = os.path.join(export_dir, filename)
                        wb.save(filepath)
                        
                        message += f'\n\nفایل اکسل ذخیره شد:\n{filename}'
                        
                    except Exception as e:
                        message += f'\n\n(هشدار: فایل اکسل ساخته نشد - {str(e)})'
                
                popup.dismiss()
                self.show_message('موفق' if success else 'خطا', message)
                if success:
                    self._apply_market_filter(None)

            mark_btn.bind(on_press=do_mark)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش نامه: {e}", error_details)

    def _build_manager_letter(self, visit):
        """ساخت متن نامه اداری از جزئیات سرکشی"""
        visit_id = visit.get('id', '')
        date = visit.get('date', '')
        time = visit.get('time', '')
        route = visit.get('route', '')
        customer = visit.get('customer', '')
        visit_type = visit.get('visit_type', '')
        visit_reason = visit.get('visit_reason', '')
        customer_status = visit.get('customer_status', '')
        shelf_status = visit.get('shelf_status', '')
        monthly_visits = visit.get('monthly_visits', '')
        visit_sufficient = visit.get('visit_sufficient', '')
        expected_purchase = visit.get('expected_purchase', '')
        inventory_status = visit.get('inventory_status', '')
        agent_behavior = visit.get('agent_behavior', '')
        distributor_behavior = visit.get('distributor_behavior', '')
        customer_satisfaction = visit.get('customer_satisfaction', '')
        target_achievement = visit.get('target_achievement', '')
        need_followup = visit.get('need_followup', '')
        next_visit_date = visit.get('next_visit_date', '')
        supervisor_note = visit.get('supervisor_note', '')
        customer_feedback = visit.get('customer_feedback', '')
        supervisor_opinion = visit.get('supervisor_opinion', '')

        letter = f"""بنام خدا

گزارش بررسی بازار - شماره: {visit_id}

تاریخ سرکشی: {date} - ساعت: {time}
مسیر: {route}
مشتری: {customer}

نحوه سرکشی: {visit_type}
علت سرکشی: {visit_reason}

وضعیت مشتری: {customer_status}
وضعیت حضور در شلف: {shelf_status}
تعداد سرکشی بازاریابان در ماه: {monthly_visits}
آیا میزان سرکشی کافیست؟: {visit_sufficient}
میزان خرید مورد انتظار: {expected_purchase}
وضعیت موجودی مشتری: {inventory_status}

نحوه برخورد بازاریابان: {agent_behavior}
نحوه برخورد موزعین: {distributor_behavior}
میزان رضایتمندی مشتری: {customer_satisfaction}
میزان تحقق هدف سرکشی: {target_achievement}

نیاز به پیگیری مجدد: {need_followup}
تاریخ مراجعه بعدی: {next_visit_date if next_visit_date else '---'}

توضیحات سوپروایزر: {supervisor_note if supervisor_note else '---'}

نظرات مشتری: {customer_feedback if customer_feedback else '---'}

نظریه نهایی سوپروایزر: {supervisor_opinion if supervisor_opinion else '---'}

----------------------------------------------
این گزارش توسط سیستم مدیریت بازاریابی تهیه شده است.
تاریخ ثبت گزارش: {get_today_jalali()}"""

        return letter

    # ============================================================
    # تب ۴: آمار و ارزیابی
    # ============================================================

    def show_stats_evaluation_tab(self):
        """تب ۴: آمار و ارزیابی - مشابه ReportScreen"""
        try:
            self.eval_filter_agent = None
            self.eval_from_date = ''
            self.eval_to_date = ''
            self.current_evaluation_data = []

            main_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint=(1, 1), scroll_type=['bars', 'content'], bar_width=dp(8)
            )

            main_content = BoxLayout(
                orientation='vertical', spacing=dp(8),
                size_hint_y=None, padding=dp(10)
            )
            main_content.bind(minimum_height=main_content.setter('height'))

            # ========== عنوان ==========
            main_content.add_widget(RTLLabel(
                text='آمار و ارزیابی',
                size_hint_y=None, height=dp(45),
                font_size=sp(22), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            # ========== فیلترها ==========
            filter_box = BoxLayout(
                orientation='vertical', size_hint_y=None,
                height=dp(220), spacing=dp(5), padding=dp(8)
            )
            with filter_box.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                RoundedRectangle(pos=filter_box.pos, size=filter_box.size, radius=[dp(8)])

            filter_box.add_widget(RTLLabel(
                text='فیلترها:', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 0.8, 0.2, 1), bold=True
            ))

            # ردیف ۱: انتخاب عامل
            row1 = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(8))
            agents = get_agents()
            agent_names = ['همه'] + [a.get('name', '') for a in agents] if agents else ['همه']
            self.eval_filter_agent = PersianComboBox(
                text='همه', values=agent_names, height=dp(55)
            )
            self.eval_filter_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.eval_filter_agent.main_btn.color = (1, 1, 1, 1)
            self.eval_filter_agent.main_btn.font_size = sp(16)
            self.eval_filter_agent.size_hint_x = 0.88
            row1.add_widget(RTLLabel(
                text='عامل:', size_hint_x=0.12, size_hint_y=None, height=dp(55),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            row1.add_widget(self.eval_filter_agent)
            filter_box.add_widget(row1)

            # ردیف ۲: از تاریخ + تا تاریخ
            row2 = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(8))
            today = get_today_jalali()
            first_day = self._get_first_day_of_month()

            self.eval_from_input = RTLTextInput(
                text=first_day, hint_text='از تاریخ', multiline=False,
                size_hint_x=0.44, size_hint_y=None, height=dp(55), font_size=sp(18)
            )
            self.eval_from_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.eval_from_input.border_color = (0.3, 0.3, 0.3, 1)
            self.eval_from_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.eval_from_input._hidden_input.foreground_color = (1, 1, 1, 1)
            row2.add_widget(RTLLabel(
                text='از:', size_hint_x=0.06, size_hint_y=None, height=dp(55),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            row2.add_widget(self.eval_from_input)

            self.eval_to_input = RTLTextInput(
                text=today, hint_text='تا تاریخ', multiline=False,
                size_hint_x=0.44, size_hint_y=None, height=dp(55), font_size=sp(18)
            )
            self.eval_to_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.eval_to_input.border_color = (0.3, 0.3, 0.3, 1)
            self.eval_to_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.eval_to_input._hidden_input.foreground_color = (1, 1, 1, 1)
            row2.add_widget(RTLLabel(
                text='تا:', size_hint_x=0.06, size_hint_y=None, height=dp(55),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            row2.add_widget(self.eval_to_input)
            filter_box.add_widget(row2)

            # ردیف ۳: دکمه‌ها
            btn_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
            apply_btn = PersianButton(
                text='اعمال فیلتر', background_color=(0.2, 0.6, 0.2, 1),
                size_hint_x=0.35, size_hint_y=None, height=dp(38),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            apply_btn.bind(on_press=self._apply_eval_filter)
            btn_row.add_widget(apply_btn)
            excel_btn = PersianButton(
                text='خروجی اکسل', background_color=(0.2, 0.7, 0.4, 1),
                size_hint_x=0.35, size_hint_y=None, height=dp(38),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            excel_btn.bind(on_press=self._export_eval_report)
            btn_row.add_widget(excel_btn)
            clear_btn = PersianButton(
                text='پاک کردن', background_color=(0.8, 0.4, 0.1, 1),
                size_hint_x=0.30, size_hint_y=None, height=dp(38),
                color=(1, 1, 1, 1), font_size=sp(14)
            )
            clear_btn.bind(on_press=self._clear_eval_filter)
            btn_row.add_widget(clear_btn)
            filter_box.add_widget(btn_row)

            main_content.add_widget(filter_box)

            # ========== محتوای ارزیابی ==========
            self.eval_content = BoxLayout(
                orientation='vertical', spacing=dp(8),
                size_hint_y=None, padding=dp(5)
            )
            self.eval_content.bind(minimum_height=self.eval_content.setter('height'))
            main_content.add_widget(self.eval_content)

            main_scroll.add_widget(main_content)
            self.content_area.add_widget(main_scroll)

            self._apply_eval_filter(None, initial_load=True)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب آمار و ارزیابی: {e}", error_details)

    def _get_first_day_of_month(self):
        """دریافت اولین روز ماه جاری"""
        try:
            today = get_today_jalali()
            parts = today.split('/')
            if len(parts) == 3:
                return f"{parts[0]}/{parts[1]}/01"
            return today
        except:
            return get_today_jalali()

    def _apply_eval_filter(self, instance, initial_load=False):
        """اعمال فیلتر و نمایش ارزیابی"""
        try:
            from utils.file_manager import get_daily_logs

            agent = self.eval_filter_agent.text if hasattr(self, 'eval_filter_agent') and self.eval_filter_agent and self.eval_filter_agent.text != 'همه' else None
            from_date = self.eval_from_input.text.strip() if hasattr(self, 'eval_from_input') else ''
            to_date = self.eval_to_input.text.strip() if hasattr(self, 'eval_to_input') else ''

            if not from_date:
                from_date = self._get_first_day_of_month()
            if not to_date:
                to_date = get_today_jalali()

            all_logs = get_daily_logs()
            settings = get_settings()
            
            # ساخت mapping: route -> agent_name
            agents = get_agents()
            route_agent_map = {}
            for a in agents:
                if isinstance(a, dict):
                    agent_routes = a.get('routes', [])
                    if isinstance(agent_routes, list):
                        for r in agent_routes:
                            route_agent_map[r] = a.get('name', '')

            # فیلتر تاریخ
            date_list = []
            for date in all_logs.keys():
                if from_date and date < from_date:
                    continue
                if to_date and date > to_date:
                    continue
                date_list.append(date)

            if not date_list:
                self.eval_content.clear_widgets()
                self.eval_content.add_widget(RTLLabel(
                    text='هیچ داده‌ای در بازه انتخابی یافت نشد',
                    size_hint_y=None, height=dp(50),
                    font_size=sp(18), color=(0.5, 0.5, 0.5, 1)
                ))
                return

            # محاسبه آمار
            supervision_rate = settings.get('supervision_rate', 70.0) / 100
            conversion_rate = settings.get('conversion_rate', 75.0) / 100
            target_units = settings.get('target_count', 100)
            target_sales = settings.get('target_amount', 50000000)
            target_cash = settings.get('target_cash_sales', 30000000)
            target_check = settings.get('target_credit_sales', 20000000)

            total_visits = 0
            total_invoices = 0
            total_units = 0
            total_sales = 0
            total_cash = 0
            total_check = 0
            total_new_customers = 0
            agents_data = {}

            for date in date_list:
                if date not in all_logs or not isinstance(all_logs[date], list):
                    continue
                for log in all_logs[date]:
                    if not isinstance(log, dict):
                        continue

                    log_route = log.get('route', '')
                    log_agent = log.get('agent_name', '')
                    
                    # اگر agent_name نداره، از route تشخیص بده
                    if not log_agent and log_route:
                        log_agent = route_agent_map.get(log_route, log_route)
                    
                    if agent and log_agent != agent:
                        continue

                    visit_status = log.get('visit_status', '')
                    sales_status = log.get('sales_status', '')
                    payment_method = log.get('payment_method', '')
                    sales_amount = log.get('sales_amount', 0)
                    units_sold = log.get('units_sold', 0)

                    if log_agent not in agents_data:
                        agents_data[log_agent] = {
                            'visits': 0, 'invoices': 0, 'units': 0,
                            'sales': 0, 'cash': 0, 'check': 0, 'new_customers': 0
                        }

                    if visit_status == 'موفق':
                        total_visits += 1
                        agents_data[log_agent]['visits'] += 1
                    if sales_status == 'موفق':
                        total_invoices += 1
                        total_units += units_sold
                        total_sales += sales_amount
                        agents_data[log_agent]['invoices'] += 1
                        agents_data[log_agent]['units'] += units_sold
                        agents_data[log_agent]['sales'] += sales_amount
                        if payment_method == 'نقد':
                            total_cash += sales_amount
                            agents_data[log_agent]['cash'] += sales_amount
                        elif payment_method == 'چک':
                            total_check += sales_amount
                            agents_data[log_agent]['check'] += sales_amount
                    if log.get('is_new_customer', False):
                        total_new_customers += 1
                        agents_data[log_agent]['new_customers'] += 1

            day_count = len(date_list)
            target_visits_day = int(supervision_rate * 50) * day_count
            target_invoices_day = int(target_visits_day * conversion_rate)
            target_units_day = target_units * day_count
            target_sales_day = target_sales * day_count
            target_cash_day = target_cash * day_count
            target_check_day = target_check * day_count

            self.eval_content.clear_widgets()

            # ========== کارت‌های آماری ==========
            stats_box = BoxLayout(size_hint_y=None, height=dp(85), spacing=dp(6), padding=dp(5))
            stats_data = [
                ('روزهای کاری', f'{day_count}', (0.2, 0.5, 0.8, 1)),
                ('کل ویزیت‌ها', f'{total_visits:,}', (0.3, 0.6, 1, 1)),
                ('فاکتورها', f'{total_invoices:,}', (0.2, 0.7, 0.2, 1)),
                ('واحد فروش', f'{total_units:,}', (0.5, 0.3, 0.7, 1)),
            ]
            for title, value, color in stats_data:
                card = BoxLayout(orientation='vertical', size_hint_x=0.25, size_hint_y=None, height=dp(80), padding=dp(4), spacing=dp(2))
                with card.canvas.before:
                    Color(*color)
                    RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(6)])
                card.add_widget(RTLLabel(text=title, size_hint_y=None, height=dp(22), font_size=sp(10), color=(1, 1, 1, 1)))
                card.add_widget(RTLLabel(text=value, size_hint_y=None, height=dp(35), font_size=sp(22), bold=True, color=(1, 1, 1, 1)))
                stats_box.add_widget(card)
            self.eval_content.add_widget(stats_box)

            stats_box2 = BoxLayout(size_hint_y=None, height=dp(85), spacing=dp(6), padding=dp(5))
            stats_data2 = [
                ('کل فروش', f'{total_sales:,}', (0.2, 0.6, 0.3, 1)),
                ('فروش نقدی', f'{total_cash:,}', (0.2, 0.5, 0.8, 1)),
                ('فروش چکی', f'{total_check:,}', (0.6, 0.3, 0.6, 1)),
                ('مشتری جدید', f'{total_new_customers:,}', (0.2, 0.8, 0.4, 1)),
            ]
            for title, value, color in stats_data2:
                card = BoxLayout(orientation='vertical', size_hint_x=0.25, size_hint_y=None, height=dp(80), padding=dp(4), spacing=dp(2))
                with card.canvas.before:
                    Color(*color)
                    RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(6)])
                card.add_widget(RTLLabel(text=title, size_hint_y=None, height=dp(22), font_size=sp(10), color=(1, 1, 1, 1)))
                card.add_widget(RTLLabel(text=value, size_hint_y=None, height=dp(35), font_size=sp(22), bold=True, color=(1, 1, 1, 1)))
                stats_box2.add_widget(card)
            self.eval_content.add_widget(stats_box2)

            # ========== جدول ارزیابی ==========
            self.eval_content.add_widget(RTLLabel(
                text='ارزیابی عملکرد',
                size_hint_y=None, height=dp(35),
                font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            header = BoxLayout(size_hint_y=None, height=dp(32), spacing=dp(3))
            for text, size in [('آیتم', 0.28), ('هدف', 0.24), ('عملکرد', 0.24), ('نتیجه', 0.24)]:
                header.add_widget(RTLLabel(
                    text=text, size_hint_x=size, size_hint_y=None, height=dp(30),
                    font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                ))
            self.eval_content.add_widget(header)

            items = [
                ('تعداد ویزیت', target_visits_day, total_visits),
                ('تعداد فاکتور', target_invoices_day, total_invoices),
                ('واحد فروش', target_units_day, total_units),
                ('مبلغ فروش', target_sales_day, total_sales),
                ('فروش نقدی', target_cash_day, total_cash),
                ('فروش چکی', target_check_day, total_check),
            ]

            total_percent = 0
            item_count = 0

            for name, target_val, actual_val in items:
                diff = actual_val - target_val
                diff_str = f'{diff:+,}'
                diff_color = (0.2, 0.8, 0.2, 1) if diff >= 0 else (0.8, 0.3, 0.3, 1)

                if target_val > 0:
                    percent = (actual_val / target_val) * 100
                    total_percent += percent
                    item_count += 1

                row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(3))
                row.add_widget(RTLLabel(text=name, size_hint_x=0.28, size_hint_y=None, height=dp(28), font_size=sp(13), color=(1, 1, 1, 1)))
                row.add_widget(RTLLabel(text=f'{target_val:,}', size_hint_x=0.24, size_hint_y=None, height=dp(28), font_size=sp(13), color=(1, 1, 1, 1), halign='center'))
                row.add_widget(RTLLabel(text=f'{actual_val:,}', size_hint_x=0.24, size_hint_y=None, height=dp(28), font_size=sp(13), color=(1, 1, 1, 1), halign='center'))
                row.add_widget(RTLLabel(text=diff_str, size_hint_x=0.24, size_hint_y=None, height=dp(28), font_size=sp(13), color=diff_color, halign='center'))
                self.eval_content.add_widget(row)

            avg_percent = total_percent / item_count if item_count > 0 else 0
            
            self.current_evaluation_data = {
                'items': items,
                'avg_percent': avg_percent,
                'agents_data': agents_data,
                'from_date': from_date,
                'to_date': to_date,
                'day_count': day_count
            }
            
            if avg_percent >= 100:
                eval_text = "عملکرد عالی"
                eval_color = (0.2, 0.8, 0.2, 1)
            elif avg_percent >= 70:
                eval_text = "عملکرد خوب"
                eval_color = (0.3, 0.6, 1, 1)
            elif avg_percent >= 50:
                eval_text = "نیاز به تلاش بیشتر"
                eval_color = (1, 0.7, 0, 1)
            else:
                eval_text = "ضعیف - نیاز به بررسی"
                eval_color = (0.8, 0.3, 0.3, 1)

            self.eval_content.add_widget(RTLLabel(
                text=f'میانگین تحقق: {avg_percent:.1f}% - {eval_text}',
                size_hint_y=None, height=dp(40),
                font_size=sp(18), bold=True, color=eval_color
            ))

            # ========== جدول عملکرد عامل‌ها ==========
            if agents_data and not agent:
                self.eval_content.add_widget(RTLLabel(
                    text='عملکرد تفکیکی عامل‌ها',
                    size_hint_y=None, height=dp(35),
                    font_size=sp(18), bold=True, color=(1, 0.5, 0, 1)
                ))

                agent_header = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(3))
                for text, size in [('عامل', 0.22), ('ویزیت', 0.16), ('فاکتور', 0.16), ('فروش', 0.23), ('امتیاز', 0.23)]:
                    agent_header.add_widget(RTLLabel(
                        text=text, size_hint_x=size, size_hint_y=None, height=dp(28),
                        font_size=sp(12), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                    ))
                self.eval_content.add_widget(agent_header)

                for ag_name, ag_data in agents_data.items():
                    ag_row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(3))
                    ag_row.add_widget(RTLLabel(text=ag_name, size_hint_x=0.22, size_hint_y=None, height=dp(26), font_size=sp(12), color=(1, 1, 1, 1)))
                    ag_row.add_widget(RTLLabel(text=str(ag_data['visits']), size_hint_x=0.16, size_hint_y=None, height=dp(26), font_size=sp(12), color=(1, 1, 1, 1), halign='center'))
                    ag_row.add_widget(RTLLabel(text=str(ag_data['invoices']), size_hint_x=0.16, size_hint_y=None, height=dp(26), font_size=sp(12), color=(1, 1, 1, 1), halign='center'))
                    ag_row.add_widget(RTLLabel(text=f'{ag_data["sales"]:,}', size_hint_x=0.23, size_hint_y=None, height=dp(26), font_size=sp(12), color=(1, 1, 1, 1), halign='center'))

                    ag_percent = (ag_data['visits'] / max(target_visits_day / max(len(agents_data), 1), 1)) * 100
                    ag_color = (0.2, 0.8, 0.2, 1) if ag_percent >= 70 else (1, 0.7, 0, 1) if ag_percent >= 50 else (0.8, 0.3, 0.3, 1)
                    ag_row.add_widget(RTLLabel(text=f'{ag_percent:.0f}%', size_hint_x=0.23, size_hint_y=None, height=dp(26), font_size=sp(12), color=ag_color, halign='center'))
                    self.eval_content.add_widget(ag_row)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال فیلتر ارزیابی: {e}", error_details)

    def _clear_eval_filter(self, instance):
        """پاک کردن فیلترهای تب ارزیابی"""
        if hasattr(self, 'eval_filter_agent') and self.eval_filter_agent:
            self.eval_filter_agent.text = 'همه'
        if hasattr(self, 'eval_from_input') and self.eval_from_input:
            self.eval_from_input.text = self._get_first_day_of_month()
        if hasattr(self, 'eval_to_input') and self.eval_to_input:
            self.eval_to_input.text = get_today_jalali()
        self._apply_eval_filter(None)

    def _export_eval_report(self, instance):
        """خروجی اکسل از داده‌های ارزیابی"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from utils.storage import get_backup_path

            if not hasattr(self, 'current_evaluation_data') or not self.current_evaluation_data:
                self.show_message('خطا', 'ابتدا فیلتر را اعمال کنید')
                return

            data = self.current_evaluation_data
            items = data.get('items', [])
            avg_percent = data.get('avg_percent', 0)
            agents_data = data.get('agents_data', {})
            from_date = data.get('from_date', '')
            to_date = data.get('to_date', '')
            day_count = data.get('day_count', 0)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "آمار و ارزیابی"
            ws.sheet_view.rightToLeft = True

            # استایل‌ها
            header_font = Font(name='B Nazanin', size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            text_font = Font(name='B Nazanin', size=11)
            bold_font = Font(name='B Nazanin', size=11, bold=True)
            green_font = Font(name='B Nazanin', size=11, color="008000")
            red_font = Font(name='B Nazanin', size=11, color="CC0000")
            orange_font = Font(name='B Nazanin', size=11, color="CC8800")
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # عنوان
            ws.merge_cells('A1:F1')
            title_cell = ws.cell(row=1, column=1, value=f'گزارش آمار و ارزیابی ({from_date} تا {to_date}) - {day_count} روز کاری')
            title_cell.font = Font(name='B Nazanin', size=16, bold=True)
            title_cell.alignment = center_align

            # هدر جدول
            headers = ['آیتم', 'هدف', 'عملکرد', 'اختلاف', 'درصد', 'وضعیت']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # داده‌ها
            for row_idx, item in enumerate(items, 4):
                name = item[0]
                target_val = item[1]
                actual_val = item[2]
                
                diff = actual_val - target_val
                diff_str = f'{diff:+,}'
                percent = (actual_val / target_val * 100) if target_val > 0 else 0
                
                if percent >= 70:
                    status = 'مطلوب'
                    status_font = green_font
                elif percent >= 50:
                    status = 'متوسط'
                    status_font = orange_font
                else:
                    status = 'ضعیف'
                    status_font = red_font

                values = [
                    name,
                    target_val,
                    actual_val,
                    diff,
                    percent,
                    status
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = text_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # فرمت‌بندی اعداد
                    if col_idx in [2, 3]:
                        cell.number_format = '#,##0'
                        cell.value = int(value) if isinstance(value, (int, float)) else value
                    elif col_idx == 4:
                        cell.number_format = '#,##0'
                        cell.value = int(value) if isinstance(value, (int, float)) else value
                        if diff < 0:
                            cell.font = red_font
                        elif diff >= 0:
                            cell.font = green_font
                    elif col_idx == 5:
                        cell.number_format = '0.0"%"'
                    elif col_idx == 6:
                        cell.font = status_font

            # میانگین
            avg_row = len(items) + 4
            ws.merge_cells(f'A{avg_row}:C{avg_row}')
            avg_label = ws.cell(row=avg_row, column=1, value='میانگین تحقق')
            avg_label.font = bold_font
            avg_label.alignment = center_align
            avg_label.border = thin_border

            ws.merge_cells(f'D{avg_row}:F{avg_row}')
            avg_value = ws.cell(row=avg_row, column=4, value=f'{avg_percent:.1f}%')
            avg_value.font = Font(name='B Nazanin', size=13, bold=True, color="008000" if avg_percent >= 70 else "CC0000")
            avg_value.alignment = center_align
            avg_value.border = thin_border

            # ارزیابی کلی
            eval_row = avg_row + 1
            ws.merge_cells(f'A{eval_row}:F{eval_row}')
            if avg_percent >= 100:
                eval_text = "عملکرد عالی - تبریک!"
            elif avg_percent >= 70:
                eval_text = "عملکرد خوب - در مسیر درست"
            elif avg_percent >= 50:
                eval_text = "نیاز به تلاش بیشتر"
            else:
                eval_text = "ضعیف - نیاز به بررسی و پیگیری"
            
            eval_cell = ws.cell(row=eval_row, column=1, value=eval_text)
            eval_cell.font = Font(name='B Nazanin', size=14, bold=True)
            eval_cell.alignment = center_align

            # جدول تفکیکی عامل‌ها
            if agents_data and len(agents_data) > 1:
                agent_start = eval_row + 2
                ws.merge_cells(f'A{agent_start}:F{agent_start}')
                agent_title = ws.cell(row=agent_start, column=1, value='عملکرد تفکیکی عامل‌ها')
                agent_title.font = Font(name='B Nazanin', size=14, bold=True)
                agent_title.alignment = center_align

                agent_headers = ['عامل', 'ویزیت', 'فاکتور', 'فروش', 'نقدی', 'چکی']
                header_row = agent_start + 1
                for col, header in enumerate(agent_headers, 1):
                    cell = ws.cell(row=header_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                for i, (ag_name, ag_data) in enumerate(agents_data.items()):
                    row = header_row + 1 + i
                    values = [
                        ag_name,
                        ag_data['visits'],
                        ag_data['invoices'],
                        ag_data['sales'],
                        ag_data['cash'],
                        ag_data['check']
                    ]
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = text_font
                        cell.alignment = center_align
                        cell.border = thin_border
                        if col >= 2:
                            cell.number_format = '#,##0'

            # عرض ستون‌ها
            column_widths = [28, 18, 18, 18, 14, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # ذخیره
            filename = f'ارزیابی_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            filepath = os.path.join(export_dir, filename)
            wb.save(filepath)

            self.show_message('موفق', f'فایل اکسل ذخیره شد:\n{filename}')

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خروجی اکسل: {e}", error_details)

    # ============================================================
    # توابع عمومی
    # ============================================================



    def show_message(self, title, message):
        """نمایش پیام با پشتیبانی از متن طولانی"""
        try:
            from kivy.uix.boxlayout import BoxLayout
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            from kivy.metrics import dp, sp
            from kivy.graphics import Color, Rectangle
            from utils.rtl_widgets import PersianPopup, PersianButton
            
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(message)
                display_text = get_display(reshaped)
            except:
                display_text = message
            
            msg_label = Label(
                text=display_text,
                font_size=sp(15),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                halign='center',
                valign='top',
                text_size=(dp(550), None),
                font_name='fonts/Amiri-Regular.ttf',
                padding=(dp(20), 0, dp(0), 0)
            )
            
            lines = message.count('\n') + 1
            if len(message) > 50 and lines == 1:
                approx_chars_per_line = 35
                lines = (len(message) // approx_chars_per_line) + 1
            
            line_height = sp(15) + dp(6)
            label_height = max(lines * line_height + dp(25), dp(50))
            label_height = min(label_height, dp(490))
            
            msg_label.height = label_height
            msg_label.text_size = (dp(550), label_height)
            
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=None,
                height=label_height + dp(10),
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )
            scroll.add_widget(msg_label)
            content.add_widget(scroll)
            
            btn = PersianButton(
                text='باشه',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(16),
                color=(1, 1, 1, 1),
                background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            
            popup = PersianPopup(
                title=title,
                content=content,
                size_hint=(0.9, None),
                height=label_height + dp(250),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            print(f"خطا در نمایش پیام: {e}")
            import traceback
            traceback.print_exc()
            try:
                from error_handler import ErrorPopup
                ErrorPopup.show_error(str(message))
            except:
                pass
            
    def go_back(self, instance):
        """بازگشت به صفحه سوپروایزر"""
        if self.manager:
            self.manager.current = 'supervisor'
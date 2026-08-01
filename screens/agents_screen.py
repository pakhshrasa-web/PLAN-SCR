# screens/agents_screen.py
# ========== صفحه ثبت ویزیت بازاریابان با لیست مشتریان ==========

import os         
import traceback
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.clock import Clock
from kivy.graphics import Color, Rectangle, RoundedRectangle
from kivy.core.window import Window
from kivy.app import App

from utils.rtl_widgets import RTLTextInput, PersianComboBox, PersianButton, RTLLabel, PersianPopup
from utils.file_manager import get_routes, get_customers, get_settings, save_daily_log, get_daily_logs, add_customer
from utils.jalali_date import get_today_jalali, get_current_time
from error_handler import ErrorPopup


class AgentsScreen(Screen):
    def __init__(self, **kwargs):
        try:
            super().__init__(**kwargs)
            with self.canvas.before:
                Color(0.08, 0.08, 0.08, 1)
                self.bg_rect = Rectangle(pos=self.pos, size=self.size)
                self.bind(pos=self._update_bg, size=self._update_bg)
            
            Window.softinput_mode = 'resize'
            self.focusable_fields = []
            
            self.amount_words_label = None
            self._last_reason_text = ''
            self._last_route_text = ''
            self.locked_route = None
            self.route_confirmed = False
            self.session_new_customers = []
            
            self.customers_list_container = GridLayout(
                cols=1, spacing=dp(6), size_hint_y=None, padding=dp(5)
            )
            self.customers_list_container.height = dp(50)
            
            self.settings = get_settings()
            self.selected_customer = None
            self.selected_route = None
            
            # متغیرهای دیالوگ تحقق
            self.fulfillment_file_path = None
            self.fulfillment_data = []
            
            # متغیرهای ریز فروش
            self._detailed_sales_data = []
            self._detailed_sales_total = 0
            
            self.build_ui()
            
            Window.bind(on_keyboard=self._on_keyboard)
            Clock.schedule_once(self._check_today_visits, 0.5)
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ساخت AgentsScreen: {e}", error_details)
            raise

    def number_to_persian_words(self, number):
        try:
            if number == 0:
                return "صفر ریال"
            ones = ['', 'یک', 'دو', 'سه', 'چهار', 'پنج', 'شش', 'هفت', 'هشت', 'نه']
            tens = ['', 'ده', 'بیست', 'سی', 'چهل', 'پنجاه', 'شصت', 'هفتاد', 'هشتاد', 'نود']
            hundreds = ['', 'یکصد', 'دویست', 'سیصد', 'چهارصد', 'پانصد', 'ششصد', 'هفتصد', 'هشتصد', 'نهصد']
            groups = ['', 'هزار', 'میلیون', 'میلیارد']
            
            def convert_three_digits(num):
                if num == 0:
                    return ''
                h = num // 100
                t = (num % 100) // 10
                o = num % 10
                result = []
                if h > 0:
                    result.append(hundreds[h])
                if t == 1:
                    if o == 0: result.append('ده')
                    elif o == 1: result.append('یازده')
                    elif o == 2: result.append('دوازده')
                    elif o == 3: result.append('سیزده')
                    elif o == 4: result.append('چهارده')
                    elif o == 5: result.append('پانزده')
                    elif o == 6: result.append('شانزده')
                    elif o == 7: result.append('هفده')
                    elif o == 8: result.append('هجده')
                    elif o == 9: result.append('نوزده')
                else:
                    if t > 0: result.append(tens[t])
                    if o > 0: result.append(ones[o])
                return ' و '.join(result)
            
            num_str = str(number)
            group_list = []
            for i in range(len(num_str), 0, -3):
                start = max(0, i - 3)
                group = int(num_str[start:i])
                group_list.insert(0, group)
            
            result_parts = []
            for i, group in enumerate(group_list):
                if group == 0: continue
                group_words = convert_three_digits(group)
                if group_words:
                    group_name = groups[len(group_list) - 1 - i]
                    if group_name: result_parts.append(f"{group_words} {group_name}")
                    else: result_parts.append(group_words)
            
            return " و ".join(result_parts) + " ریال"
        except Exception as e:
            print(f"خطا در تبدیل عدد به حروف: {e}")
            return f"{number:,} ریال"
    
    def _update_bg(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size
    
    def _on_field_focus(self, instance, value):
        if value:
            Clock.schedule_once(lambda dt: self._select_all_text(instance), 0.1)
            Clock.schedule_once(lambda dt: self._scroll_to_field(instance), 0.3)
    
    def _select_all_text(self, instance):
        if instance and hasattr(instance, 'select_all'):
            instance.select_all()
    
    def _scroll_to_field(self, instance):
        try:
            scroll = None
            for child in self.children:
                if isinstance(child, ScrollView):
                    scroll = child
                    break
            if not scroll:
                for child in self.children:
                    if hasattr(child, 'children'):
                        for sub in child.children:
                            if isinstance(sub, ScrollView):
                                scroll = sub
                                break
            if not scroll:
                return
            
            field_pos = instance.to_window(0, 0)
            field_y = field_pos[1]
            keyboard_height = 250
            window_height = Window.height
            target_y = window_height - keyboard_height - dp(80)
            content_height = scroll.children[0].height if scroll.children else 1
            scroll_height = scroll.height
            if content_height > scroll_height:
                if field_y > target_y:
                    field_ratio = (content_height - field_y) / content_height
                    scroll.scroll_y = min(0.95, max(0.05, field_ratio + 0.1))
                elif field_y < dp(50):
                    scroll.scroll_y = 0.9
        except Exception as e:
            print(f"خطا در اسکرول به فیلد: {e}")
    
    def _on_keyboard(self, window, key, *args):
        if key == 9:
            self._focus_next()
            return True
        return False
    
    def _focus_next(self):
        if not self.focusable_fields: return
        for i, field in enumerate(self.focusable_fields):
            if field.focus:
                next_i = (i + 1) % len(self.focusable_fields)
                self.focusable_fields[next_i].focus = True
                break
    
    def _check_today_visits(self, dt):
        try:
            today = get_today_jalali()
            logs = get_daily_logs()
            if today in logs and logs[today] and len(logs[today]) > 0:
                if hasattr(self, 'route_spinner'):
                    last_visit = logs[today][-1]
                    locked_route = last_visit.get('route', '')
                    if locked_route:
                        self.locked_route = locked_route
                        self.route_spinner.text = locked_route
                        self.route_spinner.main_btn.disabled = True
                        self.route_spinner.main_btn.background_color = (0.15, 0.15, 0.15, 1)
                        self.route_spinner.main_btn.color = (0.6, 0.6, 0.6, 1)
                        self.route_confirmed = True
                        Clock.schedule_once(lambda dt: self.update_customers_list(), 0.3)
                self.show_message('اطلاع', 'برای امروز ویزیت ثبت شده و مسیر قفل است.')
        except Exception as e:
            print(f"خطا در بررسی ویزیت‌های امروز: {e}")
    
    def build_ui(self):
        try:
            main_layout = BoxLayout(orientation='vertical')
            scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint=(1, 1), scroll_type=['bars', 'content'], bar_width=dp(8)
            )
            content = BoxLayout(
                orientation='vertical', padding=dp(15),
                spacing=dp(8), size_hint_y=None
            )
            content.bind(minimum_height=content.setter('height'))
            
            # ========== عنوان ==========
            content.add_widget(RTLLabel(
                text='ثبت ویزیت بازاریابان', font_size=sp(22),
                size_hint_y=None, height=dp(50), color=(1, 1, 1, 1), bold=True
            ))
            content.add_widget(Label(size_hint_y=None, height=dp(5)))
            
            # ========== تاریخ ==========
            content.add_widget(RTLLabel(
                text='تاریخ:', size_hint_y=None, height=dp(30),
                font_size=sp(16), color=(0.4, 0.7, 1, 1), bold=True
            ))
            self.date_label = RTLLabel(
                text=get_today_jalali(), size_hint_y=None,
                height=dp(40), font_size=sp(20), color=(1, 1, 1, 1)
            )
            content.add_widget(self.date_label)
            content.add_widget(Label(size_hint_y=None, height=dp(5)))
            
            # ========== ساعت ==========
            content.add_widget(RTLLabel(
                text='ساعت:', size_hint_y=None, height=dp(30),
                font_size=sp(16), color=(0.4, 0.7, 1, 1), bold=True
            ))
            self.time_label = RTLLabel(
                text=get_current_time(), size_hint_y=None,
                height=dp(40), font_size=sp(20), color=(1, 1, 1, 1)
            )
            content.add_widget(self.time_label)
            content.add_widget(Label(size_hint_y=None, height=dp(5)))
            
            # ========== مسیر ==========
            content.add_widget(RTLLabel(
                text='انتخاب مسیر:', size_hint_y=None, height=dp(30),
                font_size=sp(16), color=(0.4, 0.7, 1, 1), bold=True
            ))
            routes = get_routes()
            route_names = [r.get('name', '') for r in routes] if routes else ['']
            self.route_spinner = PersianComboBox(text='', values=route_names, height=dp(70))
            self.route_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.route_spinner.main_btn.color = (1, 1, 1, 1)
            self.route_spinner.main_btn.font_size = sp(18)
            
            today = get_today_jalali()
            logs = get_daily_logs()
            if self.locked_route and today in logs and logs[today] and len(logs[today]) > 0:
                if self.locked_route in route_names:
                    self.route_spinner.text = self.locked_route
                    self.route_spinner.main_btn.disabled = True
                    self.route_spinner.main_btn.background_color = (0.15, 0.15, 0.15, 1)
                    self.route_spinner.main_btn.color = (0.6, 0.6, 0.6, 1)
                    self.route_confirmed = True
            
            self._last_route_text = self.route_spinner.text
            Clock.schedule_interval(self._check_route_change_with_confirm, 0.3)
            content.add_widget(self.route_spinner)
            content.add_widget(Label(size_hint_y=None, height=dp(5)))
            
            # ========== دکمه افزودن مشتری ==========
            add_customer_btn = PersianButton(
                text='افزودن مشتری جدید', background_color=(0.2, 0.6, 0.2, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            add_customer_btn.bind(on_press=self.show_add_customer_dialog)
            content.add_widget(add_customer_btn)
            
            # ========== دکمه انتخاب مشتری ==========
            select_customer_btn = PersianButton(
                text='انتخاب مشتری', background_color=(0.2, 0.5, 0.9, 1),
                size_hint_y=None, height=dp(55), color=(1, 1, 1, 1),
                font_size=sp(20), bold=True
            )
            select_customer_btn.bind(on_press=self.show_customer_selection_dialog)
            content.add_widget(select_customer_btn)
            
            # ========== نمایش مشتری انتخاب شده ==========
            self.selected_customer_label = RTLLabel(
                text='مشتری انتخاب شده: هیچ', size_hint_y=None,
                height=dp(40), font_size=sp(18), color=(0.5, 0.5, 0.5, 1), bold=True
            )
            content.add_widget(self.selected_customer_label)
            
            # ========== دکمه تحقق ریزتارگت‌ها ==========
            fulfillment_btn = PersianButton(
                text='تحقق ریزتارگت‌ها', background_color=(1, 1, 1, 1),
                color=(1, 0.5, 0, 1), size_hint_y=None, height=dp(50),
                font_size=sp(18), bold=True
            )
            fulfillment_btn.bind(on_press=self.show_target_fulfillment_dialog)
            content.add_widget(fulfillment_btn)

            # ========== دکمه بروزرسانی کالاها ==========
            update_products_btn = PersianButton(
                text='بروزرسانی کالاها', background_color=(0.2, 0.5, 0.7, 1),
                color=(1, 1, 1, 1), size_hint_y=None, height=dp(45), font_size=sp(16)
            )
            update_products_btn.bind(on_press=self.show_update_products_dialog)
            content.add_widget(update_products_btn)
            
            # ========== دکمه بازگشت ==========
            back_btn = PersianButton(
                text='بازگشت', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            back_btn.bind(on_press=self.go_back)
            content.add_widget(back_btn)
            
            scroll.add_widget(content)
            main_layout.add_widget(scroll)
            self.add_widget(main_layout)
            
            Clock.schedule_interval(self.update_time, 60)
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ساخت UI AgentsScreen: {e}", error_details)
            raise
    
    def _check_route_change_with_confirm(self, dt):
        if hasattr(self, 'route_spinner') and not self.route_spinner.main_btn.disabled:
            current_text = self.route_spinner.text
            if current_text != self._last_route_text and current_text and current_text != '':
                self._last_route_text = current_text
                self.show_route_confirm_dialog(current_text)
    
    def update_time(self, dt):
        self.time_label.text = get_current_time()
    
    def update_customers_list(self):
        try:
            if not hasattr(self, 'customers_list_container') or not self.customers_list_container.parent:
                return
            self.customers_list_container.clear_widgets()
            selected_route = self.route_spinner.text
            if not selected_route:
                self.customers_list_container.add_widget(RTLLabel(
                    text='لطفاً ابتدا یک مسیر انتخاب کنید', size_hint_y=None,
                    height=dp(40), font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
                ))
                self.customers_list_container.height = dp(40)
                return
            
            all_customers = get_customers()
            search_text = self.search_input.text.strip() if hasattr(self, 'search_input') else ''
            filtered = []
            for c in all_customers:
                route_name = c.get('route_name', '').strip()
                customer_name = c.get('name', '')
                if route_name == selected_route.strip():
                    if search_text:
                        if search_text in customer_name: filtered.append(customer_name)
                    else:
                        filtered.append(customer_name)
            
            if not filtered:
                self.customers_list_container.add_widget(RTLLabel(
                    text='هیچ مشتری‌ای در این مسیر یافت نشد', size_hint_y=None,
                    height=dp(40), font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
                ))
                self.customers_list_container.height = dp(40)
                return
            
            for customer_name in filtered:
                customer_box = BoxLayout(
                    size_hint_y=None, height=dp(50), spacing=dp(5),
                    padding=[dp(8), dp(4), dp(8), dp(4)]
                )
                with customer_box.canvas.before:
                    Color(0.15, 0.15, 0.2, 1)
                    rect = Rectangle(pos=customer_box.pos, size=customer_box.size)
                    customer_box.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                                    size=lambda i, v: setattr(rect, 'size', v))
                customer_label = RTLLabel(
                    text=customer_name, size_hint_x=0.7, size_hint_y=None,
                    height=dp(45), font_size=sp(18), color=(1, 1, 1, 1), halign='right'
                )
                customer_box.add_widget(customer_label)
                visit_btn = PersianButton(
                    text='ویزیت', size_hint_x=0.3, size_hint_y=None, height=dp(40),
                    background_color=(0.2, 0.6, 0.8, 1), color=(1, 1, 1, 1), font_size=sp(15)
                )
                visit_btn.bind(on_press=lambda x, name=customer_name: self.on_customer_selected(name))
                customer_box.add_widget(visit_btn)
                self.customers_list_container.add_widget(customer_box)
            
            total_height = len(filtered) * dp(55) + dp(10)
            self.customers_list_container.height = total_height
        except Exception as e:
            print(f"خطا در بروزرسانی لیست مشتریان: {e}")
    
    # ============================================================
    # دیالوگ انتخاب مشتری
    # ============================================================
    
    def show_customer_selection_dialog(self, instance):
        try:
            selected_route = self.route_spinner.text
            if not selected_route:
                self.show_message('خطا', 'لطفاً ابتدا یک مسیر انتخاب کنید')
                return
            
            all_customers = get_customers()
            filtered_customers = []
            for c in all_customers:
                if c.get('route_name', '').strip() == selected_route.strip():
                    filtered_customers.append(c.get('name', ''))
            
            if not filtered_customers:
                self.show_message('توجه', 'هیچ مشتری‌ای در این مسیر یافت نشد')
                return
            
            today = get_today_jalali()
            logs = get_daily_logs()
            visited_today = []
            if today in logs and isinstance(logs[today], list):
                for log in logs[today]:
                    if isinstance(log, dict):
                        visited_today.append(log.get('customer', ''))
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                            size=lambda i, v: setattr(content_rect, 'size', v))
            
            title_layout = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(60))
            title_layout.add_widget(RTLLabel(
                text=f'انتخاب مشتری - {selected_route}', size_hint_y=None,
                height=dp(30), font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            title_layout.add_widget(RTLLabel(
                text=' مشتریان آبی رنگ امروز ویزیت شده‌اند', size_hint_y=None,
                height=dp(25), font_size=sp(15), color=(0.6, 0.6, 0.6, 1)
            ))
            content.add_widget(title_layout)
            
            search_input = RTLTextInput(
                hint_text='جستجوی مشتری...', multiline=False,
                size_hint_y=None, height=dp(55), font_size=sp(24)
            )
            search_input.bg_color = (0.15, 0.15, 0.15, 1)
            search_input.border_color = (0.3, 0.3, 0.3, 1)
            search_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            search_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(search_input)
            
            customers_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True, size_hint_y=0.65,
                scroll_type=['bars', 'content'], bar_width=dp(6)
            )
            customers_grid = GridLayout(cols=1, spacing=dp(4), size_hint_y=None, padding=dp(5))
            customers_grid.bind(minimum_height=customers_grid.setter('height'))
            
            def filter_customers(text):
                customers_grid.clear_widgets()
                search_text = text.strip()
                for customer in filtered_customers:
                    if search_text and search_text not in customer: continue
                    is_visited = customer in visited_today
                    customer_btn = PersianButton(
                        text=customer, size_hint_y=None, height=dp(45),
                        background_color=(0.2, 0.6, 1, 1) if is_visited else (0.2, 0.2, 0.2, 1),
                        color=(1, 1, 1, 1), font_size=sp(18)
                    )
                    customer_btn.bind(on_press=lambda x, name=customer: self._handle_customer_selection(name, content))
                    customers_grid.add_widget(customer_btn)
                customers_grid.height = len(customers_grid.children) * dp(50) + dp(10)
            
            search_input._hidden_input.bind(text=lambda i, v: filter_customers(v))
            filter_customers('')
            customers_scroll.add_widget(customers_grid)
            content.add_widget(customers_scroll)
            
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='انتخاب مشتری', content=content,
                size_hint=(0.9, 0.75), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=True
            )
            self.customer_selection_popup = popup
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ انتخاب مشتری: {e}", error_details)
    
    def _handle_customer_selection(self, customer_name, dialog_content):
        try:
            today = get_today_jalali()
            logs = get_daily_logs()
            is_visited_today = False
            if today in logs and isinstance(logs[today], list):
                for log in logs[today]:
                    if isinstance(log, dict) and log.get('customer') == customer_name:
                        is_visited_today = True
                        break
            
            if is_visited_today:
                content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
                with content.canvas.before:
                    Color(0.15, 0.15, 0.15, 1)
                    content_rect = Rectangle(pos=content.pos, size=content.size)
                    content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                                size=lambda i, v: setattr(content_rect, 'size', v))
                content.add_widget(RTLLabel(
                    text=f'مشتری "{customer_name}" امروز ویزیت شده است.از تغییر وضعیت مطمئن هستید؟',
                    size_hint_y=None, height=dp(60), font_size=sp(18), color=(1, 1, 1, 1)
                ))
                btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
                yes_btn = PersianButton(
                    text='بله', background_color=(0.2, 0.7, 0.2, 1),
                    size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
                )
                no_btn = PersianButton(
                    text='خیر', background_color=(0.8, 0.2, 0.2, 1),
                    size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
                )
                btn_layout.add_widget(yes_btn)
                btn_layout.add_widget(no_btn)
                content.add_widget(btn_layout)
                confirm_popup = PersianPopup(
                    title='تأیید تغییر', content=content,
                    size_hint=(0.85, 0.4), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
                )
                def on_yes(instance):
                    confirm_popup.dismiss()
                    if hasattr(self, 'customer_selection_popup'): self.customer_selection_popup.dismiss()
                    self.selected_customer = customer_name
                    self.selected_customer_label.text = f'مشتری انتخاب شده: {customer_name}'
                    self.selected_customer_label.color = (0.2, 0.8, 0.4, 1)
                    self.show_confirm_dialog(customer_name)
                def on_no(instance): confirm_popup.dismiss()
                yes_btn.bind(on_press=on_yes)
                no_btn.bind(on_press=on_no)
                confirm_popup.open()
            else:
                if hasattr(self, 'customer_selection_popup'): self.customer_selection_popup.dismiss()
                self.selected_customer = customer_name
                self.selected_customer_label.text = f'مشتری انتخاب شده: {customer_name}'
                self.selected_customer_label.color = (0.2, 0.8, 0.4, 1)
                self.show_confirm_dialog(customer_name)
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در انتخاب مشتری: {e}", error_details)
    
    def on_customer_selected(self, customer_name):
        if customer_name and customer_name not in ['', 'مشتری‌ای یافت نشد', 'هیچ مشتری‌ای در این مسیر یافت نشد']:
            self.selected_customer = customer_name
            self.selected_customer_label.text = f'مشتری انتخاب شده: {customer_name}'
            self.selected_customer_label.color = (0.2, 0.8, 0.4, 1)
            self.show_confirm_dialog(customer_name)
    
    # ============================================================
    # دیالوگ‌های افزودن مشتری، تأیید مسیر، ویزیت و فروش
    # ============================================================
    
    def show_add_customer_dialog(self, instance):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text='افزودن مشتری جدید', size_hint_y=None, height=dp(35),
                font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            content.add_widget(RTLLabel(
                text='انتخاب مسیر:', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            routes = get_routes()
            route_names = [r.get('name', '') for r in routes] if routes else ['ابتدا مسیر ایجاد کنید']
            default_route = self.route_spinner.text if self.route_spinner.text else route_names[0]
            customer_route_spinner = PersianComboBox(
                text=default_route if default_route in route_names else route_names[0],
                values=route_names, height=dp(55)
            )
            customer_route_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            customer_route_spinner.main_btn.color = (1, 1, 1, 1)
            customer_route_spinner.main_btn.font_size = sp(16)
            content.add_widget(customer_route_spinner)
            
            content.add_widget(RTLLabel(
                text='نام مشتری (الزامی):', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            customer_name_input = RTLTextInput(
                hint_text='نام مشتری را وارد کنید', multiline=False,
                size_hint_y=None, height=dp(55), font_size=sp(22)
            )
            customer_name_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_name_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_name_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_name_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(customer_name_input)
            
            content.add_widget(RTLLabel(
                text='نام فروشگاه:', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            customer_store_input = RTLTextInput(
                hint_text='نام فروشگاه را وارد کنید', multiline=False,
                size_hint_y=None, height=dp(55), font_size=sp(22)
            )
            customer_store_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_store_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_store_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_store_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(customer_store_input)
            
            content.add_widget(RTLLabel(
                text='موبایل (الزامی):', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            customer_mobile_input = RTLTextInput(
                hint_text='شماره موبایل را وارد کنید (11 رقم)', multiline=False,
                size_hint_y=None, height=dp(55), font_size=sp(22)
            )
            customer_mobile_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_mobile_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_mobile_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_mobile_input._hidden_input.foreground_color = (1, 1, 1, 1)
            customer_mobile_input._hidden_input.input_filter = 'int'
            content.add_widget(customer_mobile_input)
            
            content.add_widget(RTLLabel(
                text='آدرس:', size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            customer_address_input = RTLTextInput(
                hint_text='آدرس را وارد کنید', multiline=False,
                size_hint_y=None, height=dp(55), font_size=sp(22)
            )
            customer_address_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_address_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_address_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_address_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(customer_address_input)
            
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(45))
            submit_btn = PersianButton(
                text='افزودن مشتری', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(submit_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='افزودن مشتری', content=content,
                size_hint=(0.9, 0.75), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            
            def do_add_customer(instance):
                try:
                    route_name = customer_route_spinner.text
                    if route_name == 'ابتدا مسیر ایجاد کنید':
                        self.show_message('خطا', 'لطفاً ابتدا یک مسیر ایجاد کنید')
                        return
                    name = customer_name_input.text.strip()
                    if not name:
                        self.show_message('خطا', 'نام مشتری الزامی است')
                        return
                    mobile = customer_mobile_input.text.strip()
                    if not mobile:
                        self.show_message('خطا', 'شماره موبایل الزامی است')
                        return
                    mobile_clean = mobile.replace(' ', '').replace('-', '').replace('_', '')
                    if not mobile_clean.isdigit():
                        self.show_message('خطا', 'شماره موبایل باید فقط شامل عدد باشد')
                        return
                    if len(mobile_clean) != 11:
                        self.show_message('خطا', 'شماره موبایل باید ۱۱ رقم باشد')
                        return
                    if not mobile_clean.startswith('09'):
                        self.show_message('خطا', 'شماره موبایل باید با 09 شروع شود')
                        return
                    all_customers = get_customers()
                    for c in all_customers:
                        if c.get('name', '').strip() == name:
                            self.show_message('خطا', f'مشتری با نام "{name}" قبلاً ثبت شده است')
                            return
                        existing_mobile = c.get('mobile', '').strip()
                        if existing_mobile and existing_mobile == mobile_clean:
                            self.show_message('خطا', f'شماره موبایل "{mobile_clean}" قبلاً ثبت شده است')
                            return
                    customer = {
                        'name': name, 'store_name': customer_store_input.text.strip(),
                        'route_name': route_name, 'mobile': mobile_clean,
                        'address': customer_address_input.text.strip()
                    }
                    add_customer(customer)
                    self.session_new_customers.append(name)
                    popup.dismiss()
                    self.update_customers_list()
                    self.show_message('موفق', f'مشتری "{name}" با موفقیت اضافه شد')
                except Exception as e:
                    error_details = traceback.format_exc()
                    ErrorPopup.show_error(f"خطا در افزودن مشتری: {e}", error_details)
            
            def cancel_add(instance): popup.dismiss()
            submit_btn.bind(on_press=do_add_customer)
            cancel_btn.bind(on_press=cancel_add)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ افزودن مشتری: {e}", error_details)

    def show_route_confirm_dialog(self, route_name):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                        size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text=f'آیا قصد ویزیت مسیر "{route_name}" را دارید؟',
                size_hint_y=None, height=dp(50), font_size=sp(18), color=(1, 1, 1, 1)
            ))
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            yes_btn = PersianButton(
                text='بله', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            no_btn = PersianButton(
                text='خیر', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(yes_btn)
            btn_layout.add_widget(no_btn)
            content.add_widget(btn_layout)
            popup = PersianPopup(
                title='تأیید مسیر', content=content,
                size_hint=(0.85, 0.35), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def on_yes(instance):
                popup.dismiss()
                self.locked_route = route_name
                self.route_confirmed = True
                self.route_spinner.main_btn.disabled = True
                self.route_spinner.main_btn.background_color = (0.15, 0.15, 0.15, 1)
                self.route_spinner.main_btn.color = (0.6, 0.6, 0.6, 1)
                self.update_customers_list()
                self.show_message('موفق', f'مسیر "{route_name}" با موفقیت انتخاب و قفل شد.')
            def on_no(instance):
                popup.dismiss()
                self.route_spinner.text = ''
                self._last_route_text = ''
                self.route_confirmed = False
                self.update_customers_list()
            yes_btn.bind(on_press=on_yes)
            no_btn.bind(on_press=on_no)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ تأیید مسیر: {e}", error_details)

    def show_confirm_dialog(self, customer_name):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text=f'آیا برای "{customer_name}" ویزیت ثبت می‌نمایید؟',
                size_hint_y=None, height=dp(50), font_size=sp(18), color=(1, 1, 1, 1)
            ))
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            yes_btn = PersianButton(
                text='بله', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            no_btn = PersianButton(
                text='خیر', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(yes_btn)
            btn_layout.add_widget(no_btn)
            content.add_widget(btn_layout)
            popup = PersianPopup(
                title='تأیید ویزیت', content=content,
                size_hint=(0.85, 0.35), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def on_yes(instance):
                popup.dismiss()
                self.show_visit_result_dialog(customer_name)
            def on_no(instance):
                popup.dismiss()
                self.selected_customer = None
            yes_btn.bind(on_press=on_yes)
            no_btn.bind(on_press=on_no)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ تأیید: {e}", error_details)

    def show_visit_result_dialog(self, customer_name):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text=f'نتیجه ویزیت برای "{customer_name}":',
                size_hint_y=None, height=dp(40), font_size=sp(18), color=(1, 1, 1, 1)
            ))
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            success_btn = PersianButton(
                text='ویزیت موفق', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            fail_btn = PersianButton(
                text='ویزیت ناموفق', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(success_btn)
            btn_layout.add_widget(fail_btn)
            content.add_widget(btn_layout)
            back_btn = PersianButton(
                text='بازگشت', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            content.add_widget(back_btn)
            popup = PersianPopup(
                title='نتیجه ویزیت', content=content,
                size_hint=(0.85, 0.5), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def on_success(instance):
                popup.dismiss()
                self.show_sales_result_dialog(customer_name)
            def on_fail(instance):
                popup.dismiss()
                self.show_fail_reason_dialog(customer_name)
            def on_back(instance): popup.dismiss()
            success_btn.bind(on_press=on_success)
            fail_btn.bind(on_press=on_fail)
            back_btn.bind(on_press=on_back)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ نتیجه ویزیت: {e}", error_details)

    def show_fail_reason_dialog(self, customer_name):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text='علت ویزیت ناموفق را وارد کنید:',
                size_hint_y=None, height=dp(35), font_size=sp(16), color=(1, 1, 1, 1)
            ))
            reason_input = RTLTextInput(
                hint_text='متن علت...', size_hint_y=None, height=dp(100), font_size=sp(32)
            )
            reason_input.bg_color = (0.15, 0.15, 0.15, 1)
            reason_input.border_color = (0.3, 0.3, 0.3, 1)
            reason_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            reason_input._hidden_input.foreground_color = (1, 1, 1, 1)
            reason_input._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(reason_input._hidden_input)
            content.add_widget(reason_input)
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            submit_btn = PersianButton(
                text='ثبت عملیات', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            back_btn = PersianButton(
                text='بازگشت', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(submit_btn)
            btn_layout.add_widget(back_btn)
            content.add_widget(btn_layout)
            popup = PersianPopup(
                title='علت ویزیت ناموفق', content=content,
                size_hint=(0.85, 0.5), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def on_submit(instance):
                reason = reason_input.text.strip()
                if not reason:
                    ErrorPopup.show_error('لطفاً علت را وارد کنید')
                    return
                self.save_visit_log(customer_name=customer_name, visit_status='ناموفق', fail_reason=reason)
                popup.dismiss()
                self.show_message('موفق', f'ویزیت ناموفق برای "{customer_name}" ثبت شد')
                self.reset_form()
            def on_back(instance):
                popup.dismiss()
                self.show_visit_result_dialog(customer_name)
            submit_btn.bind(on_press=on_submit)
            back_btn.bind(on_press=on_back)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ علت ناموفق: {e}", error_details)

    def show_sales_result_dialog(self, customer_name):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text=f'نتیجه فروش برای "{customer_name}":',
                size_hint_y=None, height=dp(40), font_size=sp(18), color=(1, 1, 1, 1)
            ))
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            success_btn = PersianButton(
                text='فروش موفق', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            fail_btn = PersianButton(
                text='فروش ناموفق', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(success_btn)
            btn_layout.add_widget(fail_btn)
            content.add_widget(btn_layout)
            back_btn = PersianButton(
                text='بازگشت', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            content.add_widget(back_btn)
            popup = PersianPopup(
                title='نتیجه فروش', content=content,
                size_hint=(0.85, 0.45), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def on_success(instance):
                popup.dismiss()
                self.show_success_sales_dialog(customer_name)
            def on_fail(instance):
                popup.dismiss()
                self.show_fail_sales_reason_dialog(customer_name)
            def on_back(instance):
                popup.dismiss()
                self.show_visit_result_dialog(customer_name)
            success_btn.bind(on_press=on_success)
            fail_btn.bind(on_press=on_fail)
            back_btn.bind(on_press=on_back)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ نتیجه فروش: {e}", error_details)

    def show_fail_sales_reason_dialog(self, customer_name):
        try:
            fail_reasons = ['موکول به زمان دیگر', 'عدم نیاز مشتری', 'شاکی بودن مشتری',
                          'وجود مغایرت', 'عدم ایجاد ارتباط مناسب', 'سایر علل']
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text='علت فروش ناموفق:', size_hint_y=None, height=dp(35),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            reason_spinner = PersianComboBox(text=fail_reasons[0], values=fail_reasons, height=dp(65))
            reason_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            reason_spinner.main_btn.color = (1, 1, 1, 1)
            reason_spinner.main_btn.font_size = sp(18)
            content.add_widget(reason_spinner)
            description_input = RTLTextInput(
                hint_text='توضیحات (در صورت انتخاب سایر علل)', multiline=False,
                size_hint_y=None, height=dp(75), font_size=sp(22)
            )
            description_input.bg_color = (0.15, 0.15, 0.15, 1)
            description_input.border_color = (0.3, 0.3, 0.3, 1)
            description_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            description_input._hidden_input.foreground_color = (1, 1, 1, 1)
            description_input._hidden_input.disabled = True
            content.add_widget(description_input)
            self._last_reason_text = reason_spinner.text
            def check_reason_change(dt):
                if hasattr(self, '_last_reason_text'):
                    current = reason_spinner.text
                    if current != self._last_reason_text:
                        self._last_reason_text = current
                        description_input._hidden_input.disabled = (current != 'سایر علل')
                        if description_input._hidden_input.disabled: description_input.text = ''
            Clock.schedule_interval(check_reason_change, 0.3)
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            submit_btn = PersianButton(
                text='ثبت عملیات', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            back_btn = PersianButton(
                text='بازگشت', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(submit_btn)
            btn_layout.add_widget(back_btn)
            content.add_widget(btn_layout)
            popup = PersianPopup(
                title='علت فروش ناموفق', content=content,
                size_hint=(0.85, 0.6), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def on_submit(instance):
                reason = reason_spinner.text
                description = description_input.text.strip()
                if reason == 'سایر علل' and not description:
                    ErrorPopup.show_error('لطفاً توضیحات را وارد کنید')
                    return
                self.save_visit_log(
                    customer_name=customer_name, visit_status='موفق',
                    sales_status='ناموفق', fail_sales_reason=reason, sales_description=description
                )
                popup.dismiss()
                self.show_message('موفق', f'فروش ناموفق برای "{customer_name}" ثبت شد')
                self.reset_form()
            def on_back(instance):
                popup.dismiss()
                self.show_sales_result_dialog(customer_name)
            submit_btn.bind(on_press=on_submit)
            back_btn.bind(on_press=on_back)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ علت فروش ناموفق: {e}", error_details)

    def _update_amount_label(self, instance, value):
        try:
            if not self.amount_words_label: return
            amount = value.strip()
            if not amount or amount == '0':
                self.amount_words_label.set_text('صفر ریال')
                return
            number = int(amount)
            words = self.number_to_persian_words(number)
            self.amount_words_label.set_text(words)
        except ValueError:
            if self.amount_words_label: self.amount_words_label.set_text('مبلغ نامعتبر')
        except Exception as e:
            print(f"خطا در تبدیل عدد به حروف: {e}")

    # ============================================================
    # 🆕 دیالوگ فروش موفق با ریز فروش
    # ============================================================

    def show_success_sales_dialog(self, customer_name):
        """نمایش دیالوگ فروش موفق با قابلیت ریز فروش"""
        try:
            payment_methods = ['نقد', 'چک', 'اعتباری']
            
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(8))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                        size=lambda i, v: setattr(content_rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text=f'فروش موفق برای "{customer_name}"',
                size_hint_y=None, height=dp(32),
                font_size=sp(18), color=(1, 0.8, 0.2, 1), bold=True
            ))
            
            # ========== دکمه ریز فروش ==========
            if not hasattr(self, '_detailed_sales_total'):
                self._detailed_sales_total = 0
                self._detailed_sales_data = []
            
            detailed_sales_btn = PersianButton(
                text='📦 ریز فروش (کالا به کالا)',
                size_hint_y=None, height=dp(42),
                background_color=(0.6, 0.3, 0.7, 1),
                color=(1, 1, 1, 1), font_size=sp(15), bold=True
            )
            detailed_sales_btn.bind(on_press=lambda x: self._show_detailed_sales_dialog())
            content.add_widget(detailed_sales_btn)
            
            # نمایش جمع ریز فروش
            self.detailed_sales_summary = RTLLabel(
                text=f'تعداد کل: {self._detailed_sales_total}',
                size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(0.6, 0.6, 0.6, 1)
            )
            content.add_widget(self.detailed_sales_summary)
            
            # ========== تعداد واحد فروش ==========
            content.add_widget(RTLLabel(
                text='تعداد واحد فروش:', size_hint_y=None, height=dp(22),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))

            # ✅ مقدار اولیه از ریز فروش
            initial_units = str(self._detailed_sales_total) if self._detailed_sales_total > 0 else '0'

            units_input = RTLTextInput(
                text=initial_units,
                multiline=False, size_hint_y=None, height=dp(55),
                input_filter='int', font_size=sp(18)
            )
            units_input.bg_color = (0.15, 0.15, 0.15, 1)
            units_input.border_color = (0.3, 0.3, 0.3, 1)
            units_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            units_input._hidden_input.foreground_color = (1, 1, 1, 1)
            units_input._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(units_input._hidden_input)
            content.add_widget(units_input)

            # ✅ ذخیره reference برای آپدیت بعدی
            self._current_units_input = units_input

            # ✅ وقتی کاربر روی فیلد تعداد کلیک می‌کنه، جمع ریز فروش رو بذار
            def on_units_focus(inst, value):
                if value and self._detailed_sales_total > 0:
                    units_input.text = str(self._detailed_sales_total)
            units_input._hidden_input.bind(focus=on_units_focus)
            
            # ========== مبلغ فاکتور ==========
            content.add_widget(RTLLabel(
                text='مبلغ فاکتور (ریال):', size_hint_y=None, height=dp(22),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            self.amount_input = RTLTextInput(
                text='0', multiline=False, size_hint_y=None, height=dp(55),
                input_filter='int', font_size=sp(18)
            )
            self.amount_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.amount_input.border_color = (0.3, 0.3, 0.3, 1)
            self.amount_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.amount_input._hidden_input.foreground_color = (1, 1, 1, 1)
            self.amount_input._hidden_input.bind(focus=self._on_field_focus)
            self.amount_input._hidden_input.bind(text=self._update_amount_label)
            self.focusable_fields.append(self.amount_input._hidden_input)
            content.add_widget(self.amount_input)

            self.amount_words_label = RTLLabel(
                text='صفر ریال', size_hint_y=None, height=dp(35),
                font_size=sp(40), color=(0.8, 1, 0.8, 1), halign='right'
            )
            content.add_widget(self.amount_words_label)
            
            # ========== نحوه تسویه ==========
            content.add_widget(RTLLabel(
                text='نحوه تسویه:', size_hint_y=None, height=dp(22),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            payment_spinner = PersianComboBox(
                text=payment_methods[0], values=payment_methods, height=dp(50)
            )
            payment_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            payment_spinner.main_btn.color = (1, 1, 1, 1)
            payment_spinner.main_btn.font_size = sp(18)
            content.add_widget(payment_spinner)
            
            # ========== دکمه‌ها ==========
            btn_layout = BoxLayout(spacing=dp(8), size_hint_y=None, height=dp(45))
            submit_btn = PersianButton(
                text='ثبت عملیات', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            back_btn = PersianButton(
                text='بازگشت', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(submit_btn)
            btn_layout.add_widget(back_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='ثبت فروش موفق', content=content,
                size_hint=(0.9, 0.82), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            
            def on_submit(instance):
                units = units_input.text.strip()
                amount = self.amount_input.text.strip()
                payment = payment_spinner.text
                
                if not units or units == '0':
                    ErrorPopup.show_error('لطفاً تعداد واحد فروش را وارد کنید')
                    return
                if not amount or amount == '0':
                    ErrorPopup.show_error('لطفاً مبلغ فاکتور را وارد کنید')
                    return
                try:
                    units_int = int(units)
                    amount_int = int(amount)
                    if units_int <= 0 or amount_int <= 0:
                        ErrorPopup.show_error('مقادیر باید بیشتر از صفر باشند')
                        return
                except ValueError:
                    ErrorPopup.show_error('لطفاً مقادیر عددی معتبر وارد کنید')
                    return
                
                detailed_sales = getattr(self, '_detailed_sales_data', [])
                self.save_visit_log(
                    customer_name=customer_name, visit_status='موفق',
                    sales_status='موفق', units_sold=units_int,
                    sales_amount=amount_int, payment_method=payment,
                    detailed_sales=detailed_sales
                )
                popup.dismiss()
                self._detailed_sales_total = 0
                self._detailed_sales_data = []
                self._current_units_input = None  # ✅ پاکسازی
                self.show_message('موفق', f'فروش موفق برای "{customer_name}" ثبت شد')
                self.reset_form()
            
            def on_back(instance):
                popup.dismiss()
                self.show_sales_result_dialog(customer_name)
            
            submit_btn.bind(on_press=on_submit)
            back_btn.bind(on_press=on_back)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ فروش موفق: {e}", error_details)


    # ============================================================
    # 🆕 دیالوگ ریز فروش (کالا به کالا) - اصلاح‌شده
    # ============================================================

    def _show_detailed_sales_dialog(self):
        """نمایش دیالوگ ریز فروش برای انتخاب کالا و تعداد"""
        try:
            from utils.file_manager import get_product_groups, get_target_units
            
            products = get_product_groups()
            if not isinstance(products, list) or not products:
                self.show_message('خطا', 'ابتدا کالاها را بروزرسانی کنید')
                return
            
            units = get_target_units()
            if not isinstance(units, list):
                units = ["کارتن", "عدد", "شل", "بسته", "جعبه", "بانکه"]
            
            if not hasattr(self, '_detailed_sales_data'):
                self._detailed_sales_data = []
                self._detailed_sales_total = 0
            
            content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(5))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            # هدر
            header = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(3))
            header.add_widget(RTLLabel(text='تعداد', size_hint_x=0.2, font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'))
            header.add_widget(RTLLabel(text='واحد', size_hint_x=0.2, font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'))
            header.add_widget(RTLLabel(text='نام کالا', size_hint_x=0.45, font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'))
            header.add_widget(Label(size_hint_x=0.15))
            content.add_widget(header)
            
            # جدول - فضای اسکرول بیشتر
            scroll = ScrollView(size_hint_y=0.7, do_scroll_x=False)
            grid = GridLayout(cols=1, spacing=dp(4), size_hint_y=None, padding=dp(2))
            grid.bind(minimum_height=grid.setter('height'))
            
            # نمایش آیتم‌های قبلی
            for i, item in enumerate(self._detailed_sales_data):
                row = self._make_detailed_sale_row(grid, i, item['product'], item['unit'], item['count'], products, units)
                grid.add_widget(row)
            
            # ردیف جدید
            new_row = self._make_detailed_sale_row(grid, len(self._detailed_sales_data), '', units[0] if units else '', '0', products, units)
            grid.add_widget(new_row)
            
            scroll.add_widget(grid)
            content.add_widget(scroll)
            
            # دکمه افزودن ردیف جدید
            add_row_btn = PersianButton(
                text='+ افزودن کالا', size_hint_y=None, height=dp(36),
                background_color=(0.2, 0.5, 0.8, 1), color=(1, 1, 1, 1), font_size=sp(15)
            )
            add_row_btn.bind(on_press=lambda x: self._add_detailed_sale_row(grid, products, units))
            content.add_widget(add_row_btn)
            
            # جمع کل
            total_label = RTLLabel(
                text=f'جمع کل: {self._detailed_sales_total}',
                size_hint_y=None, height=dp(28),
                font_size=sp(16), bold=True, color=(0.2, 0.8, 0.2, 1)
            )
            content.add_widget(total_label)
            
            # دکمه‌ها
            btn_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
            save_btn = PersianButton(
                text='تأیید', size_hint_x=0.5, size_hint_y=None, height=dp(38),
                background_color=(0.2, 0.7, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(15)
            )
            close_btn = PersianButton(
                text='بستن', size_hint_x=0.5, size_hint_y=None, height=dp(38),
                background_color=(0.3, 0.3, 0.3, 1), color=(1, 1, 1, 1), font_size=sp(15)
            )
            btn_row.add_widget(save_btn)
            btn_row.add_widget(close_btn)
            content.add_widget(btn_row)
            
            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.94, 0.82), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            
            def do_save(inst):
                total = 0
                valid_items = []
                for child in grid.children:
                    if hasattr(child, 'children') and len(child.children) >= 4:
                        count_input = child.children[3]
                        product_combo = child.children[1]
                        unit_combo = child.children[2]
                        count_text = count_input.text.strip() if hasattr(count_input, 'text') else '0'
                        try:
                            count = int(count_text) if count_text.isdigit() else 0
                        except:
                            count = 0
                        product = product_combo.text if hasattr(product_combo, 'text') else ''
                        unit = unit_combo.text if hasattr(unit_combo, 'text') else ''
                        if count > 0 and product:
                            total += count
                            valid_items.append({'product': product, 'unit': unit, 'count': count})
                
                self._detailed_sales_data = valid_items
                self._detailed_sales_total = total
                
                # ✅ بروزرسانی فیلد تعداد در دیالوگ فروش موفق
                if hasattr(self, 'detailed_sales_summary'):
                    self.detailed_sales_summary.text = f'تعداد کل: {total}'
                
                # ✅ آپدیت مستقیم فیلد تعداد
                if hasattr(self, '_current_units_input') and self._current_units_input:
                    self._current_units_input.text = str(total)
                
                popup.dismiss()
            
            save_btn.bind(on_press=do_save)
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)


    def _make_detailed_sale_row(self, grid, index, product, unit, count, products, units):
        """ساخت یک ردیف ریز فروش"""
        row = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(3))
        
        # ✅ تعداد - فونت ۳ سایز بزرگتر (sp(19))
        count_input = RTLTextInput(
            text=str(count), multiline=False, size_hint_x=0.2,
            size_hint_y=None, height=dp(44), input_filter='int', font_size=sp(22)
        )
        count_input.bg_color = (0.18, 0.18, 0.18, 1)
        count_input.border_color = (0.4, 0.4, 0.4, 1)
        count_input.border_color_focus = (0.2, 0.5, 0.9, 1)
        count_input._hidden_input.foreground_color = (1, 1, 1, 1)
        count_input._hidden_input.halign = 'center'
        row.add_widget(count_input)
        
        # ✅ واحد - فونت بزرگتر + رنگ پس‌زمینه روشن‌تر
        unit_combo = PersianComboBox(text=unit if unit else (units[0] if units else ''), values=units, height=dp(44))
        unit_combo.main_btn.background_color = (0.25, 0.25, 0.25, 1)
        unit_combo.main_btn.color = (1, 1, 1, 1)
        unit_combo.main_btn.font_size = sp(20)
        unit_combo.size_hint_x = 0.2
        row.add_widget(unit_combo)
        
        # ✅ نام کالا - فونت بزرگتر
        product_combo = PersianComboBox(
            text=product if product else (products[0] if products else ''),
            values=products, height=dp(44)
        )
        product_combo.main_btn.background_color = (0.25, 0.25, 0.25, 1)
        product_combo.main_btn.color = (1, 1, 1, 1)
        product_combo.main_btn.font_size = sp(20)
        product_combo.size_hint_x = 0.45
        row.add_widget(product_combo)
        
        # دکمه حذف
        del_btn = PersianButton(
            text='-', size_hint_x=0.15, size_hint_y=None, height=dp(44),
            background_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(22)
        )
        del_btn.bind(on_press=lambda x: self._remove_detailed_sale_row(row, grid))
        row.add_widget(del_btn)
        
        return row


    def _add_detailed_sale_row(self, grid, products, units):
        """افزودن ردیف جدید به ریز فروش"""
        row = self._make_detailed_sale_row(grid, len(grid.children), '', units[0] if units else '', '0', products, units)
        grid.add_widget(row, index=0)


    def _remove_detailed_sale_row(self, row, grid):
        """حذف ردیف از ریز فروش"""
        if len(grid.children) > 1:
            grid.remove_widget(row)


    # ============================================================
    # 🆕 دیالوگ بروزرسانی کالاها
    # ============================================================

    def show_update_products_dialog(self, instance):
        """بروزرسانی گروه کالا و واحد از فایل اکسل ریزتارگت"""
        try:
            import openpyxl
            from utils.file_picker_import import ImportFilePicker
            from utils.file_manager import add_product_group, get_product_groups, add_target_unit        

            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text='بروزرسانی کالاها از فایل ریزتارگت',
                size_hint_y=None, height=dp(40),
                font_size=sp(20), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            
            content.add_widget(RTLLabel(
                text='فایل اکسل ریزتارگت را انتخاب کنید\n(ستون‌های "نام گروه کالا" و "واحد تارگت" بارگذاری می‌شوند)',
                size_hint_y=None, height=dp(50),
                font_size=sp(14), color=(0.6, 0.6, 0.6, 1)
            ))
            
            self.update_products_status = RTLLabel(
                text='فایلی انتخاب نشده است',
                size_hint_y=None, height=dp(35),
                font_size=sp(15), color=(0.5, 0.5, 0.5, 1)
            )
            content.add_widget(self.update_products_status)
            
            file_picker = ImportFilePicker(
                on_select=lambda filepath: self._on_products_file_selected(
                    filepath, self.update_products_status
                ),
                size_hint_y=None, height=dp(50)
            )
            content.add_widget(file_picker)
            
            existing_label = RTLLabel(
                text=self._get_current_products_summary(),
                size_hint_y=None, height=dp(60),
                font_size=sp(14), color=(0.6, 0.6, 0.6, 1)
            )
            content.add_widget(existing_label)
            
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='بروزرسانی کالاها', content=content,
                size_hint=(0.9, 0.55), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)


    def _on_products_file_selected(self, filepath, status_label):
        """پردازش فایل اکسل و بروزرسانی کالاها"""
        try:
            import openpyxl
            from utils.file_manager import add_product_group, add_target_unit 

            if not filepath: return
            
            filepath = filepath[0] if isinstance(filepath, list) else filepath
            
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            col_map = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    col_map[str(cell.value).strip()] = col_idx
            
            product_col = col_map.get('نام گروه کالا')
            unit_col = col_map.get('واحد تارگت')
            
            if not product_col:
                status_label.text = '❌ ستون "نام گروه کالا" یافت نشد'
                status_label.color = (0.8, 0.2, 0.2, 1)
                return
            
            added_products = 0
            added_units = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                
                if product_col and row[product_col - 1]:
                    product_name = str(row[product_col - 1]).strip()
                    if product_name and product_name != 'None':
                        if add_product_group(product_name):
                            added_products += 1
                
                if unit_col and row[unit_col - 1]:
                    unit_name = str(row[unit_col - 1]).strip()
                    if unit_name and unit_name != 'None':
                        from utils.file_manager import add_target_unit
                        if add_target_unit(unit_name):
                            added_units += 1
            
            wb.close()
            
            filename = os.path.basename(filepath)
            status_label.text = f'✅ {filename}\n{added_products} گروه کالا و {added_units} واحد اضافه شد'
            status_label.color = (0.2, 0.8, 0.2, 1)
            
        except Exception as e:
            status_label.text = f'❌ خطا: {str(e)[:50]}'
            status_label.color = (0.8, 0.2, 0.2, 1)


    def _get_current_products_summary(self):
        """نمایش خلاصه کالاهای موجود"""
        from utils.file_manager import get_product_groups, get_target_units
        products = get_product_groups()
        units = get_target_units()
        if not isinstance(products, list): products = []
        if not isinstance(units, list): units = []
        return f'کالاهای موجود: {len(products)} گروه | {len(units)} واحد'


    # ============================================================
    # دیالوگ تحقق ریزتارگت‌ها
    # ============================================================

    def show_target_fulfillment_dialog(self, instance):
        """نمایش دیالوگ تحقق ریزتارگت‌ها"""
        try:
            import openpyxl
            import os
            from datetime import datetime
            from utils.storage import get_backup_path
            from utils.detailed_target_manager import get_all_detailed_targets
            from utils.file_picker_import import ImportFilePicker
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            # ========== دکمه‌های عملیات ==========
            action_btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(6))
            
            self.fulfillment_file_picker = ImportFilePicker(
                on_select=lambda filepath: self._on_fulfillment_file_selected(
                    [filepath] if not isinstance(filepath, list) else filepath,
                    None, self.fulfillment_grid
                ),
                size_hint_x=0.35, size_hint_y=None, height=dp(45)
            )
            action_btn_layout.add_widget(self.fulfillment_file_picker)
            
            calc_btn = PersianButton(
                text='محاسبه خودکار',
                size_hint_x=0.3, size_hint_y=None, height=dp(45),
                background_color=(0.2, 0.6, 0.2, 1), color=(1, 1, 1, 1),
                font_size=sp(14), bold=True
            )
            calc_btn.bind(on_press=self._auto_calculate_fulfillment)
            action_btn_layout.add_widget(calc_btn)
            
            history_btn = PersianButton(
                text='تاریخچه',
                size_hint_x=0.35, size_hint_y=None, height=dp(45),
                background_color=(0.6, 0.4, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(14)
            )
            history_btn.bind(on_press=self._show_fulfillment_history)
            action_btn_layout.add_widget(history_btn)
            content.add_widget(action_btn_layout)
            
            # ========== هدر جدول ==========
            header_box = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(3))
            headers = [
                ('کسر تارگت', 0.22), ('تحقق', 0.18),
                ('تارگت روز', 0.22), ('نام گروه کالا', 0.38)
            ]
            for text, size in headers:
                header_box.add_widget(RTLLabel(
                    text=text, size_hint_x=size, size_hint_y=None, height=dp(32),
                    font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                ))
            content.add_widget(header_box)
            
            # ========== جدول داده‌ها ==========
            self.fulfillment_data = []
            self.fulfillment_inputs = []
            
            self.fulfillment_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint_y=0.55, scroll_type=['bars', 'content'], bar_width=dp(6)
            )
            self.fulfillment_grid = GridLayout(cols=1, spacing=dp(3), size_hint_y=None, padding=dp(3))
            self.fulfillment_grid.bind(minimum_height=self.fulfillment_grid.setter('height'))
            self.fulfillment_grid.add_widget(RTLLabel(
                text='لطفاً فایل اکسل ریزتارگت را انتخاب کنید',
                size_hint_y=None, height=dp(40), font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
            ))
            self.fulfillment_scroll.add_widget(self.fulfillment_grid)
            content.add_widget(self.fulfillment_scroll)
            
            # ========== دکمه‌های پایین ==========
            btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
            
            self.fulfillment_save_btn = PersianButton(
                text='ثبت ریز تارگت', size_hint_x=0.5, size_hint_y=None, height=dp(45),
                background_color=(0.2, 0.7, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(self.fulfillment_save_btn)
            
            export_btn = PersianButton(
                text='خروجی اکسل', size_hint_x=0.5, size_hint_y=None, height=dp(45),
                background_color=(0.8, 0.6, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(export_btn)
            content.add_widget(btn_layout)
            
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(16)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='تحقق ریزتارگت‌ها', content=content,
                size_hint=(0.95, 0.9), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            
            # ========== رویدادها ==========
            
            def save_fulfillment(inst):
                if not self.fulfillment_data:
                    self.show_message('خطا', 'هیچ داده‌ای برای ثبت وجود ندارد')
                    return
                
                all_targets = get_all_detailed_targets()
                self._do_save_fulfillment(all_targets, popup)
            
            self.fulfillment_save_btn.bind(on_press=save_fulfillment)
            
            def export_fulfillment(inst):
                if not self.fulfillment_data:
                    self.show_message('خطا', 'داده‌ای برای خروجی وجود ندارد')
                    return
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    from openpyxl.utils import get_column_letter
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "گزارش تحقق"
                    header_font = Font(bold=True, size=11, color="FFFFFF")
                    header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    headers = ['شناسه', 'عامل', 'گروه کالا', 'تارگت روز', 'تحقق', 'کسر تارگت', 'واحد']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                    for row, item in enumerate(self.fulfillment_data, 2):
                        daily = item.get('daily_target', 1)
                        achieved = item.get('achieved', 0)
                        remaining = max(0, daily - achieved)
                        values = [item.get('id', ''), item.get('agent_name', ''),
                                item.get('product_group', ''), daily, achieved, remaining, item.get('unit', '')]
                        for col, value in enumerate(values, 1):
                            cell = ws.cell(row=row, column=col, value=value)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = thin_border
                    column_widths = [16, 18, 18, 14, 14, 14, 12]
                    for i, width in enumerate(column_widths, 1):
                        ws.column_dimensions[get_column_letter(i)].width = width
                    today = get_today_jalali().replace('/', '-')
                    filename = f'گزارش_تحقق_{today}_{datetime.now().strftime("%H%M%S")}.xlsx'
                    export_dir = get_backup_path()
                    os.makedirs(export_dir, exist_ok=True)
                    filepath = os.path.join(export_dir, filename)
                    wb.save(filepath)
                    self.show_message('موفق', f'فایل ذخیره شد:\n{filename}')
                except ImportError:
                    self.show_message('خطا', 'ماژول openpyxl نصب نیست')
                except Exception as e:
                    self.show_message('خطا', f'خطا در خروجی: {str(e)}')
            
            export_btn.bind(on_press=export_fulfillment)
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ تحقق: {e}", error_details)


    def _do_save_fulfillment(self, all_targets, parent_popup):
        """اجرای واقعی ذخیره تحقق - به‌روزرسانی برای روز جاری"""
        today = get_today_jalali()
        updated = 0
        
        for item in self.fulfillment_data:
            target_id = item.get('id')
            achieved = item.get('achieved', 0)
            
            if target_id and achieved > 0:
                for t in all_targets:
                    if t.get('id') == target_id:
                        t['achieved_value'] = achieved
                        t['status'] = 'فعال'
                        t['last_fulfillment_date'] = today
                        updated += 1
                        break
        
        if updated > 0:
            import json
            from utils.storage import get_data_path
            path = os.path.join(get_data_path(), 'detailed_targets.json')
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(all_targets, f, ensure_ascii=False, indent=2)
            
            self.fulfillment_save_btn.disabled = True
            self.fulfillment_save_btn.text = 'ثبت شد'
            self.fulfillment_save_btn.background_color = (0.3, 0.3, 0.3, 1)
            
            self.show_message('موفق', f'{updated} ریزتارگت برای امروز ({today}) بروزرسانی شد')
        else:
            self.show_message('خطا', 'هیچ تارگتی بروزرسانی نشد')


    def _on_fulfillment_file_selected(self, selection, status_label, grid):
        """پردازش فایل اکسل انتخاب شده"""
        try:
            import openpyxl
            import os
            
            if not selection:
                return
            
            filepath = selection[0] if isinstance(selection, list) else selection
            self.fulfillment_file_path = filepath
            filename = os.path.basename(filepath)
            
            if status_label is not None:
                status_label.text = f'{filename}'
                status_label.color = (0.2, 0.8, 0.2, 1)
            
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            col_map = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    col_map[str(cell.value).strip()] = col_idx
            
            id_col = col_map.get('شناسه ریزتارگت')
            agent_col = col_map.get('نام عامل')
            product_col = col_map.get('نام گروه کالا')
            daily_col = col_map.get('تارگت روزانه')
            unit_col = col_map.get('واحد تارگت')
            
            if not all([id_col, product_col, daily_col]):
                if status_label is not None:
                    status_label.text = 'ستون‌های مورد نیاز یافت نشد'
                    status_label.color = (0.8, 0.2, 0.2, 1)
                return
            
            self.fulfillment_data = []
            self.fulfillment_inputs = []
            grid.clear_widgets()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[id_col - 1]:
                    continue
                
                target_id = str(row[id_col - 1]).strip() if row[id_col - 1] else ''
                agent_name = str(row[agent_col - 1]).strip() if agent_col and row[agent_col - 1] else ''
                product_group = str(row[product_col - 1]).strip() if row[product_col - 1] else ''
                daily_target = int(row[daily_col - 1]) if row[daily_col - 1] else 0
                unit = str(row[unit_col - 1]).strip() if unit_col and row[unit_col - 1] else ''
                
                idx = len(self.fulfillment_data)
                self.fulfillment_data.append({
                    'id': target_id, 'agent_name': agent_name,
                    'product_group': product_group, 'daily_target': daily_target,
                    'unit': unit, 'achieved': 0
                })
                
                row_box = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(3))
                
                remaining_input = RTLTextInput(
                    text=f"{daily_target:,}", multiline=False, size_hint_x=0.22,
                    size_hint_y=None, height=dp(50), font_size=sp(24)
                )
                remaining_input.bg_color = (0.15, 0.15, 0.15, 1)
                remaining_input.border_color = (0.4, 0.4, 0.4, 1)
                remaining_input.border_color_focus = (0.4, 0.4, 0.4, 1)
                remaining_input._hidden_input.foreground_color = (1, 1, 1, 1)
                remaining_input._hidden_input.disabled = True
                remaining_input._hidden_input.halign = 'center'
                row_box.add_widget(remaining_input)
                
                from kivy.uix.textinput import TextInput
                achieved_input = TextInput(
                    text='0', multiline=False, size_hint_x=0.18, size_hint_y=None, height=dp(50),
                    input_filter='int', font_size=sp(24), font_name='PersianFont',
                    halign='center', padding=[dp(2), dp(10), dp(2), dp(10)],
                    background_color=(0.3, 0.3, 0.3, 1), foreground_color=(1, 1, 1, 1),
                    cursor_color=(1, 1, 1, 1), selection_color=(0.2, 0.5, 0.8, 0.5)
                )
                row_box.add_widget(achieved_input)
                
                daily_input = RTLTextInput(
                    text=f"{daily_target:,}", multiline=False, size_hint_x=0.22,
                    size_hint_y=None, height=dp(50), font_size=sp(24)
                )
                daily_input.bg_color = (0.15, 0.15, 0.15, 1)
                daily_input.border_color = (0.4, 0.4, 0.4, 1)
                daily_input.border_color_focus = (0.4, 0.4, 0.4, 1)
                daily_input._hidden_input.foreground_color = (1, 1, 1, 1)
                daily_input._hidden_input.disabled = True
                daily_input._hidden_input.halign = 'center'
                row_box.add_widget(daily_input)
                
                product_input = RTLTextInput(
                    text=product_group, multiline=False, size_hint_x=0.38,
                    size_hint_y=None, height=dp(50), font_size=sp(22)
                )
                product_input.bg_color = (0.15, 0.15, 0.15, 1)
                product_input.border_color = (0.4, 0.4, 0.4, 1)
                product_input.border_color_focus = (0.4, 0.4, 0.4, 1)
                product_input._hidden_input.foreground_color = (1, 1, 1, 1)
                product_input._hidden_input.disabled = True
                product_input._hidden_input.halign = 'right'
                row_box.add_widget(product_input)
                
                self.fulfillment_inputs.append({
                    'achieved_input': achieved_input,
                    'remaining_input': remaining_input,
                    'daily_target': daily_target, 'index': idx
                })
                
                achieved_input.bind(
                    text=lambda inst, val, i=idx: self._update_fulfillment_remaining(i, val)
                )
                grid.add_widget(row_box)
            
            wb.close()
            
            if status_label is not None:
                status_label.text = f'{filename} ({len(self.fulfillment_data)} ردیف)'
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خواندن فایل: {e}", error_details)


    def _update_fulfillment_remaining(self, index, value):
        """بروزرسانی کسر تارگت روزانه با تغییر عدد تحقق"""
        try:
            if index >= len(self.fulfillment_data): return
            try:
                achieved = int(value) if value and str(value).strip().isdigit() else 0
            except:
                achieved = 0
            self.fulfillment_data[index]['achieved'] = achieved
            for inp_data in self.fulfillment_inputs:
                if inp_data['index'] == index:
                    daily_target = inp_data['daily_target']
                    remaining = daily_target - achieved
                    display_remaining = max(0, remaining)
                    inp_data['remaining_input']._hidden_input.text = f"{display_remaining:,}"
                    if daily_target > 0:
                        ratio = achieved / daily_target
                        if ratio < 0.25:
                            text_color = (0.8, 0.2, 0.2, 1); bg_color = (0.25, 0.08, 0.08, 1)
                        elif ratio < 0.50:
                            text_color = (1, 0.7, 0, 1); bg_color = (0.25, 0.18, 0.05, 1)
                        elif ratio < 0.75:
                            text_color = (0.2, 0.6, 1, 1); bg_color = (0.05, 0.12, 0.25, 1)
                        else:
                            text_color = (0.2, 0.8, 0.2, 1); bg_color = (0.05, 0.2, 0.05, 1)
                    else:
                        text_color = (0.5, 0.5, 0.5, 1); bg_color = (0.15, 0.15, 0.15, 1)
                    inp_data['remaining_input']._hidden_input.foreground_color = text_color
                    inp_data['remaining_input'].bg_color = bg_color
                    remaining_widget = inp_data['remaining_input']
                    remaining_widget.canvas.before.clear()
                    with remaining_widget.canvas.before:
                        Color(*bg_color)
                        RoundedRectangle(pos=remaining_widget.pos, size=remaining_widget.size, radius=[dp(3)])
                    break
        except Exception as e:
            print(f"خطا در بروزرسانی کسر تارگت: {e}")


    def _auto_calculate_fulfillment(self, instance):
        """محاسبه خودکار تحقق بر اساس daily_log.json و تاریخ امروز"""
        try:
            if not self.fulfillment_data:
                self.show_message('خطا', 'ابتدا فایل ریزتارگت را انتخاب کنید')
                return
            
            today = get_today_jalali()
            logs = get_daily_logs()
            today_logs = logs.get(today, [])
            
            if not today_logs or not isinstance(today_logs, list):
                self.show_message('اطلاع', f'هیچ ویزیتی برای تاریخ {today} ثبت نشده است')
                return
            
            self._do_auto_calculate(today_logs)
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در محاسبه خودکار: {e}", error_details)


    def _do_auto_calculate(self, today_logs):
        """اجرای واقعی محاسبه خودکار"""
        product_sales = {}
        
        for log in today_logs:
            if not isinstance(log, dict):
                continue
            if log.get('visit_status') != 'موفق' or log.get('sales_status') != 'موفق':
                continue
            
            detailed_sales = log.get('detailed_sales', [])
            if detailed_sales and isinstance(detailed_sales, list):
                for item in detailed_sales:
                    if not isinstance(item, dict):
                        continue
                    product = item.get('product', '')
                    count = item.get('count', 0)
                    if product and count > 0:
                        product_sales[product] = product_sales.get(product, 0) + count
        
        if not product_sales:
            today = get_today_jalali()
            self.show_message('اطلاع', f'برای تاریخ {today} فروشی با ریز فروش ثبت نشده است')
            return
        
        total_matched = 0
        
        for i, item in enumerate(self.fulfillment_data):
            product_group = item.get('product_group', '')
            if product_group in product_sales:
                sale_count = product_sales[product_group]
                self.fulfillment_data[i]['achieved'] = sale_count
                
                for inp_data in self.fulfillment_inputs:
                    if inp_data['index'] == i:
                        achieved_input = inp_data['achieved_input']
                        if hasattr(achieved_input, 'text'):
                            achieved_input.text = str(sale_count)
                        break
                total_matched += 1
        
        if total_matched > 0:
            self.show_message('موفق', f'{total_matched} گروه کالا بروزرسانی شد')
        else:
            self.show_message('اطلاع', 'گروه‌های کالا با ریزتارگت‌ها تطبیق نداشت')


    # ============================================================
    # تاریخچه با فیلتر ماهیانه
    # ============================================================

    def _show_fulfillment_history(self, instance):
        """نمایش تاریخچه تحقق با فیلتر ماهیانه"""
        try:
            from utils.detailed_target_manager import get_all_detailed_targets
            
            all_targets = get_all_detailed_targets()
            if not isinstance(all_targets, list) or not all_targets:
                self.show_message('اطلاع', 'هیچ ریزتارگتی ثبت نشده است')
                return
            
            fulfilled = [t for t in all_targets if isinstance(t, dict) and t.get('achieved_value', 0) > 0]
            if not fulfilled:
                self.show_message('اطلاع', 'هنوز هیچ تحققی ثبت نشده است')
                return
            
            months_set = set()
            for t in fulfilled:
                start_date = t.get('start_date', '')
                if start_date and len(start_date) >= 7:
                    months_set.add(start_date[:7])
            
            months_list = sorted(list(months_set), reverse=True)
            months_display = ['همه'] + months_list
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                h_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(h_rect, 'pos', v),
                            size=lambda i, v: setattr(h_rect, 'size', v))
            
            header_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8))
            header_row.add_widget(RTLLabel(
                text=f'تاریخچه تحقق ({len(fulfilled)} مورد)',
                size_hint_x=0.55, size_hint_y=None, height=dp(38),
                font_size=sp(17), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            
            month_filter = PersianComboBox(
                text='همه', values=months_display, height=dp(38)
            )
            month_filter.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            month_filter.main_btn.color = (1, 1, 1, 1)
            month_filter.main_btn.font_size = sp(14)
            month_filter.size_hint_x = 0.45
            header_row.add_widget(month_filter)
            
            content.add_widget(header_row)
            
            hist_header = BoxLayout(size_hint_y=None, height=dp(32), spacing=dp(2))
            hist_headers = [('کسر', 0.16), ('تحقق', 0.18), ('هدف', 0.18),
                          ('گروه کالا', 0.26), ('عامل', 0.22)]
            for text, size in hist_headers:
                hist_header.add_widget(RTLLabel(
                    text=text, size_hint_x=size, size_hint_y=None, height=dp(30),
                    font_size=sp(13), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                ))
            content.add_widget(hist_header)
            
            hist_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint_y=0.75, scroll_type=['bars', 'content'], bar_width=dp(5)
            )
            hist_grid = GridLayout(cols=1, spacing=dp(2), size_hint_y=None, padding=dp(2))
            hist_grid.bind(minimum_height=hist_grid.setter('height'))
            
            def populate_history(selected_month):
                hist_grid.clear_widgets()
                
                filtered_list = fulfilled
                if selected_month != 'همه':
                    filtered_list = [t for t in fulfilled if t.get('start_date', '').startswith(selected_month)]
                
                if not filtered_list:
                    hist_grid.add_widget(RTLLabel(
                        text='هیچ تحققی در این ماه یافت نشد',
                        size_hint_y=None, height=dp(40),
                        font_size=sp(14), color=(0.5, 0.5, 0.5, 1)
                    ))
                    return
                
                for t in filtered_list:
                    target_count = t.get('target_count', 1)
                    achieved = t.get('achieved_value', 0)
                    remaining = max(0, target_count - achieved)
                    
                    if target_count > 0:
                        percent_done = (achieved / target_count) * 100
                        if percent_done < 25: r_color = (0.8, 0.2, 0.2, 1)
                        elif percent_done < 50: r_color = (1, 0.7, 0, 1)
                        elif percent_done < 75: r_color = (0.2, 0.6, 1, 1)
                        else: r_color = (0.2, 0.8, 0.2, 1)
                    else:
                        r_color = (0.5, 0.5, 0.5, 1)
                    
                    row = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(2))
                    row.add_widget(RTLLabel(
                        text=f"{remaining:,}", size_hint_x=0.16, size_hint_y=None, height=dp(33),
                        font_size=sp(14), bold=True, color=r_color, halign='center'
                    ))
                    row.add_widget(RTLLabel(
                        text=f"{achieved:,}", size_hint_x=0.18, size_hint_y=None, height=dp(33),
                        font_size=sp(13), color=(1, 1, 1, 1), halign='center'
                    ))
                    row.add_widget(RTLLabel(
                        text=f"{target_count:,}", size_hint_x=0.18, size_hint_y=None, height=dp(33),
                        font_size=sp(13), color=(0.8, 0.8, 0.8, 1), halign='center'
                    ))
                    row.add_widget(RTLLabel(
                        text=t.get('product_group', ''), size_hint_x=0.26, size_hint_y=None, height=dp(33),
                        font_size=sp(13), color=(1, 1, 1, 1), halign='right'
                    ))
                    row.add_widget(RTLLabel(
                        text=t.get('agent_name', ''), size_hint_x=0.22, size_hint_y=None, height=dp(33),
                        font_size=sp(13), color=(0.6, 0.6, 0.6, 1), halign='right'
                    ))
                    hist_grid.add_widget(row)
            
            month_filter.bind(text=lambda inst, val: populate_history(val))
            populate_history('همه')
            
            hist_scroll.add_widget(hist_grid)
            content.add_widget(hist_scroll)
            
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(15)
            )
            content.add_widget(close_btn)
            
            hist_popup = PersianPopup(
                title='تاریخچه تحقق', content=content,
                size_hint=(0.92, 0.8), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            close_btn.bind(on_press=hist_popup.dismiss)
            hist_popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)

    # ============================================================
    # توابع عمومی
    # ============================================================

    def save_visit_log(self, **kwargs):
        try:
            today = get_today_jalali()
            logs = get_daily_logs()
            if today not in logs: logs[today] = []
            if not isinstance(logs[today], list): logs[today] = []
            
            customer_name = kwargs.get('customer_name')
            is_new_customer = customer_name in self.session_new_customers
            
            log_data = {
                'date': today,
                'route': self.route_spinner.text,
                'agent_name': App.get_running_app().current_username,
                'customer': customer_name,
                'visit_status': kwargs.get('visit_status'),
                'time': get_current_time(),
                'is_new_customer': is_new_customer
            }
            
            if kwargs.get('visit_status') == 'ناموفق':
                log_data['fail_reason'] = kwargs.get('fail_reason', '')
            elif kwargs.get('visit_status') == 'موفق':
                log_data['sales_status'] = kwargs.get('sales_status', '')
                if kwargs.get('sales_status') == 'ناموفق':
                    log_data['fail_sales_reason'] = kwargs.get('fail_sales_reason', '')
                    log_data['sales_description'] = kwargs.get('sales_description', '')
                elif kwargs.get('sales_status') == 'موفق':
                    log_data['units_sold'] = kwargs.get('units_sold', 0)
                    log_data['sales_amount'] = kwargs.get('sales_amount', 0)
                    log_data['payment_method'] = kwargs.get('payment_method', '')
                    detailed_sales = kwargs.get('detailed_sales', [])
                    if detailed_sales:
                        log_data['detailed_sales'] = detailed_sales
            
            logs[today] = [log for log in logs[today] if log.get('customer') != customer_name]
            logs[today].append(log_data)
            save_daily_log(today, logs[today])
            
            self.locked_route = self.route_spinner.text
            if hasattr(self, 'route_spinner'):
                self.route_spinner.main_btn.disabled = True
                self.route_spinner.main_btn.background_color = (0.15, 0.15, 0.15, 1)
                self.route_spinner.main_btn.color = (0.6, 0.6, 0.6, 1)
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ذخیره لاگ ویزیت: {e}", error_details)

    def reset_form(self):
        self.selected_customer = None
        self.selected_customer_label.text = 'مشتری انتخاب شده: هیچ'
        self.selected_customer_label.color = (0.5, 0.5, 0.5, 1)
        self.update_customers_list()

    def show_message(self, title, message):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text=message, size_hint_y=None, height=dp(80), font_size=sp(18), color=(1, 1, 1, 1)
            ))
            btn = PersianButton(
                text='باشه', size_hint_y=None, height=dp(50), font_size=sp(18),
                color=(1, 1, 1, 1), background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            popup = PersianPopup(
                title=title, content=content,
                size_hint=(0.85, 0.4), background_color=(0.08, 0.08, 0.08, 1)
            )
            btn.bind(on_press=popup.dismiss)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش پیام: {e}", error_details)

    def go_back(self, instance):
        today = get_today_jalali()
        logs = get_daily_logs()
        if today in logs and logs[today] and len(logs[today]) > 0:
            if hasattr(self, 'route_spinner'):
                self.locked_route = self.route_spinner.text
        else:
            self.locked_route = None
            self.route_confirmed = False
            if hasattr(self, 'route_spinner'):
                self.route_spinner.main_btn.disabled = False
                self.route_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
                self.route_spinner.main_btn.color = (1, 1, 1, 1)
                self.route_spinner.text = ''
                self._last_route_text = ''
        self.manager.current = 'user'
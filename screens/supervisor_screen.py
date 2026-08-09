# screens/supervisor_screen.py
# ========== صفحه سوپروایزر ==========

import traceback
import os
from datetime import datetime
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.graphics import Color, Rectangle, RoundedRectangle
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox

from utils.rtl_widgets import (
    RTLTextInput, PersianComboBox, PersianButton,
    RTLLabel, PersianPopup, RTLMessageLabel
)
from utils.file_manager import (
    get_agents, get_routes, get_customers, get_settings,
    get_product_groups, add_product_group,
    get_target_units, add_target_unit, update_target_unit, delete_target_unit,
    get_target_periods, add_target_period, update_target_period, delete_target_period
)
from utils.jalali_date import get_today_jalali, get_current_time, validate_jalali_date
from utils.target_manager import (
    create_target,
    get_all_targets,
    get_targets_filtered,
    get_target_statistics,
    update_target,
    delete_target,
    can_edit_target,
    export_targets_to_excel,
    get_active_targets_by_agent,
    finalize_targets,
    read_excel_summary
)
from utils.detailed_target_manager import (
    create_detailed_target,
    get_all_detailed_targets,
    update_detailed_target,
    delete_detailed_target,
    can_edit_target as can_edit_detailed_target,
    export_to_excel as export_detailed_to_excel
)
from constants import TARGET_TYPES, TARGET_STATUSES, TARGET_EXCEL_MAPPING, PERIOD_DISPLAY, PERIOD_MAPPING
from error_handler import ErrorPopup


class SupervisorScreen(Screen):
    """صفحه سوپروایزر - ترکیبی از امکانات ایجنت و ادمین"""

    def __init__(self, **kwargs):
        try:
            super().__init__(**kwargs)
            with self.canvas.before:
                Color(0.08, 0.08, 0.08, 1)
                self.bg_rect = Rectangle(pos=self.pos, size=self.size)
                self.bind(pos=self._update_bg, size=self._update_bg)

            Window.softinput_mode = 'resize'
            self.focusable_fields = []
            self.tab_buttons = []
            self.current_tab = 0
            self.fulfillment_selected = {}
            self._clock_events = []
            self._last_market_route_text = ''
            self._last_followup_text = ''
            self._last_dt_agent = ''
            self._dt_linked_ids = {}

            self.build_ui()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ساخت SupervisorScreen: {e}", error_details)
            raise

    def _update_bg(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size

    def _on_field_focus(self, instance, value):
        if value:
            Clock.schedule_once(lambda dt: self._select_all_text(instance), 0.1)
            Clock.schedule_once(lambda dt: self._scroll_to_field(instance), 0.3)

    def _select_all_text(self, instance):
        if instance and hasattr(instance, 'select_all'):
            instance.select_all()

    def _scroll_to_field(self, instance):
        try:
            scroll = None
            for child in self.content_area.children:
                if isinstance(child, ScrollView):
                    scroll = child
                    break

            if not scroll:
                return

            field_pos = instance.to_window(0, 0)
            field_y = field_pos[1]
            keyboard_height = 250
            window_height = Window.height
            target_y = window_height - keyboard_height - dp(80)

            content_height = scroll.children[0].height if scroll.children else 1
            scroll_height = scroll.height

            if content_height > scroll_height:
                if field_y > target_y:
                    field_ratio = (content_height - field_y) / content_height
                    scroll_value = min(0.95, max(0.05, field_ratio + 0.1))
                    scroll.scroll_y = scroll_value
                elif field_y < dp(50):
                    scroll.scroll_y = 0.9

        except Exception as e:
            print(f"خطا در اسکرول: {e}")

    def build_ui(self):
        try:
            layout = BoxLayout(orientation='vertical', padding=[dp(5), dp(5), dp(5), dp(5)])

            # ========== تب‌ها ==========
            tabs_layout = BoxLayout(
                size_hint_y=None,
                height=dp(40),
                spacing=dp(2)
            )

            tab_names = [
                ('هدف‌گذاری', 0),
                ('ریزتارگت', 1),
                ('تحقق تارگت', 2),
                ('بررسی بازار', 3),
                ('گزارشات', 4)
            ]

            for name, tab_id in tab_names:
                btn = PersianButton(
                    text=name,
                    background_color=(0.3, 0.5, 0.8, 0.6),
                    size_hint_y=None,
                    height=dp(36),
                    color=(1, 1, 1, 1),
                    font_size=sp(14)
                )
                btn.bind(on_press=lambda x, tid=tab_id: self.switch_tab(tid))
                tabs_layout.add_widget(btn)
                self.tab_buttons.append(btn)

            layout.add_widget(tabs_layout)

            self.content_area = BoxLayout(orientation='vertical')
            layout.add_widget(self.content_area)

            # ========== دکمه بازگشت ==========
            back_btn = PersianButton(
                text='بازگشت',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(36),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            back_btn.bind(on_press=self.go_back)
            layout.add_widget(back_btn)

            self.add_widget(layout)

            # نمایش تب پیش‌فرض
            Clock.schedule_once(lambda dt: self.switch_tab(0), 0.1)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ساخت UI SupervisorScreen: {e}", error_details)
            raise

    def switch_tab(self, tab_id):
        try:
            self._cleanup_current_tab()
            
            self.current_tab = tab_id

            for i, btn in enumerate(self.tab_buttons):
                btn.background_color = (0.3, 0.5, 0.8, 1) if i == tab_id else (0.3, 0.5, 0.8, 0.6)

            self.content_area.clear_widgets()
            self.focusable_fields = []
            self.fulfillment_selected = {}
            self._clock_events = []
            self._dt_linked_ids = {}

            if tab_id == 0:
                self.show_targeting_tab()
            elif tab_id == 1:
                self.show_detailed_target_tab()
            elif tab_id == 2:
                self.show_fulfillment_tab()
            elif tab_id == 3:
                self.show_market_check_tab()
            elif tab_id == 4:
                self.show_reports_tab()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در تغییر تب: {e}", error_details)

    def _cleanup_current_tab(self):
        """پاکسازی منابع تب فعلی قبل از switch"""
        try:
            if hasattr(self, '_clock_events'):
                for event in self._clock_events:
                    try:
                        Clock.unschedule(event)
                    except:
                        pass
            
            widget_attrs = [
                'agent_spinner', 'target_type_spinner', 'period_spinner',
                'target_value_input', 'duration_input', 'start_date_input',
                'description_input', 'filter_agent', 'filter_type', 'filter_status',
                'fulfillment_agent', 'fulfillment_start_date', 'fulfillment_end_date',
                'market_route_spinner', 'market_customer_spinner', 'market_visit_type',
                'market_visit_reason', 'market_supervisor_note', 'market_customer_status',
                'market_shelf_status', 'market_monthly_visits', 'market_visit_sufficient',
                'market_expected_purchase', 'market_inventory_status',
                'market_agent_behavior', 'market_distributor_behavior',
                'market_customer_satisfaction', 'market_customer_feedback',
                'market_target_achievement', 'market_supervisor_opinion',
                'market_need_followup', 'market_next_visit_date',
                'fulfillment_file_picker', 'fulfillment_list', 'fulfillment_list_scroll',
                'list_content', 'targets_popup'
                # ❌ dt_* ها کاملاً حذف شدن
            ]
            
            for attr in widget_attrs:
                if hasattr(self, attr):
                    try:
                        widget = getattr(self, attr)
                        if hasattr(widget, 'unbind_all'):
                            try:
                                widget.unbind_all()
                            except:
                                pass
                    except:
                        pass
                    try:
                        delattr(self, attr)
                    except:
                        pass
            
            self._last_market_route_text = ''
            self._last_followup_text = ''
            self._last_dt_agent = ''
            
        except Exception as e:
            print(f"خطا در cleanup: {e}")

    # ============================================================
    # تب ۰: هدفگذاری
    # ============================================================

    def show_targeting_tab(self):
        try:
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint=(1, 1),
                scroll_type=['bars', 'content'],
                bar_width=dp(8)
            )

            content = GridLayout(
                cols=1,
                spacing=dp(10),
                size_hint_y=None,
                padding=dp(12)
            )
            content.bind(minimum_height=content.setter('height'))

            # عنوان
            content.add_widget(RTLLabel(
                text='ثبت تارگت جدید',
                size_hint_y=None,
                height=dp(50),
                font_size=sp(22),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))

            # ========== ۱- کامبوباکس عاملین ==========
            content.add_widget(RTLLabel(
                text='انتخاب عامل:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            agents = get_agents()
            agent_names = [a.get('name', '') for a in agents] if agents else ['']

            self.agent_spinner = PersianComboBox(
                text=agent_names[0] if agent_names else '',
                values=agent_names,
                height=dp(75)
            )
            self.agent_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.agent_spinner.main_btn.color = (1, 1, 1, 1)
            self.agent_spinner.main_btn.font_size = sp(22)
            content.add_widget(self.agent_spinner)

            # ========== ۲- کامبوباکس نوع تارگت ==========
            content.add_widget(RTLLabel(
                text='نوع تارگت:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.target_type_spinner = PersianComboBox(
                text=TARGET_TYPES[0],
                values=TARGET_TYPES,
                height=dp(75)
            )
            self.target_type_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.target_type_spinner.main_btn.color = (1, 1, 1, 1)
            self.target_type_spinner.main_btn.font_size = sp(22)
            content.add_widget(self.target_type_spinner)

            # ========== ۳- میزان هدف ==========
            content.add_widget(RTLLabel(
                text='میزان هدف:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.target_value_input = RTLTextInput(
                hint_text='مقدار عددی را وارد کنید',
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                input_filter='int',
                font_size=sp(22)
            )
            self.target_value_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.target_value_input.border_color = (0.3, 0.3, 0.3, 1)
            self.target_value_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.target_value_input._hidden_input.foreground_color = (1, 1, 1, 1)
            self.target_value_input._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.target_value_input._hidden_input)
            content.add_widget(self.target_value_input)

            # ========== ۴- کامبوباکس دوره ==========
            content.add_widget(RTLLabel(
                text='دوره:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.period_spinner = PersianComboBox(
                text=PERIOD_DISPLAY[0],
                values=PERIOD_DISPLAY,
                height=dp(75)
            )
            self.period_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.period_spinner.main_btn.color = (1, 1, 1, 1)
            self.period_spinner.main_btn.font_size = sp(22)
            content.add_widget(self.period_spinner)

            # ========== ۵- مدت تارگت ==========
            content.add_widget(RTLLabel(
                text='مدت تارگت:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.duration_input = RTLTextInput(
                text='1',
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                input_filter='int',
                font_size=sp(22)
            )
            self.duration_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.duration_input.border_color = (0.3, 0.3, 0.3, 1)
            self.duration_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.duration_input._hidden_input.foreground_color = (1, 1, 1, 1)
            self.duration_input._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.duration_input._hidden_input)
            content.add_widget(self.duration_input)

            # ========== ۶- تاریخ شروع ==========
            content.add_widget(RTLLabel(
                text='تاریخ شروع (سال/ماه/روز):',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.start_date_input = RTLTextInput(
                text=get_today_jalali(),
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                font_size=sp(22)
            )
            self.start_date_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.start_date_input.border_color = (0.3, 0.3, 0.3, 1)
            self.start_date_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.start_date_input._hidden_input.foreground_color = (1, 1, 1, 1)
            self.start_date_input._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.start_date_input._hidden_input)
            content.add_widget(self.start_date_input)

            # ========== ۷- توضیحات ==========
            content.add_widget(RTLLabel(
                text='توضیحات:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.description_input = RTLTextInput(
                hint_text='توضیحات (اختیاری)',
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                font_size=sp(22)
            )
            self.description_input.bg_color = (0.15, 0.15, 0.15, 1)
            self.description_input.border_color = (0.3, 0.3, 0.3, 1)
            self.description_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.description_input._hidden_input.foreground_color = (1, 1, 1, 1)
            self.description_input._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.description_input._hidden_input)
            content.add_widget(self.description_input)

            # ========== ۸- دکمه ثبت تارگت ==========
            btn_layout = BoxLayout(size_hint_y=None, height=dp(65), spacing=dp(10))

            submit_btn = PersianButton(
                text='ثبت تارگت',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None,
                height=dp(58),
                color=(1, 1, 1, 1),
                font_size=sp(20)
            )
            submit_btn.bind(on_press=self.submit_target)
            btn_layout.add_widget(submit_btn)

            list_btn = PersianButton(
                text='نمایش لیست تارگت‌ها',
                background_color=(0.2, 0.5, 0.8, 1),
                size_hint_y=None,
                height=dp(58),
                color=(1, 1, 1, 1),
                font_size=sp(20)
            )
            list_btn.bind(on_press=self.show_targets_list)
            btn_layout.add_widget(list_btn)

            content.add_widget(btn_layout)

            scroll.add_widget(content)
            self.content_area.add_widget(scroll)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب هدفگذاری: {e}", error_details)

    def submit_target(self, instance):
        """ثبت تارگت جدید"""
        try:
            agent_name = self.agent_spinner.text
            target_type = self.target_type_spinner.text
            target_value = self.target_value_input.text.strip()
            period_display = self.period_spinner.text
            duration = self.duration_input.text.strip()
            start_date = self.start_date_input.text.strip()
            description = self.description_input.text.strip()

            # اعتبارسنجی
            if not agent_name or agent_name == '':
                self.show_message('خطا', 'لطفاً یک عامل را انتخاب کنید')
                return

            if not target_value:
                self.show_message('خطا', 'لطفاً میزان هدف را وارد کنید')
                return

            try:
                target_value_int = int(target_value)
                if target_value_int <= 0:
                    self.show_message('خطا', 'میزان هدف باید بزرگتر از صفر باشد')
                    return
            except ValueError:
                self.show_message('خطا', 'میزان هدف باید عددی باشد')
                return

            if not duration:
                self.show_message('خطا', 'لطفاً مدت تارگت را وارد کنید')
                return

            try:
                duration_int = int(duration)
                if duration_int <= 0:
                    self.show_message('خطا', 'مدت تارگت باید بزرگتر از صفر باشد')
                    return
            except ValueError:
                self.show_message('خطا', 'مدت تارگت باید عددی باشد')
                return

            if not start_date:
                self.show_message('خطا', 'لطفاً تاریخ شروع را وارد کنید')
                return

            if not validate_jalali_date(start_date):
                self.show_message('خطا', 'فرمت تاریخ باید سال/ماه/روز باشد (مثال: 1405/01/31)')
                return

            period_type = PERIOD_MAPPING.get(period_display, 'daily')

            # ========== بررسی تکراری نبودن ==========
            from utils.target_manager import check_duplicate_target
            
            is_allowed, dup_msg, existing_target = check_duplicate_target(
                agent_name, target_type, period_type
            )
            
            if not is_allowed:
                # تشخیص نوع پیام: تکراری دقیق یا دوره متفاوت
                if 'تکراری است' in dup_msg:
                    # تکراری دقیق - فقط پیام خطا
                    self.show_message('خطا', dup_msg)
                    return
                else:
                    # دوره متفاوت - نیاز به تأیید کاربر
                    content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
                    with content.canvas.before:
                        Color(0.12, 0.12, 0.12, 1)
                        content_rect = Rectangle(pos=content.pos, size=content.size)
                        content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                                size=lambda i, v: setattr(content_rect, 'size', v))
                    
                    content.add_widget(RTLLabel(
                        text=dup_msg,
                        size_hint_y=None, height=dp(60),
                        font_size=sp(18), color=(1, 0.8, 0.2, 1)
                    ))
                    
                    if existing_target:
                        content.add_widget(RTLLabel(
                            text=f'تارگت موجود: {existing_target.get("target_value", 0):,} {target_type}',
                            size_hint_y=None, height=dp(30),
                            font_size=sp(14), color=(0.6, 0.6, 0.6, 1)
                        ))
                    
                    btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
                    
                    yes_btn = PersianButton(
                        text='بله، ثبت شود',
                        background_color=(0.2, 0.7, 0.2, 1),
                        size_hint_y=None, height=dp(45),
                        color=(1, 1, 1, 1), font_size=sp(16)
                    )
                    no_btn = PersianButton(
                        text='خیر',
                        background_color=(0.8, 0.2, 0.2, 1),
                        size_hint_y=None, height=dp(45),
                        color=(1, 1, 1, 1), font_size=sp(16)
                    )
                    
                    btn_layout.add_widget(yes_btn)
                    btn_layout.add_widget(no_btn)
                    content.add_widget(btn_layout)
                    
                    popup = PersianPopup(
                        title='تأیید ایجاد تارگت',
                        content=content,
                        size_hint=(0.85, 0.4),
                        background_color=(0.08, 0.08, 0.08, 1),
                        auto_dismiss=False
                    )
                    
                    def do_create(inst):
                        popup.dismiss()
                        self._create_target(
                            agent_name, target_type, target_value_int,
                            period_type, duration_int, start_date, description
                        )
                    
                    def cancel_create(inst):
                        popup.dismiss()
                    
                    yes_btn.bind(on_press=do_create)
                    no_btn.bind(on_press=cancel_create)
                    popup.open()
                    return

            # ایجاد تارگت
            self._create_target(
                agent_name, target_type, target_value_int,
                period_type, duration_int, start_date, description
            )

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ثبت تارگت: {e}", error_details)


    def _create_target(self, agent_name, target_type, target_value_int, 
                        period_type, duration_int, start_date, description):
        """ایجاد تارگت جدید - تاریخ پایان توسط target_manager محاسبه می‌شود"""
        try:
            from kivy.app import App
            from utils.target_manager import create_target
            
            # ✅ دریافت نام کاربر جاری
            app = App.get_running_app()
            current_username = app.current_username if hasattr(app, 'current_username') else ''
            
            if not current_username:
                try:
                    from utils.user_manager import get_current_user
                    user = get_current_user()
                    if user:
                        current_username = user.get('username', '') or user.get('name', '')
                except:
                    pass
            
            if not current_username:
                current_username = 'supervisor'
            
            # ✅ ایجاد تارگت - تاریخ پایان داخل create_target محاسبه می‌شود
            success, message, target = create_target(
                agent_name=agent_name,
                target_type=target_type,
                target_value=target_value_int,
                period_type=period_type,
                duration=duration_int,
                start_date=start_date,
                description=description,
                created_by=current_username
            )

            if success:
                self.target_value_input.text = ''
                self.duration_input.text = '1'
                self.start_date_input.text = get_today_jalali()
                self.description_input.text = ''
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)
                
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ایجاد تارگت: {e}", error_details)

    # ============================================================
    # دیالوگ نمایش لیست تارگت‌ها با فیلتر
    # ============================================================

    def show_targets_list(self, instance):
        """نمایش لیست تارگت‌ها با دیالوگ فیلتردار"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))

            filter_layout = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(200))
            filter_layout.bind(minimum_height=filter_layout.setter('height'))

            filter_layout.add_widget(RTLLabel(
                text='عامل:',
                size_hint_y=None, height=dp(35), font_size=sp(16), color=(1, 1, 1, 1)
            ))

            agents = get_agents()
            agent_names = ['همه'] + [a.get('name', '') for a in agents] if agents else ['همه']
            self.filter_agent = PersianComboBox(
                text='همه', values=agent_names, height=dp(65)
            )
            self.filter_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.filter_agent.main_btn.color = (1, 1, 1, 1)
            self.filter_agent.main_btn.font_size = sp(18)
            filter_layout.add_widget(self.filter_agent)

            filter_layout.add_widget(RTLLabel(
                text='نوع تارگت:',
                size_hint_y=None, height=dp(35), font_size=sp(16), color=(1, 1, 1, 1)
            ))

            self.filter_type = PersianComboBox(
                text='همه', values=['همه'] + TARGET_TYPES, height=dp(65)
            )
            self.filter_type.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.filter_type.main_btn.color = (1, 1, 1, 1)
            self.filter_type.main_btn.font_size = sp(18)
            filter_layout.add_widget(self.filter_type)

            filter_layout.add_widget(RTLLabel(
                text='وضعیت:',
                size_hint_y=None, height=dp(35), font_size=sp(16), color=(1, 1, 1, 1)
            ))

            self.filter_status = PersianComboBox(
                text='همه', values=['همه'] + TARGET_STATUSES, height=dp(65)
            )
            self.filter_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.filter_status.main_btn.color = (1, 1, 1, 1)
            self.filter_status.main_btn.font_size = sp(18)
            filter_layout.add_widget(self.filter_status)

            content.add_widget(filter_layout)

            btn_filter_layout = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(10))

            apply_btn = PersianButton(
                text='اعمال فیلتر', background_color=(0.2, 0.6, 1, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(50),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            apply_btn.bind(on_press=self.apply_filter)
            btn_filter_layout.add_widget(apply_btn)

            export_btn = PersianButton(
                text='خروجی اکسل', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(50),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            export_btn.bind(on_press=self.export_filtered_targets)
            btn_filter_layout.add_widget(export_btn)

            content.add_widget(btn_filter_layout)

            scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.5)

            self.list_content = GridLayout(
                cols=1, spacing=dp(6), size_hint_y=None, padding=dp(5)
            )
            self.list_content.bind(minimum_height=self.list_content.setter('height'))

            targets = get_all_targets()
            self._populate_targets_list(self.list_content, targets)

            scroll.add_widget(self.list_content)
            content.add_widget(scroll)

            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(50), color=(1, 1, 1, 1), font_size=sp(18)
            )
            content.add_widget(close_btn)

            self.targets_popup = PersianPopup(
                title='لیست تارگت‌ها', content=content,
                size_hint=(0.92, 0.88), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            close_btn.bind(on_press=self.targets_popup.dismiss)
            self.targets_popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش لیست تارگت‌ها: {e}", error_details)

    def apply_filter(self, instance):
        try:
            agent = self.filter_agent.text
            target_type = self.filter_type.text
            status = self.filter_status.text

            filtered = get_targets_filtered(
                agent_name=agent if agent != 'همه' else None,
                target_type=target_type if target_type != 'همه' else None,
                status=status if status != 'همه' else None
            )

            self.list_content.clear_widgets()
            self._populate_targets_list(self.list_content, filtered)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال فیلتر: {e}", error_details)

    def export_filtered_targets(self, instance):
        try:
            agent = self.filter_agent.text
            target_type = self.filter_type.text
            status = self.filter_status.text

            filtered = get_targets_filtered(
                agent_name=agent if agent != 'همه' else None,
                target_type=target_type if target_type != 'همه' else None,
                status=status if status != 'همه' else None
            )

            if not filtered:
                self.show_message('خطا', 'هیچ تارگتی برای خروجی وجود ندارد')
                return

            success, message, filepath = export_targets_to_excel(filtered)
            if success:
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خروجی اکسل: {e}", error_details)

    def _populate_targets_list(self, list_content, targets):
        try:
            if not targets:
                list_content.add_widget(RTLLabel(
                    text='هیچ تارگتی یافت نشد',
                    size_hint_y=None, height=dp(45), font_size=sp(18), color=(0.5, 0.5, 0.5, 1)
                ))
                return

            for target in targets:
                status = target.get('status', '')
                if status == 'تکمیل شده':
                    bg_color = (0.2, 0.6, 0.2, 0.3)
                elif status == 'فعال':
                    bg_color = (0.2, 0.5, 0.8, 0.3)
                elif status == 'لغو شده':
                    bg_color = (0.8, 0.2, 0.2, 0.3)
                else:
                    bg_color = (0.8, 0.6, 0.2, 0.3)

                box = BoxLayout(
                    orientation='vertical', size_hint_y=None, height=dp(130),
                    spacing=dp(3), padding=[dp(8), dp(6), dp(8), dp(6)]
                )

                with box.canvas.before:
                    Color(*bg_color)
                    rect = Rectangle(pos=box.pos, size=box.size)
                    box.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))

                row1 = BoxLayout(size_hint_y=None, height=dp(30))
                row1.add_widget(RTLLabel(
                    text=f"{target.get('target_id', '')} | {target.get('agent_name', '')}",
                    size_hint_x=0.6, font_size=sp(16), color=(1, 1, 1, 1)
                ))
                row1.add_widget(RTLLabel(
                    text=status, size_hint_x=0.4, font_size=sp(16), color=(1, 1, 1, 1), halign='right'
                ))
                box.add_widget(row1)

                row2 = BoxLayout(size_hint_y=None, height=dp(30))
                row2.add_widget(RTLLabel(
                    text=f"{target.get('target_type', '')}: {target.get('target_value', 0):,}",
                    size_hint_x=0.5, font_size=sp(15), color=(0.8, 0.8, 0.8, 1)
                ))
                row2.add_widget(RTLLabel(
                    text=f"{target.get('start_date', '')} -> {target.get('end_date', '')}",
                    size_hint_x=0.5, font_size=sp(15), color=(0.8, 0.8, 0.8, 1), halign='right'
                ))
                box.add_widget(row2)

                row3 = BoxLayout(size_hint_y=None, height=dp(25))
                achieved = target.get('achieved_value', 0)
                if achieved > 0:
                    row3.add_widget(RTLLabel(
                        text=f"محقق شده: {achieved:,}", size_hint_x=1,
                        font_size=sp(14), color=(0.2, 0.8, 0.2, 1)
                    ))
                else:
                    row3.add_widget(RTLLabel(
                        text="محقق شده: ۰", size_hint_x=1,
                        font_size=sp(14), color=(0.5, 0.5, 0.5, 1)
                    ))
                box.add_widget(row3)

                row4 = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(5))

                if can_edit_target(target):
                    edit_btn = PersianButton(
                        text='ویرایش', size_hint_x=0.33, size_hint_y=None, height=dp(30),
                        background_color=(0.8, 0.6, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(13)
                    )
                    edit_btn.bind(on_press=lambda x, t=target: self._edit_target(t))
                    row4.add_widget(edit_btn)
                else:
                    edit_btn = PersianButton(
                        text='ویرایش', size_hint_x=0.33, size_hint_y=None, height=dp(30),
                        background_color=(0.3, 0.3, 0.3, 1), color=(0.5, 0.5, 0.5, 1),
                        font_size=sp(13), disabled=True
                    )
                    row4.add_widget(edit_btn)

                if status != 'تکمیل شده':
                    delete_btn = PersianButton(
                        text='حذف', size_hint_x=0.33, size_hint_y=None, height=dp(30),
                        background_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(13)
                    )
                    delete_btn.bind(on_press=lambda x, t=target: self._delete_target(t))
                    row4.add_widget(delete_btn)
                else:
                    delete_btn = PersianButton(
                        text='حذف', size_hint_x=0.33, size_hint_y=None, height=dp(30),
                        background_color=(0.3, 0.3, 0.3, 1), color=(0.5, 0.5, 0.5, 1),
                        font_size=sp(13), disabled=True
                    )
                    row4.add_widget(delete_btn)

                row4.add_widget(Label(size_hint_x=0.34))
                box.add_widget(row4)

                list_content.add_widget(box)

        except Exception as e:
            print(f"خطا در پر کردن لیست تارگت‌ها: {e}")

    def _edit_target(self, target):
        try:
            from constants import TARGET_TYPES, TARGET_STATUSES

            if not can_edit_target(target):
                self.show_message('خطا', 'این تارگت قابل ویرایش نیست')
                return

            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'ویرایش تارگت - {target.get("target_id", "")}',
                size_hint_y=None, height=dp(35), font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            content.add_widget(RTLLabel(
                text='نوع تارگت:', size_hint_y=None, height=dp(25), font_size=sp(22), color=(1, 1, 1, 1)
            ))
            edit_type = PersianComboBox(
                text=target.get('target_type', ''), values=TARGET_TYPES, height=dp(55)
            )
            edit_type.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            edit_type.main_btn.color = (1, 1, 1, 1)
            edit_type.main_btn.font_size = sp(22)
            content.add_widget(edit_type)

            content.add_widget(RTLLabel(
                text='میزان هدف:', size_hint_y=None, height=dp(25), font_size=sp(22), color=(1, 1, 1, 1)
            ))
            edit_value = RTLTextInput(
                text=str(target.get('target_value', 0)), multiline=False,
                size_hint_y=None, height=dp(55), input_filter='int', font_size=sp(22)
            )
            edit_value.bg_color = (0.15, 0.15, 0.15, 1)
            edit_value.border_color = (0.3, 0.3, 0.3, 1)
            edit_value.border_color_focus = (0.2, 0.5, 0.9, 1)
            edit_value._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(edit_value)

            content.add_widget(RTLLabel(
                text='مدت (روز):', size_hint_y=None, height=dp(25), font_size=sp(22), color=(1, 1, 1, 1)
            ))
            edit_duration = RTLTextInput(
                text=str(target.get('duration', 0)), multiline=False,
                size_hint_y=None, height=dp(55), input_filter='int', font_size=sp(22)
            )
            edit_duration.bg_color = (0.15, 0.15, 0.15, 1)
            edit_duration.border_color = (0.3, 0.3, 0.3, 1)
            edit_duration.border_color_focus = (0.2, 0.5, 0.9, 1)
            edit_duration._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(edit_duration)

            content.add_widget(RTLLabel(
                text='تاریخ شروع:', size_hint_y=None, height=dp(25), font_size=sp(22), color=(1, 1, 1, 1)
            ))
            edit_start_date = RTLTextInput(
                text=target.get('start_date', ''), multiline=False,
                size_hint_y=None, height=dp(55), font_size=sp(22)
            )
            edit_start_date.bg_color = (0.15, 0.15, 0.15, 1)
            edit_start_date.border_color = (0.3, 0.3, 0.3, 1)
            edit_start_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            edit_start_date._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(edit_start_date)

            content.add_widget(RTLLabel(
                text='وضعیت:', size_hint_y=None, height=dp(25), font_size=sp(22), color=(1, 1, 1, 1)
            ))
            edit_status = PersianComboBox(
                text=target.get('status', ''), values=TARGET_STATUSES, height=dp(55)
            )
            edit_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            edit_status.main_btn.color = (1, 1, 1, 1)
            edit_status.main_btn.font_size = sp(16)
            content.add_widget(edit_status)

            content.add_widget(RTLLabel(
                text='توضیحات:', size_hint_y=None, height=dp(25), font_size=sp(20), color=(1, 1, 1, 1)
            ))
            edit_description = RTLTextInput(
                text=target.get('description', ''), multiline=True,
                size_hint_y=None, height=dp(60), font_size=sp(22)
            )
            edit_description.bg_color = (0.15, 0.15, 0.15, 1)
            edit_description.border_color = (0.3, 0.3, 0.3, 1)
            edit_description.border_color_focus = (0.2, 0.5, 0.9, 1)
            edit_description._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(edit_description)

            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))

            save_btn = PersianButton(
                text='ذخیره تغییرات', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )

            btn_layout.add_widget(save_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='ویرایش تارگت', content=content,
                size_hint=(0.92, 0.85), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_save(instance):
                try:
                    from utils.jalali_date import validate_jalali_date

                    target_id = target.get('target_id')
                    updates = {
                        'target_type': edit_type.text,
                        'target_value': int(edit_value.text) if edit_value.text else 0,
                        'duration': int(edit_duration.text) if edit_duration.text else 0,
                        'start_date': edit_start_date.text,
                        'status': edit_status.text,
                        'description': edit_description.text
                    }

                    if updates['target_value'] <= 0:
                        self.show_message('خطا', 'میزان هدف باید بزرگتر از صفر باشد')
                        return

                    if updates['duration'] <= 0:
                        self.show_message('خطا', 'مدت باید بزرگتر از صفر باشد')
                        return

                    if not validate_jalali_date(updates['start_date']):
                        self.show_message('خطا', 'تاریخ شروع نامعتبر است')
                        return

                    success, message = update_target(target_id, updates)
                    popup.dismiss()

                    if success:
                        self.show_message('موفق', message)
                        self.show_targets_list(None)
                    else:
                        self.show_message('خطا', message)

                except Exception as e:
                    error_details = traceback.format_exc()
                    ErrorPopup.show_error(f"خطا در ذخیره تغییرات: {e}", error_details)

            save_btn.bind(on_press=do_save)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ویرایش تارگت: {e}", error_details)

    def _delete_target(self, target):
        try:
            from utils.target_manager import delete_target

            status = target.get('status', '')
            if status == 'تکمیل شده':
                self.show_message('خطا', 'تارگت‌های نهایی شده قابل حذف نیستند')
                return

            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                           size=lambda i, v: setattr(content_rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'آیا از حذف تارگت "{target.get("target_id", "")}" اطمینان دارید؟',
                size_hint_y=None, height=dp(45), font_size=sp(18), color=(1, 0.8, 0.2, 1)
            ))

            content.add_widget(RTLLabel(
                text=f'عامل: {target.get("agent_name", "")}\nنوع: {target.get("target_type", "")}\nمیزان: {target.get("target_value", 0):,}',
                size_hint_y=None, height=dp(50), font_size=sp(14), color=(0.8, 0.8, 0.8, 1)
            ))

            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))

            confirm_btn = PersianButton(
                text='بله، حذف شود', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )

            btn_layout.add_widget(confirm_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='تأیید حذف', content=content,
                size_hint=(0.85, 0.45), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_delete(instance):
                popup.dismiss()
                target_id = target.get('target_id')
                success, message = delete_target(target_id)
                if success:
                    self.show_message('موفق', message)
                    self.show_targets_list(None)
                else:
                    self.show_message('خطا', message)

            def cancel_delete(instance):
                popup.dismiss()

            confirm_btn.bind(on_press=do_delete)
            cancel_btn.bind(on_press=cancel_delete)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در حذف تارگت: {e}", error_details)

    # ============================================================
    # تب ۱: ریزتارگت (Detailed Target)
    # ============================================================

    def show_detailed_target_tab(self):
        """نمایش تب ریزتارگت"""
        try:
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint=(1, 1),
                scroll_type=['bars', 'content'],
                bar_width=dp(8)
            )

            content = GridLayout(
                cols=1,
                spacing=dp(10),
                size_hint_y=None,
                padding=dp(12)
            )
            content.bind(minimum_height=content.setter('height'))

            # ========== عنوان ==========
            content.add_widget(RTLLabel(
                text='ثبت ریزتارگت جدید',
                size_hint_y=None,
                height=dp(50),
                font_size=sp(22),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))

            # ========== 🆕 فیلتر لیست ==========
            content.add_widget(RTLLabel(
                text='فیلتر نمایش:',
                size_hint_y=None,
                height=dp(30),
                font_size=sp(16),
                color=(1, 0.8, 0.2, 1),
                bold=True
            ))

            filter_row = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(5))

            # فیلتر عامل
            agents = get_agents()
            agent_names = ['همه'] + [a.get('name', '') for a in agents] if agents else ['همه']
            
            self.dt_filter_agent = PersianComboBox(
                text='همه',
                values=agent_names,
                height=dp(55)
            )
            self.dt_filter_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_filter_agent.main_btn.color = (1, 1, 1, 1)
            self.dt_filter_agent.main_btn.font_size = sp(14)
            self.dt_filter_agent.size_hint_x = 0.4
            filter_row.add_widget(self.dt_filter_agent)

            # فیلتر گروه کالا
            product_groups = get_product_groups()
            if not isinstance(product_groups, list):
                product_groups = []
            if not product_groups:
                product_groups = ['']
            product_filter_list = ['همه'] + product_groups
            
            self.dt_filter_product = PersianComboBox(
                text='همه',
                values=product_filter_list,
                height=dp(55)
            )
            self.dt_filter_product.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_filter_product.main_btn.color = (1, 1, 1, 1)
            self.dt_filter_product.main_btn.font_size = sp(14)
            self.dt_filter_product.size_hint_x = 0.4
            filter_row.add_widget(self.dt_filter_product)

            # دکمه اعمال فیلتر
            apply_filter_btn = PersianButton(
                text='اعمال',
                size_hint_x=0.2,
                size_hint_y=None,
                height=dp(55),
                background_color=(0.2, 0.6, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(14)
            )
            apply_filter_btn.bind(on_press=self._show_filtered_detailed_targets)
            filter_row.add_widget(apply_filter_btn)

            content.add_widget(filter_row)

            # ========== ۱- انتخاب عامل ==========
            content.add_widget(RTLLabel(
                text='انتخاب عامل:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.dt_agent_spinner = PersianComboBox(
                text=agent_names[0] if agent_names else '',
                values=agent_names,
                height=dp(75)
            )
            self.dt_agent_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_agent_spinner.main_btn.color = (1, 1, 1, 1)
            self.dt_agent_spinner.main_btn.font_size = sp(22)
            content.add_widget(self.dt_agent_spinner)

            # ========== ۲- گروه کالا (با دکمه افزودن) ==========
            content.add_widget(RTLLabel(
                text='گروه کالا:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            product_row = BoxLayout(
                size_hint_y=None,
                height=dp(75),
                spacing=dp(5)
            )

            product_groups = get_product_groups()
            if not isinstance(product_groups, list):
                product_groups = []
            if not product_groups:
                product_groups = ['']

            self.dt_product_spinner = PersianComboBox(
                text=product_groups[0] if product_groups else '',
                values=product_groups,
                height=dp(75)
            )
            self.dt_product_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_product_spinner.main_btn.color = (1, 1, 1, 1)
            self.dt_product_spinner.main_btn.font_size = sp(22)
            self.dt_product_spinner.size_hint_x = 0.85
            product_row.add_widget(self.dt_product_spinner)

            add_product_btn = PersianButton(
                text='+',
                size_hint_x=0.15,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.2, 0.6, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(28),
                bold=True
            )
            add_product_btn.bind(on_press=self._show_add_product_dialog)
            product_row.add_widget(add_product_btn)

            content.add_widget(product_row)

            # ========== ۳- تعداد هدف ==========
            content.add_widget(RTLLabel(
                text='تعداد هدف:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.dt_target_count = RTLTextInput(
                text='0',
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                input_filter='int',
                font_size=sp(22)
            )
            self.dt_target_count.bg_color = (0.15, 0.15, 0.15, 1)
            self.dt_target_count.border_color = (0.3, 0.3, 0.3, 1)
            self.dt_target_count.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.dt_target_count._hidden_input.foreground_color = (1, 1, 1, 1)
            self.dt_target_count._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.dt_target_count._hidden_input)
            content.add_widget(self.dt_target_count)

            # ========== ۴- واحد تارگت (با دکمه‌های مدیریت) ==========
            content.add_widget(RTLLabel(
                text='واحد تارگت:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            unit_row = BoxLayout(
                size_hint_y=None,
                height=dp(75),
                spacing=dp(5)
            )

            units = get_target_units()
            if not isinstance(units, list):
                units = ["کارتن", "عدد", "شل", "بسته", "جعبه", "بانکه"]
            
            self.dt_unit_spinner = PersianComboBox(
                text=units[0] if units else '',
                values=units,
                height=dp(75)
            )
            self.dt_unit_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_unit_spinner.main_btn.color = (1, 1, 1, 1)
            self.dt_unit_spinner.main_btn.font_size = sp(22)
            self.dt_unit_spinner.size_hint_x = 0.6
            unit_row.add_widget(self.dt_unit_spinner)

            add_unit_btn = PersianButton(
                text='+',
                size_hint_x=0.13,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.2, 0.6, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(28),
                bold=True
            )
            add_unit_btn.bind(on_press=self._show_add_unit_dialog)
            unit_row.add_widget(add_unit_btn)

            edit_unit_btn = PersianButton(
                text='<>',
                size_hint_x=0.13,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.8, 0.6, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(24)
            )
            edit_unit_btn.bind(on_press=self._show_edit_unit_dialog)
            unit_row.add_widget(edit_unit_btn)

            del_unit_btn = PersianButton(
                text='-',
                size_hint_x=0.14,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.8, 0.2, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(24)
            )
            del_unit_btn.bind(on_press=self._show_delete_unit_dialog)
            unit_row.add_widget(del_unit_btn)

            content.add_widget(unit_row)

            # ========== ۵- دوره تارگت (با دکمه‌های مدیریت) ==========
            content.add_widget(RTLLabel(
                text='دوره تارگت:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            period_row = BoxLayout(
                size_hint_y=None,
                height=dp(75),
                spacing=dp(5)
            )

            periods = get_target_periods()
            if not isinstance(periods, list):
                periods = ["روزانه", "ماهانه", "فصلی", "سالیانه"]
            
            self.dt_period_spinner = PersianComboBox(
                text=periods[0] if periods else '',
                values=periods,
                height=dp(75)
            )
            self.dt_period_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_period_spinner.main_btn.color = (1, 1, 1, 1)
            self.dt_period_spinner.main_btn.font_size = sp(22)
            self.dt_period_spinner.size_hint_x = 0.6
            period_row.add_widget(self.dt_period_spinner)

            add_period_btn = PersianButton(
                text='+',
                size_hint_x=0.13,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.2, 0.6, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(28),
                bold=True
            )
            add_period_btn.bind(on_press=self._show_add_period_dialog)
            period_row.add_widget(add_period_btn)

            edit_period_btn = PersianButton(
                text='<>',
                size_hint_x=0.13,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.8, 0.6, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(24)
            )
            edit_period_btn.bind(on_press=self._show_edit_period_dialog)
            period_row.add_widget(edit_period_btn)

            del_period_btn = PersianButton(
                text='-',
                size_hint_x=0.14,
                size_hint_y=None,
                height=dp(75),
                background_color=(0.8, 0.2, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(24)
            )
            del_period_btn.bind(on_press=self._show_delete_period_dialog)
            period_row.add_widget(del_period_btn)

            content.add_widget(period_row)

            # ========== ۶- مدت ==========
            content.add_widget(RTLLabel(
                text='مدت (بر اساس دوره):',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            self.dt_duration = RTLTextInput(
                text='1',
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                input_filter='int',
                font_size=sp(22)
            )
            self.dt_duration.bg_color = (0.15, 0.15, 0.15, 1)
            self.dt_duration.border_color = (0.3, 0.3, 0.3, 1)
            self.dt_duration.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.dt_duration._hidden_input.foreground_color = (1, 1, 1, 1)
            self.dt_duration._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.dt_duration._hidden_input)
            content.add_widget(self.dt_duration)

            # ========== ۷- تاریخ شروع ==========
            content.add_widget(RTLLabel(
                text='تاریخ شروع (سال/ماه/روز):',
                size_hint_y=None, height=dp(35), font_size=sp(18), color=(1, 1, 1, 1)
            ))

            self.dt_start_date = RTLTextInput(
                text=get_today_jalali(),
                multiline=False,
                size_hint_y=None,
                height=dp(80),
                font_size=sp(22)
            )
            self.dt_start_date.bg_color = (0.15, 0.15, 0.15, 1)
            self.dt_start_date.border_color = (0.3, 0.3, 0.3, 1)
            self.dt_start_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.dt_start_date._hidden_input.foreground_color = (1, 1, 1, 1)
            self.dt_start_date._hidden_input.bind(focus=self._on_field_focus)
            self.focusable_fields.append(self.dt_start_date._hidden_input)
            content.add_widget(self.dt_start_date)

            # ========== ۸- پیوند به تارگت اصلی ==========
            content.add_widget(RTLLabel(
                text='پیوند به تارگت اصلی:',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))

            all_targets = get_all_targets()
            if not isinstance(all_targets, list):
                all_targets = []
            
            agent_name = self.dt_agent_spinner.text
            unfulfilled = [t for t in all_targets 
                        if isinstance(t, dict)
                        and t.get('agent_name') == agent_name 
                        and t.get('status') in ['در انتظار', 'فعال']]
            
            target_labels = [f"{t.get('target_id')} | {t.get('target_type')} | {t.get('target_value'):,}" 
                            for t in unfulfilled] if unfulfilled else ['هیچ تارگت فعالی نیست']
            
            self.dt_linked_target = PersianComboBox(
                text=target_labels[0],
                values=target_labels,
                height=dp(75)
            )
            self.dt_linked_target.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.dt_linked_target.main_btn.color = (1, 1, 1, 1)
            self.dt_linked_target.main_btn.font_size = sp(18)
            content.add_widget(self.dt_linked_target)

            self._dt_linked_ids = {}
            for t, label in zip(unfulfilled, target_labels):
                self._dt_linked_ids[label] = t.get('target_id')

            # ========== ۹- دکمه‌ها ==========
            btn_layout = BoxLayout(size_hint_y=None, height=dp(65), spacing=dp(10))

            submit_btn = PersianButton(
                text='ثبت ریزتارگت',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.33,
                size_hint_y=None,
                height=dp(58),
                color=(1, 1, 1, 1),
                font_size=sp(18)
            )
            submit_btn.bind(on_press=self._submit_detailed_target)
            btn_layout.add_widget(submit_btn)

            list_btn = PersianButton(
                text='لیست ریزتارگت‌ها',
                background_color=(0.2, 0.5, 0.8, 1),
                size_hint_x=0.33,
                size_hint_y=None,
                height=dp(58),
                color=(1, 1, 1, 1),
                font_size=sp(18)
            )
            list_btn.bind(on_press=self._show_detailed_targets_list)
            btn_layout.add_widget(list_btn)

            excel_btn = PersianButton(
                text='خروجی اکسل',
                background_color=(0.2, 0.7, 0.4, 1),
                size_hint_x=0.34,
                size_hint_y=None,
                height=dp(58),
                color=(1, 1, 1, 1),
                font_size=sp(18)
            )
            excel_btn.bind(on_press=self._export_detailed_targets_excel)
            btn_layout.add_widget(excel_btn)

            content.add_widget(btn_layout)

            scroll.add_widget(content)
            self.content_area.add_widget(scroll)

            # بروزرسانی تارگت‌های پیوند با تغییر عامل
            self._last_dt_agent = self.dt_agent_spinner.text
            event = Clock.schedule_interval(self._check_dt_agent_change, 0.5)
            if not hasattr(self, '_clock_events'):
                self._clock_events = []
            self._clock_events.append(event)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب ریزتارگت: {e}", error_details)

    def _check_dt_agent_change(self, dt):
        """بررسی تغییر عامل در تب ریزتارگت"""
        if not hasattr(self, 'dt_agent_spinner') or not hasattr(self, 'dt_linked_target'):
            return
        try:
            current = self.dt_agent_spinner.text
            if current != self._last_dt_agent:
                self._last_dt_agent = current
                self._update_dt_linked_targets()
        except Exception as e:
            print(f"خطا در _check_dt_agent_change: {e}")


    def _update_dt_linked_targets(self):
        """بروزرسانی لیست تارگت‌های پیوند بر اساس عامل انتخاب‌شده"""
        try:
            if not hasattr(self, 'dt_linked_target') or not hasattr(self, 'dt_agent_spinner'):
                return
            
            agent_name = self.dt_agent_spinner.text
            all_targets = get_all_targets()
            if not isinstance(all_targets, list):
                all_targets = []
            
            unfulfilled = [t for t in all_targets 
                        if isinstance(t, dict)
                        and t.get('agent_name') == agent_name 
                        and t.get('status') in ['در انتظار', 'فعال']]
            
            target_labels = [f"{t.get('target_id')} | {t.get('target_type')} | {t.get('target_value'):,}" 
                            for t in unfulfilled] if unfulfilled else ['هیچ تارگت فعالی نیست']
            
            self.dt_linked_target.values = target_labels
            self.dt_linked_target.text = target_labels[0]
            
            self._dt_linked_ids = {}
            for t, label in zip(unfulfilled, target_labels):
                self._dt_linked_ids[label] = t.get('target_id')
                
        except Exception as e:
            print(f"خطا در بروزرسانی تارگت‌های پیوند: {e}")


    # ============================================================
    # دیالوگ‌های مدیریت گروه کالا
    # ============================================================

    def _show_add_product_dialog(self, instance):
        """دیالوگ افزودن گروه کالا"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))

            content.add_widget(RTLLabel(
                text='افزودن گروه کالای جدید',
                size_hint_y=None, height=dp(40),
                font_size=sp(20), bold=True, color=(1, 1, 1, 1)
            ))

            name_input = RTLTextInput(
                hint_text='نام گروه کالا',
                multiline=False, size_hint_y=None, height=dp(65), font_size=sp(20)
            )
            name_input.bg_color = (0.15, 0.15, 0.15, 1)
            name_input.border_color = (0.3, 0.3, 0.3, 1)
            name_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            name_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(name_input)

            btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
            save_btn = PersianButton(
                text='ذخیره', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(save_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='گروه کالا', content=content,
                size_hint=(0.85, 0.4), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_save(inst):
                name = name_input.text.strip()
                if not name:
                    self.show_message('خطا', 'نام گروه کالا را وارد کنید')
                    return
                if add_product_group(name):
                    self.show_message('موفق', f'گروه کالا "{name}" اضافه شد')
                    popup.dismiss()
                    self.switch_tab(1)
                else:
                    self.show_message('خطا', 'این گروه کالا قبلاً وجود دارد')

            save_btn.bind(on_press=do_save)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)

    # ============================================================
    # دیالوگ‌های عمومی مدیریت (واحد و دوره)
    # ============================================================
    def _show_filtered_detailed_targets(self, instance):
        """نمایش لیست فیلتر شده ریزتارگت‌ها"""
        try:
            from utils.detailed_target_manager import (
                get_detailed_targets_filtered, can_edit_target as can_edit_dt,
                export_to_excel
            )
            
            # خواندن فیلترها
            agent_filter = self.dt_filter_agent.text if hasattr(self, 'dt_filter_agent') else 'همه'
            product_filter = self.dt_filter_product.text if hasattr(self, 'dt_filter_product') else 'همه'
            status_filter = self.dt_filter_status.text if hasattr(self, 'dt_filter_status') else 'همه'
            id_search = self.dt_filter_id.text.strip() if hasattr(self, 'dt_filter_id') else ''
            start_date = self.dt_filter_start.text.strip() if hasattr(self, 'dt_filter_start') else ''
            end_date = self.dt_filter_end.text.strip() if hasattr(self, 'dt_filter_end') else ''
            
            # ✅ استفاده از تابع فیلتر جدید
            filtered = get_detailed_targets_filtered(
                agent_name=agent_filter if agent_filter != 'همه' else None,
                product_group=product_filter if product_filter != 'همه' else None,
                status=status_filter if status_filter != 'همه' else None,
                target_id=id_search if id_search else None,
                linked_target_id=id_search if id_search else None,
                start_date=start_date if start_date else None,
                end_date=end_date if end_date else None
            )
            
            # ========== ساخت دیالوگ ==========
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))

            # ✅ عنوان با خلاصه فیلترها
            title_parts = ['لیست ریزتارگت‌ها']
            if agent_filter != 'همه':
                title_parts.append(f'عامل: {agent_filter}')
            if product_filter != 'همه':
                title_parts.append(f'گروه: {product_filter}')
            if status_filter != 'همه':
                title_parts.append(status_filter)
            if id_search:
                title_parts.append(f'جستجو: {id_search}')
            
            content.add_widget(RTLLabel(
                text=f'{" | ".join(title_parts)} ({len(filtered)} مورد)',
                size_hint_y=None, height=dp(40),
                font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint_y=0.75, scroll_type=['bars', 'content'], bar_width=dp(6)
            )

            list_content = GridLayout(
                cols=1, spacing=dp(10), size_hint_y=None, padding=dp(5)
            )
            list_content.bind(minimum_height=list_content.setter('height'))

            if not filtered:
                list_content.add_widget(RTLLabel(
                    text='هیچ ریزتارگتی با این فیلترها یافت نشد',
                    size_hint_y=None, height=dp(45),
                    font_size=sp(18), color=(0.5, 0.5, 0.5, 1)
                ))
            else:
                # ✅ تابع کمکی
                def make_field(text, size_x, font_sz=16, fg=(1,1,1,1), halign='center'):
                    fld = RTLTextInput(
                        text=str(text) if text else '', multiline=False,
                        size_hint_x=size_x, size_hint_y=None, height=dp(36),
                        font_size=sp(font_sz)
                    )
                    fld.bg_color = (0.1, 0.1, 0.1, 1)
                    fld.border_color = (0.3, 0.3, 0.3, 1)
                    fld.border_color_focus = (0.3, 0.3, 0.3, 1)
                    fld._hidden_input.foreground_color = fg
                    fld._hidden_input.disabled = True
                    fld._hidden_input.halign = halign
                    return fld

                for target in filtered:
                    if not isinstance(target, dict):
                        continue
                        
                    status = target.get('status', '')
                    editable = can_edit_dt(target)
                    is_completed = status == 'تکمیل شده'

                    # ... (دقیقاً مثل _show_detailed_targets_list - ساختار کارت)
                    # [اینجا کد کارت مثل تابع بالا قرار بگیره]

            scroll.add_widget(list_content)
            content.add_widget(scroll)

            # دکمه‌های پایین
            btn_row = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
            export_btn = PersianButton(
                text='📥 خروجی اکسل از این لیست', background_color=(0.2, 0.7, 0.4, 1),
                size_hint_x=0.6, size_hint_y=None, height=dp(42),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_row.add_widget(export_btn)
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.4, size_hint_y=None, height=dp(42),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_row.add_widget(close_btn)
            content.add_widget(btn_row)

            popup = PersianPopup(
                title='ریزتارگت‌ها', content=content,
                size_hint=(0.94, 0.88), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def export_filtered(inst):
                success, message, filepath = export_to_excel(filtered)
                if success:
                    self.show_message('موفق', message)
                else:
                    self.show_message('خطا', message)

            export_btn.bind(on_press=export_filtered)
            close_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش لیست: {e}", error_details)


    def _show_manage_item_dialog(self, title, hint, save_func, success_msg, error_msg, default_text=''):
        """دیالوگ عمومی برای افزودن/ویرایش"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))

            content.add_widget(RTLLabel(
                text=title, size_hint_y=None, height=dp(40),
                font_size=sp(20), bold=True, color=(1, 1, 1, 1)
            ))

            name_input = RTLTextInput(
                text=default_text, hint_text=hint,
                multiline=False, size_hint_y=None, height=dp(65), font_size=sp(20)
            )
            name_input.bg_color = (0.15, 0.15, 0.15, 1)
            name_input.border_color = (0.3, 0.3, 0.3, 1)
            name_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            name_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(name_input)

            btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
            save_btn = PersianButton(
                text='ذخیره', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(save_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.85, 0.4), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_save(inst):
                name = name_input.text.strip()
                if not name:
                    self.show_message('خطا', 'مقدار را وارد کنید')
                    return
                if save_func(name):
                    self.show_message('موفق', success_msg)
                    popup.dismiss()
                    self.switch_tab(1)
                else:
                    self.show_message('خطا', error_msg)

            save_btn.bind(on_press=do_save)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)

    def _show_delete_item_dialog(self, title, item_name, delete_func, success_msg, error_msg):
        """دیالوگ عمومی برای حذف"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'آیا از حذف "{item_name}" اطمینان دارید؟',
                size_hint_y=None, height=dp(50),
                font_size=sp(18), color=(1, 0.8, 0.2, 1)
            ))

            btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
            confirm_btn = PersianButton(
                text='بله، حذف شود', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(confirm_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.8, 0.35), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_delete(inst):
                if delete_func():
                    self.show_message('موفق', success_msg)
                    popup.dismiss()
                    self.switch_tab(1)
                else:
                    self.show_message('خطا', error_msg)

            confirm_btn.bind(on_press=do_delete)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)

    # ============================================================
    # دیالوگ‌های واحد تارگت
    # ============================================================

    def _show_add_unit_dialog(self, instance):
        self._show_manage_item_dialog(
            title='افزودن واحد جدید',
            hint='نام واحد',
            save_func=lambda name: add_target_unit(name),
            success_msg='واحد اضافه شد',
            error_msg='این واحد قبلاً وجود دارد'
        )

    def _show_edit_unit_dialog(self, instance):
        old_name = self.dt_unit_spinner.text
        self._show_manage_item_dialog(
            title='ویرایش واحد',
            hint='نام جدید واحد',
            default_text=old_name,
            save_func=lambda new_name: update_target_unit(old_name, new_name),
            success_msg='واحد ویرایش شد',
            error_msg='خطا در ویرایش'
        )

    def _show_delete_unit_dialog(self, instance):
        name = self.dt_unit_spinner.text
        self._show_delete_item_dialog(
            title='حذف واحد',
            item_name=name,
            delete_func=lambda: delete_target_unit(name),
            success_msg='واحد حذف شد',
            error_msg='حداقل ۲ واحد باید باقی بماند'
        )

    # ============================================================
    # دیالوگ‌های دوره تارگت
    # ============================================================

    def _show_add_period_dialog(self, instance):
        self._show_manage_item_dialog(
            title='افزودن دوره جدید',
            hint='نام دوره',
            save_func=lambda name: add_target_period(name),
            success_msg='دوره اضافه شد',
            error_msg='این دوره قبلاً وجود دارد'
        )

    def _show_edit_period_dialog(self, instance):
        old_name = self.dt_period_spinner.text
        self._show_manage_item_dialog(
            title='ویرایش دوره',
            hint='نام جدید دوره',
            default_text=old_name,
            save_func=lambda new_name: update_target_period(old_name, new_name),
            success_msg='دوره ویرایش شد',
            error_msg='خطا در ویرایش'
        )

    def _show_delete_period_dialog(self, instance):
        name = self.dt_period_spinner.text
        self._show_delete_item_dialog(
            title='حذف دوره',
            item_name=name,
            delete_func=lambda: delete_target_period(name),
            success_msg='دوره حذف شد',
            error_msg='حداقل ۲ دوره باید باقی بماند'
        )

    # ============================================================
    # ثبت، لیست و خروجی ریزتارگت
    # ============================================================

    def _submit_detailed_target(self, instance):
        """ثبت ریزتارگت جدید"""
        try:
            # ========== بررسی وجود تمام فیلدهای ضروری ==========
            required_attrs = {
                'dt_agent_spinner': 'انتخاب عامل',
                'dt_product_spinner': 'گروه کالا',
                'dt_target_count': 'تعداد هدف',
                'dt_unit_spinner': 'واحد تارگت',
                'dt_period_spinner': 'دوره تارگت',
                'dt_duration': 'مدت',
                'dt_linked_target': 'پیوند به تارگت اصلی',
                'dt_start_date': 'تاریخ شروع'
            }
            
            for attr, name in required_attrs.items():
                if not hasattr(self, attr):
                    self.show_message('خطا', f'فیلد "{name}" در دسترس نیست. لطفاً دوباره تلاش کنید.')
                    return
            
            agent_name = self.dt_agent_spinner.text
            product_group = self.dt_product_spinner.text
            target_count = self.dt_target_count.text.strip()
            unit = self.dt_unit_spinner.text
            period = self.dt_period_spinner.text
            duration = self.dt_duration.text.strip()
            linked_label = self.dt_linked_target.text
            start_date = self.dt_start_date.text.strip()

            if not agent_name:
                self.show_message('خطا', 'لطفاً یک عامل را انتخاب کنید')
                return

            if not product_group or product_group == '':
                self.show_message('خطا', 'لطفاً یک گروه کالا را انتخاب کنید')
                return

            try:
                target_count_int = int(target_count) if target_count else 0
                if target_count_int <= 0:
                    self.show_message('خطا', 'تعداد هدف باید بزرگتر از صفر باشد')
                    return
            except ValueError:
                self.show_message('خطا', 'تعداد هدف باید عددی باشد')
                return

            if not unit:
                self.show_message('خطا', 'لطفاً واحد تارگت را انتخاب کنید')
                return

            try:
                duration_int = int(duration) if duration else 1
                if duration_int <= 0:
                    self.show_message('خطا', 'مدت باید بزرگتر از صفر باشد')
                    return
            except ValueError:
                self.show_message('خطا', 'مدت باید عددی باشد')
                return

            if not start_date:
                self.show_message('خطا', 'لطفاً تاریخ شروع را وارد کنید')
                return

            if not validate_jalali_date(start_date):
                self.show_message('خطا', 'فرمت تاریخ شروع نامعتبر است (مثال: 1405/01/31)')
                return

            linked_id = self._dt_linked_ids.get(linked_label, '') if hasattr(self, '_dt_linked_ids') else ''
            if not linked_id or linked_label == 'هیچ تارگت فعالی نیست':
                self.show_message('خطا', 'لطفاً یک تارگت اصلی معتبر انتخاب کنید')
                return

            # ========== بررسی تکراری نبودن ==========
            from utils.detailed_target_manager import check_duplicate_detailed_target
            
            is_allowed, dup_msg, existing_target = check_duplicate_detailed_target(
                agent_name, product_group, period, linked_id
            )
            
            if not is_allowed:
                if 'تکراری است' in dup_msg:
                    # تکراری دقیق
                    self.show_message('خطا', dup_msg)
                    return
                else:
                    # دوره متفاوت - نیاز به تأیید کاربر
                    content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
                    with content.canvas.before:
                        Color(0.12, 0.12, 0.12, 1)
                        content_rect = Rectangle(pos=content.pos, size=content.size)
                        content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                                size=lambda i, v: setattr(content_rect, 'size', v))
                    
                    content.add_widget(RTLLabel(
                        text=dup_msg,
                        size_hint_y=None, height=dp(60),
                        font_size=sp(18), color=(1, 0.8, 0.2, 1)
                    ))
                    
                    if existing_target:
                        content.add_widget(RTLLabel(
                            text=f'تارگت موجود: {existing_target.get("target_count", 0):,} {existing_target.get("unit", "")}',
                            size_hint_y=None, height=dp(30),
                            font_size=sp(14), color=(0.6, 0.6, 0.6, 1)
                        ))
                    
                    btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
                    
                    yes_btn = PersianButton(
                        text='بله، ثبت شود',
                        background_color=(0.2, 0.7, 0.2, 1),
                        size_hint_y=None, height=dp(45),
                        color=(1, 1, 1, 1), font_size=sp(16)
                    )
                    no_btn = PersianButton(
                        text='خیر',
                        background_color=(0.8, 0.2, 0.2, 1),
                        size_hint_y=None, height=dp(45),
                        color=(1, 1, 1, 1), font_size=sp(16)
                    )
                    
                    btn_layout.add_widget(yes_btn)
                    btn_layout.add_widget(no_btn)
                    content.add_widget(btn_layout)
                    
                    popup = PersianPopup(
                        title='تأیید ایجاد ریزتارگت',
                        content=content,
                        size_hint=(0.85, 0.4),
                        background_color=(0.08, 0.08, 0.08, 1),
                        auto_dismiss=False
                    )
                    
                    def do_create(inst):
                        popup.dismiss()
                        self._create_detailed_target(
                            agent_name, product_group, target_count_int,
                            unit, period, duration_int, linked_id, start_date
                        )
                    
                    def cancel_create(inst):
                        popup.dismiss()
                    
                    yes_btn.bind(on_press=do_create)
                    no_btn.bind(on_press=cancel_create)
                    popup.open()
                    return

            # ایجاد ریزتارگت
            self._create_detailed_target(
                agent_name, product_group, target_count_int,
                unit, period, duration_int, linked_id, start_date
            )

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ثبت ریزتارگت: {e}", error_details)


    def _create_detailed_target(self, agent_name, product_group, target_count_int,
                                unit, period, duration_int, linked_id, start_date):
        """ایجاد ریزتارگت جدید با نام کاربر جاری"""
        try:
            from kivy.app import App
            
            # ✅ دریافت نام کاربر جاری
            app = App.get_running_app()
            current_username = app.current_username if hasattr(app, 'current_username') else ''
            
            if not current_username:
                try:
                    from utils.user_manager import get_current_user
                    user = get_current_user()
                    if user:
                        current_username = user.get('username', '') or user.get('name', '')
                except:
                    pass
            
            if not current_username:
                current_username = 'supervisor'
            
            success, message, target = create_detailed_target(
                agent_name=agent_name,
                product_group=product_group,
                target_count=target_count_int,
                unit=unit,
                period=period,
                duration=duration_int,
                linked_target_id=linked_id,
                start_date=start_date,
                created_by=current_username  # ✅ اضافه شد
            )

            if success:
                self.dt_target_count.text = '0'
                self.dt_duration.text = '1'
                self.dt_start_date.text = get_today_jalali()
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)
                
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ایجاد ریزتارگت: {e}", error_details)

    def _show_detailed_targets_list(self, instance):
        """نمایش لیست ریزتارگت‌ها با فیلتر در دیالوگ"""
        try:
            from utils.detailed_target_manager import (
                get_detailed_targets_filtered, can_edit_target as can_edit_dt,
                export_to_excel
            )

            all_targets = get_all_detailed_targets()
            if not isinstance(all_targets, list):
                all_targets = []

            content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(5))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))

            # ========== فیلترها - ردیف ۱ ==========
            filter_row1 = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(3))
            
            agents = get_agents()
            agent_names = ['همه'] + [a.get('name', '') for a in agents] if agents else ['همه']
            filter_agent = PersianComboBox(text='همه', values=agent_names, height=dp(42))
            filter_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            filter_agent.main_btn.color = (1, 1, 1, 1)
            filter_agent.main_btn.font_size = sp(16)
            filter_agent.size_hint_x = 0.33
            filter_row1.add_widget(filter_agent)

            product_groups = get_product_groups()
            if not isinstance(product_groups, list):
                product_groups = []
            product_filter_list = ['همه'] + product_groups if product_groups else ['همه']
            filter_product = PersianComboBox(text='همه', values=product_filter_list, height=dp(42))
            filter_product.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            filter_product.main_btn.color = (1, 1, 1, 1)
            filter_product.main_btn.font_size = sp(16)
            filter_product.size_hint_x = 0.33
            filter_row1.add_widget(filter_product)

            status_list = ['همه', 'در انتظار', 'فعال', 'تکمیل شده', 'لغو شده']
            filter_status = PersianComboBox(text='همه', values=status_list, height=dp(42))
            filter_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            filter_status.main_btn.color = (1, 1, 1, 1)
            filter_status.main_btn.font_size = sp(16)
            filter_status.size_hint_x = 0.34
            filter_row1.add_widget(filter_status)
            
            content.add_widget(filter_row1)

            # ========== فیلترها - ردیف ۲ ==========
            filter_row2 = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(3))
            
            filter_id = RTLTextInput(
                hint_text='شناسه...', multiline=False,
                size_hint_x=0.25, size_hint_y=None, height=dp(42), font_size=sp(18)
            )
            filter_id.bg_color = (0.15, 0.15, 0.15, 1)
            filter_id.border_color = (0.3, 0.3, 0.3, 1)
            filter_id.border_color_focus = (0.2, 0.5, 0.9, 1)
            filter_id._hidden_input.foreground_color = (1, 1, 1, 1)
            filter_row2.add_widget(filter_id)

            filter_start = RTLTextInput(
                hint_text='از تاریخ', multiline=False,
                size_hint_x=0.25, size_hint_y=None, height=dp(42), font_size=sp(18)
            )
            filter_start.bg_color = (0.15, 0.15, 0.15, 1)
            filter_start.border_color = (0.3, 0.3, 0.3, 1)
            filter_start.border_color_focus = (0.2, 0.5, 0.9, 1)
            filter_start._hidden_input.foreground_color = (1, 1, 1, 1)
            filter_row2.add_widget(filter_start)

            filter_end = RTLTextInput(
                hint_text='تا تاریخ', multiline=False,
                size_hint_x=0.25, size_hint_y=None, height=dp(42), font_size=sp(18)
            )
            filter_end.bg_color = (0.15, 0.15, 0.15, 1)
            filter_end.border_color = (0.3, 0.3, 0.3, 1)
            filter_end.border_color_focus = (0.2, 0.5, 0.9, 1)
            filter_end._hidden_input.foreground_color = (1, 1, 1, 1)
            filter_row2.add_widget(filter_end)

            apply_btn = PersianButton(
                text='اعمال', size_hint_x=0.25, size_hint_y=None, height=dp(42),
                background_color=(0.2, 0.6, 0.2, 1), color=(1, 1, 1, 1),
                font_size=sp(15), bold=True
            )
            filter_row2.add_widget(apply_btn)
            
            content.add_widget(filter_row2)

            # ========== جدول ==========
            scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint_y=0.65, scroll_type=['bars', 'content'], bar_width=dp(5)
            )

            filtered_list_content = GridLayout(
                cols=1, spacing=dp(6), size_hint_y=None, padding=dp(2)
            )
            filtered_list_content.bind(minimum_height=filtered_list_content.setter('height'))

            # ========== تابع کمکی با فونت درشت‌تر ==========
            def make_field(text, size_x, font_sz=16, fg=(1,1,1,1), halign='center', bg=None):
                fld = RTLTextInput(
                    text=str(text) if text else '', multiline=False,
                    size_hint_x=size_x, size_hint_y=None, height=dp(34),
                    font_size=sp(font_sz)
                )
                fld.bg_color = bg if bg else (0.1, 0.1, 0.1, 1)
                fld.border_color = (0.3, 0.3, 0.3, 1)
                fld.border_color_focus = (0.3, 0.3, 0.3, 1)
                fld._hidden_input.foreground_color = fg
                fld._hidden_input.disabled = True
                fld._hidden_input.halign = halign
                return fld

            # ========== تابع نمایش لیست ==========
            def populate_list(filtered_targets):
                nonlocal filtered_list_content
                filtered_list_content.clear_widgets()
                
                if not filtered_targets:
                    filtered_list_content.add_widget(RTLLabel(
                        text='هیچ ریزتارگتی با این فیلترها یافت نشد',
                        size_hint_y=None, height=dp(40),
                        font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
                    ))
                    return

                for target in filtered_targets:
                    if not isinstance(target, dict):
                        continue
                        
                    status = target.get('status', '')
                    editable = can_edit_dt(target)
                    is_completed = status == 'تکمیل شده'

                    # رنگ کارت و فیلد وضعیت
                    if status == 'تکمیل شده':
                        card_bg = (0.12, 0.28, 0.12, 1)
                        status_text = 'تکمیل شده'
                        status_bg = (0.15, 0.5, 0.15, 1)
                        status_fg = (1, 1, 1, 1)
                    elif status == 'فعال':
                        card_bg = (0.12, 0.18, 0.32, 1)
                        status_text = 'فعال'
                        status_bg = (0.35, 0.35, 0.08, 1)
                        status_fg = (0, 0, 0, 1)
                    elif status == 'در انتظار':
                        card_bg = (0.28, 0.18, 0.08, 1)
                        status_text = 'در انتظار'
                        status_bg = (0.9, 0.9, 0.9, 1)
                        status_fg = (0, 0, 0, 1)
                    else:
                        card_bg = (0.15, 0.15, 0.15, 1)
                        status_text = status
                        status_bg = (0.3, 0.3, 0.3, 1)
                        status_fg = (1, 1, 1, 1)

                    card = BoxLayout(
                        orientation='vertical', size_hint_y=None, height=dp(155),
                        spacing=dp(3), padding=[dp(8), dp(4), dp(8), dp(4)]
                    )
                    with card.canvas.before:
                        Color(*card_bg)
                        RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(6)])

                    # ردیف ۱: شناسه | عامل | گروه کالا | وضعیت (رنگی)
                    row1 = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(3))
                    row1.add_widget(make_field(target.get('id', ''), 0.22, 16))
                    row1.add_widget(make_field(target.get('agent_name', ''), 0.26, 16))
                    row1.add_widget(make_field(target.get('product_group', ''), 0.28, 16))
                    # فیلد وضعیت با رنگ
                    row1.add_widget(make_field(status_text, 0.24, 16, status_fg, 'center', status_bg))
                    card.add_widget(row1)

                    # ردیف ۲: تارگت | واحد | دوره | مدت | روزانه | محقق
                    row2 = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(3))
                    row2.add_widget(make_field(f"{target.get('target_count', 0):,}", 0.17, 16))
                    row2.add_widget(make_field(target.get('unit', ''), 0.13, 16))
                    row2.add_widget(make_field(target.get('period', ''), 0.14, 16))
                    row2.add_widget(make_field(str(target.get('duration', 0)), 0.08, 16))
                    row2.add_widget(make_field(f"{target.get('daily_target', 0):,}", 0.17, 16))
                    achieved = target.get('achieved_value', 0)
                    ach_color = (0.2, 0.8, 0.2, 1) if achieved > 0 else (0.5, 0.5, 0.5, 1)
                    row2.add_widget(make_field(f"{achieved:,}", 0.17, 16, ach_color))
                    # فیلتر وضعیت اینجا خالی
                    row2.add_widget(Label(size_hint_x=0.14))
                    card.add_widget(row2)

                    # ردیف ۳: تاریخ شروع | تاریخ پایان | پیوند
                    row3 = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(3))
                    row3.add_widget(make_field(target.get('start_date', ''), 0.22, 16))
                    row3.add_widget(make_field(target.get('end_date', ''), 0.22, 16))
                    row3.add_widget(make_field(f"{target.get('linked_target_id', '')}", 0.56, 15, (0.5, 0.7, 1, 1)))
                    card.add_widget(row3)

                    # ردیف ۴: دکمه‌ها
                    row4 = BoxLayout(size_hint_y=None, height=dp(33), spacing=dp(5))
                    if is_completed:
                        row4.add_widget(PersianButton(
                            text='نهایی شده', size_hint_x=1, size_hint_y=None, height=dp(28),
                            background_color=(0.2, 0.6, 0.2, 1), color=(1, 1, 1, 1),
                            font_size=sp(14), disabled=True
                        ))
                    elif editable:
                        edit_btn = PersianButton(
                            text='ویرایش', size_hint_x=0.5, size_hint_y=None, height=dp(28),
                            background_color=(0.8, 0.6, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(14)
                        )
                        edit_btn.bind(on_press=lambda x, t=target: self._edit_detailed_target(t))
                        row4.add_widget(edit_btn)
                        delete_btn = PersianButton(
                            text='حذف', size_hint_x=0.5, size_hint_y=None, height=dp(28),
                            background_color=(0.8, 0.2, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(14)
                        )
                        delete_btn.bind(on_press=lambda x, t=target: self._delete_detailed_target(t))
                        row4.add_widget(delete_btn)
                    else:
                        row4.add_widget(PersianButton(
                            text='منقضی شده', size_hint_x=1, size_hint_y=None, height=dp(28),
                            background_color=(0.3, 0.3, 0.3, 1), color=(0.5, 0.5, 0.5, 1),
                            font_size=sp(14), disabled=True
                        ))
                    card.add_widget(row4)
                    filtered_list_content.add_widget(card)

            # ========== رویداد اعمال فیلتر ==========
            def apply_filter(inst):
                a = filter_agent.text
                p = filter_product.text
                s = filter_status.text
                ids = filter_id.text.strip()
                sd = filter_start.text.strip()
                ed = filter_end.text.strip()

                filtered = get_detailed_targets_filtered(
                    agent_name=a if a != 'همه' else None,
                    product_group=p if p != 'همه' else None,
                    status=s if s != 'همه' else None,
                    target_id=ids if ids else None,
                    linked_target_id=ids if ids else None,
                    start_date=sd if sd else None,
                    end_date=ed if ed else None
                )

                populate_list(filtered)

            apply_btn.bind(on_press=apply_filter)

            # بارگذاری اولیه
            populate_list(all_targets)

            scroll.add_widget(filtered_list_content)
            content.add_widget(scroll)

            # ========== دکمه‌های پایین ==========
            btn_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8))
            
            export_btn = PersianButton(
                text='خروجی اکسل', background_color=(0.2, 0.7, 0.4, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(36),
                color=(1, 1, 1, 1), font_size=sp(15)
            )
            btn_row.add_widget(export_btn)

            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(36),
                color=(1, 1, 1, 1), font_size=sp(15)
            )
            btn_row.add_widget(close_btn)
            content.add_widget(btn_row)

            popup = PersianPopup(
                title='', title_size=0, content=content,
                size_hint=(0.96, 0.92), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def export_current(inst):
                a = filter_agent.text
                p = filter_product.text
                s = filter_status.text
                ids = filter_id.text.strip()
                sd = filter_start.text.strip()
                ed = filter_end.text.strip()

                filtered = get_detailed_targets_filtered(
                    agent_name=a if a != 'همه' else None,
                    product_group=p if p != 'همه' else None,
                    status=s if s != 'همه' else None,
                    target_id=ids if ids else None,
                    linked_target_id=ids if ids else None,
                    start_date=sd if sd else None,
                    end_date=ed if ed else None
                )
                
                success, message, filepath = export_to_excel(filtered)
                if success:
                    self.show_message('موفق', message)
                else:
                    self.show_message('خطا', message)

            export_btn.bind(on_press=export_current)
            close_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش لیست ریزتارگت‌ها: {e}", error_details)

    def _edit_detailed_target(self, target):
        """ویرایش ریزتارگت"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'ویرایش ریزتارگت - {target.get("id", "")}',
                size_hint_y=None, height=dp(35),
                font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            # گروه کالا
            content.add_widget(RTLLabel(
                text='گروه کالا:', size_hint_y=None, height=dp(25),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            edit_product = PersianComboBox(
                text=target.get('product_group', ''),
                values=get_product_groups(),
                height=dp(55)
            )
            edit_product.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            edit_product.main_btn.color = (1, 1, 1, 1)
            edit_product.main_btn.font_size = sp(18)
            content.add_widget(edit_product)

            # تعداد هدف
            content.add_widget(RTLLabel(
                text='تعداد هدف:', size_hint_y=None, height=dp(25),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            edit_count = RTLTextInput(
                text=str(target.get('target_count', 0)),
                multiline=False, size_hint_y=None, height=dp(55),
                input_filter='int', font_size=sp(18)
            )
            edit_count.bg_color = (0.15, 0.15, 0.15, 1)
            edit_count.border_color = (0.3, 0.3, 0.3, 1)
            edit_count.border_color_focus = (0.2, 0.5, 0.9, 1)
            edit_count._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(edit_count)

            # واحد
            content.add_widget(RTLLabel(
                text='واحد:', size_hint_y=None, height=dp(25),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            edit_unit = PersianComboBox(
                text=target.get('unit', ''),
                values=get_target_units(),
                height=dp(55)
            )
            edit_unit.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            edit_unit.main_btn.color = (1, 1, 1, 1)
            edit_unit.main_btn.font_size = sp(18)
            content.add_widget(edit_unit)

            # دوره
            content.add_widget(RTLLabel(
                text='دوره:', size_hint_y=None, height=dp(25),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            edit_period = PersianComboBox(
                text=target.get('period', ''),
                values=get_target_periods(),
                height=dp(55)
            )
            edit_period.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            edit_period.main_btn.color = (1, 1, 1, 1)
            edit_period.main_btn.font_size = sp(18)
            content.add_widget(edit_period)

            # مدت
            content.add_widget(RTLLabel(
                text='مدت:', size_hint_y=None, height=dp(25),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            edit_duration = RTLTextInput(
                text=str(target.get('duration', 1)),
                multiline=False, size_hint_y=None, height=dp(55),
                input_filter='int', font_size=sp(18)
            )
            edit_duration.bg_color = (0.15, 0.15, 0.15, 1)
            edit_duration.border_color = (0.3, 0.3, 0.3, 1)
            edit_duration.border_color_focus = (0.2, 0.5, 0.9, 1)
            edit_duration._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(edit_duration)

            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))

            save_btn = PersianButton(
                text='ذخیره', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )

            btn_layout.add_widget(save_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='ویرایش ریزتارگت', content=content,
                size_hint=(0.92, 0.8), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_save(inst):
                updates = {
                    'product_group': edit_product.text,
                    'target_count': int(edit_count.text) if edit_count.text else 0,
                    'unit': edit_unit.text,
                    'period': edit_period.text,
                    'duration': int(edit_duration.text) if edit_duration.text else 1
                }

                if updates['target_count'] <= 0:
                    self.show_message('خطا', 'تعداد هدف باید بزرگتر از صفر باشد')
                    return

                success, message = update_detailed_target(target.get('id'), updates)
                popup.dismiss()
                if success:
                    self.show_message('موفق', message)
                    self._show_detailed_targets_list(None)
                else:
                    self.show_message('خطا', message)

            save_btn.bind(on_press=do_save)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ویرایش ریزتارگت: {e}", error_details)

    def _delete_detailed_target(self, target):
        """حذف ریزتارگت"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'آیا از حذف ریزتارگت "{target.get("id", "")}" اطمینان دارید؟',
                size_hint_y=None, height=dp(45),
                font_size=sp(18), color=(1, 0.8, 0.2, 1)
            ))

            content.add_widget(RTLLabel(
                text=f'{target.get("product_group", "")}: {target.get("target_count", 0):,} {target.get("unit", "")}',
                size_hint_y=None, height=dp(35),
                font_size=sp(14), color=(0.8, 0.8, 0.8, 1)
            ))

            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))

            confirm_btn = PersianButton(
                text='بله، حذف شود', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None, height=dp(45), color=(1, 1, 1, 1), font_size=sp(16)
            )

            btn_layout.add_widget(confirm_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='تأیید حذف', content=content,
                size_hint=(0.85, 0.45), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_delete(inst):
                popup.dismiss()
                success, message = delete_detailed_target(target.get('id'))
                if success:
                    self.show_message('موفق', message)
                    self._show_detailed_targets_list(None)
                else:
                    self.show_message('خطا', message)

            confirm_btn.bind(on_press=do_delete)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در حذف ریزتارگت: {e}", error_details)

    def _export_detailed_targets_excel(self, instance):
        """خروجی اکسل ریزتارگت‌ها با قابلیت فیلتر بر اساس عامل"""
        try:
            from utils.detailed_target_manager import get_all_detailed_targets, export_to_excel as export_detailed_to_excel
            
            all_targets = get_all_detailed_targets()
            
            if not all_targets:
                self.show_message('خطا', 'هیچ ریزتارگتی برای خروجی وجود ندارد')
                return
            
            # ========== دیالوگ انتخاب عامل ==========
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text='خروجی اکسل ریزتارگت‌ها',
                size_hint_y=None, height=dp(35),
                font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            
            content.add_widget(RTLLabel(
                text=f'کل ریزتارگت‌ها: {len(all_targets)} مورد',
                size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(0.6, 0.6, 0.6, 1)
            ))
            
            # استخراج لیست عامل‌ها از ریزتارگت‌ها
            agent_names = list(set(t.get('agent_name', '') for t in all_targets if t.get('agent_name')))
            agent_names.sort()
            agent_names = ['همه'] + agent_names
            
            content.add_widget(RTLLabel(
                text='انتخاب عامل:',
                size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            
            agent_spinner = PersianComboBox(
                text='همه',
                values=agent_names,
                height=dp(55)
            )
            agent_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            agent_spinner.main_btn.color = (1, 1, 1, 1)
            agent_spinner.main_btn.font_size = sp(16)
            content.add_widget(agent_spinner)
            
            # استخراج لیست گروه‌های کالا
            product_groups = list(set(t.get('product_group', '') for t in all_targets if t.get('product_group')))
            product_groups.sort()
            product_groups = ['همه'] + product_groups
            
            content.add_widget(RTLLabel(
                text='انتخاب گروه کالا:',
                size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(1, 1, 1, 1)
            ))
            
            product_spinner = PersianComboBox(
                text='همه',
                values=product_groups,
                height=dp(55)
            )
            product_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            product_spinner.main_btn.color = (1, 1, 1, 1)
            product_spinner.main_btn.font_size = sp(16)
            content.add_widget(product_spinner)
            
            # نمایش تعداد فیلتر شده
            count_label = RTLLabel(
                text=f'تعداد: {len(all_targets)}',
                size_hint_y=None, height=dp(25),
                font_size=sp(14), color=(0.2, 0.8, 0.4, 1)
            )
            content.add_widget(count_label)
            
            # آپدیت تعداد با تغییر فیلتر
            def update_count(*args):
                agent = agent_spinner.text
                product = product_spinner.text
                
                count = len(all_targets)
                for t in all_targets:
                    if agent != 'همه' and t.get('agent_name') != agent:
                        count -= 1
                    elif product != 'همه' and t.get('product_group') != product:
                        count -= 1
                
                count_label.text = f'تعداد: {count}'
            
            agent_spinner.bind(text=update_count)
            product_spinner.bind(text=update_count)
            
            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
            
            export_btn = PersianButton(
                text='خروجی اکسل',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='انصراف',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            
            btn_layout.add_widget(export_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='خروجی ریزتارگت', content=content,
                size_hint=(0.88, 0.55), background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )
            
            def do_export(inst):
                agent = agent_spinner.text
                product = product_spinner.text
                
                # فیلتر کردن
                filtered = all_targets
                if agent != 'همه':
                    filtered = [t for t in filtered if t.get('agent_name') == agent]
                if product != 'همه':
                    filtered = [t for t in filtered if t.get('product_group') == product]
                
                if not filtered:
                    self.show_message('خطا', 'هیچ ریزتارگتی با این فیلتر وجود ندارد')
                    return
                
                popup.dismiss()
                success, message, filepath = export_detailed_to_excel(filtered)
                if success:
                    self.show_message('موفق', message)
                else:
                    self.show_message('خطا', message)
            
            export_btn.bind(on_press=do_export)
            cancel_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خروجی اکسل: {e}", error_details)

    # ============================================================
    # تب ۲: تحقق تارگت
    # ============================================================

    def show_fulfillment_tab(self):
        """نمایش تب تحقق تارگت"""
        try:
            from utils.file_picker_import import ImportFilePicker

            scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint=(1, 1), scroll_type=['bars', 'content'], bar_width=dp(8)
            )

            content = GridLayout(
                cols=1, spacing=dp(10), size_hint_y=None, padding=dp(12)
            )
            content.bind(minimum_height=content.setter('height'))

            # ========== عنوان ==========
            content.add_widget(RTLLabel(
                text='تحقق تارگت',
                size_hint_y=None, height=dp(50),
                font_size=sp(22), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            # ========== فیلترها ==========
            filter_layout = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(250))
            filter_layout.bind(minimum_height=filter_layout.setter('height'))

            filter_layout.add_widget(RTLLabel(
                text='از تاریخ:', size_hint_y=None, height=dp(35),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            self.fulfillment_start_date = RTLTextInput(
                text=get_today_jalali(), multiline=False,
                size_hint_y=None, height=dp(65), font_size=sp(22)
            )
            self.fulfillment_start_date.bg_color = (0.15, 0.15, 0.15, 1)
            self.fulfillment_start_date.border_color = (0.3, 0.3, 0.3, 1)
            self.fulfillment_start_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.fulfillment_start_date._hidden_input.foreground_color = (1, 1, 1, 1)
            filter_layout.add_widget(self.fulfillment_start_date)

            filter_layout.add_widget(RTLLabel(
                text='تا تاریخ:', size_hint_y=None, height=dp(35),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            self.fulfillment_end_date = RTLTextInput(
                text=get_today_jalali(), multiline=False,
                size_hint_y=None, height=dp(65), font_size=sp(22)
            )
            self.fulfillment_end_date.bg_color = (0.15, 0.15, 0.15, 1)
            self.fulfillment_end_date.border_color = (0.3, 0.3, 0.3, 1)
            self.fulfillment_end_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.fulfillment_end_date._hidden_input.foreground_color = (1, 1, 1, 1)
            filter_layout.add_widget(self.fulfillment_end_date)

            filter_layout.add_widget(RTLLabel(
                text='انتخاب عامل:', size_hint_y=None, height=dp(35),
                font_size=sp(16), color=(1, 1, 1, 1)
            ))
            agents = get_agents()
            agent_names = [a.get('name', '') for a in agents] if agents else ['']
            self.fulfillment_agent = PersianComboBox(
                text=agent_names[0] if agent_names else '', values=agent_names, height=dp(65)
            )
            self.fulfillment_agent.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.fulfillment_agent.main_btn.color = (1, 1, 1, 1)
            self.fulfillment_agent.main_btn.font_size = sp(18)
            filter_layout.add_widget(self.fulfillment_agent)

            content.add_widget(filter_layout)

            # ========== بخش تارگت‌های اصلی ==========
            content.add_widget(RTLLabel(
                text='─ تارگت‌های اصلی ─', size_hint_y=None, height=dp(30),
                font_size=sp(16), color=(0.4, 0.7, 1, 1), bold=True, halign='center'
            ))

            show_btn = PersianButton(
                text='نمایش تارگت‌ها', background_color=(0.2, 0.6, 1, 1),
                size_hint_y=None, height=dp(50), color=(1, 1, 1, 1), font_size=sp(18)
            )
            show_btn.bind(on_press=self.show_fulfillment_targets)
            content.add_widget(show_btn)

            self.fulfillment_list_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True, size_hint_y=0.25,
                scroll_type=['bars', 'content'], bar_width=dp(6)
            )
            self.fulfillment_list = GridLayout(cols=1, spacing=dp(6), size_hint_y=None, padding=dp(5))
            self.fulfillment_list.bind(minimum_height=self.fulfillment_list.setter('height'))
            self.fulfillment_list_scroll.add_widget(self.fulfillment_list)
            content.add_widget(self.fulfillment_list_scroll)

            self.fulfillment_file_picker = ImportFilePicker(
                on_select=self.on_fulfillment_file_selected, size_hint_y=None, height=dp(60)
            )
            content.add_widget(self.fulfillment_file_picker)

            # ========== بخش ریزتارگت‌ها ==========
            content.add_widget(RTLLabel(
                text='─ تحقق ریزتارگت‌ها ─', size_hint_y=None, height=dp(30),
                font_size=sp(16), color=(1, 0.5, 0, 1), bold=True, halign='center'
            ))

            dt_fulfill_btn = PersianButton(
                text='آپلود فایل و تحقق ریزتارگت‌ها', background_color=(1, 0.5, 0, 1),
                size_hint_y=None, height=dp(50), color=(1, 1, 1, 1), font_size=sp(18), bold=True
            )
            dt_fulfill_btn.bind(on_press=self._show_detailed_fulfillment_dialog)
            content.add_widget(dt_fulfill_btn)

            scroll.add_widget(content)
            self.content_area.add_widget(scroll)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب تحقق تارگت: {e}", error_details)


    # ============================================================
    # دیالوگ تحقق ریزتارگت‌ها (سوپروایزر)
    # ============================================================

    def _show_detailed_fulfillment_dialog(self, instance):
        """نمایش دیالوگ تحقق ریزتارگت‌ها با آپلود فایل اکسل"""
        try:
            import os
            from datetime import datetime
            from utils.storage import get_backup_path
            from utils.detailed_target_manager import get_all_detailed_targets
            from utils.file_picker_import import ImportFilePicker

            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))

            # ========== دکمه بارگذاری فایل ==========
            file_btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))

            self.dt_fulfill_status = RTLLabel(
                text='فایلی انتخاب نشده است', size_hint_x=0.6, size_hint_y=None, height=dp(45),
                font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
            )
            file_btn_layout.add_widget(self.dt_fulfill_status)

            self.dt_fulfill_file_picker = ImportFilePicker(
                on_select=lambda filepath: self._on_dt_fulfill_file_selected(
                    [filepath] if not isinstance(filepath, list) else filepath,
                    self.dt_fulfill_status, self.dt_fulfill_grid
                ),
                size_hint_x=0.4, size_hint_y=None, height=dp(45)
            )
            file_btn_layout.add_widget(self.dt_fulfill_file_picker)
            content.add_widget(file_btn_layout)

            # ========== هدر جدول ==========
            header_box = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(3))
            headers = [
                ('درصد تحقق', 0.22), ('تحقق', 0.18),
                ('تارگت روز', 0.22), ('گروه کالا', 0.38)
            ]
            for text, size in headers:
                header_box.add_widget(RTLLabel(
                    text=text, size_hint_x=size, size_hint_y=None, height=dp(32),
                    font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                ))
            content.add_widget(header_box)

            # ========== جدول داده‌ها ==========
            self.dt_fulfill_data = []
            self.dt_fulfill_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint_y=0.55, scroll_type=['bars', 'content'], bar_width=dp(6)
            )
            self.dt_fulfill_grid = GridLayout(cols=1, spacing=dp(3), size_hint_y=None, padding=dp(3))
            self.dt_fulfill_grid.bind(minimum_height=self.dt_fulfill_grid.setter('height'))
            self.dt_fulfill_grid.add_widget(RTLLabel(
                text='لطفاً فایل اکسل خروجی تحقق ایجنت را انتخاب کنید',
                size_hint_y=None, height=dp(40), font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
            ))
            self.dt_fulfill_scroll.add_widget(self.dt_fulfill_grid)
            content.add_widget(self.dt_fulfill_scroll)

            # ========== دکمه‌های پایین ==========
            btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))

            self.dt_finalize_btn = PersianButton(
                text='نهایی‌سازی', size_hint_x=0.33, size_hint_y=None, height=dp(45),
                background_color=(0.2, 0.7, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(self.dt_finalize_btn)

            history_btn = PersianButton(
                text='تاریخچه', size_hint_x=0.33, size_hint_y=None, height=dp(45),
                background_color=(0.6, 0.4, 0.2, 1), color=(1, 1, 1, 1), font_size=sp(16)
            )
            history_btn.bind(on_press=self._show_fulfillment_history)
            btn_layout.add_widget(history_btn)

            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.34, size_hint_y=None, height=dp(45),
                color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(close_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='تحقق ریزتارگت‌ها', content=content,
                size_hint=(0.95, 0.88), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )

            # ========== رویدادها ==========
            def do_finalize(inst):
                if not self.dt_fulfill_data:
                    self.show_message('خطا', 'هیچ داده‌ای برای نهایی‌سازی وجود ندارد')
                    return

                all_targets = get_all_detailed_targets()
                temp_updates = []

                for item in self.dt_fulfill_data:
                    target_id = item.get('id')
                    achieved = item.get('achieved', 0)
                    if target_id and achieved > 0:
                        for t in all_targets:
                            if t.get('id') == target_id:
                                current = t.get('achieved_value', 0)
                                temp_updates.append({
                                    'target': t,
                                    'new_achieved': current + achieved,
                                    'added': achieved
                                })
                                break

                if not temp_updates:
                    self.show_message('خطا', 'هیچ تارگتی بروزرسانی نشد')
                    return

                self._show_dt_status_selection_dialog(temp_updates, all_targets, popup)

            self.dt_finalize_btn.bind(on_press=do_finalize)
            close_btn.bind(on_press=popup.dismiss)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ تحقق: {e}", error_details)


    def _show_dt_status_selection_dialog(self, temp_updates, all_targets, parent_popup):
        """نمایش دیالوگ انتخاب وضعیت برای ریزتارگت‌ها (سوپروایزر)"""
        try:
            from kivy.app import App
            
            # ✅ دریافت نام کاربر جاری
            app = App.get_running_app()
            current_username = app.current_username if hasattr(app, 'current_username') else ''
            
            if not current_username:
                try:
                    from utils.user_manager import get_current_user
                    user = get_current_user()
                    if user:
                        current_username = user.get('username', '') or user.get('name', '')
                except:
                    pass
            
            if not current_username:
                current_username = 'supervisor'
            
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                            size=lambda i, v: setattr(rect, 'size', v))

            content.add_widget(RTLLabel(
                text=f'انتخاب وضعیت برای {len(temp_updates)} ریزتارگت',
                size_hint_y=None, height=dp(40),
                font_size=sp(18), bold=True, color=(0.4, 0.7, 1, 1)
            ))

            scroll = ScrollView(size_hint_y=0.7, do_scroll_x=False)
            status_grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None, padding=dp(5))
            status_grid.bind(minimum_height=status_grid.setter('height'))
            status_selections = []

            for update in temp_updates:
                t = update['target']
                target_id = t.get('id', '')
                product = t.get('product_group', '')
                new_achieved = update['new_achieved']
                target_count = t.get('target_count', 0)

                card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(70),
                            spacing=dp(3), padding=[dp(10), dp(3), dp(10), dp(3)])
                with card.canvas.before:
                    Color(0.15, 0.15, 0.2, 1)
                    RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(6)])

                info_row = BoxLayout(size_hint_y=None, height=dp(35))
                info_row.add_widget(RTLLabel(
                    text=f"{target_id} | {product} | {new_achieved:,}/{target_count:,}",
                    size_hint_x=0.6, font_size=sp(14), color=(1, 1, 1, 1)
                ))
                status_combo = PersianComboBox(
                    text='فعال', values=['فعال', 'تکمیل شده'], height=dp(38)
                )
                status_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
                status_combo.main_btn.color = (1, 1, 1, 1)
                status_combo.main_btn.font_size = sp(14)
                status_combo.size_hint_x = 0.4
                info_row.add_widget(status_combo)
                card.add_widget(info_row)

                status_selections.append({'target_id': target_id, 'combo': status_combo})
                status_grid.add_widget(card)

            scroll.add_widget(status_grid)
            content.add_widget(scroll)

            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
            save_status_btn = PersianButton(
                text='بروزرسانی وضعیت', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.6, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(15)
            )
            cancel_btn = PersianButton(
                text='انصراف', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.4, size_hint_y=None, height=dp(40),
                color=(1, 1, 1, 1), font_size=sp(15)
            )
            btn_layout.add_widget(save_status_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            status_popup = PersianPopup(
                title='بروزرسانی وضعیت', content=content,
                size_hint=(0.9, 0.7), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )

            def save_statuses(inst):
                import json
                from utils.storage import get_data_path

                today = get_today_jalali()
                filename = os.path.basename(self.dt_fulfill_file_path) if hasattr(self, 'dt_fulfill_file_path') and self.dt_fulfill_file_path else ''
                
                # ✅ چک تکراری
                if filename:
                    for update in temp_updates:
                        t = update['target']
                        processed = t.get('processed_files', {})
                        if processed.get(today) == filename:
                            self.show_message('خطا', f'این فایل قبلاً برای امروز ({today}) ثبت شده است.')
                            return
                
                # ✅ ثبت با نام کاربر
                for update in temp_updates:
                    t = update['target']
                    target_id = t.get('id', '')
                    added = update['added']
                    
                    selected_status = 'فعال'
                    for sel in status_selections:
                        if sel['target_id'] == target_id:
                            selected_status = sel['combo'].text
                            break

                    if 'daily_achievements' not in t:
                        t['daily_achievements'] = {}
                    t['daily_achievements'][today] = added
                    
                    if filename:
                        if 'processed_files' not in t:
                            t['processed_files'] = {}
                        t['processed_files'][today] = filename
                    
                    t['achieved_value'] = sum(t['daily_achievements'].values())
                    t['status'] = selected_status
                    t['fulfilled_by'] = current_username  # ✅ ذخیره نام کاربر
                    t['fulfilled_date'] = today  # ✅ ذخیره تاریخ تحقق

                path = os.path.join(get_data_path(), 'detailed_targets.json')
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(all_targets, f, ensure_ascii=False, indent=2)

                status_popup.dismiss()
                if parent_popup:
                    parent_popup.dismiss()
                self.show_message('موفق', f'{len(temp_updates)} ریزتارگت نهایی‌سازی شد')

            save_status_btn.bind(on_press=save_statuses)
            cancel_btn.bind(on_press=status_popup.dismiss)
            status_popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)


    def _on_dt_fulfill_file_selected(self, selection, status_label, grid):
        """پردازش فایل اکسل تحقق ایجنت و تطبیق با شناسه ریزتارگت"""
        try:
            import openpyxl
            import os

            if not selection:
                return

            filepath = selection[0] if isinstance(selection, list) else selection
            self.dt_fulfill_file_path = filepath  
            filename = os.path.basename(filepath)
            status_label.text = f'{filename}'
            status_label.color = (0.2, 0.8, 0.2, 1)

            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            col_map = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    col_map[str(cell.value).strip()] = col_idx

            id_col = col_map.get('شناسه')
            product_col = col_map.get('گروه کالا')
            achieved_col = col_map.get('تحقق')
            daily_col = col_map.get('تارگت روز')
            unit_col = col_map.get('واحد')
            agent_col = col_map.get('عامل')

            if not all([id_col, achieved_col]):
                status_label.text = 'ستون شناسه یا تحقق یافت نشد'
                status_label.color = (0.8, 0.2, 0.2, 1)
                return

            from utils.detailed_target_manager import get_all_detailed_targets
            all_targets = get_all_detailed_targets()
            if not isinstance(all_targets, list):
                all_targets = []

            targets_dict = {t.get('id', ''): t for t in all_targets if isinstance(t, dict)}

            self.dt_fulfill_data = []
            grid.clear_widgets()
            matched = 0
            not_found = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[id_col - 1]:
                    continue

                target_id = str(row[id_col - 1]).strip() if row[id_col - 1] else ''
                achieved = int(row[achieved_col - 1]) if row[achieved_col - 1] else 0

                if not target_id or achieved <= 0:
                    continue

                matched_target = targets_dict.get(target_id)

                if matched_target:
                    matched += 1
                    product_group = matched_target.get('product_group', '')
                    daily_target = matched_target.get('daily_target', 0)
                    unit = matched_target.get('unit', '')
                    agent_name = matched_target.get('agent_name', '')
                    target_count = matched_target.get('target_count', 0)

                    percent = (achieved / daily_target * 100) if daily_target > 0 else 0
                    if percent < 65:
                        p_color = (0.8, 0.2, 0.2, 1)
                    elif percent < 85:
                        p_color = (1, 0.7, 0, 1)
                    else:
                        p_color = (0.2, 0.8, 0.2, 1)

                    self.dt_fulfill_data.append({
                        'id': target_id, 'agent_name': agent_name,
                        'product_group': product_group, 'daily_target': daily_target,
                        'target_count': target_count, 'unit': unit, 'achieved': achieved
                    })

                    row_box = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(3))
                    row_box.add_widget(RTLLabel(
                        text=f"{percent:.0f}%", size_hint_x=0.22, size_hint_y=None, height=dp(48),
                        font_size=sp(18), bold=True, color=p_color, halign='center'
                    ))
                    row_box.add_widget(RTLLabel(
                        text=f"{achieved:,}", size_hint_x=0.18, size_hint_y=None, height=dp(48),
                        font_size=sp(18), color=(1, 1, 1, 1), halign='center'
                    ))
                    row_box.add_widget(RTLLabel(
                        text=f"{daily_target:,}", size_hint_x=0.22, size_hint_y=None, height=dp(48),
                        font_size=sp(18), color=(1, 1, 1, 1), halign='center'
                    ))
                    row_box.add_widget(RTLLabel(
                        text=product_group, size_hint_x=0.38, size_hint_y=None, height=dp(48),
                        font_size=sp(16), color=(1, 1, 1, 1), halign='right'
                    ))
                    grid.add_widget(row_box)
                else:
                    not_found += 1

            wb.close()

            total = matched + not_found
            status_label.text = f'{filename} ({matched} تطبیق یافته'
            if not_found > 0:
                status_label.text += f' - {not_found} یافت نشد'
            status_label.text += ')'
            status_label.color = (0.2, 0.8, 0.2, 1) if matched > 0 else (0.8, 0.2, 0.2, 1)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در خواندن فایل: {e}", error_details)

    # ============================================================
    # متدهای تارگت اصلی (بدون تغییر)
    # ============================================================

    def show_fulfillment_targets(self, instance):
        """نمایش تارگت‌ها در یک دیالوگ جداگانه"""
        try:
            agent = self.fulfillment_agent.text
            start_date = self.fulfillment_start_date.text.strip()
            end_date = self.fulfillment_end_date.text.strip()

            if not agent or agent == '':
                self.show_message('خطا', 'لطفاً یک عامل را انتخاب کنید')
                return
            if start_date and not validate_jalali_date(start_date):
                self.show_message('خطا', 'تاریخ شروع نامعتبر است')
                return
            if end_date and not validate_jalali_date(end_date):
                self.show_message('خطا', 'تاریخ پایان نامعتبر است')
                return

            targets = get_active_targets_by_agent(agent, start_date, end_date)

            dialog_content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
            with dialog_content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=dialog_content.pos, size=dialog_content.size)
                dialog_content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                                size=lambda i, v: setattr(content_rect, 'size', v))

            if not targets:
                dialog_content.add_widget(RTLLabel(
                    text='هیچ تارگت فعالی در بازه انتخابی یافت نشد',
                    size_hint_y=None, height=dp(45), font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
                ))
            else:
                list_scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.6)
                list_content = GridLayout(cols=1, spacing=dp(4), size_hint_y=None, padding=dp(5))
                list_content.bind(minimum_height=list_content.setter('height'))
                self.fulfillment_selected = {}

                for target in targets:
                    box = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(5),
                                padding=[dp(5), dp(2), dp(5), dp(2)])
                    check = CheckBox(size_hint_x=0.1, size_hint_y=None, height=dp(40), color=(0.4, 0.7, 1, 1))
                    check.active = False
                    target_id = target.get('target_id')
                    check.bind(active=lambda cb, value, tid=target_id: self._toggle_fulfillment_selection(tid, value))
                    box.add_widget(check)
                    self.fulfillment_selected[target_id] = False
                    info = RTLLabel(
                        text=f"{target.get('target_id', '')} | {target.get('target_type', '')} | {target.get('target_value', 0):,}",
                        size_hint_x=0.9, size_hint_y=None, height=dp(40), font_size=sp(13), color=(1, 1, 1, 1)
                    )
                    box.add_widget(info)
                    list_content.add_widget(box)

                list_scroll.add_widget(list_content)
                dialog_content.add_widget(list_scroll)

                select_all_btn = PersianButton(
                    text='انتخاب همه', background_color=(0.2, 0.5, 0.8, 1),
                    size_hint_y=None, height=dp(35), color=(1, 1, 1, 1), font_size=sp(14)
                )
                select_all_btn.bind(on_press=self._select_all_fulfillment_targets)
                dialog_content.add_widget(select_all_btn)

            btn_row = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
            confirm_btn = PersianButton(
                text='تأیید و اعمال تحقق', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(14)
            )
            confirm_btn.bind(on_press=lambda x: self._apply_fulfillment_from_dialog(dialog_popup))
            btn_row.add_widget(confirm_btn)
            close_btn = PersianButton(
                text='بستن', background_color=(0.3, 0.3, 0.3, 1),
                size_hint_x=0.5, size_hint_y=None, height=dp(40), color=(1, 1, 1, 1), font_size=sp(14)
            )
            btn_row.add_widget(close_btn)
            dialog_content.add_widget(btn_row)

            dialog_popup = PersianPopup(
                title='انتخاب تارگت برای تحقق', content=dialog_content,
                size_hint=(0.92, 0.8), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            close_btn.bind(on_press=dialog_popup.dismiss)
            dialog_popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تارگت‌ها: {e}", error_details)


    def _select_all_fulfillment_targets(self, instance):
        try:
            if hasattr(self, 'fulfillment_selected'):
                count = 0
                for key in self.fulfillment_selected:
                    self.fulfillment_selected[key] = True
                    count += 1
                self.show_message('اطلاع', f'{count} تارگت انتخاب شد، برای نهایی سازی فایل اکسل را انتخاب نمایید')
        except Exception as e:
            print(f"خطا در انتخاب همه: {e}")


    def _apply_fulfillment_from_dialog(self, popup):
        try:
            selected_targets = [tid for tid, selected in self.fulfillment_selected.items() if selected]
            if not selected_targets:
                self.show_message('خطا', 'هیچ تارگتی انتخاب نشده است')
                return
            agent_name = self.fulfillment_agent.text
            self.show_message('اطلاع', f'{len(selected_targets)} تارگت انتخاب شد، برای نهایی سازی فایل اکسل را انتخاب نمایید')
            popup.dismiss()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال تحقق: {e}", error_details)


    def _toggle_fulfillment_selection(self, target_id, value):
        if hasattr(self, 'fulfillment_selected'):
            self.fulfillment_selected[target_id] = value


    def on_fulfillment_file_selected(self, filepath):
        try:
            if not filepath:
                self.show_message('خطا', 'فایلی انتخاب نشده است')
                return
            summary_data = read_excel_summary(filepath)
            if not summary_data:
                self.show_message('خطا', 'داده‌های خلاصه در فایل اکسل یافت نشد')
                return
            selected_targets = [tid for tid, selected in self.fulfillment_selected.items() if selected]
            if not selected_targets:
                self.show_message('خطا', 'هیچ تارگتی انتخاب نشده است')
                return
            all_targets = get_all_targets()
            achieved_values = {}
            target_details = []
            agent_name = self.fulfillment_agent.text
            for target in all_targets:
                target_id = target.get('target_id')
                if target_id in selected_targets:
                    target_type = target.get('target_type', '')
                    excel_key = TARGET_EXCEL_MAPPING.get(target_type)
                    if excel_key and excel_key in summary_data:
                        achieved_values[target_id] = summary_data[excel_key]
                        target_details.append({
                            'id': target_id, 'type': target_type,
                            'achieved': summary_data[excel_key],
                            'target': target.get('target_value', 0)
                        })
                    else:
                        self.show_message('خطا', f'نوع تارگت "{target_type}" در فایل اکسل یافت نشد')
                        return
            if not achieved_values:
                self.show_message('خطا', 'هیچ داده‌ای برای تطبیق با تارگت‌ها یافت نشد')
                return
            self._show_fulfillment_confirm_dialog(selected_targets, achieved_values, target_details, agent_name)
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در اعمال تحقق: {e}", error_details)


    def _show_fulfillment_confirm_dialog(self, target_ids, achieved_values, target_details, agent_name):
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                        size=lambda i, v: setattr(content_rect, 'size', v))
            content.add_widget(RTLLabel(
                text=f'کاربر گرامی از انتخاب این فایل اکسل برای {agent_name} مطمئن هستید؟',
                size_hint_y=None, height=dp(50), font_size=sp(18), color=(1, 0.8, 0.2, 1)
            ))
            list_scroll = ScrollView(size_hint_y=0.5, do_scroll_x=False)
            list_content = GridLayout(cols=1, spacing=dp(4), size_hint_y=None)
            list_content.bind(minimum_height=list_content.setter('height'))
            for detail in target_details:
                list_content.add_widget(RTLLabel(
                    text=f"{detail['id']} | {detail['type']} | هدف: {detail['target']:,} | تحقق: {detail['achieved']:,}",
                    size_hint_y=None, height=dp(30), font_size=sp(14), color=(1, 1, 1, 1)
                ))
            list_scroll.add_widget(list_content)
            content.add_widget(list_scroll)
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(55))
            confirm_btn = PersianButton(
                text='بله', background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None, height=dp(50), color=(1, 1, 1, 1), font_size=sp(16)
            )
            cancel_btn = PersianButton(
                text='خیر', background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None, height=dp(50), color=(1, 1, 1, 1), font_size=sp(16)
            )
            btn_layout.add_widget(confirm_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)
            popup = PersianPopup(
                title='تأیید نهایی‌سازی', content=content,
                size_hint=(0.9, 0.6), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            def do_finalize(instance):
                popup.dismiss()
                self._perform_fulfillment(target_ids, achieved_values)
            def cancel_finalize(instance):
                popup.dismiss()
            confirm_btn.bind(on_press=do_finalize)
            cancel_btn.bind(on_press=cancel_finalize)
            popup.open()
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ تأیید: {e}", error_details)


    def _show_fulfillment_history(self, instance):
        """نمایش تاریخچه تحقق روزانه با فیلتر ماهیانه و عامل"""
        try:
            from utils.detailed_target_manager import get_all_detailed_targets, export_to_excel
            
            all_targets = get_all_detailed_targets()
            if not isinstance(all_targets, list) or not all_targets:
                self.show_message('اطلاع', 'هیچ ریزتارگتی ثبت نشده است')
                return
            
            # جمع‌آوری همه تحقق‌های روزانه
            daily_records = []
            agents_set = set()
            
            for t in all_targets:
                if not isinstance(t, dict):
                    continue
                daily_achievements = t.get('daily_achievements', {})
                if not isinstance(daily_achievements, dict):
                    continue
                
                agent_name = t.get('agent_name', '')
                if agent_name:
                    agents_set.add(agent_name)
                
                for date, value in daily_achievements.items():
                    if value > 0:
                        daily_records.append({
                            'date': date,
                            'target_id': t.get('id', ''),
                            'agent_name': agent_name,
                            'product_group': t.get('product_group', ''),
                            'daily_target': t.get('daily_target', 0),
                            'target_count': t.get('target_count', 0),
                            'unit': t.get('unit', ''),
                            'achieved': value
                        })
            
            if not daily_records:
                self.show_message('اطلاع', 'هنوز هیچ تحققی ثبت نشده است')
                return
            
            # استخراج ماه‌های موجود
            months_set = set()
            for r in daily_records:
                date = r.get('date', '')
                if date and len(date) >= 7:
                    months_set.add(date[:7])
            
            months_list = sorted(list(months_set), reverse=True)
            months_display = ['همه'] + months_list
            
            agents_list = sorted(list(agents_set))
            agents_display = ['همه'] + agents_list
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                h_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(h_rect, 'pos', v),
                            size=lambda i, v: setattr(h_rect, 'size', v))
            
            # ========== فیلترها ==========
            filter_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
            
            # فیلتر ماه
            month_filter = PersianComboBox(
                text='همه', values=months_display, height=dp(38)
            )
            month_filter.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            month_filter.main_btn.color = (1, 1, 1, 1)
            month_filter.main_btn.font_size = sp(14)
            month_filter.size_hint_x = 0.3
            filter_row.add_widget(month_filter)
            
            # فیلتر عامل
            agent_filter = PersianComboBox(
                text='همه', values=agents_display, height=dp(38)
            )
            agent_filter.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            agent_filter.main_btn.color = (1, 1, 1, 1)
            agent_filter.main_btn.font_size = sp(14)
            agent_filter.size_hint_x = 0.35
            filter_row.add_widget(agent_filter)
            
            # عنوان با تعداد
            count_label = RTLLabel(
                text=f'{len(daily_records)} ثبت',
                size_hint_x=0.35, size_hint_y=None, height=dp(38),
                font_size=sp(16), bold=True, color=(0.4, 0.7, 1, 1)
            )
            filter_row.add_widget(count_label)
            
            content.add_widget(filter_row)
            
            # ========== هدر جدول ==========
            hist_header = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(2))
            hist_headers = [
                ('تاریخ', 0.14), ('تحقق', 0.14), ('تارگت روز', 0.14),
                ('گروه کالا', 0.25), ('عامل', 0.20), ('شناسه', 0.13)
            ]
            for text, size in hist_headers:
                hist_header.add_widget(RTLLabel(
                    text=text, size_hint_x=size, size_hint_y=None, height=dp(28),
                    font_size=sp(11), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
                ))
            content.add_widget(hist_header)
            
            # ========== جدول ==========
            hist_scroll = ScrollView(
                do_scroll_x=False, do_scroll_y=True,
                size_hint_y=0.65, scroll_type=['bars', 'content'], bar_width=dp(5)
            )
            hist_grid = GridLayout(cols=1, spacing=dp(2), size_hint_y=None, padding=dp(2))
            hist_grid.bind(minimum_height=hist_grid.setter('height'))
            
            # داده فیلترشده برای خروجی اکسل
            current_filtered_data = []
            
            def populate_history():
                nonlocal current_filtered_data
                
                selected_month = month_filter.text or 'همه'
                selected_agent = agent_filter.text or 'همه'
                
                hist_grid.clear_widgets()
                
                filtered_list = daily_records
                
                if selected_month != 'همه':
                    filtered_list = [r for r in filtered_list if r.get('date', '').startswith(selected_month)]
                
                if selected_agent != 'همه':
                    filtered_list = [r for r in filtered_list if r.get('agent_name', '') == selected_agent]
                
                filtered_list.sort(key=lambda x: x.get('date', ''), reverse=True)
                current_filtered_data = filtered_list
                
                count_label.text = f'{len(filtered_list)} ثبت'
                
                if not filtered_list:
                    hist_grid.add_widget(RTLLabel(
                        text='هیچ تحققی با این فیلترها یافت نشد',
                        size_hint_y=None, height=dp(40),
                        font_size=sp(14), color=(0.5, 0.5, 0.5, 1)
                    ))
                    return
                
                for r in filtered_list:
                    daily_target = r.get('daily_target', 1)
                    achieved = r.get('achieved', 0)
                    
                    if daily_target > 0:
                        percent = (achieved / daily_target) * 100
                        if percent < 50: date_color = (0.8, 0.3, 0.3, 1)
                        elif percent < 75: date_color = (1, 0.7, 0, 1)
                        elif percent < 100: date_color = (0.3, 0.6, 1, 1)
                        else: date_color = (0.2, 0.8, 0.2, 1)
                    else:
                        date_color = (0.5, 0.5, 0.5, 1)
                    
                    row = BoxLayout(size_hint_y=None, height=dp(33), spacing=dp(2))
                    row.add_widget(RTLLabel(
                        text=r.get('date', ''), size_hint_x=0.14, size_hint_y=None, height=dp(31),
                        font_size=sp(12), bold=True, color=date_color, halign='center'
                    ))
                    row.add_widget(RTLLabel(
                        text=f"{achieved:,}", size_hint_x=0.14, size_hint_y=None, height=dp(31),
                        font_size=sp(12), color=(1, 1, 1, 1), halign='center'
                    ))
                    row.add_widget(RTLLabel(
                        text=f"{daily_target:,}", size_hint_x=0.14, size_hint_y=None, height=dp(31),
                        font_size=sp(12), color=(0.8, 0.8, 0.8, 1), halign='center'
                    ))
                    row.add_widget(RTLLabel(
                        text=r.get('product_group', ''), size_hint_x=0.25, size_hint_y=None, height=dp(31),
                        font_size=sp(12), color=(1, 1, 1, 1), halign='right'
                    ))
                    row.add_widget(RTLLabel(
                        text=r.get('agent_name', ''), size_hint_x=0.20, size_hint_y=None, height=dp(31),
                        font_size=sp(12), color=(0.6, 0.6, 0.6, 1), halign='right'
                    ))
                    row.add_widget(RTLLabel(
                        text=r.get('target_id', ''), size_hint_x=0.13, size_hint_y=None, height=dp(31),
                        font_size=sp(11), color=(0.4, 0.7, 1, 1), halign='center'
                    ))
                    hist_grid.add_widget(row)
                
                # جمع کل
                total_achieved = sum(r.get('achieved', 0) for r in filtered_list)
                summary_row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(2))
                summary_row.add_widget(RTLLabel(
                    text=f'جمع کل: {total_achieved:,}', size_hint_x=1, size_hint_y=None, height=dp(28),
                    font_size=sp(13), bold=True, color=(0.2, 0.8, 0.2, 1), halign='center'
                ))
                hist_grid.add_widget(summary_row)


            # ✅ bind ساده - فقط populate رو صدا کن
            month_filter.bind(text=lambda inst, val: populate_history())
            agent_filter.bind(text=lambda inst, val: populate_history())
            populate_history()
            
            hist_scroll.add_widget(hist_grid)
            content.add_widget(hist_scroll)
            
            # ========== دکمه‌های پایین ==========
            btn_row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
            
            export_btn = PersianButton(
                text='خروجی اکسل', size_hint_x=0.5, size_hint_y=None, height=dp(38),
                background_color=(0.2, 0.7, 0.4, 1), color=(1, 1, 1, 1), font_size=sp(15)
            )
            btn_row.add_widget(export_btn)
            
            close_btn = PersianButton(
                text='بستن', size_hint_x=0.5, size_hint_y=None, height=dp(38),
                background_color=(0.3, 0.3, 0.3, 1), color=(1, 1, 1, 1), font_size=sp(15)
            )
            btn_row.add_widget(close_btn)
            content.add_widget(btn_row)
            
            popup = PersianPopup(
                title='تاریخچه تحقق روزانه', content=content,
                size_hint=(0.94, 0.85), background_color=(0.08, 0.08, 0.08, 1), auto_dismiss=False
            )
            
            def export_current(inst):
                if not current_filtered_data:
                    self.show_message('خطا', 'داده‌ای برای خروجی وجود ندارد')
                    return
                
                # تبدیل به فرمت مناسب برای export_to_excel
                # چون export_to_excel فرمت detailed_target می‌خواد، یه خروجی ساده می‌سازیم
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    from openpyxl.utils import get_column_letter
                    from utils.storage import get_backup_path
                    from datetime import datetime
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "تاریخچه تحقق روزانه"
                    
                    header_font = Font(bold=True, size=10, color="FFFFFF")
                    header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    headers = ['تاریخ', 'شناسه', 'عامل', 'گروه کالا', 'تارگت روز', 'تحقق', 'درصد', 'واحد']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                    
                    for row, r in enumerate(current_filtered_data, 2):
                        daily = r.get('daily_target', 1)
                        achieved = r.get('achieved', 0)
                        percent = f"{(achieved / daily * 100):.0f}%" if daily > 0 else "0%"
                        
                        values = [
                            r.get('date', ''), r.get('target_id', ''),
                            r.get('agent_name', ''), r.get('product_group', ''),
                            daily, achieved, percent, r.get('unit', '')
                        ]
                        for col, value in enumerate(values, 1):
                            cell = ws.cell(row=row, column=col, value=value)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = thin_border
                    
                    column_widths = [14, 16, 18, 20, 14, 14, 10, 12]
                    for i, width in enumerate(column_widths, 1):
                        ws.column_dimensions[get_column_letter(i)].width = width
                    
                    today = get_today_jalali().replace('/', '-')
                    filename = f'تاریخچه_تحقق_{today}_{datetime.now().strftime("%H%M%S")}.xlsx'
                    export_dir = get_backup_path()
                    os.makedirs(export_dir, exist_ok=True)
                    filepath = os.path.join(export_dir, filename)
                    wb.save(filepath)
                    self.show_message('موفق', f'فایل ذخیره شد:\n{filename}')
                    
                except ImportError:
                    self.show_message('خطا', 'ماژول openpyxl نصب نیست')
                except Exception as e:
                    self.show_message('خطا', f'خطا در خروجی: {str(e)}')
            
            export_btn.bind(on_press=export_current)
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)

    def _perform_fulfillment(self, target_ids, achieved_values):
        """نهایی‌سازی تارگت‌های اصلی با نام کاربر"""
        try:
            from kivy.app import App
            
            # ✅ دریافت نام کاربر جاری
            app = App.get_running_app()
            current_username = app.current_username if hasattr(app, 'current_username') else ''
            
            if not current_username:
                try:
                    from utils.user_manager import get_current_user
                    user = get_current_user()
                    if user:
                        current_username = user.get('username', '') or user.get('name', '')
                except:
                    pass
            
            if not current_username:
                current_username = 'supervisor'
            
            # ارسال نام کاربر به تابع finalize_targets
            success, message = finalize_targets(target_ids, achieved_values, finalized_by=current_username)
            
            if success:
                self.show_message('موفق', 'عملیات نهایی سازی با موفقیت انجام شد')
                self.show_fulfillment_targets(None)
            else:
                self.show_message('خطا', message)
                
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نهایی‌سازی تارگت‌ها: {e}", error_details)
            
    # ============================================================
    # تب ۳: بررسی بازار - با ساختار جدید مثل ایجنت اسکرین
    # ============================================================

    def show_market_check_tab(self):
        """نمایش تب بررسی بازار با ساختار جدید (مثل ایجنت اسکرین)"""
        try:
            from utils.supervisor_visits_manager import create_supervisor_visit
            from constants import (
                VISIT_TYPES, VISIT_REASONS, CUSTOMER_STATUSES,
                SHELF_STATUSES, MONTHLY_VISITS, VISIT_SUFFICIENT,
                EXPECTED_PURCHASE, INVENTORY_STATUSES, BEHAVIOR_RATINGS,
                SATISFACTION_RATINGS, TARGET_ACHIEVEMENTS, YES_NO_OPTIONS
            )

            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint=(1, 1),
                scroll_type=['bars', 'content'],
                bar_width=dp(8)
            )

            content = BoxLayout(
                orientation='vertical',
                padding=dp(15),
                spacing=dp(8),
                size_hint_y=None
            )
            content.bind(minimum_height=content.setter('height'))

            # ========== عنوان ==========
            content.add_widget(RTLLabel(
                text='بررسی بازار',
                font_size=sp(22),
                size_hint_y=None,
                height=dp(50),
                color=(1, 1, 1, 1),
                bold=True
            ))
            content.add_widget(Label(size_hint_y=None, height=dp(5)))

            # ========== تاریخ و ساعت ==========
            date_time_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(10))
            date_time_layout.add_widget(RTLLabel(
                text=f'تاریخ: {get_today_jalali()}',
                size_hint_x=0.5,
                size_hint_y=None,
                height=dp(40),
                font_size=sp(18),
                color=(0.4, 0.7, 1, 1)
            ))
            self.market_time_label = RTLLabel(
                text=f'ساعت: {get_current_time()}',
                size_hint_x=0.5,
                size_hint_y=None,
                height=dp(40),
                font_size=sp(18),
                color=(0.4, 0.7, 1, 1)
            )
            date_time_layout.add_widget(self.market_time_label)
            content.add_widget(date_time_layout)
            content.add_widget(Label(size_hint_y=None, height=dp(5)))

            # ========== مسیر ==========
            content.add_widget(RTLLabel(
                text='انتخاب مسیر:',
                size_hint_y=None,
                height=dp(30),
                font_size=sp(16),
                color=(0.4, 0.7, 1, 1),
                bold=True
            ))

            routes = get_routes()
            route_names = [r.get('name', '') for r in routes] if routes else ['']

            self.market_route_spinner = PersianComboBox(
                text='',
                values=route_names,
                height=dp(70)
            )
            self.market_route_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_route_spinner.main_btn.color = (1, 1, 1, 1)
            self.market_route_spinner.main_btn.font_size = sp(18)
            content.add_widget(self.market_route_spinner)
            content.add_widget(Label(size_hint_y=None, height=dp(5)))

            # ========== دکمه افزودن مشتری ==========
            add_customer_btn = PersianButton(
                text='افزودن مشتری جدید',
                background_color=(0.2, 0.6, 0.2, 1),
                size_hint_y=None,
                height=dp(40),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            add_customer_btn.bind(on_press=self.show_add_customer_dialog)
            content.add_widget(add_customer_btn)

            # ========== دکمه انتخاب مشتری ==========
            select_customer_btn = PersianButton(
                text='انتخاب مشتری',
                background_color=(0.2, 0.5, 0.9, 1),
                size_hint_y=None,
                height=dp(55),
                color=(1, 1, 1, 1),
                font_size=sp(20),
                bold=True
            )
            select_customer_btn.bind(on_press=self.show_market_customer_selection_dialog)
            content.add_widget(select_customer_btn)

            # ========== نمایش مشتری انتخاب شده ==========
            self.market_selected_customer_label = RTLLabel(
                text='مشتری انتخاب شده: هیچ',
                size_hint_y=None,
                height=dp(40),
                font_size=sp(18),
                color=(0.5, 0.5, 0.5, 1),
                bold=True
            )
            content.add_widget(self.market_selected_customer_label)

            # ========== دکمه بازگشت ==========
            back_btn = PersianButton(
                text='بازگشت',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            back_btn.bind(on_press=self.go_back)
            content.add_widget(back_btn)

            scroll.add_widget(content)
            self.content_area.add_widget(scroll)

            Clock.schedule_interval(self.update_market_time, 60)

            # تنظیم Clock برای تغییر مسیر
            self._last_market_route_text = self.market_route_spinner.text
            event = Clock.schedule_interval(self._check_market_route_change, 0.5)
            if not hasattr(self, '_clock_events'):
                self._clock_events = []
            self._clock_events.append(event)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش تب بررسی بازار: {e}", error_details)


    def update_market_time(self, dt):
        """بروزرسانی ساعت در تب بازار"""
        if hasattr(self, 'market_time_label'):
            self.market_time_label.text = f'ساعت: {get_current_time()}'


    def _check_market_route_change(self, dt):
        """بررسی تغییر مسیر در تب بررسی بازار - بدون قفل"""
        if not hasattr(self, 'market_route_spinner'):
            return
        try:
            current_text = self.market_route_spinner.text
            if current_text != self._last_market_route_text and current_text and current_text != '':
                self._last_market_route_text = current_text
                # ✅ فقط مسیر را بروزرسانی کن، بدون قفل و بدون دیالوگ تأیید
                self.show_message('اطلاع', f'مسیر "{current_text}" انتخاب شد.')
        except Exception as e:
            print(f"خطا در _check_market_route_change: {e}")

    # ============================================================
    # دیالوگ افزودن مشتری برای سوپروایزر
    # ============================================================

    def show_add_customer_dialog(self, instance):
        """دیالوگ افزودن مشتری جدید برای سوپروایزر"""
        try:
            from utils.file_manager import get_routes, get_customers, add_customer

            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                        size=lambda i, v: setattr(content_rect, 'size', v))

            content.add_widget(RTLLabel(
                text='افزودن مشتری جدید',
                size_hint_y=None,
                height=dp(40),
                font_size=sp(22),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))

            content.add_widget(RTLLabel(
                text='انتخاب مسیر:',
                size_hint_y=None,
                height=dp(28),
                font_size=sp(16),
                color=(1, 1, 1, 1)
            ))

            routes = get_routes()
            if routes:
                route_names = [r.get('name', '') for r in routes]
            else:
                route_names = ['ابتدا مسیر ایجاد کنید']

            default_route = self.market_route_spinner.text if hasattr(self, 'market_route_spinner') and self.market_route_spinner.text else route_names[0]

            customer_route_spinner = PersianComboBox(
                text=default_route if default_route in route_names else route_names[0],
                values=route_names,
                height=dp(60)
            )
            customer_route_spinner.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            customer_route_spinner.main_btn.color = (1, 1, 1, 1)
            customer_route_spinner.main_btn.font_size = sp(18)
            content.add_widget(customer_route_spinner)

            content.add_widget(RTLLabel(
                text='نام مشتری (الزامی):',
                size_hint_y=None,
                height=dp(28),
                font_size=sp(16),
                color=(1, 1, 1, 1)
            ))
            customer_name_input = RTLTextInput(
                hint_text='نام مشتری را وارد کنید',
                multiline=False,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(22)
            )
            customer_name_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_name_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_name_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_name_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(customer_name_input)

            content.add_widget(RTLLabel(
                text='نام فروشگاه:',
                size_hint_y=None,
                height=dp(28),
                font_size=sp(16),
                color=(1, 1, 1, 1)
            ))
            customer_store_input = RTLTextInput(
                hint_text='نام فروشگاه را وارد کنید (اختیاری)',
                multiline=False,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(22)
            )
            customer_store_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_store_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_store_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_store_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(customer_store_input)

            content.add_widget(RTLLabel(
                text='موبایل (الزامی):',
                size_hint_y=None,
                height=dp(28),
                font_size=sp(16),
                color=(1, 1, 1, 1)
            ))
            customer_mobile_input = RTLTextInput(
                hint_text='شماره موبایل را وارد کنید (11 رقم)',
                multiline=False,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(22)
            )
            customer_mobile_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_mobile_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_mobile_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_mobile_input._hidden_input.foreground_color = (1, 1, 1, 1)
            customer_mobile_input._hidden_input.input_filter = 'int'
            content.add_widget(customer_mobile_input)

            content.add_widget(RTLLabel(
                text='آدرس:',
                size_hint_y=None,
                height=dp(28),
                font_size=sp(16),
                color=(1, 1, 1, 1)
            ))
            customer_address_input = RTLTextInput(
                hint_text='آدرس را وارد کنید (اختیاری)',
                multiline=False,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(22)
            )
            customer_address_input.bg_color = (0.15, 0.15, 0.15, 1)
            customer_address_input.border_color = (0.3, 0.3, 0.3, 1)
            customer_address_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            customer_address_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(customer_address_input)

            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))

            submit_btn = PersianButton(
                text='افزودن مشتری',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(18)
            )
            cancel_btn = PersianButton(
                text='انصراف',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(18)
            )

            btn_layout.add_widget(submit_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)

            popup = PersianPopup(
                title='افزودن مشتری',
                content=content,
                size_hint=(0.9, 0.75),
                background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            def do_add_customer(instance):
                try:
                    route_name = customer_route_spinner.text
                    if route_name == 'ابتدا مسیر ایجاد کنید':
                        self.show_message('خطا', 'لطفاً ابتدا یک مسیر ایجاد کنید')
                        return

                    name = customer_name_input.text.strip()
                    if not name:
                        self.show_message('خطا', 'نام مشتری الزامی است')
                        customer_name_input._hidden_input.focus = True
                        return

                    mobile = customer_mobile_input.text.strip()
                    if not mobile:
                        self.show_message('خطا', 'شماره موبایل الزامی است')
                        customer_mobile_input._hidden_input.focus = True
                        return

                    mobile_clean = mobile.replace(' ', '').replace('-', '').replace('_', '')
                    if not mobile_clean.isdigit():
                        self.show_message('خطا', 'شماره موبایل باید فقط شامل عدد باشد')
                        customer_mobile_input._hidden_input.focus = True
                        return
                    
                    if len(mobile_clean) != 11:
                        self.show_message('خطا', 'شماره موبایل باید ۱۱ رقم باشد')
                        customer_mobile_input._hidden_input.focus = True
                        return
                    
                    if not mobile_clean.startswith('09'):
                        self.show_message('خطا', 'شماره موبایل باید با 09 شروع شود')
                        customer_mobile_input._hidden_input.focus = True
                        return

                    all_customers = get_customers()
                    for c in all_customers:
                        if c.get('name', '').strip() == name:
                            self.show_message('خطا', f'مشتری با نام "{name}" قبلاً ثبت شده است')
                            customer_name_input._hidden_input.focus = True
                            return
                        
                        existing_mobile = c.get('mobile', '').strip()
                        if existing_mobile and existing_mobile == mobile_clean:
                            self.show_message('خطا', f'شماره موبایل "{mobile_clean}" قبلاً برای مشتری "{c.get("name")}" ثبت شده است')
                            customer_mobile_input._hidden_input.focus = True
                            return

                    customer = {
                        'name': name,
                        'store_name': customer_store_input.text.strip(),
                        'route_name': route_name,
                        'mobile': mobile_clean,
                        'address': customer_address_input.text.strip()
                    }
                    
                    add_customer(customer)
                    popup.dismiss()
                    self.show_message('موفق', f'مشتری "{name}" با موفقیت اضافه شد')
                    
                except Exception as e:
                    error_details = traceback.format_exc()
                    ErrorPopup.show_error(f"خطا در افزودن مشتری: {e}", error_details)

            def cancel_add(instance):
                popup.dismiss()

            submit_btn.bind(on_press=do_add_customer)
            cancel_btn.bind(on_press=cancel_add)
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ افزودن مشتری: {e}", error_details)

    def show_market_route_confirm_dialog(self, route_name):
        """دیالوگ تأیید مسیر برای بررسی بازار - بدون قفل شدن"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.15, 0.15, 0.15, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                            size=lambda i, v: setattr(content_rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text=f'آیا قصد بررسی مسیر "{route_name}" را دارید؟',
                size_hint_y=None,
                height=dp(50),
                font_size=sp(18),
                color=(1, 1, 1, 1)
            ))
            
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(50))
            yes_btn = PersianButton(
                text='بله',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            no_btn = PersianButton(
                text='خیر',
                background_color=(0.8, 0.2, 0.2, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            btn_layout.add_widget(yes_btn)
            btn_layout.add_widget(no_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='تأیید مسیر',
                content=content,
                size_hint=(0.85, 0.35),
                background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )
            
            def on_yes(instance):
                popup.dismiss()
                # ✅ حذف قفل کردن مسیر - سوپروایزر می‌تواند چند مسیر را بررسی کند
                self.show_message('اطلاع', f'مسیر "{route_name}" انتخاب شد. حالا مشتری را انتخاب کنید.')
            
            def on_no(instance):
                popup.dismiss()
                self.market_route_spinner.text = ''
                self._last_market_route_text = ''
            
            yes_btn.bind(on_press=on_yes)
            no_btn.bind(on_press=on_no)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ تأیید مسیر: {e}", error_details)


    def show_market_customer_selection_dialog(self, instance):
        """نمایش دیالوگ انتخاب مشتری برای بررسی بازار"""
        try:
            selected_route = self.market_route_spinner.text
            
            if not selected_route:
                self.show_message('خطا', 'لطفاً ابتدا یک مسیر انتخاب کنید')
                return
            
            all_customers = get_customers()
            filtered_customers = []
            for c in all_customers:
                if c.get('route_name', '').strip() == selected_route.strip():
                    filtered_customers.append(c.get('name', ''))
            
            if not filtered_customers:
                self.show_message('توجه', 'هیچ مشتری‌ای در این مسیر یافت نشد')
                return
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                content_rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(content_rect, 'pos', v),
                            size=lambda i, v: setattr(content_rect, 'size', v))
            
            title_layout = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(60))
            title_layout.add_widget(RTLLabel(
                text=f'انتخاب مشتری - {selected_route}',
                size_hint_y=None,
                height=dp(30),
                font_size=sp(18),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))
            title_layout.add_widget(RTLLabel(
                text='مشتریان آبی رنگ امروز بررسی شده‌اند',
                size_hint_y=None,
                height=dp(25),
                font_size=sp(15),
                color=(0.6, 0.6, 0.6, 1)
            ))
            content.add_widget(title_layout)
            
            search_input = RTLTextInput(
                hint_text='جستجوی مشتری...',
                multiline=False,
                size_hint_y=None,
                height=dp(55),
                font_size=sp(24)
            )
            search_input.bg_color = (0.15, 0.15, 0.15, 1)
            search_input.border_color = (0.3, 0.3, 0.3, 1)
            search_input.border_color_focus = (0.2, 0.5, 0.9, 1)
            search_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(search_input)
            
            customers_scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=0.65,
                scroll_type=['bars', 'content'],
                bar_width=dp(6)
            )
            customers_grid = GridLayout(cols=1, spacing=dp(4), size_hint_y=None, padding=dp(5))
            customers_grid.bind(minimum_height=customers_grid.setter('height'))
            
            def filter_customers(text):
                customers_grid.clear_widgets()
                search_text = text.strip()
                for customer in filtered_customers:
                    if search_text and search_text not in customer:
                        continue
                    customer_btn = PersianButton(
                        text=customer,
                        size_hint_y=None,
                        height=dp(45),
                        background_color=(0.2, 0.2, 0.2, 1),
                        color=(1, 1, 1, 1),
                        font_size=sp(18)
                    )
                    customer_btn.bind(
                        on_press=lambda x, name=customer: self._handle_market_customer_selection(name)
                    )
                    customers_grid.add_widget(customer_btn)
                customers_grid.height = len(customers_grid.children) * dp(50) + dp(10)
            
            search_input._hidden_input.bind(text=lambda i, v: filter_customers(v))
            filter_customers('')
            customers_scroll.add_widget(customers_grid)
            content.add_widget(customers_scroll)
            
            close_btn = PersianButton(
                text='بستن',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='انتخاب مشتری',
                content=content,
                size_hint=(0.9, 0.75),
                background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=True
            )
            self.market_customer_popup = popup
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ انتخاب مشتری: {e}", error_details)


    def _handle_market_customer_selection(self, customer_name):
        """مدیریت انتخاب مشتری و نمایش دیالوگ بررسی بازار"""
        try:
            if hasattr(self, 'market_customer_popup'):
                self.market_customer_popup.dismiss()
            
            self.market_selected_customer = customer_name
            self.market_selected_customer_label.text = f'مشتری انتخاب شده: {customer_name}'
            self.market_selected_customer_label.color = (0.2, 0.8, 0.4, 1)
            
            self.show_market_check_dialog(customer_name)
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در انتخاب مشتری: {e}", error_details)


    def show_market_check_dialog(self, customer_name):
        """نمایش دیالوگ بررسی بازار با اسکرول و ساختار جابجا شده"""
        try:
            from constants import (
                VISIT_TYPES, VISIT_REASONS, CUSTOMER_STATUSES,
                SHELF_STATUSES, MONTHLY_VISITS, VISIT_SUFFICIENT,
                EXPECTED_PURCHASE, INVENTORY_STATUSES, BEHAVIOR_RATINGS,
                SATISFACTION_RATINGS, TARGET_ACHIEVEMENTS, YES_NO_OPTIONS
            )

            # ========== کانتینر اصلی با اسکرول ==========
            main_container = BoxLayout(orientation='vertical', spacing=dp(3))
            
            # ========== عنوان (کوچک‌تر) ==========
            title_box = BoxLayout(
                size_hint_y=None, 
                height=dp(32), 
                spacing=dp(5),
                padding=[dp(5), dp(2), dp(5), dp(2)]
            )
            with title_box.canvas.before:
                Color(0.15, 0.2, 0.3, 1)
                rect = Rectangle(pos=title_box.pos, size=title_box.size)
                title_box.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                            size=lambda i, v: setattr(rect, 'size', v))
            
            title_box.add_widget(RTLLabel(
                text=f'📋 بررسی بازار',
                size_hint_x=0.4,
                size_hint_y=None,
                height=dp(30),
                font_size=sp(16),
                bold=True,
                color=(0.4, 0.8, 1, 1)
            ))
            
            title_box.add_widget(RTLLabel(
                text=f'{customer_name}',
                size_hint_x=0.6,
                size_hint_y=None,
                height=dp(30),
                font_size=sp(14),
                color=(1, 1, 1, 1),
                halign='left'
            ))
            
            main_container.add_widget(title_box)

            # ========== اسکرول ==========
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint=(1, 1),
                scroll_type=['bars', 'content'],
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )

            content = BoxLayout(
                orientation='vertical',
                padding=dp(8),
                spacing=dp(5),
                size_hint_y=None
            )
            content.bind(minimum_height=content.setter('height'))
            
            # ============================================================
            # ساختار با جابجایی لیبل و فیلد (فیلد در سمت راست، لیبل در سمت چپ)
            # ============================================================
            
            def make_row(label_text, widget, height=dp(48)):
                """ساخت یک ردیف با لیبل و فیلد"""
                row = BoxLayout(size_hint_y=None, height=height, spacing=dp(6))
                row.add_widget(widget)  # فیلد اول
                row.add_widget(RTLLabel(
                    text=label_text,
                    size_hint_x=0.28,
                    size_hint_y=None,
                    height=height,
                    font_size=sp(12),
                    color=(1, 1, 1, 1),
                    halign='right'
                ))
                return row

            # ۱- نحوه سرکشی
            self.market_dialog_visit_type = PersianComboBox(
                text=VISIT_TYPES[0],
                values=VISIT_TYPES,
                height=dp(44)
            )
            self.market_dialog_visit_type.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_visit_type.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_visit_type.main_btn.font_size = sp(15)
            self.market_dialog_visit_type.size_hint_x = 0.72
            content.add_widget(make_row('نحوه سرکشی:', self.market_dialog_visit_type))

            # ۲- علت سرکشی
            self.market_dialog_visit_reason = PersianComboBox(
                text=VISIT_REASONS[0],
                values=VISIT_REASONS,
                height=dp(44)
            )
            self.market_dialog_visit_reason.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_visit_reason.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_visit_reason.main_btn.font_size = sp(15)
            self.market_dialog_visit_reason.size_hint_x = 0.72
            content.add_widget(make_row('علت سرکشی:', self.market_dialog_visit_reason))

            # ۳- توضیحات سوپروایزر
            self.market_dialog_supervisor_note = RTLTextInput(
                hint_text='توضیحات را وارد کنید...',
                multiline=True,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(15)
            )
            self.market_dialog_supervisor_note.bg_color = (0.15, 0.15, 0.15, 1)
            self.market_dialog_supervisor_note.border_color = (0.3, 0.3, 0.3, 1)
            self.market_dialog_supervisor_note.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.market_dialog_supervisor_note._hidden_input.foreground_color = (1, 1, 1, 1)
            self.market_dialog_supervisor_note.size_hint_x = 0.72
            content.add_widget(make_row('توضیحات سوپروایزر:', self.market_dialog_supervisor_note, height=dp(70)))

            # ۴- وضعیت مشتری
            self.market_dialog_customer_status = PersianComboBox(
                text=CUSTOMER_STATUSES[0],
                values=CUSTOMER_STATUSES,
                height=dp(44)
            )
            self.market_dialog_customer_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_customer_status.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_customer_status.main_btn.font_size = sp(15)
            self.market_dialog_customer_status.size_hint_x = 0.72
            content.add_widget(make_row('وضعیت مشتری:', self.market_dialog_customer_status))

            # ۵- وضعیت حضور در شلف
            self.market_dialog_shelf_status = PersianComboBox(
                text=SHELF_STATUSES[0],
                values=SHELF_STATUSES,
                height=dp(44)
            )
            self.market_dialog_shelf_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_shelf_status.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_shelf_status.main_btn.font_size = sp(15)
            self.market_dialog_shelf_status.size_hint_x = 0.72
            content.add_widget(make_row('وضعیت حضور در شلف:', self.market_dialog_shelf_status))

            # ۶- تعداد سرکشی بازاریابان در ماه
            self.market_dialog_monthly_visits = PersianComboBox(
                text=MONTHLY_VISITS[0],
                values=MONTHLY_VISITS,
                height=dp(44)
            )
            self.market_dialog_monthly_visits.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_monthly_visits.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_monthly_visits.main_btn.font_size = sp(15)
            self.market_dialog_monthly_visits.size_hint_x = 0.72
            content.add_widget(make_row('تعداد سرکشی در ماه:', self.market_dialog_monthly_visits))

            # ۷- آیا میزان سرکشی کافیست؟
            self.market_dialog_visit_sufficient = PersianComboBox(
                text=VISIT_SUFFICIENT[0],
                values=VISIT_SUFFICIENT,
                height=dp(44)
            )
            self.market_dialog_visit_sufficient.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_visit_sufficient.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_visit_sufficient.main_btn.font_size = sp(15)
            self.market_dialog_visit_sufficient.size_hint_x = 0.72
            content.add_widget(make_row('آیا سرکشی کافیست؟:', self.market_dialog_visit_sufficient))

            # ۸- میزان خرید مورد انتظار
            self.market_dialog_expected_purchase = PersianComboBox(
                text=EXPECTED_PURCHASE[0],
                values=EXPECTED_PURCHASE,
                height=dp(44)
            )
            self.market_dialog_expected_purchase.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_expected_purchase.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_expected_purchase.main_btn.font_size = sp(15)
            self.market_dialog_expected_purchase.size_hint_x = 0.72
            content.add_widget(make_row('خرید مورد انتظار:', self.market_dialog_expected_purchase))

            # ۹- وضعیت موجودی
            self.market_dialog_inventory_status = PersianComboBox(
                text=INVENTORY_STATUSES[0],
                values=INVENTORY_STATUSES,
                height=dp(44)
            )
            self.market_dialog_inventory_status.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_inventory_status.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_inventory_status.main_btn.font_size = sp(15)
            self.market_dialog_inventory_status.size_hint_x = 0.72
            content.add_widget(make_row('وضعیت موجودی:', self.market_dialog_inventory_status))

            # ۱۰- نحوه برخورد بازاریابان
            self.market_dialog_agent_behavior = PersianComboBox(
                text=BEHAVIOR_RATINGS[0],
                values=BEHAVIOR_RATINGS,
                height=dp(44)
            )
            self.market_dialog_agent_behavior.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_agent_behavior.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_agent_behavior.main_btn.font_size = sp(15)
            self.market_dialog_agent_behavior.size_hint_x = 0.72
            content.add_widget(make_row('برخورد بازاریاب:', self.market_dialog_agent_behavior))

            # ۱۱- نحوه برخورد موزعین
            self.market_dialog_distributor_behavior = PersianComboBox(
                text=BEHAVIOR_RATINGS[0],
                values=BEHAVIOR_RATINGS,
                height=dp(44)
            )
            self.market_dialog_distributor_behavior.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_distributor_behavior.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_distributor_behavior.main_btn.font_size = sp(15)
            self.market_dialog_distributor_behavior.size_hint_x = 0.72
            content.add_widget(make_row('برخورد موزع:', self.market_dialog_distributor_behavior))

            # ۱۲- میزان رضایتمندی مشتری
            self.market_dialog_customer_satisfaction = PersianComboBox(
                text=SATISFACTION_RATINGS[0],
                values=SATISFACTION_RATINGS,
                height=dp(44)
            )
            self.market_dialog_customer_satisfaction.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_customer_satisfaction.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_customer_satisfaction.main_btn.font_size = sp(15)
            self.market_dialog_customer_satisfaction.size_hint_x = 0.72
            content.add_widget(make_row('رضایتمندی مشتری:', self.market_dialog_customer_satisfaction))

            # ۱۳- نظرات مشتری
            self.market_dialog_customer_feedback = RTLTextInput(
                hint_text='حداکثر ۱۰۰۰ کاراکتر...',
                multiline=True,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(15)
            )
            self.market_dialog_customer_feedback.bg_color = (0.15, 0.15, 0.15, 1)
            self.market_dialog_customer_feedback.border_color = (0.3, 0.3, 0.3, 1)
            self.market_dialog_customer_feedback.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.market_dialog_customer_feedback._hidden_input.foreground_color = (1, 1, 1, 1)
            self.market_dialog_customer_feedback.size_hint_x = 0.72
            content.add_widget(make_row('نظرات مشتری:', self.market_dialog_customer_feedback, height=dp(70)))

            # ۱۴- میزان تحقق هدف سرکشی
            self.market_dialog_target_achievement = PersianComboBox(
                text=TARGET_ACHIEVEMENTS[0],
                values=TARGET_ACHIEVEMENTS,
                height=dp(44)
            )
            self.market_dialog_target_achievement.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_target_achievement.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_target_achievement.main_btn.font_size = sp(15)
            self.market_dialog_target_achievement.size_hint_x = 0.72
            content.add_widget(make_row('تحقق هدف سرکشی:', self.market_dialog_target_achievement))

            # ۱۵- نظریه سوپروایزر
            self.market_dialog_supervisor_opinion = RTLTextInput(
                hint_text='حداکثر ۱۵۰۰ کاراکتر...',
                multiline=True,
                size_hint_y=None,
                height=dp(60),
                font_size=sp(15)
            )
            self.market_dialog_supervisor_opinion.bg_color = (0.15, 0.15, 0.15, 1)
            self.market_dialog_supervisor_opinion.border_color = (0.3, 0.3, 0.3, 1)
            self.market_dialog_supervisor_opinion.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.market_dialog_supervisor_opinion._hidden_input.foreground_color = (1, 1, 1, 1)
            self.market_dialog_supervisor_opinion.size_hint_x = 0.72
            content.add_widget(make_row('نظریه سوپروایزر:', self.market_dialog_supervisor_opinion, height=dp(70)))

            # ۱۶- آیا پیگیری مجدد نیاز است؟
            self.market_dialog_need_followup = PersianComboBox(
                text=YES_NO_OPTIONS[0],
                values=YES_NO_OPTIONS,
                height=dp(44)
            )
            self.market_dialog_need_followup.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.market_dialog_need_followup.main_btn.color = (1, 1, 1, 1)
            self.market_dialog_need_followup.main_btn.font_size = sp(15)
            self.market_dialog_need_followup.size_hint_x = 0.72
            content.add_widget(make_row('نیاز به پیگیری:', self.market_dialog_need_followup))

            # ۱۷- تاریخ مراجعه بعدی
            self.market_dialog_next_visit_date = RTLTextInput(
                text='',
                hint_text='سال/ماه/روز',
                multiline=False,
                size_hint_y=None,
                height=dp(44),
                font_size=sp(18)
            )
            self.market_dialog_next_visit_date.bg_color = (0.15, 0.15, 0.15, 1)
            self.market_dialog_next_visit_date.border_color = (0.3, 0.3, 0.3, 1)
            self.market_dialog_next_visit_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.market_dialog_next_visit_date._hidden_input.foreground_color = (1, 1, 1, 1)
            self.market_dialog_next_visit_date.size_hint_x = 0.72
            content.add_widget(make_row('تاریخ مراجعه بعدی:', self.market_dialog_next_visit_date))

            # ============================================================
            # افزودن محتوا به اسکرول
            # ============================================================
            scroll.add_widget(content)
            main_container.add_widget(scroll)

            # ========== دکمه‌ها (خارج از اسکرول) ==========
            btn_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8), padding=dp(4))

            submit_btn = PersianButton(
                text='ثبت سرکشی',
                background_color=(0.2, 0.7, 0.2, 1),
                size_hint_y=None,
                height=dp(40),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            submit_btn.bind(on_press=lambda x: self._submit_market_check_dialog(customer_name))
            btn_layout.add_widget(submit_btn)

            cancel_btn = PersianButton(
                text='انصراف',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(40),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            cancel_btn.bind(on_press=self._close_market_dialog)
            btn_layout.add_widget(cancel_btn)

            main_container.add_widget(btn_layout)

            popup = PersianPopup(
                title='بررسی بازار',
                content=main_container,
                size_hint=(0.92, 0.82),
                background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )

            self._market_dialog_popup = popup
            popup.open()

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ بررسی بازار: {e}", error_details)


    def _submit_market_check_dialog(self, customer_name):
        """ثبت سرکشی از دیالوگ بررسی بازار"""
        try:
            from utils.supervisor_visits_manager import create_supervisor_visit
            from kivy.app import App
            from utils.jalali_date import validate_jalali_date

            # دریافت نام کاربر جاری
            app = App.get_running_app()
            current_username = app.current_username if hasattr(app, 'current_username') else ''

            if not current_username:
                try:
                    from utils.user_manager import get_current_user
                    user = get_current_user()
                    if user:
                        current_username = user.get('username', '') or user.get('name', '')
                except:
                    pass

            if not current_username:
                current_username = 'supervisor'

            data = {
                'route': self.market_route_spinner.text,
                'customer': customer_name,
                'visit_type': self.market_dialog_visit_type.text,
                'visit_reason': self.market_dialog_visit_reason.text,
                'supervisor_note': self.market_dialog_supervisor_note.text.strip(),
                'customer_status': self.market_dialog_customer_status.text,
                'shelf_status': self.market_dialog_shelf_status.text,
                'monthly_visits': self.market_dialog_monthly_visits.text,
                'visit_sufficient': self.market_dialog_visit_sufficient.text,
                'expected_purchase': self.market_dialog_expected_purchase.text,
                'inventory_status': self.market_dialog_inventory_status.text,
                'agent_behavior': self.market_dialog_agent_behavior.text,
                'distributor_behavior': self.market_dialog_distributor_behavior.text,
                'customer_satisfaction': self.market_dialog_customer_satisfaction.text,
                'customer_feedback': self.market_dialog_customer_feedback.text.strip(),
                'target_achievement': self.market_dialog_target_achievement.text,
                'supervisor_opinion': self.market_dialog_supervisor_opinion.text.strip(),
                'need_followup': self.market_dialog_need_followup.text,
                'next_visit_date': self.market_dialog_next_visit_date.text.strip(),
                'created_by': current_username,
                'agent_name': current_username
            }

            if not data['route'] or data['route'] == '':
                self.show_message('خطا', 'لطفاً یک مسیر را انتخاب کنید')
                return

            if data['need_followup'] == 'بله':
                if not data['next_visit_date']:
                    self.show_message('خطا', 'در صورت نیاز به پیگیری، تاریخ مراجعه بعدی را وارد کنید')
                    return
                if not validate_jalali_date(data['next_visit_date']):
                    self.show_message('خطا', 'فرمت تاریخ مراجعه بعدی نامعتبر است (مثال: 1405/01/31)')
                    return

            success, message, visit = create_supervisor_visit(data)

            if success:
                if hasattr(self, '_market_dialog_popup'):
                    self._market_dialog_popup.dismiss()
                
                self.market_selected_customer = None
                self.market_selected_customer_label.text = 'مشتری انتخاب شده: هیچ'
                self.market_selected_customer_label.color = (0.5, 0.5, 0.5, 1)
                
                self.show_message('موفق', message)
            else:
                self.show_message('خطا', message)

        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ثبت سرکشی: {e}", error_details)


    def _close_market_dialog(self, instance):
        """بستن دیالوگ بررسی بازار"""
        if hasattr(self, '_market_dialog_popup'):
            self._market_dialog_popup.dismiss()

    # ============================================================
    # تب ۴: گزارشات
    # ============================================================

    def show_reports_tab(self):
        """تب ۴: گزارشات - هدایت به صفحه گزارش سوپروایزر"""
        self.manager.current = 'supervisor_report'

    # ============================================================
    # توابع عمومی
    # ============================================================

    def go_back(self, instance):
        """بازگشت به صفحه ورود"""
        self._cleanup_current_tab()
        self.manager.current = 'login'

    def show_message(self, title, message):
        """نمایش پیام با پشتیبانی از متن طولانی"""
        try:
            # جلوگیری از نمایش همزمان
            if hasattr(self, '_active_message_popup') and self._active_message_popup:
                try:
                    self._active_message_popup.dismiss()
                except:
                    pass
                self._active_message_popup = None
            
            from kivy.uix.boxlayout import BoxLayout
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            from kivy.metrics import dp, sp
            from kivy.graphics import Color, Rectangle
            from utils.rtl_widgets import PersianPopup, PersianButton
            
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(message)
                display_text = get_display(reshaped)
            except:
                display_text = message
            
            msg_label = Label(
                text=display_text,
                font_size=sp(15),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                halign='center',
                valign='top',
                text_size=(dp(550), None),
                font_name='fonts/Amiri-Regular.ttf',
                padding=(dp(20), 0, dp(0), 0)
            )
            
            lines = message.count('\n') + 1
            if len(message) > 50 and lines == 1:
                approx_chars_per_line = 35
                lines = (len(message) // approx_chars_per_line) + 1
            
            line_height = sp(15) + dp(6)
            label_height = max(lines * line_height + dp(25), dp(50))
            label_height = min(label_height, dp(490))
            
            msg_label.height = label_height
            msg_label.text_size = (dp(550), label_height)
            
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=None,
                height=label_height + dp(10),
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )
            scroll.add_widget(msg_label)
            content.add_widget(scroll)
            
            btn = PersianButton(
                text='باشه',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(16),
                color=(1, 1, 1, 1),
                background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            
            popup = PersianPopup(
                title=title,
                content=content,
                size_hint=(0.9, None),
                height=label_height + dp(250),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            
            self._active_message_popup = popup
            
            def dismiss_and_clean(instance):
                popup.dismiss()
                self._active_message_popup = None
            
            btn.bind(on_press=dismiss_and_clean)
            popup.open()
            
        except Exception as e:
            self._active_message_popup = None
            print(f"خطا در نمایش پیام: {e}")
            import traceback
            traceback.print_exc()
            try:
                from error_handler import ErrorPopup
                ErrorPopup.show_error(str(message))
            except:
                pass
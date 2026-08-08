# screens/total_report_screen.py
# ========== صفحه گزارش عملکرد روزانه ==========

import os
import json
import traceback
import threading
from datetime import datetime
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.checkbox import CheckBox
from kivy.graphics import Color, Rectangle
from kivy.core.window import Window
from kivy.clock import Clock

from utils.rtl_widgets import (
    PersianButton, RTLLabel, PersianPopup, 
    RTLTextInput, PersianComboBox
)
from utils.jalali_date import get_today_jalali
from utils.user_manager import get_users, login
from utils.attendance_manager import AttendanceManager
from utils.storage import get_data_path, get_backup_path
from error_handler import ErrorPopup

# ✅ import توابع کمکی اصلی
from utils.excel_exporter import export_to_excel
from utils.excel_exporter_distributor import export_distributor_to_excel
from utils.collection_manager import get_collections
from utils.delivery_manager import get_all_deliveries
from utils.detailed_target_manager import get_all_detailed_targets
from utils.target_manager import get_targets_filtered, export_targets_to_excel
from utils.supervisor_visits_manager import get_visits_filtered, export_visits_to_excel
from utils.file_manager import get_daily_logs
from utils.file_manager import get_do_missions


class TotalReportScreen(Screen):
    """صفحه گزارش عملکرد روزانه - با لاگین جداگانه"""
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        with self.canvas.before:
            Color(0.08, 0.08, 0.08, 1)
            self.bg_rect = Rectangle(pos=self.pos, size=self.size)
            self.bind(pos=self._update_bg, size=self._update_bg)
        
        self.current_user = None
        self.report_checkboxes = {}
        self.build_ui()
    
    def _update_bg(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size
    
    def build_ui(self):
        """ساخت رابط کاربری"""
        self.clear_widgets()
        
        layout = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(5))
        
        # ========== هدر ==========
        header = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
        header.add_widget(RTLLabel(
            text='گزارش عملکرد روزانه',
            font_size=sp(20),
            bold=True,
            color=(0.4, 0.8, 1, 1),
            size_hint_x=0.5
        ))
        header.add_widget(RTLLabel(
            text=get_today_jalali(),
            font_size=sp(14),
            color=(0.6, 0.6, 0.6, 1),
            size_hint_x=0.3,
            halign='center'
        ))
        if self.current_user:
            header.add_widget(RTLLabel(
                text=self.current_user.get('name', ''),
                font_size=sp(14),
                color=(1, 1, 1, 1),
                size_hint_x=0.2,
                halign='left'
            ))
        layout.add_widget(header)
        
        # ========== محتوای اصلی ==========
        self.content_area = BoxLayout(orientation='vertical', size_hint_y=1)
        layout.add_widget(self.content_area)
        
        # ========== دکمه بازگشت ==========
        back_btn = PersianButton(
            text='بازگشت',
            size_hint_y=None,
            height=dp(40),
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(14)
        )
        back_btn.bind(on_press=self.go_back)
        layout.add_widget(back_btn)
        
        self.add_widget(layout)
        
        # نمایش فرم لاگین
        self.show_login_dialog()
    
    def go_back(self, instance):
        """بازگشت به صفحه ورود"""
        self.manager.current = 'login'
    
    def show_login_dialog(self):
        """نمایش دیالوگ لاگین گزارش عملکرد"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text='ورود به گزارش عملکرد روزانه',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                bold=True,
                color=(0.4, 0.8, 1, 1)
            ))
            
            content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
            
            # کامبوباکس کاربران
            content.add_widget(RTLLabel(
                text='انتخاب کاربر:',
                size_hint_y=None,
                height=dp(22),
                font_size=sp(13),
                color=(1, 1, 1, 1)
            ))
            
            users = get_users()
            user_names = []
            for u in users:
                name = u.get('name', '') or u.get('username', '')
                if name:
                    user_names.append(name)
            
            if not user_names:
                user_names = ['هیچ کاربری ثبت نشده']
            
            self.login_user_combo = PersianComboBox(
                text=user_names[0] if user_names else '',
                values=user_names,
                height=dp(50)
            )
            self.login_user_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
            self.login_user_combo.main_btn.color = (1, 1, 1, 1)
            self.login_user_combo.main_btn.font_size = sp(16)
            content.add_widget(self.login_user_combo)
            
            content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
            
            # رمز عبور
            content.add_widget(RTLLabel(
                text='رمز عبور:',
                size_hint_y=None,
                height=dp(22),
                font_size=sp(13),
                color=(1, 1, 1, 1)
            ))
            
            from kivy.uix.textinput import TextInput
            
            self.login_password = TextInput(
                hint_text='رمز عبور را وارد کنید',
                password=True,
                multiline=False,
                size_hint_y=None,
                height=dp(48),
                font_size=sp(16),
                halign='right',
                font_name='PersianFont',
                background_color=(0.15, 0.15, 0.15, 1),
                foreground_color=(1, 1, 1, 1),
                cursor_color=(0.2, 0.5, 0.9, 1),
                padding=[dp(10), dp(10), dp(10), dp(10)]
            )
            content.add_widget(self.login_password)
            
            content.add_widget(BoxLayout(size_hint_y=None, height=dp(5)))
            
            # دکمه‌ها
            btn_layout = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(8))
            
            login_btn = PersianButton(
                text='ورود',
                background_color=(0.2, 0.6, 0.2, 1),
                size_hint_y=None,
                height=dp(38),
                color=(1, 1, 1, 1),
                font_size=sp(15),
                bold=True
            )
            
            cancel_btn = PersianButton(
                text='انصراف',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(38),
                color=(1, 1, 1, 1),
                font_size=sp(15)
            )
            
            btn_layout.add_widget(login_btn)
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='ورود',
                content=content,
                size_hint=(0.85, None),
                height=dp(300),
                background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )
            
            def do_login(instance):
                selected_name = self.login_user_combo.text
                
                users = get_users()
                user = None
                for u in users:
                    name = u.get('name', '') or u.get('username', '')
                    if name == selected_name:
                        user = u
                        break
                
                if not user:
                    self.show_message('خطا', 'کاربر انتخاب شده یافت نشد')
                    return
                
                username = user.get('username', '')
                password = self.login_password.text.strip()
                
                if not username or not password:
                    self.show_message('خطا', 'لطفاً نام کاربری و رمز عبور را وارد کنید')
                    return
                
                logged_in_user = login(username, password)
                
                if logged_in_user:
                    popup.dismiss()
                    self.current_user = logged_in_user
                    self.show_report_tab()
                else:
                    self.show_message('خطا', 'نام کاربری یا رمز عبور اشتباه است')
                    self.login_password.text = ''
                    Clock.schedule_once(lambda dt: setattr(self.login_password, 'focus', True), 0.1)
            
            def on_cancel(instance):
                popup.dismiss()
                self.go_back(None)
            
            login_btn.bind(on_press=do_login)
            cancel_btn.bind(on_press=on_cancel)
            self.login_password.bind(on_text_validate=do_login)
            
            popup.open()
            
            Clock.schedule_once(lambda dt: setattr(self.login_user_combo.main_btn, 'focus', True), 0.2)
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)
    
    def show_report_tab(self):
        """نمایش فرم اصلی گزارش"""
        try:
            self.content_area.clear_widgets()
            
            main_scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint=(1, 1),
                scroll_type=['bars', 'content'],
                bar_width=dp(6)
            )
            
            main_content = BoxLayout(
                orientation='vertical',
                spacing=dp(6),
                size_hint_y=None,
                padding=dp(8),
                width=dp(350)
            )
            main_content.bind(minimum_height=main_content.setter('height'))
            
            # ========== عنوان ==========
            main_content.add_widget(RTLLabel(
                text=f'گزارش عملکرد روزانه',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(18),
                bold=True,
                color=(0.4, 0.8, 1, 1)
            ))
            
            main_content.add_widget(RTLLabel(
                text=self.current_user.get('name', ''),
                size_hint_y=None,
                height=dp(20),
                font_size=sp(13),
                color=(0.6, 0.6, 0.6, 1)
            ))
            
            main_content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
            
            # ========== فیلتر بازه زمانی ==========
            filter_box = BoxLayout(
                orientation='vertical',
                size_hint_y=None,
                height=dp(110),
                spacing=dp(4),
                padding=dp(6)
            )
            with filter_box.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=filter_box.pos, size=filter_box.size)
                filter_box.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                              size=lambda i, v: setattr(rect, 'size', v))
            
            filter_box.add_widget(RTLLabel(
                text='بازه زمانی:',
                size_hint_y=None,
                height=dp(20),
                font_size=sp(12),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))
            
            today = get_today_jalali()
            
            date_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5))
            
            date_row.add_widget(RTLLabel(
                text='از:',
                size_hint_x=0.12,
                size_hint_y=None,
                height=dp(32),
                font_size=sp(12),
                color=(1, 1, 1, 1)
            ))
            self.from_date = RTLTextInput(
                text=today,
                multiline=False,
                size_hint_x=0.35,
                size_hint_y=None,
                height=dp(32),
                font_size=sp(14)
            )
            self.from_date.bg_color = (0.15, 0.15, 0.15, 1)
            self.from_date.border_color = (0.3, 0.3, 0.3, 1)
            self.from_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.from_date._hidden_input.foreground_color = (1, 1, 1, 1)
            date_row.add_widget(self.from_date)
            
            date_row.add_widget(RTLLabel(
                text='تا:',
                size_hint_x=0.12,
                size_hint_y=None,
                height=dp(32),
                font_size=sp(12),
                color=(1, 1, 1, 1)
            ))
            self.to_date = RTLTextInput(
                text=today,
                multiline=False,
                size_hint_x=0.35,
                size_hint_y=None,
                height=dp(32),
                font_size=sp(14)
            )
            self.to_date.bg_color = (0.15, 0.15, 0.15, 1)
            self.to_date.border_color = (0.3, 0.3, 0.3, 1)
            self.to_date.border_color_focus = (0.2, 0.5, 0.9, 1)
            self.to_date._hidden_input.foreground_color = (1, 1, 1, 1)
            date_row.add_widget(self.to_date)
            
            filter_box.add_widget(date_row)
            
            btn_row = BoxLayout(size_hint_y=None, height=dp(32), spacing=dp(5))
            today_btn = PersianButton(
                text='امروز',
                background_color=(0.2, 0.5, 0.9, 1),
                size_hint_x=0.25,
                size_hint_y=None,
                height=dp(28),
                color=(1, 1, 1, 1),
                font_size=sp(12)
            )
            today_btn.bind(on_press=self._set_today_filter)
            btn_row.add_widget(today_btn)
            
            btn_row.add_widget(Label(size_hint_x=0.75))
            
            filter_box.add_widget(btn_row)
            
            main_content.add_widget(filter_box)
            
            main_content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
            
            # ========== گزینه‌های گزارش (چک‌باکس) ==========
            main_content.add_widget(RTLLabel(
                text='انتخاب گزارش:',
                size_hint_y=None,
                height=dp(22),
                font_size=sp(13),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))
            
            options_box = BoxLayout(
                orientation='vertical',
                size_hint_y=None,
                spacing=dp(2),
                padding=dp(5)
            )
            with options_box.canvas.before:
                Color(0.1, 0.1, 0.1, 1)
                rect = Rectangle(pos=options_box.pos, size=options_box.size)
                options_box.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                                size=lambda i, v: setattr(rect, 'size', v))
            
            # ✅ لیست کامل گزارش‌ها
            report_options = self._get_report_options_by_role()
            
            self.report_checkboxes = {}
            for key, label in report_options:
                row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(6))
                
                cb = CheckBox(
                    active=True,
                    size_hint_x=0.1,
                    size_hint_y=None,
                    height=dp(24),
                    color=(0.4, 0.7, 1, 1)
                )
                row.add_widget(cb)
                
                row.add_widget(RTLLabel(
                    text=label,
                    size_hint_x=0.9,
                    font_size=sp(12),
                    color=(1, 1, 1, 1)
                ))
                
                self.report_checkboxes[key] = cb
                options_box.add_widget(row)
            
            main_content.add_widget(options_box)
            
            main_content.add_widget(BoxLayout(size_hint_y=None, height=dp(5)))
            
            # ========== دکمه ایجاد گزارش ==========
            generate_btn = PersianButton(
                text='ایجاد گزارش',
                size_hint_y=None,
                height=dp(45),
                background_color=(0.2, 0.7, 0.2, 1),
                color=(1, 1, 1, 1),
                font_size=sp(16),
                bold=True
            )
            generate_btn.bind(on_press=self._generate_reports)
            main_content.add_widget(generate_btn)
            
            main_content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
            
            # ========== دکمه تاریخچه ==========
            history_btn = PersianButton(
                text='تاریخچه',
                size_hint_y=None,
                height=dp(40),
                background_color=(0.4, 0.3, 0.6, 1),
                color=(1, 1, 1, 1),
                font_size=sp(14)
            )
            history_btn.bind(on_press=self._show_history)
            main_content.add_widget(history_btn)
            
            main_content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
            
            # ========== دکمه بازگشت ==========
            back_btn = PersianButton(
                text='بازگشت',
                size_hint_y=None,
                height=dp(40),
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(14)
            )
            back_btn.bind(on_press=self.go_back)
            main_content.add_widget(back_btn)
            
            main_scroll.add_widget(main_content)
            self.content_area.add_widget(main_scroll)
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا: {e}", error_details)
    
    def _get_report_options_by_role(self):
        """دریافت گزینه‌های گزارش بر اساس نقش کاربر"""
        if not self.current_user:
            return []
        
        role = self.current_user.get('role', '')
        
        # ============================================================
        # ✅ گزارش‌های عمومی (همه نقش‌ها)
        # ============================================================
        options = [
            ('attendance', 'ورود و خروج (حضور و غیاب)'),
            ('leave', 'مرخصی'),
            ('mission', 'ماموریت'),
        ]
        
        # ============================================================
        # ✅ گزارش‌های تخصصی بر اساس نقش
        # ============================================================
        
        if role == 'بازاریاب':
            options.extend([
                ('daily_visits', 'گزارش فروش (ویزیت روزانه)'),
                ('collection', 'گزارش وصول'),
                ('detailed_targets', 'گزارش تحقق ریزتارگت'),
            ])
        
        elif role == 'سوپروایزر':
            options.extend([
                ('targets', 'گزارش تارگت‌ها'),
                ('detailed_targets', 'گزارش ریزتارگت‌ها'),
                ('supervisor_visits', 'گزارش سرکشی بازار'),
                ('evaluation', 'گزارش ارزیابی'),
                ('market_report', 'گزارش بازاری (ارسال به مدیر)'),
            ])
        
        elif role == 'موزع':
            options.extend([
                ('delivery', 'گزارش توزیع'),
                # ('collection', 'گزارش وصول'),  # در آینده اضافه می‌شود
            ])
        
        elif role == 'مدیر':
            # مدیر میتواند همه گزارش‌ها را ببیند
            options.extend([
                ('daily_visits', 'گزارش فروش (ویزیت روزانه)'),
                ('collection', 'گزارش وصول'),
                ('detailed_targets', 'گزارش تحقق ریزتارگت'),
                ('targets', 'گزارش تارگت‌ها'),
                ('supervisor_visits', 'گزارش سرکشی بازار'),
                ('evaluation', 'گزارش ارزیابی'),
                ('market_report', 'گزارش بازاری (ارسال به مدیر)'),
                ('delivery', 'گزارش توزیع'),
            ])
        
        elif role == 'سرپرست':
            options.extend([
                ('daily_visits', 'گزارش فروش (ویزیت روزانه)'),
                ('collection', 'گزارش وصول'),
                ('detailed_targets', 'گزارش تحقق ریزتارگت'),
                ('targets', 'گزارش تارگت‌ها'),
                ('supervisor_visits', 'گزارش سرکشی بازار'),
                ('evaluation', 'گزارش ارزیابی'),
                ('market_report', 'گزارش بازاری (ارسال به مدیر)'),
                ('delivery', 'گزارش توزیع'),
            ])
        
        return options
    
    def _set_today_filter(self, instance):
        """تنظیم فیلتر به امروز"""
        today = get_today_jalali()
        self.from_date.text = today
        self.to_date.text = today
    
    def _get_reports_dir(self):
        """دریافت مسیر پوشه گزارشات"""
        today = get_today_jalali().replace('/', '-')
        base_dir = os.path.join(get_backup_path(), 'daily_reports', today)
        return base_dir
    
    def _generate_reports(self, instance):
        """ایجاد گزارش‌های انتخاب شده"""
        try:
            selected = []
            for key, cb in self.report_checkboxes.items():
                if cb.active:
                    selected.append(key)
            
            if not selected:
                self.show_message('خطا', 'هیچ گزارشی انتخاب نشده است')
                return
            
            from_date = self.from_date.text.strip()
            to_date = self.to_date.text.strip()
            
            if not from_date or not to_date:
                self.show_message('خطا', 'لطفاً بازه زمانی را مشخص کنید')
                return
            
            reports_dir = self._get_reports_dir()
            os.makedirs(reports_dir, exist_ok=True)
            
            user_name = self.current_user.get('name', '') or self.current_user.get('username', '')
            user_name_clean = user_name.replace(' ', '_')
            
            generated = []
            failed = []
            
            for report_type in selected:
                try:
                    result = self._generate_single_report(
                        report_type, from_date, to_date, reports_dir, user_name_clean
                    )
                    if result:
                        generated.append(result)
                    else:
                        failed.append(report_type)
                except Exception as e:
                    failed.append(report_type)
                    print(f"خطا در تولید {report_type}: {e}")
            
            if generated:
                msg = f'{len(generated)} گزارش با موفقیت ایجاد شد:\n'
                for g in generated:
                    msg += f'• {g}\n'
                if failed:
                    msg += f'\nخطا در {len(failed)} مورد'
                self.show_message('موفق', msg)
            else:
                self.show_message('خطا', 'هیچ گزارشی ایجاد نشد')
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در ایجاد گزارش: {e}", error_details)
    
    def _generate_single_report(self, report_type, from_date, to_date, reports_dir, user_name):
        """تولید یک گزارش با استفاده از توابع اصلی"""
        try:
            user_id = self.current_user.get('id')
            agent_name = self.current_user.get('name', '') or self.current_user.get('username', '')
            role = self.current_user.get('role', '')
            
            # ============================================================
            # گزارش‌های عمومی
            # ============================================================
            if report_type == 'attendance':
                return self._export_attendance_report(user_id, from_date, to_date, reports_dir, user_name)
            elif report_type == 'leave':
                return self._export_leave_report(user_id, from_date, to_date, reports_dir, user_name)
            elif report_type == 'mission':
                return self._export_mission_report(agent_name, from_date, to_date, reports_dir, user_name)
            
            # ============================================================
            # گزارش‌های تخصصی بازاریاب
            # ============================================================
            elif report_type == 'daily_visits':
                return self._export_daily_visits_report(agent_name, from_date, to_date, reports_dir, user_name)
            elif report_type == 'collection':
                return self._export_collection_report(agent_name, from_date, to_date, reports_dir, user_name)
            elif report_type == 'detailed_targets':
                # ✅ برای سوپروایزر از تابع جداگانه استفاده کن
                if role == 'سوپروایزر':
                    return self._export_detailed_targets_report_supervisor(agent_name, from_date, to_date, reports_dir, user_name)
                else:
                    return self._export_detailed_targets_report(agent_name, from_date, to_date, reports_dir, user_name)
            
            # ============================================================
            # گزارش‌های تخصصی سوپروایزر
            # ============================================================
            elif report_type == 'targets':
                # ✅ برای سوپروایزر از تابع جداگانه استفاده کن
                if role == 'سوپروایزر':
                    return self._export_targets_report_supervisor(agent_name, from_date, to_date, reports_dir, user_name)
                else:
                    return self._export_targets_report(agent_name, from_date, to_date, reports_dir, user_name)
            
            elif report_type == 'supervisor_visits':
                return self._export_supervisor_visits_report(agent_name, from_date, to_date, reports_dir, user_name)
            elif report_type == 'evaluation':
                return self._export_evaluation_report(agent_name, from_date, to_date, reports_dir, user_name)
            elif report_type == 'market_report':
                return self._export_market_report(agent_name, from_date, to_date, reports_dir, user_name)
            
            # ============================================================
            # گزارش‌های تخصصی موزع
            # ============================================================
            elif report_type == 'delivery':
                return self._export_delivery_report(agent_name, from_date, to_date, reports_dir, user_name)
            
            return None
            
        except Exception as e:
            print(f"خطا در تولید {report_type}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    # ============================================================
    # ۱. گزارش حضور و غیاب (ورود و خروج) - مطابق با خروجی برنامه
    # ============================================================
    
    def _export_attendance_report(self, user_id, from_date, to_date, reports_dir, user_name):
        """خروجی حضور و غیاب - مطابق با خروجی برنامه"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            records = AttendanceManager.load_attendance()
            
            filtered = []
            for r in records:
                if r.get('user_id') != user_id:
                    continue
                r_date = r.get('date', '')
                if from_date <= r_date <= to_date:
                    filtered.append(r)
            
            if not filtered:
                return None
            
            filename = f'گزارش_ورود_خروج_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش ورود و خروج"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ✅ هدرهای مطابق با خروجی برنامه
            headers = [
                'ردیف', 'تاریخ', 'وضعیت', 'اولین ورود', 'آخرین خروج',
                'تعداد ورود', 'تعداد خروج', 'کارکرد (دقیقه)', 'کارکرد (ساعت)',
                'اضافه کار (دقیقه)', 'اضافه کار (ساعت)', 'کسر کار (دقیقه)', 'کسر کار (ساعت)'
            ]
            
            col_widths = [6, 12, 10, 12, 12, 12, 12, 16, 14, 16, 14, 16, 14]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            total_work_minutes = 0
            total_overtime_minutes = 0
            total_shortage_minutes = 0
            
            for idx, record in enumerate(filtered, 1):
                date = record.get('date', '')
                check_in = record.get('check_in', '')
                check_out = record.get('check_out', '')
                
                # محاسبه کارکرد
                work_minutes = 0
                if check_in and check_out:
                    try:
                        in_h, in_m = map(int, check_in.split(':'))
                        out_h, out_m = map(int, check_out.split(':'))
                        work_minutes = (out_h - in_h) * 60 + (out_m - in_m)
                    except:
                        pass
                
                # کارکرد (ساعت)
                work_hours = f"{work_minutes // 60:02d}:{work_minutes % 60:02d}" if work_minutes > 0 else "00:00"
                
                # اضافه کار (فرضی)
                overtime_minutes = 0
                overtime_hours = "00:00"
                
                # کسر کار (فرضی)
                shortage_minutes = max(0, 480 - work_minutes) if work_minutes > 0 else 0
                shortage_hours = f"{shortage_minutes // 60:02d}:{shortage_minutes % 60:02d}" if shortage_minutes > 0 else "00:00"
                
                total_work_minutes += work_minutes
                total_overtime_minutes += overtime_minutes
                total_shortage_minutes += shortage_minutes
                
                values = [
                    idx,
                    date,
                    'حضور' if check_in else 'غیبت',
                    check_in,
                    check_out,
                    1 if check_in else 0,
                    1 if check_out else 0,
                    work_minutes,
                    work_hours,
                    overtime_minutes,
                    overtime_hours,
                    shortage_minutes,
                    shortage_hours
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== خلاصه در انتها ==========
            total_work_hours = f"{total_work_minutes // 60:02d}:{total_work_minutes % 60:02d}"
            total_overtime_hours = f"{total_overtime_minutes // 60:02d}:{total_overtime_minutes % 60:02d}"
            total_shortage_hours = f"{total_shortage_minutes // 60:02d}:{total_shortage_minutes % 60:02d}"
            
            summary_row = len(filtered) + 3
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
            ws.cell(row=summary_row, column=1, value='خلاصه:').font = Font(bold=True, size=12, color="FFD700")
            
            ws.cell(row=summary_row, column=8, value=f'کل کارکرد: {total_work_hours}')
            ws.cell(row=summary_row, column=8).font = Font(bold=True, size=11, color="00FF00")
            
            ws.cell(row=summary_row + 1, column=8, value=f'اضافه کار: {total_overtime_hours}')
            ws.cell(row=summary_row + 1, column=8).font = Font(bold=True, size=11, color="00BFFF")
            
            ws.cell(row=summary_row + 2, column=8, value=f'کسر کار: {total_shortage_hours}')
            ws.cell(row=summary_row + 2, column=8).font = Font(bold=True, size=11, color="FF6B6B")
            
            ws.cell(row=summary_row + 3, column=8, value=f'روزهای کاری: {len(filtered)}')
            ws.cell(row=summary_row + 3, column=8).font = Font(bold=True, size=11, color="FFFFFF")
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی حضور و غیاب: {e}")
            return None
    
    # ============================================================
    # ۲. گزارش مرخصی - مطابق با خروجی برنامه
    # ============================================================
    
    def _export_leave_report(self, user_id, from_date, to_date, reports_dir, user_name):
        """خروجی مرخصی - مطابق با خروجی برنامه"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            if not os.path.exists(file_path):
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_requests = json.load(f)
            
            filtered = []
            for r in all_requests:
                if r.get('user_id') != user_id:
                    continue
                r_date = r.get('created_at', '')
                if from_date <= r_date <= to_date:
                    filtered.append(r)
            
            if not filtered:
                return None
            
            # دریافت نام کاربر
            user_name_full = self.current_user.get('name', '') or self.current_user.get('username', '')
            
            filename = f'گزارش_مرخصی_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش مرخصی"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ✅ هدرهای مطابق با خروجی برنامه
            headers = ['ردیف', 'عامل', 'نوع مرخصی', 'مدت', 'تاریخ شروع', 'تاریخ پایان', 'وضعیت', 'تاریخ ثبت']
            col_widths = [6, 20, 16, 12, 14, 14, 12, 16]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, req in enumerate(filtered, 1):
                values = [
                    idx,
                    user_name_full,
                    req.get('leave_type', ''),
                    req.get('duration_display', ''),
                    req.get('start_date', ''),
                    req.get('end_date', ''),
                    req.get('status', ''),
                    req.get('created_at', '')
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی مرخصی: {e}")
            return None
    
    # ============================================================
    # ۳. گزارش ماموریت - مطابق با خروجی برنامه
    # ============================================================
    
    def _export_mission_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """خروجی ماموریت - با ساختار جدید مطابق با خروجی برنامه"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            file_path = os.path.join(get_data_path(), 'do_missions.json')
            if not os.path.exists(file_path):
                print(f"⚠️ فایل do_missions.json یافت نشد: {file_path}")
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_missions = json.load(f)
            
            print(f"📋 نوع داده ماموریت‌ها: {type(all_missions)}")
            
            # اگر دیکشنری بود تبدیل به لیست
            if isinstance(all_missions, dict):
                missions_list = []
                for date, missions in all_missions.items():
                    if isinstance(missions, list):
                        for m in missions:
                            if isinstance(m, dict):
                                m['date'] = date
                                missions_list.append(m)
                all_missions = missions_list
                print(f"📋 تبدیل دیکشنری به لیست: {len(all_missions)} ماموریت")
            
            if not isinstance(all_missions, list):
                print(f"⚠️ داده ماموریت‌ها لیست نیست: {type(all_missions)}")
                all_missions = []
            
            print(f"📋 تعداد کل ماموریت‌ها: {len(all_missions)}")
            
            filtered = []
            for m in all_missions:
                if not isinstance(m, dict):
                    continue
                m_agent = m.get('agent_name', '')
                if agent_name not in m_agent and m_agent not in agent_name:
                    continue
                m_date = m.get('start_date', '') or m.get('date', '')
                if from_date <= m_date <= to_date:
                    filtered.append(m)
            
            print(f"📋 ماموریت‌های فیلتر شده برای {agent_name}: {len(filtered)}")
            
            if not filtered:
                print(f"⚠️ هیچ ماموریتی برای {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            filename = f'گزارش_ماموریت_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش ماموریت"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ✅ هدرهای جدید مطابق با خروجی برنامه
            headers = [
                'ردیف', 'شناسه', 'نوع', 'روش', 'تاریخ شروع', 'تاریخ پایان',
                'مدت', 'امتیاز ماموریت', 'امتیاز کسب شده', 'هدف', 'وضعیت', 'توضیحات'
            ]
            col_widths = [6, 12, 14, 12, 14, 14, 10, 16, 16, 16, 14, 30]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            total_score = 0
            total_achieved = 0
            total_target = 0
            
            for idx, mission in enumerate(filtered, 1):
                score = mission.get('score', 0)
                achieved = mission.get('achieved_score', 0)
                target = mission.get('target', 0)
                
                total_score += score
                total_achieved += achieved
                total_target += target
                
                status = mission.get('status', '')
                status_display = status
                if status == 'موفق':
                    status_display = '✅ موفق'
                elif status == 'ناموفق':
                    status_display = '❌ ناموفق'
                elif status == 'در انتظار':
                    status_display = '⏳ در انتظار'
                
                values = [
                    idx,
                    mission.get('id', ''),
                    mission.get('type', ''),
                    mission.get('method', ''),
                    mission.get('start_date', ''),
                    mission.get('end_date', ''),
                    mission.get('duration', 0),
                    score,
                    achieved,
                    target,
                    status_display,
                    mission.get('description', '')
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== خلاصه ==========
            summary_row = len(filtered) + 3
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
            ws.cell(row=summary_row, column=1, value='خلاصه:').font = Font(bold=True, size=12, color="FFD700")
            
            ws.cell(row=summary_row, column=8, value=f'امتیاز ماموریت: {total_score}')
            ws.cell(row=summary_row, column=8).font = Font(bold=True, size=11, color="FFFFFF")
            
            ws.cell(row=summary_row + 1, column=8, value=f'امتیاز کسب شده: {total_achieved}')
            ws.cell(row=summary_row + 1, column=8).font = Font(bold=True, size=11, color="00FF00")
            
            ws.cell(row=summary_row + 2, column=8, value=f'هدف کل: {total_target:,}')
            ws.cell(row=summary_row + 2, column=8).font = Font(bold=True, size=11, color="FFD700")
            
            wb.save(filepath)
            print(f"✅ گزارش ماموریت ساخته شد: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی ماموریت: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    # ============================================================
    # ۴. گزارش ویزیت روزانه - ✅ قبلاً کامل است
    # ============================================================
    
    def _export_daily_visits_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """خروجی ویزیت روزانه - با ۳ شیت"""
        try:
            all_logs = get_daily_logs()
            
            # فیلتر بر اساس تاریخ و عامل
            filtered = {}
            for date, logs in all_logs.items():
                if from_date <= date <= to_date:
                    if isinstance(logs, list):
                        for log in logs:
                            if isinstance(log, dict):
                                log_agent = log.get('agent_name', '')
                                if agent_name in log_agent or log_agent in agent_name:
                                    if date not in filtered:
                                        filtered[date] = []
                                    filtered[date].append(log)
            
            if not filtered:
                return None
            
            # ✅ استفاده از تابع اصلی export_to_excel
            success, result = export_to_excel(filtered)
            
            if success:
                import shutil
                filename = f'گزارش_ویزیت_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
                filepath = os.path.join(reports_dir, filename)
                if os.path.exists(result):
                    shutil.copy2(result, filepath)
                    return filename
            
            return None
            
        except Exception as e:
            print(f"خطا در خروجی ویزیت روزانه: {e}")
            return None
    
    # ============================================================
    # ۵. گزارش وصول - مطابق با خروجی برنامه
    # ============================================================
    
    def _export_collection_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """خروجی وصول - مطابق با خروجی برنامه"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            file_path = os.path.join(get_data_path(), 'collections.json')
            if not os.path.exists(file_path):
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_collections = json.load(f)
            
            if not isinstance(all_collections, list):
                all_collections = []
            
            # فیلتر بر اساس نام عامل و تاریخ
            filtered = []
            for c in all_collections:
                if not isinstance(c, dict):
                    continue
                c_agent = c.get('agent_name', '')
                if agent_name not in c_agent and c_agent not in agent_name:
                    continue
                c_date = c.get('date', '')
                if from_date <= c_date <= to_date:
                    filtered.append(c)
            
            if not filtered:
                return None
            
            filename = f'گزارش_وصول_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش وصول"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ✅ هدرهای مطابق با خروجی برنامه
            headers = [
                'ردیف', 'شناسه', 'تاریخ', 'عامل', 'مشتری', 'مسیر', 'وضعیت',
                'نوع پرداخت', 'نوع وصول نقدی', 'مبلغ نقد (ریال)', 'کسورات (ریال)',
                'خالص نقد (ریال)', 'بانک', 'شماره پیگیری', 'تعداد چک',
                'جمع چک ها (ریال)', 'جمع کل (ریال)', 'علت عدم وصول',
                'تاریخ پیگیری بعدی', 'توضیحات'
            ]
            
            col_widths = [6, 14, 12, 18, 22, 16, 12, 14, 18, 18, 14, 18, 14, 16, 12, 18, 18, 20, 16, 20]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            total_cash = 0
            total_check = 0
            total_collection = 0
            success_count = 0
            fail_count = 0
            
            for idx, col in enumerate(filtered, 1):
                status = col.get('status', '')
                
                if status == 'موفق':
                    success_count += 1
                    cash_amount = col.get('cash_amount', 0)
                    deductions = col.get('cash_deductions', 0)
                    net_cash = col.get('net_cash', 0)
                    checks = col.get('checks', [])
                    check_count = len(checks)
                    check_amount = col.get('total_check_amount', 0)
                    total = col.get('total_collection', 0)
                    payment_type = 'نقد' if col.get('has_cash') else 'چک' if col.get('has_check') else ''
                    cash_type = col.get('cash_type', '')
                    bank = col.get('bank', '')
                    tracking = col.get('tracking_number', '')
                    
                    total_cash += net_cash
                    total_check += check_amount
                    total_collection += total
                    
                    fail_reason = ''
                    next_follow = ''
                else:
                    fail_count += 1
                    cash_amount = 0
                    deductions = 0
                    net_cash = 0
                    checks = []
                    check_count = 0
                    check_amount = 0
                    total = 0
                    payment_type = '-'
                    cash_type = '-'
                    bank = '-'
                    tracking = '-'
                    fail_reason = col.get('fail_reason', '')
                    next_follow = col.get('next_follow_up_date', '')
                
                values = [
                    idx,
                    col.get('id', ''),
                    col.get('date', ''),
                    col.get('agent_name', ''),
                    col.get('customer', ''),
                    col.get('route', ''),
                    status,
                    payment_type,
                    cash_type,
                    cash_amount,
                    deductions,
                    net_cash,
                    bank,
                    tracking,
                    check_count,
                    check_amount,
                    total,
                    fail_reason,
                    next_follow,
                    col.get('description', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
                    if col_idx == 7:
                        if status == 'موفق':
                            cell.fill = PatternFill(start_color="235347", end_color="235347", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
                        else:
                            cell.fill = PatternFill(start_color="78281F", end_color="78281F", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی وصول: {e}")
            return None
    
    # ============================================================
    # ۶. گزارش ریزتارگت - مطابق با خروجی برنامه
    # ============================================================
    
    def _export_detailed_targets_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """
        خروجی ریزتارگت‌ها - برای بازاریاب (نمایش تمام ریزتارگت‌های فعال در بازه زمانی)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            file_path = os.path.join(get_data_path(), 'detailed_targets.json')
            if not os.path.exists(file_path):
                print(f"⚠️ فایل detailed_targets.json یافت نشد")
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_targets = json.load(f)
            
            if not isinstance(all_targets, list):
                all_targets = []
            
            print(f"📋 تعداد کل ریزتارگت‌ها: {len(all_targets)}")
            
            # ✅ فیلتر بر اساس نام عامل و تاریخ شروع (برای بازاریاب)
            filtered = []
            for t in all_targets:
                if not isinstance(t, dict):
                    continue
                
                t_agent = t.get('agent_name', '')
                if agent_name not in t_agent and t_agent not in agent_name:
                    continue
                
                # ✅ بررسی تاریخ شروع (start_date) - چون تاریخ شروع تارگت‌ها 1405/05/01 است
                start_date = t.get('start_date', '')
                if from_date <= start_date <= to_date:
                    filtered.append(t)
                else:
                    # اگر تاریخ شروع در بازه نیست، بررسی کن که آیا تاریخ تحقق در بازه است
                    fulfillment_date = t.get('last_fulfillment_date', '')
                    if fulfillment_date and from_date <= fulfillment_date <= to_date:
                        filtered.append(t)
            
            print(f"📋 ریزتارگت‌های فیلتر شده برای {agent_name}: {len(filtered)}")
            
            if not filtered:
                print(f"⚠️ هیچ ریزتارگتی برای {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            # ========== مرتب‌سازی بر اساس تاریخ ==========
            filtered.sort(key=lambda x: x.get('start_date', ''), reverse=False)
            
            filename = f'ریزتارگت_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش تحقق ریزتارگت"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ✅ هدرهای کامل
            headers = [
                'ردیف', 'شناسه', 'گروه کالا', 'هدف کل', 'واحد',
                'تارگت روزانه', 'تحقق کل', 'کسر تارگت', 'درصد پیشرفت',
                'وضعیت', 'تاریخ شروع', 'تاریخ پایان', 'تاریخ آخرین تحقق'
            ]
            
            col_widths = [6, 16, 22, 14, 10, 16, 14, 14, 14, 14, 14, 14, 18]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            total_target = 0
            total_achieved = 0
            total_remaining = 0
            
            for idx, target in enumerate(filtered, 1):
                target_count = target.get('target_count', 0)
                achieved = target.get('achieved_value', 0)
                daily_target = target.get('daily_target', 0)
                remaining = max(0, target_count - achieved)
                percent = (achieved / target_count * 100) if target_count > 0 else 0
                
                total_target += target_count
                total_achieved += achieved
                total_remaining += remaining
                
                # رنگ درصد
                if percent >= 100:
                    pct_color = "00CC44"
                elif percent >= 50:
                    pct_color = "FFAA00"
                else:
                    pct_color = "CC3333"
                
                status = target.get('status', '')
                if status == 'تکمیل شده':
                    status_color = "00CC44"
                elif status == 'فعال':
                    status_color = "3399FF"
                elif status == 'در انتظار':
                    status_color = "FFCC00"
                else:
                    status_color = "888888"
                
                # تاریخ آخرین تحقق
                last_fulfillment = target.get('last_fulfillment_date', '')
                
                values = [
                    idx,
                    target.get('id', ''),
                    target.get('product_group', ''),
                    target_count,
                    target.get('unit', ''),
                    daily_target,
                    achieved,
                    remaining,
                    f"{percent:.1f}%",
                    status,
                    target.get('start_date', ''),
                    target.get('end_date', ''),
                    last_fulfillment
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
                    if col_idx == 9:  # درصد پیشرفت
                        cell.font = Font(color=pct_color, bold=True)
                    elif col_idx == 10:  # وضعیت
                        cell.font = Font(color=status_color, bold=True)
                    elif col_idx == 8:  # کسر تارگت
                        if remaining > 0:
                            cell.font = Font(color="CC3333", bold=True)
                        else:
                            cell.font = Font(color="00CC44", bold=True)
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== خلاصه در انتها ==========
            avg_percent = (total_achieved / total_target * 100) if total_target > 0 else 0
            completed_count = len([t for t in filtered if t.get('status') == 'تکمیل شده'])
            active_count = len([t for t in filtered if t.get('status') == 'فعال'])
            
            summary_row = len(filtered) + 3
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
            ws.cell(row=summary_row, column=1, value='📊 خلاصه عملکرد:').font = Font(bold=True, size=12, color="FFD700")
            
            ws.cell(row=summary_row, column=8, value=f'مجموع هدف: {total_target:,}')
            ws.cell(row=summary_row, column=8).font = Font(bold=True, size=11, color="FFFFFF")
            
            ws.cell(row=summary_row + 1, column=8, value=f'مجموع تحقق: {total_achieved:,}')
            ws.cell(row=summary_row + 1, column=8).font = Font(bold=True, size=11, color="00CC44")
            
            ws.cell(row=summary_row + 2, column=8, value=f'مجموع کسر: {total_remaining:,}')
            ws.cell(row=summary_row + 2, column=8).font = Font(bold=True, size=11, color="CC3333")
            
            ws.cell(row=summary_row + 3, column=8, value=f'میانگین پیشرفت: {avg_percent:.1f}%')
            ws.cell(row=summary_row + 3, column=8).font = Font(bold=True, size=11, color="FFD700")
            
            ws.cell(row=summary_row + 4, column=8, value=f'تکمیل شده: {completed_count} | فعال: {active_count}')
            ws.cell(row=summary_row + 4, column=8).font = Font(bold=True, size=11, color="FFFFFF")
            
            wb.save(filepath)
            print(f"✅ گزارش ریزتارگت ساخته شد: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی ریزتارگت‌ها: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    # ============================================================
    # ۷. گزارش توزیع - مطابق با خروجی برنامه (۳ شیت)
    # ============================================================
    
    def _export_delivery_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """خروجی توزیع - با ۳ شیت (مطابق با خروجی برنامه)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            all_deliveries_dict = get_all_deliveries()
            
            if not all_deliveries_dict:
                return None
            
            filtered = []
            for date, deliveries in all_deliveries_dict.items():
                if from_date <= date <= to_date:
                    if isinstance(deliveries, list):
                        for d in deliveries:
                            if not isinstance(d, dict):
                                continue
                            d_agent = d.get('distributor_name', '') or d.get('agent_name', '') or d.get('user_name', '')
                            if agent_name in d_agent or d_agent in agent_name:
                                d['date'] = date
                                filtered.append(d)
            
            if not filtered:
                return None
            
            filtered.sort(key=lambda x: x.get('date', ''), reverse=False)
            
            filename = f'گزارش_توزیع_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            
            # ========== شیت ۱: گزارش توزیع‌ها ==========
            ws1 = wb.active
            ws1.title = "گزارش توزیع‌ها"
            ws1.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = [
                'ردیف', 'تاریخ', 'مسیر', 'مشتری', 'شماره فاکتور',
                'وضعیت توزیع', 'تحویل کامل', 'مبلغ فاکتور',
                'تعداد برگشتی', 'مبلغ برگشتی', 'علت برگشتی',
                'مبلغ نقدی', 'مبلغ چکی', 'مبلغ نسیه',
                'جمع دریافتی', 'مانده',
                'درصد تخفیف', 'مبلغ تخفیف', 'سایر کسورات',
                'نوع تسویه', 'توضیحات', 'ساعت', 'شناسه'
            ]
            col_widths = [6, 12, 16, 22, 16, 14, 14, 18, 14, 18, 20, 16, 16, 16, 16, 16, 14, 16, 16, 14, 20, 12, 18]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            total_invoice = 0
            total_cash = 0
            total_check = 0
            total_credit = 0
            total_received = 0
            total_remaining = 0
            success_count = 0
            fail_count = 0
            
            for idx, delivery in enumerate(filtered, 1):
                status = delivery.get('delivery_status', '')
                
                if status == 'موفق':
                    success_count += 1
                    full_delivery = 'بله' if delivery.get('full_delivery', True) else 'خیر'
                    fail_reason = ''
                    settlement_type = delivery.get('settlement_type', '')
                else:
                    fail_count += 1
                    full_delivery = '-'
                    fail_reason = delivery.get('fail_reason', '')
                    settlement_type = '-'
                
                remaining = delivery.get('remaining_amount', 0)
                
                values = [
                    idx,
                    delivery.get('date', ''),
                    delivery.get('route', ''),
                    delivery.get('customer_name', ''),
                    delivery.get('invoice_number', ''),
                    status,
                    full_delivery,
                    delivery.get('invoice_amount', 0),
                    delivery.get('returned_quantity', 0),
                    delivery.get('returned_amount', 0),
                    delivery.get('return_reason', ''),
                    delivery.get('cash_amount', 0),
                    delivery.get('check_amount', 0),
                    remaining,
                    delivery.get('total_received', 0),
                    remaining,
                    delivery.get('discount_percent', 0),
                    delivery.get('discount_amount', 0),
                    delivery.get('other_deductions_total', 0),
                    settlement_type,
                    delivery.get('description', '') or fail_reason,
                    delivery.get('timestamp', '').split(' ')[1] if ' ' in delivery.get('timestamp', '') else '',
                    delivery.get('id', '')
                ]
                
                total_invoice += delivery.get('invoice_amount', 0)
                total_cash += delivery.get('cash_amount', 0)
                total_check += delivery.get('check_amount', 0)
                total_credit += remaining
                total_received += delivery.get('total_received', 0)
                total_remaining += remaining
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws1.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
                    if col_idx == 6:
                        if status == 'موفق':
                            cell.fill = PatternFill(start_color="235347", end_color="235347", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
                        else:
                            cell.fill = PatternFill(start_color="78281F", end_color="78281F", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
                    
                    if col_idx == 16:
                        if remaining > 0:
                            cell.font = Font(color="FF6B6B", bold=True)
                        else:
                            cell.font = Font(color="00CC44", bold=True)
            
            for i, width in enumerate(col_widths, 1):
                ws1.column_dimensions[get_column_letter(i)].width = width
            
            # ========== شیت ۲: خلاصه آمار ==========
            ws2 = wb.create_sheet("خلاصه آمار")
            ws2.right_to_left = True
            
            summary_headers = ['شاخص', 'مقدار']
            for col_idx, header in enumerate(summary_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            full_deliveries = len([d for d in filtered if d.get('full_delivery', False) and d.get('delivery_status') == 'موفق'])
            partial_deliveries = len([d for d in filtered if not d.get('full_delivery', True) and d.get('delivery_status') == 'موفق'])
            total_return_qty = sum(d.get('returned_quantity', 0) for d in filtered)
            total_return_amount = sum(d.get('returned_amount', 0) for d in filtered)
            avg_amount = total_invoice / len(filtered) if len(filtered) > 0 else 0
            
            stats = [
                ('کل مبلغ توزیع (ریال)', f'{total_invoice:,.0f}'),
                ('مبلغ نقدی (ریال)', f'{total_cash:,.0f}'),
                ('مبلغ چکی (ریال)', f'{total_check:,.0f}'),
                ('مبلغ نسیه (ریال)', f'{total_credit:,.0f}'),
                ('تعداد کل توزیع‌ها', str(len(filtered))),
                ('تعداد تحویل کامل', str(full_deliveries)),
                ('تعداد تحویل ناقص', str(partial_deliveries)),
                ('تعداد برگشتی', str(total_return_qty)),
                ('مبلغ برگشتی (ریال)', f'{total_return_amount:,.0f}'),
                ('تعداد روزهای کاری', '1'),
                ('میانگین مبلغ هر توزیع', f'{avg_amount:,.0f}')
            ]
            
            for row_idx, (label, value) in enumerate(stats, 2):
                ws2.cell(row=row_idx, column=1, value=label).alignment = Alignment(horizontal="right", vertical="center")
                ws2.cell(row=row_idx, column=1).border = thin_border
                ws2.cell(row=row_idx, column=2, value=value).alignment = Alignment(horizontal="center", vertical="center")
                ws2.cell(row=row_idx, column=2).border = thin_border
            
            ws2.column_dimensions['A'].width = 30
            ws2.column_dimensions['B'].width = 25
            
            # ========== شیت ۳: آمار روزانه ==========
            ws3 = wb.create_sheet("آمار روزانه")
            ws3.right_to_left = True
            
            daily_headers = [
                'تاریخ', 'تعداد توزیع', 'تحویل کامل', 'تحویل ناقص',
                'مبلغ کل', 'مبلغ نقدی', 'مبلغ چکی', 'مبلغ نسیه',
                'تعداد برگشتی', 'مبلغ برگشتی'
            ]
            
            for col_idx, header in enumerate(daily_headers, 1):
                cell = ws3.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            daily_stats = {}
            for d in filtered:
                date = d.get('date', '')
                if date not in daily_stats:
                    daily_stats[date] = {
                        'count': 0, 'full': 0, 'partial': 0,
                        'invoice': 0, 'cash': 0, 'check': 0, 'credit': 0,
                        'return_qty': 0, 'return_amount': 0
                    }
                
                stats = daily_stats[date]
                stats['count'] += 1
                if d.get('delivery_status') == 'موفق':
                    if d.get('full_delivery', False):
                        stats['full'] += 1
                    else:
                        stats['partial'] += 1
                stats['invoice'] += d.get('invoice_amount', 0)
                stats['cash'] += d.get('cash_amount', 0)
                stats['check'] += d.get('check_amount', 0)
                stats['credit'] += d.get('remaining_amount', 0)
                stats['return_qty'] += d.get('returned_quantity', 0)
                stats['return_amount'] += d.get('returned_amount', 0)
            
            for row_idx, (date, stats) in enumerate(daily_stats.items(), 2):
                values = [
                    date,
                    stats['count'],
                    stats['full'],
                    stats['partial'],
                    stats['invoice'],
                    stats['cash'],
                    stats['check'],
                    stats['credit'],
                    stats['return_qty'],
                    stats['return_amount']
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            daily_widths = [12, 14, 14, 14, 16, 16, 16, 16, 14, 16]
            for i, width in enumerate(daily_widths, 1):
                ws3.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(filepath)
            return filename
            
        except Exception as e:
            print(f"خطا در خروجی توزیع: {e}")
            return None
    
    # ============================================================
    # ۸. گزارش تارگت‌ها - جدید
    # ============================================================
    
    def _export_targets_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """
        خروجی تارگت‌ها - برای سوپروایزر بر اساس created_by
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from utils.target_manager import get_all_targets, export_targets_to_excel
            
            all_targets = get_all_targets()
            
            if not all_targets:
                print(f"⚠️ هیچ تارگتی یافت نشد")
                return None
            
            # ✅ فیلتر بر اساس created_by (برای سوپروایزر) یا agent_name (برای بازاریاب)
            filtered = []
            for t in all_targets:
                if not isinstance(t, dict):
                    continue
                
                # بررسی تاریخ
                t_date = t.get('start_date', '')
                if from_date <= t_date <= to_date:
                    # بررسی نام: اول created_by، سپس agent_name
                    t_creator = t.get('created_by', '')
                    t_agent = t.get('agent_name', '')
                    
                    # اگر created_by با نام کاربر مطابقت دارد یا agent_name مطابقت دارد
                    if agent_name in t_creator or t_creator in agent_name:
                        filtered.append(t)
                    elif agent_name in t_agent or t_agent in agent_name:
                        filtered.append(t)
            
            print(f"📋 تارگت‌های فیلتر شده برای {agent_name}: {len(filtered)}")
            
            if not filtered:
                print(f"⚠️ هیچ تارگتی برای {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            # استفاده از تابع اصلی export_targets_to_excel
            success, message, filepath = export_targets_to_excel(filtered)
            
            if success:
                import shutil
                filename = f'تارگت_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
                dest_path = os.path.join(reports_dir, filename)
                if os.path.exists(filepath):
                    shutil.copy2(filepath, dest_path)
                    print(f"✅ گزارش تارگت ساخته شد: {filename}")
                    return filename
            
            return None
            
        except Exception as e:
            print(f"❌ خطا در خروجی تارگت‌ها: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _export_targets_report_supervisor(self, agent_name, from_date, to_date, reports_dir, user_name):
        """
        خروجی تارگت‌ها - مخصوص سوپروایزر (بر اساس created_by و تاریخ ایجاد)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from utils.target_manager import get_all_targets
            from utils.jalali_date import convert_to_jalali
            
            all_targets = get_all_targets()
            
            if not all_targets:
                print(f"⚠️ هیچ تارگتی یافت نشد")
                return None
            
            # ✅ فیلتر بر اساس created_by و تاریخ ایجاد
            filtered = []
            for t in all_targets:
                if not isinstance(t, dict):
                    continue
                
                t_creator = t.get('created_by', '')
                
                # اگر created_by با نام کاربر مطابقت دارد
                if agent_name in t_creator or t_creator in agent_name:
                    # ✅ بررسی تاریخ ایجاد (created_at)
                    created_at = t.get('created_at', '')
                    if created_at:
                        # استخراج تاریخ از created_at (فرمت: 2026-08-08T13:49:43.129990)
                        if 'T' in created_at:
                            created_date = created_at.split('T')[0]
                            # تبدیل به شمسی
                            try:
                                created_date_jalali = convert_to_jalali(created_date)
                                if from_date <= created_date_jalali <= to_date:
                                    filtered.append(t)
                            except:
                                # اگر تبدیل نشد، بر اساس تاریخ شروع فیلتر کن
                                t_date = t.get('start_date', '')
                                if from_date <= t_date <= to_date:
                                    filtered.append(t)
                        else:
                            # اگر فرمت T نداشت، خود تاریخ را بررسی کن
                            if from_date <= created_at <= to_date:
                                filtered.append(t)
                    else:
                        # اگر created_at نبود، بر اساس تاریخ شروع فیلتر کن
                        t_date = t.get('start_date', '')
                        if from_date <= t_date <= to_date:
                            filtered.append(t)
            
            print(f"📋 تارگت‌های فیلتر شده برای سوپروایزر {agent_name}: {len(filtered)}")
            
            if not filtered:
                print(f"⚠️ هیچ تارگتی برای سوپروایزر {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            filename = f'تارگت_سوپروایزر_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "تارگت‌ها"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = [
                'شناسه', 'عامل', 'نوع تارگت', 'میزان هدف', 'دوره', 'مدت (روز)',
                'تاریخ شروع', 'تاریخ پایان', 'وضعیت', 'مقدار محقق شده', 'توضیحات', 'ایجاد شده توسط', 'تاریخ ایجاد'
            ]
            
            col_widths = [14, 20, 16, 16, 14, 12, 14, 14, 14, 18, 30, 18, 18]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            period_map = {
                'daily': 'روزانه',
                'weekly': 'هفتگی',
                'monthly': 'ماهانه',
                'seasonal': 'فصلی',
                'yearly': 'سالانه'
            }
            
            for idx, target in enumerate(filtered, 1):
                period_display = period_map.get(target.get('period_type', ''), target.get('period_type', ''))
                
                # تاریخ ایجاد به فرمت شمسی
                created_at = target.get('created_at', '')
                created_display = ''
                if created_at and 'T' in created_at:
                    created_date = created_at.split('T')[0]
                    try:
                        created_display = convert_to_jalali(created_date)
                    except:
                        created_display = created_date
                else:
                    created_display = created_at
                
                values = [
                    target.get('target_id', ''),
                    target.get('agent_name', ''),
                    target.get('target_type', ''),
                    target.get('target_value', 0),
                    period_display,
                    target.get('duration', 0),
                    target.get('start_date', ''),
                    target.get('end_date', ''),
                    target.get('status', ''),
                    target.get('achieved_value', 0),
                    target.get('description', ''),
                    target.get('created_by', ''),
                    created_display
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== خلاصه در انتها ==========
            summary_row = len(filtered) + 3
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
            ws.cell(row=summary_row, column=1, value='📊 خلاصه تارگت‌های سوپروایزر:').font = Font(bold=True, size=12, color="FFD700")
            
            total_target = sum(t.get('target_value', 0) for t in filtered)
            total_achieved = sum(t.get('achieved_value', 0) for t in filtered)
            completed_count = len([t for t in filtered if t.get('status') == 'تکمیل شده'])
            active_count = len([t for t in filtered if t.get('status') == 'فعال'])
            
            ws.cell(row=summary_row, column=10, value=f'مجموع هدف: {total_target:,}')
            ws.cell(row=summary_row, column=10).font = Font(bold=True, size=11, color="FFFFFF")
            
            ws.cell(row=summary_row + 1, column=10, value=f'مجموع تحقق: {total_achieved:,}')
            ws.cell(row=summary_row + 1, column=10).font = Font(bold=True, size=11, color="00CC44")
            
            ws.cell(row=summary_row + 2, column=10, value=f'تکمیل شده: {completed_count} | فعال: {active_count}')
            ws.cell(row=summary_row + 2, column=10).font = Font(bold=True, size=11, color="FFFFFF")
            
            wb.save(filepath)
            print(f"✅ گزارش تارگت سوپروایزر ساخته شد: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی تارگت‌ها (سوپروایزر): {e}")
            import traceback
            traceback.print_exc()
            return None
        
    def _export_detailed_targets_report_supervisor(self, agent_name, from_date, to_date, reports_dir, user_name):
        """
        خروجی ریزتارگت‌ها - مخصوص سوپروایزر (بر اساس created_by و تاریخ ایجاد)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from utils.jalali_date import convert_to_jalali
            
            file_path = os.path.join(get_data_path(), 'detailed_targets.json')
            if not os.path.exists(file_path):
                print(f"⚠️ فایل detailed_targets.json یافت نشد")
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_targets = json.load(f)
            
            if not isinstance(all_targets, list):
                all_targets = []
            
            print(f"📋 تعداد کل ریزتارگت‌ها: {len(all_targets)}")
            
            # ✅ فیلتر بر اساس created_by و تاریخ ایجاد
            filtered = []
            for t in all_targets:
                if not isinstance(t, dict):
                    continue
                
                t_creator = t.get('created_by', '')
                
                # اگر created_by با نام کاربر مطابقت دارد
                if agent_name in t_creator or t_creator in agent_name:
                    # ✅ بررسی تاریخ ایجاد (created_at)
                    created_at = t.get('created_at', '')
                    if created_at:
                        # استخراج تاریخ از created_at (فرمت: 2026-08-08T13:49:43.129990)
                        if 'T' in created_at:
                            created_date = created_at.split('T')[0]
                            # تبدیل به شمسی
                            try:
                                created_date_jalali = convert_to_jalali(created_date)
                                if from_date <= created_date_jalali <= to_date:
                                    filtered.append(t)
                            except:
                                # اگر تبدیل نشد، بر اساس تاریخ شروع فیلتر کن
                                t_date = t.get('start_date', '')
                                if from_date <= t_date <= to_date:
                                    filtered.append(t)
                        else:
                            # اگر فرمت T نداشت، خود تاریخ را بررسی کن
                            if from_date <= created_at <= to_date:
                                filtered.append(t)
                    else:
                        # اگر created_at نبود، بر اساس تاریخ شروع فیلتر کن
                        t_date = t.get('start_date', '')
                        if from_date <= t_date <= to_date:
                            filtered.append(t)
            
            print(f"📋 ریزتارگت‌های فیلتر شده برای سوپروایزر {agent_name}: {len(filtered)}")
            
            if not filtered:
                print(f"⚠️ هیچ ریزتارگتی برای سوپروایزر {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            filename = f'ریزتارگت_سوپروایزر_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ریزتارگت‌ها"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = [
                'شناسه ریزتارگت', 'نام عامل', 'دوره تارگت', 'تاریخ شروع', 'تاریخ پایان',
                'تاریخ ایجاد', 'نام گروه کالا', 'تعداد تارگت', 'واحد تارگت',
                'پیوند با تارگت مادر', 'تارگت روزانه', 'مقدار محقق شده', 'وضعیت',
                'ایجاد شده توسط'
            ]
            
            col_widths = [18, 20, 14, 14, 14, 14, 22, 14, 14, 20, 16, 16, 14, 18]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for idx, target in enumerate(filtered, 1):
                # تاریخ ایجاد به فرمت شمسی
                created_at = target.get('created_at', '')
                created_display = ''
                if created_at and 'T' in created_at:
                    created_date = created_at.split('T')[0]
                    try:
                        created_display = convert_to_jalali(created_date)
                    except:
                        created_display = created_date
                else:
                    created_display = created_at
                
                values = [
                    target.get('id', ''),
                    target.get('agent_name', ''),
                    target.get('period', ''),
                    target.get('start_date', ''),
                    target.get('end_date', ''),
                    created_display,
                    target.get('product_group', ''),
                    target.get('target_count', 0),
                    target.get('unit', ''),
                    target.get('linked_target_id', ''),
                    target.get('daily_target', 0),
                    target.get('achieved_value', 0),
                    target.get('status', ''),
                    target.get('created_by', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== خلاصه در انتها ==========
            summary_row = len(filtered) + 3
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
            ws.cell(row=summary_row, column=1, value='📊 خلاصه ریزتارگت‌های سوپروایزر:').font = Font(bold=True, size=12, color="FFD700")
            
            total_target = sum(t.get('target_count', 0) for t in filtered)
            total_achieved = sum(t.get('achieved_value', 0) for t in filtered)
            completed_count = len([t for t in filtered if t.get('status') == 'تکمیل شده'])
            active_count = len([t for t in filtered if t.get('status') == 'فعال'])
            
            ws.cell(row=summary_row, column=10, value=f'مجموع هدف: {total_target:,}')
            ws.cell(row=summary_row, column=10).font = Font(bold=True, size=11, color="FFFFFF")
            
            ws.cell(row=summary_row + 1, column=10, value=f'مجموع تحقق: {total_achieved:,}')
            ws.cell(row=summary_row + 1, column=10).font = Font(bold=True, size=11, color="00CC44")
            
            ws.cell(row=summary_row + 2, column=10, value=f'تکمیل شده: {completed_count} | فعال: {active_count}')
            ws.cell(row=summary_row + 2, column=10).font = Font(bold=True, size=11, color="FFFFFF")
            
            wb.save(filepath)
            print(f"✅ گزارش ریزتارگت سوپروایزر ساخته شد: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی ریزتارگت‌ها (سوپروایزر): {e}")
            import traceback
            traceback.print_exc()
            return None

    # ============================================================
    # ۹. گزارش سرکشی (بررسی بازار) - جدید
    # ============================================================
    
    def _export_supervisor_visits_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """خروجی سرکشی‌ها - با خواندن مستقیم فایل supervisor_visits.json"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # ✅ خواندن مستقیم فایل supervisor_visits.json
            file_path = os.path.join(get_data_path(), 'supervisor_visits.json')
            if not os.path.exists(file_path):
                print(f"⚠️ فایل supervisor_visits.json یافت نشد: {file_path}")
                return None
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_visits = json.load(f)
            
            if not isinstance(all_visits, list):
                all_visits = []
            
            print(f"📋 تعداد کل سرکشی‌ها: {len(all_visits)}")
            
            # ✅ فیلتر بر اساس تاریخ و نام کاربر
            filtered = []
            for v in all_visits:
                if not isinstance(v, dict):
                    continue
                
                # بررسی تاریخ
                v_date = v.get('date', '')
                if from_date <= v_date <= to_date:
                    # ✅ بررسی نام کاربر (created_by یا agent_name)
                    v_agent = v.get('created_by', '') or v.get('agent_name', '')
                    
                    # تطابق جزئی نام عامل
                    if agent_name in v_agent or v_agent in agent_name:
                        filtered.append(v)
                    elif not v_agent:
                        # اگر نام عامل خالی بود، از created_by استفاده کن
                        created_by = v.get('created_by', '')
                        if agent_name in created_by or created_by in agent_name:
                            filtered.append(v)
            
            print(f"📋 سرکشی‌های فیلتر شده برای {agent_name}: {len(filtered)}")
            
            if not filtered:
                print(f"⚠️ هیچ سرکشی برای {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            # ========== مرتب‌سازی بر اساس تاریخ ==========
            filtered.sort(key=lambda x: x.get('date', ''), reverse=False)
            
            filename = f'بررسی_بازار_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "بررسی بازار"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ✅ هدرهای کامل مطابق با فایل نمونه
            headers = [
                'شناسه', 'تاریخ', 'ساعت', 'مسیر', 'مشتری', 'نحوه سرکشی',
                'علت سرکشی', 'وضعیت مشتری', 'وضعیت حضور در شلف',
                'تعداد سرکشی در ماه', 'آیا سرکشی کافیست؟', 'خرید مورد انتظار',
                'وضعیت موجودی', 'برخورد بازاریاب', 'برخورد موزع',
                'رضایتمندی مشتری', 'نظرات مشتری', 'تحقق هدف سرکشی',
                'نظریه سوپروایزر', 'نیاز به پیگیری', 'تاریخ مراجعه بعدی'
            ]
            
            col_widths = [12, 12, 10, 14, 22, 14, 18, 14, 18, 16, 18, 20, 14, 16, 14, 16, 25, 16, 25, 14, 16]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # ========== پر کردن داده‌ها ==========
            total_visits = len(filtered)
            need_followup = 0
            reported_count = 0
            
            for idx, visit in enumerate(filtered, 1):
                need_followup += 1 if visit.get('need_followup') == 'بله' else 0
                reported_count += 1 if visit.get('reported_to_manager', False) else 0
                
                values = [
                    visit.get('id', ''),
                    visit.get('date', ''),
                    visit.get('time', ''),
                    visit.get('route', ''),
                    visit.get('customer', ''),
                    visit.get('visit_type', ''),
                    visit.get('visit_reason', ''),
                    visit.get('customer_status', ''),
                    visit.get('shelf_status', ''),
                    visit.get('monthly_visits', ''),
                    visit.get('visit_sufficient', ''),
                    visit.get('expected_purchase', ''),
                    visit.get('inventory_status', ''),
                    visit.get('agent_behavior', ''),
                    visit.get('distributor_behavior', ''),
                    visit.get('customer_satisfaction', ''),
                    visit.get('customer_feedback', ''),
                    visit.get('target_achievement', ''),
                    visit.get('supervisor_opinion', ''),
                    visit.get('need_followup', ''),
                    visit.get('next_visit_date', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=idx+1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== خلاصه در انتها ==========
            summary_row = len(filtered) + 3
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
            ws.cell(row=summary_row, column=1, value='📊 خلاصه سرکشی‌ها:').font = Font(bold=True, size=12, color="FFD700")
            
            ws.cell(row=summary_row, column=10, value=f'تعداد کل سرکشی‌ها: {total_visits}')
            ws.cell(row=summary_row, column=10).font = Font(bold=True, size=11, color="FFFFFF")
            
            ws.cell(row=summary_row + 1, column=10, value=f'نیاز به پیگیری: {need_followup}')
            ws.cell(row=summary_row + 1, column=10).font = Font(bold=True, size=11, color="FF6B6B")
            
            ws.cell(row=summary_row + 2, column=10, value=f'گزارش شده به مدیر: {reported_count}')
            ws.cell(row=summary_row + 2, column=10).font = Font(bold=True, size=11, color="00CC44")
            
            wb.save(filepath)
            print(f"✅ گزارش سرکشی ساخته شد: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی سرکشی‌ها: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _export_market_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """
        خروجی گزارش بازاری (نامه اداری) - ارسال به مدیر
        
        Args:
            agent_name: نام سوپروایزر
            from_date: تاریخ شروع
            to_date: تاریخ پایان
            reports_dir: مسیر ذخیره
            user_name: نام کاربر
        
        Returns:
            str: نام فایل یا None در صورت خطا
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from utils.supervisor_visits_manager import get_visits_filtered, get_visits_by_creator
            from utils.jalali_date import get_today_jalali
            
            # ============================================================
            # دریافت سرکشی‌های فیلتر شده
            # ============================================================
            # تلاش با get_visits_by_creator برای دریافت سرکشی‌های سوپروایزر
            visits = get_visits_by_creator(agent_name)
            
            # اگر چیزی پیدا نشد، با get_visits_filtered امتحان کن
            if not visits:
                visits = get_visits_filtered(start_date=from_date, end_date=to_date)
                # فیلتر دستی بر اساس created_by
                visits = [v for v in visits if v.get('created_by') == agent_name or v.get('agent_name') == agent_name]
            
            # فیلتر بر اساس تاریخ
            visits = [v for v in visits if from_date <= v.get('date', '') <= to_date]
            
            if not visits:
                print(f"⚠️ هیچ سرکشی برای {agent_name} در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            print(f"📋 تعداد سرکشی‌های {agent_name}: {len(visits)}")
            
            # ============================================================
            # ساخت فایل اکسل با فرمت نامه اداری
            # ============================================================
            filename = f'گزارش_بازاری_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            
            # ============================================================
            # شیت ۱: نامه اداری (برای هر سرکشی یک نامه جداگانه)
            # ============================================================
            for idx, visit in enumerate(visits):
                if idx == 0:
                    ws = wb.active
                    ws.title = f"نامه {idx+1}"
                else:
                    ws = wb.create_sheet(f"نامه {idx+1}")
                
                ws.sheet_view.rightToLeft = True
                
                # استایل‌ها
                title_font = Font(name='B Nazanin', size=16, bold=True)
                text_font = Font(name='B Nazanin', size=12)
                bold_font = Font(name='B Nazanin', size=12, bold=True)
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # ============================================================
                # عنوان
                # ============================================================
                ws.merge_cells('A1:B1')
                title_cell = ws.cell(row=1, column=1, value='بنام خدا')
                title_cell.font = title_font
                title_cell.alignment = center_align
                
                ws.merge_cells('A2:B2')
                visit_id = visit.get('id', '')
                ws.cell(row=2, column=1, value=f'گزارش بررسی بازار - شماره: {visit_id}').font = title_font
                ws.cell(row=2, column=1).alignment = center_align
                
                # ============================================================
                # اطلاعات اصلی
                # ============================================================
                row_num = 4
                
                info_fields = [
                    ('تاریخ سرکشی', f"{visit.get('date', '')} - ساعت: {visit.get('time', '')}"),
                    ('مسیر', visit.get('route', '')),
                    ('مشتری', visit.get('customer', '')),
                    ('', ''),
                    ('نحوه سرکشی', visit.get('visit_type', '')),
                    ('علت سرکشی', visit.get('visit_reason', '')),
                    ('', ''),
                    ('وضعیت مشتری', visit.get('customer_status', '')),
                    ('وضعیت حضور در شلف', visit.get('shelf_status', '')),
                    ('تعداد سرکشی بازاریابان در ماه', visit.get('monthly_visits', '')),
                    ('آیا میزان سرکشی کافیست؟', visit.get('visit_sufficient', '')),
                    ('میزان خرید مورد انتظار', visit.get('expected_purchase', '')),
                    ('وضعیت موجودی مشتری', visit.get('inventory_status', '')),
                    ('', ''),
                    ('نحوه برخورد بازاریابان', visit.get('agent_behavior', '')),
                    ('نحوه برخورد موزعین', visit.get('distributor_behavior', '')),
                    ('میزان رضایتمندی مشتری', visit.get('customer_satisfaction', '')),
                    ('میزان تحقق هدف سرکشی', visit.get('target_achievement', '')),
                    ('', ''),
                    ('نیاز به پیگیری مجدد', visit.get('need_followup', '')),
                    ('تاریخ مراجعه بعدی', visit.get('next_visit_date', '---')),
                    ('', ''),
                    ('توضیحات سوپروایزر', visit.get('supervisor_note', '---')),
                    ('نظرات مشتری', visit.get('customer_feedback', '---')),
                    ('نظریه نهایی سوپروایزر', visit.get('supervisor_opinion', '---')),
                ]
                
                for label, value in info_fields:
                    if label == '':
                        row_num += 1
                        continue
                    
                    # لیبل
                    cell_a = ws.cell(row=row_num, column=1, value=f'{label}:')
                    cell_a.font = bold_font
                    cell_a.alignment = right_align
                    cell_a.border = thin_border
                    
                    # مقدار
                    cell_b = ws.cell(row=row_num, column=2, value=str(value) if value else '---')
                    cell_b.font = text_font
                    cell_b.alignment = right_align
                    cell_b.border = thin_border
                    
                    row_num += 1
                
                # ============================================================
                # خط جداکننده
                # ============================================================
                row_num += 1
                ws.merge_cells(f'A{row_num}:B{row_num}')
                sep_cell = ws.cell(row=row_num, column=1, value='─' * 50)
                sep_cell.alignment = center_align
                sep_cell.font = text_font
                sep_cell.border = thin_border
                
                # ============================================================
                # اطلاعات ثبت
                # ============================================================
                row_num += 1
                ws.merge_cells(f'A{row_num}:B{row_num}')
                ws.cell(row=row_num, column=1, value='این گزارش توسط سیستم مدیریت بازاریابی تهیه شده است.').font = text_font
                ws.cell(row=row_num, column=1).alignment = center_align
                ws.cell(row=row_num, column=1).border = thin_border
                
                row_num += 1
                ws.merge_cells(f'A{row_num}:B{row_num}')
                today = get_today_jalali()
                created_by = visit.get('created_by', 'supervisor')
                ws.cell(row=row_num, column=1, value=f'تاریخ ثبت گزارش: {today} | ثبت شده توسط: {created_by}').font = text_font
                ws.cell(row=row_num, column=1).alignment = center_align
                ws.cell(row=row_num, column=1).border = thin_border
                
                # ============================================================
                # تنظیم عرض ستون‌ها
                # ============================================================
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 55
                
                # تنظیم ارتفاع ردیف‌ها
                ws.row_dimensions[1].height = 30
                ws.row_dimensions[2].height = 30
                for r in range(4, row_num + 1):
                    ws.row_dimensions[r].height = 24
            
            # ============================================================
            # شیت ۲: خلاصه گزارشات
            # ============================================================
            ws_summary = wb.create_sheet("خلاصه گزارشات")
            ws_summary.sheet_view.rightToLeft = True
            
            # عنوان
            ws_summary.merge_cells('A1:D1')
            ws_summary.cell(row=1, column=1, value='خلاصه گزارشات بازاری').font = Font(bold=True, size=14)
            ws_summary.cell(row=1, column=1).alignment = center_align
            
            # هدر
            headers = ['ردیف', 'شناسه', 'تاریخ', 'مشتری', 'مسیر', 'وضعیت']
            header_row = 3
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            
            # داده‌ها
            for idx, visit in enumerate(visits, 1):
                row = header_row + idx
                reported = visit.get('reported_to_manager', False)
                status = 'ارسال شده' if reported else 'در انتظار ارسال'
                status_color = "00CC44" if reported else "FFAA00"
                
                ws_summary.cell(row=row, column=1, value=idx).alignment = center_align
                ws_summary.cell(row=row, column=1).border = thin_border
                
                ws_summary.cell(row=row, column=2, value=visit.get('id', '')).alignment = center_align
                ws_summary.cell(row=row, column=2).border = thin_border
                
                ws_summary.cell(row=row, column=3, value=visit.get('date', '')).alignment = center_align
                ws_summary.cell(row=row, column=3).border = thin_border
                
                ws_summary.cell(row=row, column=4, value=visit.get('customer', '')).alignment = center_align
                ws_summary.cell(row=row, column=4).border = thin_border
                
                ws_summary.cell(row=row, column=5, value=visit.get('route', '')).alignment = center_align
                ws_summary.cell(row=row, column=5).border = thin_border
                
                status_cell = ws_summary.cell(row=row, column=6, value=status)
                status_cell.alignment = center_align
                status_cell.border = thin_border
                status_cell.font = Font(color=status_color, bold=True)
            
            # عرض ستون‌ها
            col_widths = [8, 14, 14, 22, 16, 16]
            for i, width in enumerate(col_widths, 1):
                ws_summary.column_dimensions[get_column_letter(i)].width = width
            
            # ============================================================
            # ذخیره فایل
            # ============================================================
            wb.save(filepath)
            print(f"✅ گزارش بازاری ساخته شد: {filename} ({len(visits)} نامه)")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی گزارش بازاری: {e}")
            import traceback
            traceback.print_exc()
            return None
        

    def _export_evaluation_report(self, agent_name, from_date, to_date, reports_dir, user_name):
        """خروجی گزارش ارزیابی - مطابق با supervisor_report_screen"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from utils.file_manager import get_daily_logs, get_settings, get_agents
            from utils.jalali_date import get_today_jalali
            
            # ============================================================
            # دریافت داده‌ها
            # ============================================================
            all_logs = get_daily_logs()
            settings = get_settings()
            
            # ساخت mapping route -> agent
            agents = get_agents()
            route_agent_map = {}
            for a in agents:
                if isinstance(a, dict):
                    agent_routes = a.get('routes', [])
                    if isinstance(agent_routes, list):
                        for r in agent_routes:
                            route_agent_map[r] = a.get('name', '')
            
            # فیلتر تاریخ
            date_list = []
            for date in all_logs.keys():
                if from_date <= date <= to_date:
                    date_list.append(date)
            
            if not date_list:
                print(f"⚠️ هیچ داده‌ای برای ارزیابی در بازه {from_date} تا {to_date} یافت نشد")
                return None
            
            # ============================================================
            # محاسبه آمار
            # ============================================================
            supervision_rate = settings.get('supervision_rate', 70.0) / 100
            conversion_rate = settings.get('conversion_rate', 75.0) / 100
            target_units = settings.get('target_count', 100)
            target_sales = settings.get('target_amount', 50000000)
            target_cash = settings.get('target_cash_sales', 30000000)
            target_check = settings.get('target_credit_sales', 20000000)
            
            total_visits = 0
            total_invoices = 0
            total_units = 0
            total_sales = 0
            total_cash = 0
            total_check = 0
            total_new_customers = 0
            agents_data = {}
            
            for date in date_list:
                if date not in all_logs or not isinstance(all_logs[date], list):
                    continue
                for log in all_logs[date]:
                    if not isinstance(log, dict):
                        continue
                    
                    log_route = log.get('route', '')
                    log_agent = log.get('agent_name', '')
                    
                    if not log_agent and log_route:
                        log_agent = route_agent_map.get(log_route, log_route)
                    
                    if agent_name and log_agent != agent_name:
                        continue
                    
                    visit_status = log.get('visit_status', '')
                    sales_status = log.get('sales_status', '')
                    payment_method = log.get('payment_method', '')
                    sales_amount = log.get('sales_amount', 0)
                    units_sold = log.get('units_sold', 0)
                    
                    if log_agent not in agents_data:
                        agents_data[log_agent] = {
                            'visits': 0, 'invoices': 0, 'units': 0,
                            'sales': 0, 'cash': 0, 'check': 0, 'new_customers': 0
                        }
                    
                    if visit_status == 'موفق':
                        total_visits += 1
                        agents_data[log_agent]['visits'] += 1
                    if sales_status == 'موفق':
                        total_invoices += 1
                        total_units += units_sold
                        total_sales += sales_amount
                        agents_data[log_agent]['invoices'] += 1
                        agents_data[log_agent]['units'] += units_sold
                        agents_data[log_agent]['sales'] += sales_amount
                        if payment_method == 'نقد':
                            total_cash += sales_amount
                            agents_data[log_agent]['cash'] += sales_amount
                        elif payment_method == 'چک':
                            total_check += sales_amount
                            agents_data[log_agent]['check'] += sales_amount
                    if log.get('is_new_customer', False):
                        total_new_customers += 1
                        agents_data[log_agent]['new_customers'] += 1
            
            day_count = len(date_list)
            target_visits_day = int(supervision_rate * 50) * day_count
            target_invoices_day = int(target_visits_day * conversion_rate)
            target_units_day = target_units * day_count
            target_sales_day = target_sales * day_count
            target_cash_day = target_cash * day_count
            target_check_day = target_check * day_count
            
            # ============================================================
            # ساخت فایل اکسل
            # ============================================================
            filename = f'ارزیابی_{user_name}_{from_date.replace("/", "-")}_تا_{to_date.replace("/", "-")}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "آمار و ارزیابی"
            ws.sheet_view.rightToLeft = True
            
            # استایل‌ها
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ============================================================
            # شیت ۱: آمار و ارزیابی
            # ============================================================
            
            # عنوان
            ws.merge_cells('A1:F1')
            title_cell = ws.cell(row=1, column=1, value=f'گزارش آمار و ارزیابی ({from_date} تا {to_date}) - {day_count} روز کاری')
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # هدر جدول
            headers = ['آیتم', 'هدف', 'عملکرد', 'اختلاف', 'درصد', 'وضعیت']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # داده‌ها
            items = [
                ('تعداد ویزیت', target_visits_day, total_visits),
                ('تعداد فاکتور', target_invoices_day, total_invoices),
                ('واحد فروش', target_units_day, total_units),
                ('مبلغ فروش', target_sales_day, total_sales),
                ('فروش نقدی', target_cash_day, total_cash),
                ('فروش چکی', target_check_day, total_check),
            ]
            
            total_percent = 0
            item_count = 0
            
            for row_idx, (name, target_val, actual_val) in enumerate(items, 4):
                diff = actual_val - target_val
                percent = (actual_val / target_val * 100) if target_val > 0 else 0
                
                if percent >= 70:
                    status = 'مطلوب'
                    status_color = "00CC44"
                elif percent >= 50:
                    status = 'متوسط'
                    status_color = "FFAA00"
                else:
                    status = 'ضعیف'
                    status_color = "CC3333"
                
                if target_val > 0:
                    total_percent += percent
                    item_count += 1
                
                values = [name, target_val, actual_val, diff, percent, status]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    if col_idx == 6:
                        cell.font = Font(color=status_color, bold=True)
                    if col_idx in [2, 3, 4]:
                        cell.number_format = '#,##0'
            
            # میانگین
            avg_percent = total_percent / item_count if item_count > 0 else 0
            avg_row = len(items) + 4
            
            ws.merge_cells(f'A{avg_row}:C{avg_row}')
            ws.cell(row=avg_row, column=1, value='میانگین تحقق').font = Font(bold=True)
            ws.cell(row=avg_row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=avg_row, column=1).border = thin_border
            
            ws.merge_cells(f'D{avg_row}:F{avg_row}')
            avg_cell = ws.cell(row=avg_row, column=4, value=f'{avg_percent:.1f}%')
            avg_cell.font = Font(bold=True, size=12, color="00CC44" if avg_percent >= 70 else "CC3333")
            avg_cell.alignment = Alignment(horizontal='center')
            avg_cell.border = thin_border
            
            # ارزیابی کلی
            eval_row = avg_row + 1
            ws.merge_cells(f'A{eval_row}:F{eval_row}')
            if avg_percent >= 100:
                eval_text = "عملکرد عالی - تبریک!"
            elif avg_percent >= 70:
                eval_text = "عملکرد خوب - در مسیر درست"
            elif avg_percent >= 50:
                eval_text = "نیاز به تلاش بیشتر"
            else:
                eval_text = "ضعیف - نیاز به بررسی و پیگیری"
            
            eval_cell = ws.cell(row=eval_row, column=1, value=eval_text)
            eval_cell.font = Font(bold=True, size=12)
            eval_cell.alignment = Alignment(horizontal='center')
            
            # ============================================================
            # شیت ۲: عملکرد تفکیکی عامل‌ها (اگر بیش از یک عامل وجود داشته باشد)
            # ============================================================
            if len(agents_data) > 1:
                ws2 = wb.create_sheet("عملکرد تفکیکی")
                ws2.sheet_view.rightToLeft = True
                
                # عنوان
                ws2.merge_cells('A1:F1')
                ws2.cell(row=1, column=1, value='عملکرد تفکیکی عامل‌ها').font = Font(bold=True, size=14)
                ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                # هدر
                agent_headers = ['عامل', 'ویزیت', 'فاکتور', 'فروش', 'نقدی', 'چکی']
                header_row = 3
                for col, header in enumerate(agent_headers, 1):
                    cell = ws2.cell(row=header_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # داده‌ها
                for i, (ag_name, ag_data) in enumerate(agents_data.items()):
                    row = header_row + 1 + i
                    values = [
                        ag_name,
                        ag_data['visits'],
                        ag_data['invoices'],
                        ag_data['sales'],
                        ag_data['cash'],
                        ag_data['check']
                    ]
                    for col, value in enumerate(values, 1):
                        cell = ws2.cell(row=row, column=col, value=value)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        if col >= 2:
                            cell.number_format = '#,##0'
                
                # عرض ستون‌ها
                col_widths = [20, 14, 14, 18, 18, 18]
                for i, width in enumerate(col_widths, 1):
                    ws2.column_dimensions[get_column_letter(i)].width = width
            
            # ============================================================
            # تنظیم عرض ستون‌های شیت اول
            # ============================================================
            col_widths = [28, 18, 18, 18, 14, 18]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(filepath)
            print(f"✅ گزارش ارزیابی ساخته شد: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ خطا در خروجی ارزیابی: {e}")
            import traceback
            traceback.print_exc()
            return None

    # ============================================================
    # توابع کمکی (تاریخچه، نمایش پیام، باز کردن فایل)
    # ============================================================
    
    def _show_history(self, instance):
        """نمایش تاریخچه گزارشات"""
        try:
            reports_dir = self._get_reports_dir()
            
            if not os.path.exists(reports_dir):
                self.show_message('توجه', 'هیچ گزارشی یافت نشد')
                return
            
            files = os.listdir(reports_dir)
            excel_files = [f for f in files if f.endswith('.xlsx')]
            
            if not excel_files:
                self.show_message('توجه', 'هیچ فایل اکسل در تاریخچه یافت نشد')
                return
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(6))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text=f'تاریخچه گزارشات ({len(excel_files)} فایل)',
                size_hint_y=None,
                height=dp(30),
                font_size=sp(16),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))
            
            scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.7)
            grid = GridLayout(cols=1, spacing=dp(2), size_hint_y=None, padding=dp(3))
            grid.bind(minimum_height=grid.setter('height'))
            
            for f in sorted(excel_files, reverse=True)[:20]:
                row = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(4))
                
                display_name = f
                if len(display_name) > 30:
                    display_name = display_name[:27] + '...'
                
                row.add_widget(RTLLabel(
                    text=display_name,
                    size_hint_x=0.75,
                    font_size=sp(11),
                    color=(1, 1, 1, 1)
                ))
                
                open_btn = PersianButton(
                    text='باز',
                    size_hint_x=0.25,
                    size_hint_y=None,
                    height=dp(26),
                    background_color=(0.2, 0.5, 0.9, 1),
                    color=(1, 1, 1, 1),
                    font_size=sp(11)
                )
                open_btn.bind(on_press=lambda x, fname=f: self._open_file(os.path.join(reports_dir, fname)))
                row.add_widget(open_btn)
                
                grid.add_widget(row)
            
            scroll.add_widget(grid)
            content.add_widget(scroll)
            
            close_btn = PersianButton(
                text='بستن',
                size_hint_y=None,
                height=dp(36),
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(14)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='تاریخچه',
                content=content,
                size_hint=(0.9, None),
                height=dp(400),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            print(f"خطا: {e}")
            self.show_message('خطا', f'خطا در نمایش تاریخچه: {str(e)}')
    
    def _open_file(self, filepath):
        """باز کردن فایل با برنامه پیش‌فرض"""
        try:
            import subprocess
            if os.name == 'nt':
                os.startfile(filepath)
            else:
                subprocess.Popen(['xdg-open', filepath])
        except Exception as e:
            self.show_message('خطا', f'خطا در باز کردن فایل: {str(e)}')
    
    def show_message(self, title, message):
        """نمایش پیام با پشتیبانی از متن طولانی"""
        try:
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            from kivy.uix.boxlayout import BoxLayout
            from kivy.metrics import dp, sp
            from kivy.graphics import Color, Rectangle
            from utils.rtl_widgets import PersianPopup, PersianButton
            
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(message)
                display_text = get_display(reshaped)
            except:
                display_text = message
            
            msg_label = Label(
                text=display_text,
                font_size=sp(15),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                halign='center',
                valign='top',
                text_size=(dp(550), None),
                font_name='fonts/Amiri-Regular.ttf',
                padding=(dp(20), 0, dp(0), 0)
            )
            
            lines = message.count('\n') + 1
            if len(message) > 50 and lines == 1:
                approx_chars_per_line = 35
                lines = (len(message) // approx_chars_per_line) + 1
            
            line_height = sp(15) + dp(6)
            label_height = max(lines * line_height + dp(25), dp(50))
            label_height = min(label_height, dp(490))
            
            msg_label.height = label_height
            msg_label.text_size = (dp(550), label_height)
            
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=None,
                height=label_height + dp(10),
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )
            scroll.add_widget(msg_label)
            content.add_widget(scroll)
            
            btn = PersianButton(
                text='باشه',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(16),
                color=(1, 1, 1, 1),
                background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            
            popup = PersianPopup(
                title=title,
                content=content,
                size_hint=(0.9, None),
                height=label_height + dp(250),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            print(f"خطا در نمایش پیام: {e}")
            import traceback
            traceback.print_exc()
            try:
                from error_handler import ErrorPopup
                ErrorPopup.show_error(str(message))
            except:
                pass
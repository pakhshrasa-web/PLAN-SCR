# screens/report_attendance_screen.py
# ========== صفحه گزارشات حضور و غیاب ==========

import os
import json
from datetime import datetime, timedelta
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.metrics import dp, sp
from kivy.graphics import Color, Rectangle
from kivy.clock import Clock

from utils.rtl_widgets import PersianButton, RTLLabel, PersianPopup, RTLTextInput, PersianComboBox
from utils.attendance_manager import AttendanceManager
from utils.jalali_date import get_today_jalali, get_jalali_months
from utils.file_manager import get_settings
from utils.storage import get_data_path
from error_handler import ErrorPopup


class ReportAttendanceScreen(BoxLayout):
    """صفحه گزارشات حضور و غیاب"""
    
    def __init__(self, current_user=None, **kwargs):
        super().__init__(**kwargs)
        self.current_user = current_user
        self.today = get_today_jalali()
        self.current_report_tab = 1
        
        # تنظیمات ساعات کاری
        settings = get_settings()
        self.work_start_time = settings.get('work_start_time', '08:00')
        self.work_end_time = settings.get('work_end_time', '17:00')
        self.min_daily_hours = float(settings.get('min_daily_hours', 7))
        
        # بارگذاری امتیازات
        self.points_config = self._load_points_config()
        
        with self.canvas.before:
            Color(0.08, 0.08, 0.08, 1)
            self.bg_rect = Rectangle(pos=self.pos, size=self.size)
            self.bind(pos=self._update_bg, size=self._update_bg)
        
        self.build_ui()
        self.switch_report_tab(1)
    
    def _update_bg(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size
    
    def set_user(self, user):
        self.current_user = user
    
    def _load_points_config(self):
        """بارگذاری تنظیمات امتیاز"""
        try:
            file_path = os.path.join(get_data_path(), 'points.json')
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {
                'hourly_score': 10,
                'overtime_multiplier': 1.5,
                'deduction_multiplier': 1.5
            }
        except:
            return {
                'hourly_score': 10,
                'overtime_multiplier': 1.5,
                'deduction_multiplier': 1.5
            }
    
    def _get_months_list(self):
        """دریافت لیست ماه‌های سال با فرمت 1404/01"""
        months = []
        current_year = int(self.today.split('/')[0])
        current_month = int(self.today.split('/')[1])
        
        # ۱۲ ماه اخیر
        for i in range(12):
            year = current_year
            month = current_month - i
            if month <= 0:
                month += 12
                year -= 1
            months.append(f"{year}/{month:02d}")
        
        return months
    
    def build_ui(self):
        """ساخت رابط کاربری"""
        self.clear_widgets()
        
        layout = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(6))
        
        # عنوان
        header = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8))
        header.add_widget(RTLLabel(
            text='گزارشات',
            font_size=sp(20),
            bold=True,
            color=(0.4, 0.8, 1, 1),
            size_hint_x=0.4
        ))
        header.add_widget(RTLLabel(
            text=self.today,
            font_size=sp(14),
            color=(0.6, 0.6, 0.6, 1),
            size_hint_x=0.3,
            halign='center'
        ))
        if self.current_user:
            header.add_widget(RTLLabel(
                text=self.current_user.get('name', ''),
                font_size=sp(14),
                color=(1, 1, 1, 1),
                size_hint_x=0.3,
                halign='left'
            ))
        layout.add_widget(header)
        
        # ========== نوار تب‌های گزارش ==========
        tabs = BoxLayout(size_hint_y=None, height=dp(38), spacing=dp(3))
        
        self.report_tab1 = PersianButton(
            text='ورود و خروج',
            size_hint_x=0.33,
            background_color=(0.2, 0.5, 0.9, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        self.report_tab1.bind(on_press=lambda x: self.switch_report_tab(1))
        tabs.add_widget(self.report_tab1)
        
        self.report_tab2 = PersianButton(
            text='مرخصی',
            size_hint_x=0.33,
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        self.report_tab2.bind(on_press=lambda x: self.switch_report_tab(2))
        tabs.add_widget(self.report_tab2)
        
        self.report_tab3 = PersianButton(
            text='ماموریت',
            size_hint_x=0.34,
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        self.report_tab3.bind(on_press=lambda x: self.switch_report_tab(3))
        tabs.add_widget(self.report_tab3)
        
        layout.add_widget(tabs)
        
        # ========== محتوای تب گزارش ==========
        self.report_content = BoxLayout(orientation='vertical', size_hint_y=1)
        layout.add_widget(self.report_content)
        
       
        self.add_widget(layout)
        
        # نمایش تب اول
        self.switch_report_tab(1)
    
    def switch_report_tab(self, tab_num):
        """تغییر تب گزارش"""
        self.current_report_tab = tab_num
        
        self.report_tab1.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 1 else (0.3, 0.3, 0.3, 1)
        self.report_tab2.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 2 else (0.3, 0.3, 0.3, 1)
        self.report_tab3.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 3 else (0.3, 0.3, 0.3, 1)
        
        self.report_content.clear_widgets()
        
        if tab_num == 1:
            self.show_attendance_report()
        elif tab_num == 2:
            self.show_leave_report()
        elif tab_num == 3:
            self.show_mission_report()
    
    # ============================================================
    # گزارش ورود و خروج
    # ============================================================
    
    def show_attendance_report(self):
        """نمایش گزارش ورود و خروج"""
        content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(6))
        
        # ========== ردیف فیلترها ==========
        filter_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        
        # کامبوباکس ماه
        months = self._get_months_list()
        self.month_combo = PersianComboBox(
            text=months[0] if months else '',
            values=months,
            height=dp(36),
            size_hint_x=0.33
        )
        self.month_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.month_combo.main_btn.color = (1, 1, 1, 1)
        self.month_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.month_combo)
        

        
        # کامبوباکس نوع گزارش
        self.report_type_combo = PersianComboBox(
            text='روزانه',
            values=['روزانه', 'تجمیعی', 'تفکیک ماهانه'],
            height=dp(36),
            size_hint_x=0.34
        )
        self.report_type_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.report_type_combo.main_btn.color = (1, 1, 1, 1)
        self.report_type_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.report_type_combo)
        

        
        # کامبوباکس خروجی
        self.export_type_combo = PersianComboBox(
            text='همه رکوردها',
            values=['همه رکوردها', 'خلاصه'],
            height=dp(36),
            size_hint_x=0.33
        )
        self.export_type_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.export_type_combo.main_btn.color = (1, 1, 1, 1)
        self.export_type_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.export_type_combo)
        
        
        content.add_widget(filter_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== ردیف دکمه‌ها ==========
        btn_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        
        show_btn = PersianButton(
            text='نمایش',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.2, 0.6, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        show_btn.bind(on_press=self._show_attendance_report_data)
        btn_row.add_widget(show_btn)
        
        excel_btn = PersianButton(
            text='خروجی اکسل',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        excel_btn.bind(on_press=self._export_attendance_excel)
        btn_row.add_widget(excel_btn)
        
        reset_btn = PersianButton(
            text='بازنشانی',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.8, 0.5, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        reset_btn.bind(on_press=self._reset_attendance_filters)
        btn_row.add_widget(reset_btn)
        
        content.add_widget(btn_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== لیست گزارش ==========
        self.att_report_scroll = ScrollView(
            do_scroll_x=True,
            do_scroll_y=True,
            size_hint_y=1,
            scroll_type=['bars', 'content'],
            bar_width=dp(5)
        )
        
        self.att_report_container = GridLayout(
            cols=1,
            spacing=dp(1),
            size_hint_y=None,
            size_hint_x=None,
            width=dp(900),
            padding=dp(2)
        )
        self.att_report_container.bind(minimum_height=self.att_report_container.setter('height'))
        self.att_report_scroll.add_widget(self.att_report_container)
        content.add_widget(self.att_report_scroll)
        
        self.report_content.add_widget(content)
        
        # بارگذاری اولیه
        Clock.schedule_once(lambda dt: self._show_attendance_report_data(None), 0.2)

    def _reset_attendance_filters(self, instance):
        """بازنشانی فیلترهای گزارش ورود و خروج"""
        try:
            # تنظیم مجدد ماه به ماه جاری
            months = self._get_months_list()
            if months:
                self.month_combo.text = months[0]
            
            # تنظیم مجدد نوع گزارش به روزانه
            self.report_type_combo.text = 'روزانه'
            
            # تنظیم مجدد نوع خروجی به همه رکوردها
            self.export_type_combo.text = 'همه رکوردها'
            
            # نمایش مجدد گزارش
            self._show_attendance_report_data(None)
            
        except Exception as e:
            print(f"خطا در بازنشانی فیلترها: {e}")
            self.show_message('خطا', f'خطا در بازنشانی فیلترها: {str(e)}')

    def _show_attendance_report_data(self, instance):
        """نمایش داده‌های گزارش ورود و خروج"""
        try:
            self.att_report_container.clear_widgets()
            
            # دریافت ماه انتخاب شده
            selected_month = self.month_combo.text
            if not selected_month:
                self._show_error('لطفاً یک ماه را انتخاب کنید')
                return
            
            report_type = self.report_type_combo.text
            export_type = self.export_type_combo.text
            
            # دریافت رکوردها
            records = AttendanceManager.load_attendance()
            user_id = self.current_user.get('id') if self.current_user else None
            
            # فیلتر بر اساس کاربر و ماه
            filtered = []
            year, month = selected_month.split('/')
            for r in records:
                if user_id and r.get('user_id') != user_id:
                    continue
                r_date = r.get('date', '')
                if r_date.startswith(f"{year}/{month}"):
                    filtered.append(r)
            
            if not filtered:
                self.att_report_container.add_widget(RTLLabel(
                    text='هیچ داده‌ای در ماه انتخاب شده یافت نشد',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                self._update_summary({})
                return
            
            # محاسبه آمار
            stats = self._calculate_attendance_stats(filtered, report_type)
            
            # به‌روزرسانی خلاصه
            self._update_summary(stats)
            
            # نمایش بر اساس نوع گزارش
            if report_type == 'روزانه':
                self._show_daily_report(filtered, stats)
            elif report_type == 'تجمیعی':
                self._show_summary_report(stats)
            else:  # تفکیک ماهانه
                self._show_monthly_report(filtered)
            
        except Exception as e:
            print(f"خطا در نمایش گزارش: {e}")
            import traceback
            traceback.print_exc()
            self._show_error(f'خطا در نمایش گزارش: {str(e)}')
    
    def _calculate_attendance_stats(self, records, report_type):
        """محاسبه آمار حضور و غیاب"""
        stats = {
            'total_hours': 0,
            'total_minutes': 0,
            'overtime_hours': 0,
            'overtime_minutes': 0,
            'deduction_hours': 0,
            'deduction_minutes': 0,
            'points': 0,
            'working_days': 0,
            'absent_days': 0,
            'daily_stats': []
        }
        
        daily_work_minutes = int(self.min_daily_hours * 60)
        
        # گروه‌بندی بر اساس روز
        daily_data = {}
        for r in records:
            date = r.get('date', '')
            if date not in daily_data:
                daily_data[date] = []
            daily_data[date].append(r)
        
        # محاسبه هر روز
        for date, day_records in daily_data.items():
            day_total_minutes = 0
            check_in_count = 0
            check_out_count = 0
            first_check_in = None
            last_check_out = None
            
            for r in day_records:
                check_in = r.get('check_in')
                check_out = r.get('check_out')
                status = r.get('status', '')
                
                if check_in:
                    check_in_count += 1
                    if not first_check_in or check_in < first_check_in:
                        first_check_in = check_in
                
                if check_out:
                    check_out_count += 1
                    if not last_check_out or check_out > last_check_out:
                        last_check_out = check_out
                
                if check_in and check_out:
                    in_h, in_m = map(int, check_in.split(':'))
                    out_h, out_m = map(int, check_out.split(':'))
                    diff = (out_h - in_h) * 60 + (out_m - in_m)
                    if diff > 0:
                        day_total_minutes += diff
            
            # محاسبه اضافه کار و کسر کار
            overtime = 0
            deduction = 0
            if day_total_minutes > daily_work_minutes:
                overtime = day_total_minutes - daily_work_minutes
            elif day_total_minutes < daily_work_minutes and day_total_minutes > 0:
                deduction = daily_work_minutes - day_total_minutes
            
            stats['total_minutes'] += day_total_minutes
            stats['overtime_minutes'] += overtime
            stats['deduction_minutes'] += deduction
            
            if day_total_minutes > 0:
                stats['working_days'] += 1
            
            # ذخیره آمار روزانه
            stats['daily_stats'].append({
                'date': date,
                'total_minutes': day_total_minutes,
                'overtime': overtime,
                'deduction': deduction,
                'first_check_in': first_check_in or '-',
                'last_check_out': last_check_out or '-',
                'check_in_count': check_in_count,
                'check_out_count': check_out_count
            })
        
        # محاسبه روزهای بدون کارکرد
        # اینجا باید کل روزهای ماه رو در نظر بگیریم و ببینیم کدوم روزها رکورد ندارن
        # برای سادگی، روزهایی که working_days محاسبه نشده رو بدون کارکرد در نظر میگیریم
        # (در نسخه کامل باید تمام روزهای ماه رو بررسی کنیم)
        
        # تبدیل به ساعت و دقیقه
        stats['total_hours'] = stats['total_minutes'] // 60
        stats['total_minutes'] = stats['total_minutes'] % 60
        stats['overtime_hours'] = stats['overtime_minutes'] // 60
        stats['overtime_minutes'] = stats['overtime_minutes'] % 60
        stats['deduction_hours'] = stats['deduction_minutes'] // 60
        stats['deduction_minutes'] = stats['deduction_minutes'] % 60
        
        # محاسبه امتیاز
        hourly_score = self.points_config.get('hourly_score', 10)
        overtime_multiplier = self.points_config.get('overtime_multiplier', 1.5)
        deduction_multiplier = self.points_config.get('deduction_multiplier', 1.5)
        
        total_hours_float = stats['total_hours'] + stats['total_minutes'] / 60
        overtime_hours_float = stats['overtime_hours'] + stats['overtime_minutes'] / 60
        deduction_hours_float = stats['deduction_hours'] + stats['deduction_minutes'] / 60
        
        points = (total_hours_float * hourly_score) + \
                 (overtime_hours_float * hourly_score * overtime_multiplier) - \
                 (deduction_hours_float * hourly_score * deduction_multiplier)
        
        stats['points'] = round(points, 2)
        
        # تعداد روزهای بدون کارکرد (تخمینی)
        # اینجا باید دقیق‌تر محاسبه بشه، فعلاً ساده
        total_days_in_month = 30  # تخمین
        stats['absent_days'] = total_days_in_month - stats['working_days']
        
        return stats
    
    def _update_summary(self, stats):
        """به‌روزرسانی خلاصه آمار - غیرفعال شده"""
        pass
    
    def _show_daily_report(self, records, stats):
        """نمایش گزارش روزانه"""
        # هدر
        header = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(1), size_hint_x=None, width=dp(900))
        headers = ['تاریخ', 'وضعیت', 'اولین ورود', 'آخرین خروج', 'تعداد ورود', 'تعداد خروج', 'کارکرد', 'اضافه کار', 'کسر کار']
        sizes = [0.11, 0.12, 0.11, 0.11, 0.10, 0.10, 0.12, 0.11, 0.12]
        
        for i, (text, size) in enumerate(zip(headers, sizes)):
            header.add_widget(RTLLabel(
                text=text,
                size_hint_x=size,
                font_size=sp(9),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))
        self.att_report_container.add_widget(header)
        
        # ردیف‌ها
        for idx, day in enumerate(stats['daily_stats'], 1):
            row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(1), size_hint_x=None, width=dp(900))
            with row.canvas.before:
                Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                rect = Rectangle(pos=row.pos, size=row.size)
                row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                        size=lambda i, v, r=rect: setattr(r, 'size', v))
            
            total_h = day['total_minutes'] // 60
            total_m = day['total_minutes'] % 60
            over_h = day['overtime'] // 60
            over_m = day['overtime'] % 60
            ded_h = day['deduction'] // 60
            ded_m = day['deduction'] % 60
            
            status = 'حضور' if day['total_minutes'] > 0 else 'غیبت'
            status_color = (0.2, 0.9, 0.2, 1) if day['total_minutes'] > 0 else (0.9, 0.2, 0.2, 1)
            
            row.add_widget(RTLLabel(
                text=day['date'],
                size_hint_x=0.11,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=status,
                size_hint_x=0.12,
                font_size=sp(9),
                bold=True,
                color=status_color
            ))
            row.add_widget(RTLLabel(
                text=day['first_check_in'],
                size_hint_x=0.11,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=day['last_check_out'],
                size_hint_x=0.11,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=str(day['check_in_count']),
                size_hint_x=0.10,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=str(day['check_out_count']),
                size_hint_x=0.10,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{total_h:02d}:{total_m:02d}',
                size_hint_x=0.12,
                font_size=sp(9),
                color=(0.8, 0.8, 0.2, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{over_h:02d}:{over_m:02d}' if day['overtime'] > 0 else '-',
                size_hint_x=0.11,
                font_size=sp(9),
                color=(0.2, 0.9, 0.2, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{ded_h:02d}:{ded_m:02d}' if day['deduction'] > 0 else '-',
                size_hint_x=0.12,
                font_size=sp(9),
                color=(0.9, 0.2, 0.2, 1)
            ))
            
            self.att_report_container.add_widget(row)
    
    def _show_summary_report(self, stats):
        """نمایش گزارش تجمیعی"""
        # یک ردیف خلاصه
        row = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(2), size_hint_x=None, width=dp(500))
        with row.canvas.before:
            Color(0.15, 0.2, 0.3, 1)
            rect = Rectangle(pos=row.pos, size=row.size)
            row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                    size=lambda i, v, r=rect: setattr(r, 'size', v))
        
        items = [
            (f'کل کارکرد: {stats["total_hours"]:02d}:{stats["total_minutes"]:02d}', 0.2),
            (f'اضافه کار: {stats["overtime_hours"]:02d}:{stats["overtime_minutes"]:02d}', 0.2),
            (f'کسر کار: {stats["deduction_hours"]:02d}:{stats["deduction_minutes"]:02d}', 0.2),
            (f'امتیاز: {stats["points"]:.1f}', 0.2),
            (f'روزهای کاری: {stats["working_days"]}', 0.1),
            (f'غیبت: {stats.get("absent_days", 0)}', 0.1),
        ]
        
        for text, size in items:
            row.add_widget(RTLLabel(
                text=text,
                size_hint_x=size,
                font_size=sp(12),
                color=(1, 1, 1, 1)
            ))
        
        self.att_report_container.add_widget(row)
    
    def _show_monthly_report(self, records):
        """نمایش گزارش تفکیک ماهانه"""
        # گروه‌بندی بر اساس ماه
        monthly_data = {}
        for r in records:
            date = r.get('date', '')
            month_key = date[:7]  # YYYY/MM
            if month_key not in monthly_data:
                monthly_data[month_key] = []
            monthly_data[month_key].append(r)
        
        # هدر
        header = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(1), size_hint_x=None, width=dp(700))
        headers = ['ماه', 'تعداد روز کاری', 'کل کارکرد', 'اضافه کار', 'کسر کار', 'امتیاز']
        sizes = [0.16, 0.16, 0.16, 0.16, 0.18, 0.18]
        
        for i, (text, size) in enumerate(zip(headers, sizes)):
            header.add_widget(RTLLabel(
                text=text,
                size_hint_x=size,
                font_size=sp(9),
                bold=True,
                color=(0.4, 0.7, 1, 1)
            ))
        self.att_report_container.add_widget(header)
        
        # ردیف‌ها
        for idx, (month, month_records) in enumerate(monthly_data.items(), 1):
            # محاسبه آمار ماه
            month_stats = self._calculate_attendance_stats(month_records, 'تجمیعی')
            
            row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(1), size_hint_x=None, width=dp(700))
            with row.canvas.before:
                Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                rect = Rectangle(pos=row.pos, size=row.size)
                row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                        size=lambda i, v, r=rect: setattr(r, 'size', v))
            
            row.add_widget(RTLLabel(
                text=month,
                size_hint_x=0.16,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=str(month_stats.get('working_days', 0)),
                size_hint_x=0.16,
                font_size=sp(9),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{month_stats["total_hours"]:02d}:{month_stats["total_minutes"]:02d}',
                size_hint_x=0.16,
                font_size=sp(9),
                color=(0.8, 0.8, 0.2, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{month_stats["overtime_hours"]:02d}:{month_stats["overtime_minutes"]:02d}',
                size_hint_x=0.16,
                font_size=sp(9),
                color=(0.2, 0.9, 0.2, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{month_stats["deduction_hours"]:02d}:{month_stats["deduction_minutes"]:02d}',
                size_hint_x=0.18,
                font_size=sp(9),
                color=(0.9, 0.2, 0.2, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'{month_stats["points"]:.1f}',
                size_hint_x=0.18,
                font_size=sp(9),
                color=(0.8, 0.8, 0.2, 1)
            ))
            
            self.att_report_container.add_widget(row)
    
    def _show_error(self, message):
        """نمایش خطا"""
        self.att_report_container.clear_widgets()
        self.att_report_container.add_widget(RTLLabel(
            text=message,
            size_hint_y=None,
            height=dp(35),
            font_size=sp(14),
            color=(0.9, 0.2, 0.2, 1)
        ))
    
    def _export_attendance_excel(self, instance):
        """خروجی اکسل گزارش ورود و خروج"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # دریافت داده‌های فعلی از container
            # ابتدا داده‌ها رو دوباره محاسبه میکنیم تا مطمئن باشیم
            selected_month = self.month_combo.text
            if not selected_month:
                self.show_message('خطا', 'لطفاً یک ماه را انتخاب کنید')
                return
            
            report_type = self.report_type_combo.text
            export_type = self.export_type_combo.text
            
            # دریافت رکوردها
            records = AttendanceManager.load_attendance()
            user_id = self.current_user.get('id') if self.current_user else None
            user_name = self.current_user.get('name', '') if self.current_user else ''
            
            # فیلتر بر اساس کاربر و ماه
            year, month = selected_month.split('/')
            filtered = []
            for r in records:
                if user_id and r.get('user_id') != user_id:
                    continue
                r_date = r.get('date', '')
                if r_date.startswith(f"{year}/{month}"):
                    filtered.append(r)
            
            if not filtered:
                self.show_message('خطا', 'هیچ داده‌ای برای خروجی وجود ندارد')
                return
            
            # محاسبه آمار
            stats = self._calculate_attendance_stats(filtered, report_type)
            
            # ایجاد کتاب کار
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش ورود و خروج"
            ws.right_to_left = True
            
            # استایل‌ها
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            success_fill = PatternFill(start_color="235347", end_color="235347", fill_type="solid")
            fail_fill = PatternFill(start_color="78281F", end_color="78281F", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ========== اطلاعات هدر ==========
            ws.merge_cells('A1:K1')
            title_cell = ws.cell(row=1, column=1, value=f'گزارش ورود و خروج - {selected_month}')
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # اطلاعات کاربر
            ws.cell(row=2, column=1, value='نام کاربر:')
            ws.cell(row=2, column=2, value=user_name)
            ws.cell(row=2, column=1).font = Font(bold=True, size=11)
            ws.cell(row=2, column=3, value='نوع گزارش:')
            ws.cell(row=2, column=4, value=report_type)
            ws.cell(row=2, column=3).font = Font(bold=True, size=11)
            ws.cell(row=2, column=5, value='تاریخ تولید:')
            ws.cell(row=2, column=6, value=datetime.now().strftime('%Y/%m/%d %H:%M'))
            ws.cell(row=2, column=5).font = Font(bold=True, size=11)
            
            row_start = 4
            
            # ========== بر اساس نوع گزارش ==========
            if report_type == 'روزانه':
                # هدر جدول روزانه
                headers = ['ردیف', 'تاریخ', 'وضعیت', 'اولین ورود', 'آخرین خروج', 
                        'تعداد ورود', 'تعداد خروج', 'کارکرد (دقیقه)', 'کارکرد (ساعت)', 
                        'اضافه کار (دقیقه)', 'اضافه کار (ساعت)', 
                        'کسر کار (دقیقه)', 'کسر کار (ساعت)']
                col_widths = [6, 12, 10, 12, 12, 10, 10, 12, 12, 14, 14, 14, 14]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_start, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                
                # پر کردن داده‌ها
                for idx, day in enumerate(stats['daily_stats'], 1):
                    row = row_start + idx
                    total_h = day['total_minutes'] // 60
                    total_m = day['total_minutes'] % 60
                    over_h = day['overtime'] // 60
                    over_m = day['overtime'] % 60
                    ded_h = day['deduction'] // 60
                    ded_m = day['deduction'] % 60
                    
                    status = 'حضور' if day['total_minutes'] > 0 else 'غیبت'
                    
                    values = [
                        idx,
                        day['date'],
                        status,
                        day['first_check_in'],
                        day['last_check_out'],
                        day['check_in_count'],
                        day['check_out_count'],
                        day['total_minutes'],
                        f'{total_h:02d}:{total_m:02d}',
                        day['overtime'],
                        f'{over_h:02d}:{over_m:02d}' if day['overtime'] > 0 else '-',
                        day['deduction'],
                        f'{ded_h:02d}:{ded_m:02d}' if day['deduction'] > 0 else '-'
                    ]
                    
                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        if status == 'حضور':
                            cell.fill = success_fill
                        else:
                            cell.fill = fail_fill
                        cell.font = Font(color="FFFFFF")
                
                # جمع‌بندی
                summary_row = row_start + len(stats['daily_stats']) + 2
                ws.cell(row=summary_row, column=1, value='خلاصه:')
                ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="FFD700")
                ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
                
                ws.cell(row=summary_row, column=7, value=f'کل کارکرد: {stats["total_hours"]:02d}:{stats["total_minutes"]:02d}')
                ws.cell(row=summary_row, column=7).font = Font(bold=True, size=11, color="FFFFFF")
                
                ws.cell(row=summary_row + 1, column=7, value=f'اضافه کار: {stats["overtime_hours"]:02d}:{stats["overtime_minutes"]:02d}')
                ws.cell(row=summary_row + 1, column=7).font = Font(bold=True, size=11, color="00FF00")
                
                ws.cell(row=summary_row + 2, column=7, value=f'کسر کار: {stats["deduction_hours"]:02d}:{stats["deduction_minutes"]:02d}')
                ws.cell(row=summary_row + 2, column=7).font = Font(bold=True, size=11, color="FF0000")
                
                ws.cell(row=summary_row + 3, column=7, value=f'امتیاز: {stats["points"]:.1f}')
                ws.cell(row=summary_row + 3, column=7).font = Font(bold=True, size=11, color="FFD700")
                
                ws.cell(row=summary_row + 4, column=7, value=f'روزهای کاری: {stats["working_days"]}')
                ws.cell(row=summary_row + 4, column=7).font = Font(bold=True, size=11, color="FFFFFF")
                
                ws.cell(row=summary_row + 5, column=7, value=f'روزهای بدون کارکرد: {stats.get("absent_days", 0)}')
                ws.cell(row=summary_row + 5, column=7).font = Font(bold=True, size=11, color="FFA500")
                
                # تنظیم عرض ستون‌ها
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
            
            elif report_type == 'تجمیعی':
                # هدر جدول تجمیعی
                headers = ['عنوان', 'مقدار']
                col_widths = [25, 20]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_start, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                
                # داده‌های تجمیعی
                summary_data = [
                    ('تعداد روزهای کاری', stats['working_days']),
                    ('روزهای بدون کارکرد', stats.get('absent_days', 0)),
                    ('کل کارکرد (ساعت)', f'{stats["total_hours"]:02d}:{stats["total_minutes"]:02d}'),
                    ('کل کارکرد (دقیقه)', stats['total_minutes'] + (stats['total_hours'] * 60)),
                    ('اضافه کار (ساعت)', f'{stats["overtime_hours"]:02d}:{stats["overtime_minutes"]:02d}'),
                    ('اضافه کار (دقیقه)', stats['overtime_minutes'] + (stats['overtime_hours'] * 60)),
                    ('کسر کار (ساعت)', f'{stats["deduction_hours"]:02d}:{stats["deduction_minutes"]:02d}'),
                    ('کسر کار (دقیقه)', stats['deduction_minutes'] + (stats['deduction_hours'] * 60)),
                    ('امتیاز', f'{stats["points"]:.1f}'),
                ]
                
                for idx, (title, value) in enumerate(summary_data, 1):
                    row = row_start + idx
                    ws.cell(row=row, column=1, value=title)
                    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
                    ws.cell(row=row, column=2, value=value)
                    ws.cell(row=row, column=2).font = Font(size=11, color="FFFFFF")
                    ws.cell(row=row, column=1).border = thin_border
                    ws.cell(row=row, column=2).border = thin_border
                    ws.cell(row=row, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
                
                # تنظیم عرض ستون‌ها
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
            
            else:  # تفکیک ماهانه
                # گروه‌بندی بر اساس ماه
                monthly_data = {}
                for r in filtered:
                    date = r.get('date', '')
                    month_key = date[:7]
                    if month_key not in monthly_data:
                        monthly_data[month_key] = []
                    monthly_data[month_key].append(r)
                
                # هدر
                headers = ['ماه', 'تعداد روز کاری', 'کل کارکرد (ساعت)', 'اضافه کار (ساعت)', 'کسر کار (ساعت)', 'امتیاز']
                col_widths = [14, 16, 16, 16, 16, 14]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_start, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                
                # پر کردن داده‌ها
                for idx, (month, month_records) in enumerate(monthly_data.items(), 1):
                    month_stats = self._calculate_attendance_stats(month_records, 'تجمیعی')
                    row = row_start + idx
                    
                    values = [
                        month,
                        month_stats.get('working_days', 0),
                        f'{month_stats["total_hours"]:02d}:{month_stats["total_minutes"]:02d}',
                        f'{month_stats["overtime_hours"]:02d}:{month_stats["overtime_minutes"]:02d}',
                        f'{month_stats["deduction_hours"]:02d}:{month_stats["deduction_minutes"]:02d}',
                        f'{month_stats["points"]:.1f}'
                    ]
                    
                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        cell.fill = PatternFill(
                            start_color="235347" if idx % 2 == 0 else "1A3A2D",
                            end_color="235347" if idx % 2 == 0 else "1A3A2D",
                            fill_type="solid"
                        )
                        cell.font = Font(color="FFFFFF")
                
                # تنظیم عرض ستون‌ها
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== ذخیره فایل ==========
            from utils.storage import get_backup_path
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            today = get_today_jalali().replace('/', '-')
            timestamp = datetime.now().strftime("%H%M%S")
            filename = f'گزارش_ورود_خروج_{selected_month.replace("/", "-")}_{today}_{timestamp}.xlsx'
            filepath = os.path.join(export_dir, filename)
            
            wb.save(filepath)
            self.show_message('موفق', f'فایل اکسل گزارش ذخیره شد:\n{filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی اکسل: {str(e)}')
            import traceback
            traceback.print_exc()
    
    # ============================================================
    # گزارش مرخصی - کامل بر اساس تنظیمات
    # ============================================================

    def show_leave_report(self):
        """نمایش گزارش مرخصی - الگوبرداری از گزارش ورود و خروج"""
        content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(6))
        
        # ========== ردیف فیلترها ==========
        filter_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        
        # 1. کامبوباکس عاملین (از فایل users.json)
        users = self._load_users()
        user_names = ['همه'] + [u.get('name', '') for u in users if u.get('name')]
        self.leave_user_combo = PersianComboBox(
            text=user_names[0] if user_names else 'همه',
            values=user_names,
            height=dp(36),
            size_hint_x=0.34
        )
        self.leave_user_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.leave_user_combo.main_btn.color = (1, 1, 1, 1)
        self.leave_user_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.leave_user_combo)
        

        
        # 2. کامبوباکس ماه‌ها
        months = self._get_months_list()
        self.leave_month_combo = PersianComboBox(
            text=months[0] if months else '',
            values=months,
            height=dp(36),
            size_hint_x=0.33
        )
        self.leave_month_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.leave_month_combo.main_btn.color = (1, 1, 1, 1)
        self.leave_month_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.leave_month_combo)
        

        
        # 3. کامبوباکس نوع مرخصی (از تنظیمات)
        config = self._load_leave_config()
        leave_types = ['همه'] + config.get('leave_types', ['ساعتی', 'استحقاقی', 'استعلاجی', 'اضطراری', 'بدون حقوق'])
        self.leave_type_combo = PersianComboBox(
            text=leave_types[0] if leave_types else 'همه',
            values=leave_types,
            height=dp(36),
            size_hint_x=0.33
        )
        self.leave_type_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.leave_type_combo.main_btn.color = (1, 1, 1, 1)
        self.leave_type_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.leave_type_combo)
        

        
        content.add_widget(filter_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== ردیف نمایش سقف مرخصی ==========
        limit_row = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(6), padding=[dp(4), dp(2), dp(4), dp(2)])
        with limit_row.canvas.before:
            Color(0.12, 0.15, 0.2, 1)
            rect = Rectangle(pos=limit_row.pos, size=limit_row.size)
            limit_row.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
        
        self.annual_limit_display = RTLTextInput(
            text='سقف سالانه: ۰ روز',
            multiline=False,
            size_hint_x=0.30,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(13),
            disabled=True
        )
        self.annual_limit_display.bg_color = (0.08, 0.08, 0.08, 1)
        self.annual_limit_display.border_color = (0, 0, 0, 0)
        limit_row.add_widget(self.annual_limit_display)
        
        self.dynamic_limit_display = RTLTextInput(
            text='سقف مجاز (تا امروز): ۰ روز',
            multiline=False,
            size_hint_x=0.35,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(13),
            disabled=True
        )
        self.dynamic_limit_display.bg_color = (0.08, 0.08, 0.08, 1)
        self.dynamic_limit_display.border_color = (0, 0, 0, 0)
        limit_row.add_widget(self.dynamic_limit_display)
        
        self.monthly_hourly_limit_display = RTLTextInput(
            text='سقف ماهانه ساعتی: ۰۰:۰۰',
            multiline=False,
            size_hint_x=0.35,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(13),
            disabled=True
        )
        self.monthly_hourly_limit_display.bg_color = (0.08, 0.08, 0.08, 1)
        self.monthly_hourly_limit_display.border_color = (0, 0, 0, 0)
        limit_row.add_widget(self.monthly_hourly_limit_display)
        
        content.add_widget(limit_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== ردیف دکمه‌ها ==========
        btn_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        
        show_btn = PersianButton(
            text='نمایش',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.2, 0.6, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        show_btn.bind(on_press=self._show_leave_report_data)
        btn_row.add_widget(show_btn)
        
        excel_btn = PersianButton(
            text='خروجی اکسل',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        excel_btn.bind(on_press=self._export_leave_excel)
        btn_row.add_widget(excel_btn)
        
        refresh_btn = PersianButton(
            text='بروزرسانی',
            size_hint_x=0.34,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.3, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        refresh_btn.bind(on_press=self._refresh_leave_report)
        btn_row.add_widget(refresh_btn)
        
        content.add_widget(btn_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== لیست گزارش ==========
        self.leave_report_scroll = ScrollView(
            do_scroll_x=True,
            do_scroll_y=True,
            size_hint_y=1,
            scroll_type=['bars', 'content'],
            bar_width=dp(5)
        )
        
        self.leave_report_container = GridLayout(
            cols=1,
            spacing=dp(1),
            size_hint_y=None,
            size_hint_x=None,
            width=dp(800),
            padding=dp(2)
        )
        self.leave_report_container.bind(minimum_height=self.leave_report_container.setter('height'))
        self.leave_report_scroll.add_widget(self.leave_report_container)
        content.add_widget(self.leave_report_scroll)
        
        self.report_content.add_widget(content)
        
        # بارگذاری اولیه
        Clock.schedule_once(lambda dt: self._show_leave_report_data(None), 0.2)


    # ============================================================
    # توابع کمکی گزارش مرخصی
    # ============================================================

    def _load_users(self):
        """بارگذاری لیست کاربران - ابتدا از current_user استفاده کن"""
        # اگر current_user موجود است، آن را برگردان
        if hasattr(self, 'current_user') and self.current_user:
            # اگر فقط یک کاربر داریم، لیست یک عضوی برگردان
            return [self.current_user]
        
        # اگر current_user وجود نداشت، از فایل بخوان
        try:
            file_path = os.path.join(get_data_path(), 'users.json')
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('users', [])
        except Exception as e:
            print(f"خطا در بارگذاری کاربران: {e}")
        
        return []


    def _load_leave_config(self):
        """بارگذاری تنظیمات مرخصی"""
        try:
            file_path = os.path.join(get_data_path(), 'attendance_config.json')
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            print(f"خطا در بارگذاری تنظیمات مرخصی: {e}")
        return {
            'leave_types': ['ساعتی', 'استحقاقی', 'استعلاجی', 'اضطراری', 'بدون حقوق'],
            'annual_leave_limit': 30,
            'monthly_hourly_leave_limit': '12:30',
            'hourly_to_daily_ratio': 5
        }


    def _show_leave_report_data(self, instance):
        """نمایش داده‌های گزارش مرخصی با اعمال فیلترها"""
        try:
            # ========== دیباگ ==========
            self._debug_leave_data()
            
            self.leave_report_container.clear_widgets()
            
            # ========== خواندن داده‌ها ==========
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            
            if not os.path.exists(file_path):
                self.leave_report_container.add_widget(RTLLabel(
                    text='هیچ داده‌ای یافت نشد',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                self._update_leave_stats([], {})
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_requests = json.load(f)
            
            print(f"📊 تعداد کل مرخصی‌ها: {len(all_requests)}")
            
            # ========== بارگذاری کاربران ==========
            users = self._load_users()
            user_map = {}
            
            print(f"👤 تعداد کاربران بارگذاری شده: {len(users)}")
            
            for u in users:
                user_id = u.get('id')
                user_name = u.get('name', '') or u.get('username', '')
                if user_id is not None:
                    user_map[user_id] = user_name
                    print(f"   user_id={user_id} -> {user_name}")
            
            # ========== اعمال فیلترها ==========
            
            # فیلتر 1: کاربر
            selected_user = self.leave_user_combo.text
            print(f"🔍 فیلتر کاربر: {selected_user}")
            
            if selected_user != 'همه':
                selected_user_id = None
                for u in users:
                    user_name = u.get('name', '') or u.get('username', '')
                    if user_name == selected_user:
                        selected_user_id = u.get('id')
                        break
                
                if selected_user_id is not None:
                    all_requests = [r for r in all_requests if r.get('user_id') == selected_user_id]
                    print(f"   بعد از فیلتر کاربر: {len(all_requests)} مرخصی")
            
            # فیلتر 2: ماه
            selected_month = self.leave_month_combo.text
            print(f"🔍 فیلتر ماه: {selected_month}")
            if selected_month:
                all_requests = [r for r in all_requests if r.get('created_at', '').startswith(selected_month)]
                print(f"   بعد از فیلتر ماه: {len(all_requests)} مرخصی")
            
            # فیلتر 3: نوع مرخصی
            selected_type = self.leave_type_combo.text
            print(f"🔍 فیلتر نوع: {selected_type}")
            if selected_type != 'همه':
                all_requests = [r for r in all_requests if r.get('leave_type') == selected_type]
                print(f"   بعد از فیلتر نوع: {len(all_requests)} مرخصی")
            
            # ========== نمایش داده‌ها ==========
            if not all_requests:
                self.leave_report_container.add_widget(RTLLabel(
                    text='هیچ درخواست مرخصی با این فیلترها یافت نشد',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                self._update_leave_stats([], self._load_leave_config())
                return
            
            # ========== به‌روزرسانی سقف‌ها ==========
            self._update_leave_limits()
            
            # ========== هدر جدول ==========
            header = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(1), size_hint_x=None, width=dp(800))
            headers = ['ردیف', 'عامل', 'نوع مرخصی', 'مدت', 'تاریخ شروع', 'تاریخ پایان', 'وضعیت', 'تاریخ ثبت']
            sizes = [0.06, 0.16, 0.14, 0.10, 0.14, 0.14, 0.12, 0.14]
            
            for i, (text, size) in enumerate(zip(headers, sizes)):
                header.add_widget(RTLLabel(
                    text=text,
                    size_hint_x=size,
                    font_size=sp(9),
                    bold=True,
                    color=(0.4, 0.7, 1, 1)
                ))
            self.leave_report_container.add_widget(header)
            
            # ========== ردیف‌های داده ==========
            print(f"\n📝 ساخت ردیف‌های جدول ({len(all_requests)} ردیف):")
            
            for idx, req in enumerate(all_requests, 1):
                row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(1), size_hint_x=None, width=dp(800))
                with row.canvas.before:
                    Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                    rect = Rectangle(pos=row.pos, size=row.size)
                    row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                            size=lambda i, v, r=rect: setattr(r, 'size', v))
                
                status = req.get('status', '')
                status_color = (0.2, 0.9, 0.2, 1) if status == 'فعال' else (0.6, 0.6, 0.6, 1)
                
                # ✅ دریافت نام کاربر
                user_id = req.get('user_id')
                user_name = user_map.get(user_id, '')
                
                # اگر پیدا نشد، جستجوی مستقیم
                if not user_name:
                    for u in users:
                        if u.get('id') == user_id:
                            user_name = u.get('name', '') or u.get('username', '')
                            if not user_name:
                                user_name = f"کاربر {user_id}"
                            break
                    if user_name:
                        print(f"   ✅ ردیف {idx}: user_id={user_id} -> نام پیدا شد: {user_name}")
                    else:
                        print(f"   ❌ ردیف {idx}: user_id={user_id} -> نام پیدا نشد!")
                
                row.add_widget(RTLLabel(
                    text=str(idx),
                    size_hint_x=0.06,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=user_name or 'نامشخص',
                    size_hint_x=0.16,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=req.get('leave_type', ''),
                    size_hint_x=0.14,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=req.get('duration_display', ''),
                    size_hint_x=0.10,
                    font_size=sp(9),
                    color=(0.8, 0.8, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=req.get('start_date', ''),
                    size_hint_x=0.14,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=req.get('end_date', ''),
                    size_hint_x=0.14,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=status,
                    size_hint_x=0.12,
                    font_size=sp(9),
                    bold=True,
                    color=status_color
                ))
                row.add_widget(RTLLabel(
                    text=req.get('created_at', ''),
                    size_hint_x=0.14,
                    font_size=sp(9),
                    color=(0.6, 0.6, 0.6, 1)
                ))
                
                self.leave_report_container.add_widget(row)
            
            # ========== به‌روزرسانی آمار ==========
            self._update_leave_stats(all_requests, self._load_leave_config())
            
        except Exception as e:
            print(f"❌ خطا در نمایش گزارش مرخصی: {e}")
            import traceback
            traceback.print_exc()
            self.leave_report_container.clear_widgets()
            self.leave_report_container.add_widget(RTLLabel(
                text=f'خطا: {str(e)}',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(14),
                color=(0.9, 0.2, 0.2, 1)
            ))

    def _debug_leave_data(self):
        """بررسی داده‌های مرخصی و کاربران برای دیباگ"""
        print("\n" + "="*50)
        print("🔍 دیباگ گزارش مرخصی")
        print("="*50)
        
        # 1. بررسی فایل مرخصی
        file_path = os.path.join(get_data_path(), 'leave_requests.json')
        print(f"📁 مسیر فایل مرخصی: {file_path}")
        print(f"📁 فایل وجود دارد: {os.path.exists(file_path)}")
        
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as f:
                requests = json.load(f)
            print(f"📊 تعداد مرخصی‌ها: {len(requests)}")
            if requests:
                print(f"📊 اولین مرخصی: {requests[0]}")
                print(f"📊 user_id در مرخصی: {requests[0].get('user_id')}")
        else:
            print("❌ فایل مرخصی وجود ندارد!")
            return
        
        # 2. بررسی فایل کاربران
        users_path = os.path.join(get_data_path(), 'users.json')
        print(f"\n👤 مسیر فایل کاربران: {users_path}")
        print(f"👤 فایل وجود دارد: {os.path.exists(users_path)}")
        
        if os.path.exists(users_path):
            with open(users_path, 'r', encoding='utf-8') as f:
                users_data = json.load(f)
            users = users_data.get('users', [])
            print(f"👤 تعداد کاربران: {len(users)}")
            for u in users:
                print(f"👤 کاربر: id={u.get('id')}, name={u.get('name')}, username={u.get('username')}")
            
            # 3. بررسی تطابق
            print("\n🔗 بررسی تطابق user_id ها:")
            for req in requests[:3]:  # فقط ۳ تا اول
                user_id = req.get('user_id')
                print(f"   مرخصی: user_id={user_id}")
                found = False
                for u in users:
                    if u.get('id') == user_id:
                        print(f"      ✅ پیدا شد: {u.get('name') or u.get('username')}")
                        found = True
                        break
                if not found:
                    print(f"      ❌ پیدا نشد!")
        else:
            print("❌ فایل کاربران وجود ندارد!")
        
        print("="*50 + "\n")

    def _update_leave_limits(self):
        """به‌روزرسانی نمایش سقف‌های مرخصی"""
        try:
            config = self._load_leave_config()
            annual_limit = config.get('annual_leave_limit', 30)
            monthly_hourly_limit = config.get('monthly_hourly_leave_limit', '12:30')
            
            # محاسبه سقف داینامیک
            dynamic_limit = self._calculate_dynamic_leave_limit()
            
            self.annual_limit_display.text = f'سقف سالانه: {annual_limit} روز'
            self.dynamic_limit_display.text = f'سقف مجاز (تا امروز): {dynamic_limit} روز'
            self.monthly_hourly_limit_display.text = f'سقف ماهانه ساعتی: {monthly_hourly_limit}'
            
        except Exception as e:
            print(f"خطا در به‌روزرسانی سقف‌ها: {e}")


    def _calculate_dynamic_leave_limit(self):
        """محاسبه سقف داینامیک مرخصی"""
        try:
            config = self._load_leave_config()
            annual_limit = config.get('annual_leave_limit', 30)
            
            today = get_today_jalali()
            parts = today.split('/')
            if len(parts) != 3:
                return annual_limit
            
            current_month = int(parts[1])
            current_day = int(parts[2])
            
            # ماه‌های گذشته (فروردین = ۱)
            months_passed = current_month
            if current_day < 15 and months_passed > 0:
                months_passed -= 0.5
            
            monthly_limit = annual_limit / 12.0
            dynamic_limit = monthly_limit * months_passed
            
            import math
            dynamic_limit = math.floor(dynamic_limit)
            
            return max(0, dynamic_limit)
            
        except Exception as e:
            print(f"خطا در محاسبه سقف داینامیک: {e}")
            return 30


    def _update_leave_stats(self, requests, config):
        """به‌روزرسانی آمار مرخصی - در جای مناسب نمایش داده می‌شود"""
        # آمار در هدر جدول نمایش داده نمی‌شود، اما برای استفاده در آینده نگهداری می‌شود
        pass


    def _refresh_leave_report(self, instance):
        """بروزرسانی گزارش مرخصی"""
        try:
            # بروزرسانی سقف‌ها
            self._update_leave_limits()
            
            # نمایش مجدد داده‌ها
            self._show_leave_report_data(None)
            
            self.show_message('توجه', 'گزارش مرخصی بروزرسانی شد')
            
        except Exception as e:
            print(f"خطا در بروزرسانی: {e}")
            self.show_message('خطا', f'خطا در بروزرسانی: {str(e)}')


    def _export_leave_excel(self, instance):
        """خروجی اکسل گزارش مرخصی"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # ========== دریافت داده‌های فیلتر شده ==========
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            if not os.path.exists(file_path):
                self.show_message('خطا', 'هیچ داده‌ای برای خروجی وجود ندارد')
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_requests = json.load(f)
            
            # اعمال فیلترها (همانند نمایش)
            selected_user = self.leave_user_combo.text
            if selected_user != 'همه':
                users = self._load_users()
                selected_user_id = None
                for u in users:
                    if u.get('name') == selected_user:
                        selected_user_id = u.get('id')
                        break
                if selected_user_id is not None:
                    all_requests = [r for r in all_requests if r.get('user_id') == selected_user_id]
            
            selected_month = self.leave_month_combo.text
            if selected_month:
                all_requests = [r for r in all_requests if r.get('created_at', '').startswith(selected_month)]
            
            selected_type = self.leave_type_combo.text
            if selected_type != 'همه':
                all_requests = [r for r in all_requests if r.get('leave_type') == selected_type]
            
            if not all_requests:
                self.show_message('خطا', 'هیچ داده‌ای برای خروجی وجود ندارد')
                return
            
            # ========== ایجاد کتاب کار ==========
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش مرخصی"
            ws.right_to_left = True
            
            # استایل‌ها
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ========== اطلاعات هدر ==========
            ws.merge_cells('A1:H1')
            title_cell = ws.cell(row=1, column=1, value=f'گزارش مرخصی - {selected_month if selected_month else "همه ماه‌ها"}')
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # اطلاعات کاربر
            user_name = self.current_user.get('name', '') if self.current_user else ''
            ws.cell(row=2, column=1, value='کاربر:')
            ws.cell(row=2, column=2, value=user_name)
            ws.cell(row=2, column=1).font = Font(bold=True, size=11)
            ws.cell(row=2, column=4, value='تاریخ تولید:')
            ws.cell(row=2, column=5, value=datetime.now().strftime('%Y/%m/%d %H:%M'))
            ws.cell(row=2, column=4).font = Font(bold=True, size=11)
            
            row_start = 4
            
            # ========== هدر جدول ==========
            headers = ['ردیف', 'عامل', 'نوع مرخصی', 'مدت', 'تاریخ شروع', 'تاریخ پایان', 'وضعیت', 'تاریخ ثبت']
            col_widths = [8, 18, 16, 12, 16, 16, 14, 16]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_start, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # ========== پر کردن داده‌ها ==========
            users = self._load_users()
            user_map = {u.get('id'): u.get('name', '') for u in users}
            
            for idx, req in enumerate(all_requests, 1):
                row = row_start + idx
                user_name = user_map.get(req.get('user_id'), '')
                
                values = [
                    idx,
                    user_name,
                    req.get('leave_type', ''),
                    req.get('duration_display', ''),
                    req.get('start_date', ''),
                    req.get('end_date', ''),
                    req.get('status', ''),
                    req.get('created_at', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    if idx % 2 == 0:
                        cell.fill = PatternFill(start_color="235347", end_color="235347", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="1A3A2D", end_color="1A3A2D", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
            
            # ========== تنظیم عرض ستون‌ها ==========
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== ذخیره فایل ==========
            from utils.storage import get_backup_path
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            today = get_today_jalali().replace('/', '-')
            timestamp = datetime.now().strftime("%H%M%S")
            filename = f'گزارش_مرخصی_{today}_{timestamp}.xlsx'
            filepath = os.path.join(export_dir, filename)
            
            wb.save(filepath)
            self.show_message('موفق', f'فایل اکسل گزارش مرخصی ذخیره شد:\n{filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی اکسل: {str(e)}')
            import traceback
            traceback.print_exc()
    
    # ============================================================
    # گزارش ماموریت - کامل با الگوی گزارش‌های قبلی
    # ============================================================

    def show_mission_report(self):
        """نمایش گزارش ماموریت‌ها - الگوبرداری از گزارش ورود و خروج و مرخصی"""
        content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(6))
        
        # ========== ردیف فیلترها ==========
        filter_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        
        # 1. کامبوباکس عاملین
        users = self._load_users()
        user_names = ['همه'] + [u.get('name', '') for u in users if u.get('name')]
        self.mission_user_combo = PersianComboBox(
            text=user_names[0] if user_names else 'همه',
            values=user_names,
            height=dp(36),
            size_hint_x=0.33
        )
        self.mission_user_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.mission_user_combo.main_btn.color = (1, 1, 1, 1)
        self.mission_user_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.mission_user_combo)
        

        
        # 2. کامبوباکس ماه‌ها
        months = self._get_months_list()
        self.mission_month_combo = PersianComboBox(
            text=months[0] if months else '',
            values=months,
            height=dp(36),
            size_hint_x=0.33
        )
        self.mission_month_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.mission_month_combo.main_btn.color = (1, 1, 1, 1)
        self.mission_month_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.mission_month_combo)
        

        
        # 3. کامبوباکس وضعیت
        self.mission_status_combo = PersianComboBox(
            text='همه',
            values=['همه', '⏳ در انتظار', '✅ موفق', '❌ ناموفق'],
            height=dp(36),
            size_hint_x=0.34
        )
        self.mission_status_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.mission_status_combo.main_btn.color = (1, 1, 1, 1)
        self.mission_status_combo.main_btn.font_size = sp(14)
        filter_row.add_widget(self.mission_status_combo)
        

        
        content.add_widget(filter_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== ردیف دکمه‌ها ==========
        btn_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        
        show_btn = PersianButton(
            text='نمایش',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.2, 0.6, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        show_btn.bind(on_press=self._show_mission_report_data)
        btn_row.add_widget(show_btn)
        
        excel_btn = PersianButton(
            text='خروجی اکسل',
            size_hint_x=0.33,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        excel_btn.bind(on_press=self._export_mission_excel)
        btn_row.add_widget(excel_btn)
        
        refresh_btn = PersianButton(
            text='بروزرسانی',
            size_hint_x=0.34,
            size_hint_y=None,
            height=dp(38),
            background_color=(0.3, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        refresh_btn.bind(on_press=self._refresh_mission_report)
        btn_row.add_widget(refresh_btn)
        
        content.add_widget(btn_row)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== خلاصه آمار (نوار آبی) ==========
        from utils.persian_text import PersianLabel
        
        stats_layout = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(4), padding=[dp(6), dp(3), dp(6), dp(3)])
        with stats_layout.canvas.before:
            Color(0.15, 0.2, 0.3, 1)
            rect = Rectangle(pos=stats_layout.pos, size=stats_layout.size)
            stats_layout.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                            size=lambda i, v: setattr(rect, 'size', v))
        
        self.mission_total_label = PersianLabel(
            text='تعداد کل: ۰',
            size_hint_x=0.14,
            font_size=sp(11),
            color=(255, 255, 255, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_total_label)
        
        self.mission_pending_label = PersianLabel(
            text='در انتظار: ۰',
            size_hint_x=0.14,
            font_size=sp(11),
            color=(230, 204, 51, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_pending_label)
        
        self.mission_success_label = PersianLabel(
            text='موفق: ۰',
            size_hint_x=0.14,
            font_size=sp(11),
            color=(51, 230, 51, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_success_label)
        
        self.mission_failed_label = PersianLabel(
            text='ناموفق: ۰',
            size_hint_x=0.14,
            font_size=sp(11),
            color=(230, 51, 51, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_failed_label)
        
        # ✅ امتیاز ماموریت (مجموع)
        self.mission_score_label = PersianLabel(
            text='امتیاز: ۰',
            size_hint_x=0.14,
            font_size=sp(11),
            color=(102, 178, 255, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_score_label)
        
        # ✅ امتیاز کسب شده (فقط موفق‌ها)
        self.mission_earned_label = PersianLabel(
            text='کسب شده: ۰',
            size_hint_x=0.15,
            font_size=sp(11),
            color=(51, 230, 51, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_earned_label)
        
        self.mission_target_label = PersianLabel(
            text='هدف: ۰',
            size_hint_x=0.15,
            font_size=sp(11),
            color=(204, 204, 51, 255),
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.mission_target_label)
        
        content.add_widget(stats_layout)
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== لیست گزارش ==========
        self.mission_report_scroll = ScrollView(
            do_scroll_x=True,
            do_scroll_y=True,
            size_hint_y=1,
            scroll_type=['bars', 'content'],
            bar_width=dp(5)
        )
        
        self.mission_report_container = GridLayout(
            cols=1,
            spacing=dp(1),
            size_hint_y=None,
            size_hint_x=None,
            width=dp(950),
            padding=dp(2)
        )
        self.mission_report_container.bind(minimum_height=self.mission_report_container.setter('height'))
        self.mission_report_scroll.add_widget(self.mission_report_container)
        content.add_widget(self.mission_report_scroll)
        
        self.report_content.add_widget(content)
        
        # بارگذاری اولیه
        Clock.schedule_once(lambda dt: self._show_mission_report_data(None), 0.2)


    # ============================================================
    # توابع کمکی گزارش ماموریت
    # ============================================================

    def _show_mission_report_data(self, instance):
        """نمایش داده‌های گزارش ماموریت با اعمال فیلترها"""
        try:
            self.mission_report_container.clear_widgets()
            
            # ========== خواندن داده‌ها ==========
            file_path = os.path.join(get_data_path(), 'do_missions.json')
            
            if not os.path.exists(file_path):
                self.mission_report_container.add_widget(RTLLabel(
                    text='هیچ داده‌ای یافت نشد',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                self._update_mission_stats([])
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_missions = json.load(f)
            
            # ========== اعمال فیلترها ==========
            
            # فیلتر 1: کاربر (با agent_name)
            selected_user = self.mission_user_combo.text
            if selected_user != 'همه':
                all_missions = [m for m in all_missions if m.get('agent_name') == selected_user]
            
            # فیلتر 2: ماه (بر اساس تاریخ شروع)
            selected_month = self.mission_month_combo.text
            if selected_month:
                all_missions = [m for m in all_missions if m.get('start_date', '').startswith(selected_month)]
            
            # فیلتر 3: وضعیت
            selected_status = self.mission_status_combo.text
            if selected_status != 'همه':
                all_missions = [m for m in all_missions if m.get('status') == selected_status]
            
            # ========== نمایش داده‌ها ==========
            if not all_missions:
                self.mission_report_container.add_widget(RTLLabel(
                    text='هیچ ماموریتی با این فیلترها یافت نشد',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                self._update_mission_stats([])
                return
            
            # ========== هدر جدول ==========
            header = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(1), size_hint_x=None, width=dp(950))
            headers = ['ردیف', 'شناسه', 'نوع', 'روش', 'تاریخ شروع', 'تاریخ پایان', 'مدت', 'امتیاز', 'هدف', 'وضعیت', 'توضیحات']
            sizes = [0.05, 0.08, 0.08, 0.07, 0.10, 0.10, 0.06, 0.06, 0.12, 0.10, 0.18]
            
            for i, (text, size) in enumerate(zip(headers, sizes)):
                header.add_widget(RTLLabel(
                    text=text,
                    size_hint_x=size,
                    font_size=sp(9),
                    bold=True,
                    color=(0.4, 0.7, 1, 1)
                ))
            self.mission_report_container.add_widget(header)
            
            # ========== ردیف‌های داده ==========
            for idx, mission in enumerate(all_missions, 1):
                row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(1), size_hint_x=None, width=dp(950))
                with row.canvas.before:
                    Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                    rect = Rectangle(pos=row.pos, size=row.size)
                    row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                            size=lambda i, v, r=rect: setattr(r, 'size', v))
                
                status = mission.get('status', '')
                status_color = self._get_status_color(status)
                
                # نمایش هدف با کاما
                target = mission.get('target', 0)
                target_str = f"{target:,.0f}" if target else "۰"
                
                # توضیحات مختصر
                desc = mission.get('description', '')
                if len(desc) > 25:
                    desc = desc[:25] + '...'
                
                row.add_widget(RTLLabel(
                    text=str(idx),
                    size_hint_x=0.05,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('id', ''),
                    size_hint_x=0.08,
                    font_size=sp(9),
                    color=(0.6, 0.6, 0.6, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('type', ''),
                    size_hint_x=0.08,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('method', ''),
                    size_hint_x=0.07,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('start_date', ''),
                    size_hint_x=0.10,
                    font_size=sp(9),
                    color=(0.8, 0.8, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('end_date', ''),
                    size_hint_x=0.10,
                    font_size=sp(9),
                    color=(0.8, 0.8, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=str(mission.get('duration', '')),
                    size_hint_x=0.06,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=str(mission.get('score', '')),
                    size_hint_x=0.06,
                    font_size=sp(9),
                    color=(0.8, 0.8, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=target_str,
                    size_hint_x=0.12,
                    font_size=sp(9),
                    color=(0.2, 0.9, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=status,
                    size_hint_x=0.10,
                    font_size=sp(9),
                    bold=True,
                    color=status_color
                ))
                row.add_widget(RTLLabel(
                    text=desc,
                    size_hint_x=0.18,
                    font_size=sp(8),
                    color=(0.6, 0.6, 0.6, 1)
                ))
                
                self.mission_report_container.add_widget(row)
            
            # ========== به‌روزرسانی آمار ==========
            self._update_mission_stats(all_missions)
            
        except Exception as e:
            print(f"خطا در نمایش گزارش ماموریت: {e}")
            import traceback
            traceback.print_exc()
            self.mission_report_container.clear_widgets()
            self.mission_report_container.add_widget(RTLLabel(
                text=f'خطا: {str(e)}',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(14),
                color=(0.9, 0.2, 0.2, 1)
            ))


    def _get_status_color(self, status):
        """دریافت رنگ متناسب با وضعیت ماموریت"""
        if '✅ موفق' in status or status == 'موفق':
            return (0.2, 0.9, 0.2, 1)  # سبز
        elif '❌ ناموفق' in status or status == 'ناموفق':
            return (0.9, 0.2, 0.2, 1)  # قرمز
        elif '⏳ در انتظار' in status or status == 'در انتظار':
            return (0.9, 0.8, 0.2, 1)  # زرد
        else:
            return (0.6, 0.6, 0.6, 1)  # خاکستری


    def _update_mission_stats(self, missions):
        """به‌روزرسانی آمار ماموریت‌ها - شامل امتیاز ماموریت و امتیاز کسب شده"""
        try:
            total = len(missions)
            pending = len([m for m in missions if 'در انتظار' in m.get('status', '')])
            success = len([m for m in missions if 'موفق' in m.get('status', '') and 'ناموفق' not in m.get('status', '')])
            failed = len([m for m in missions if 'ناموفق' in m.get('status', '')])
            
            # ✅ امتیاز ماموریت: مجموع score همه ماموریت‌ها
            total_score = sum([m.get('score', 0) for m in missions])
            
            # ✅ امتیاز کسب شده: فقط ماموریت‌های موفق
            earned_score = 0
            for m in missions:
                status = m.get('status', '')
                if 'موفق' in status and 'ناموفق' not in status:
                    earned_score += m.get('score', 0)
            
            total_target = sum([m.get('target', 0) for m in missions])
            
            # ✅ به‌روزرسانی با set_text
            if hasattr(self, 'mission_total_label') and self.mission_total_label:
                self.mission_total_label.set_text(f'تعداد کل: {total}')
            
            if hasattr(self, 'mission_pending_label') and self.mission_pending_label:
                self.mission_pending_label.set_text(f'در انتظار: {pending}')
            
            if hasattr(self, 'mission_success_label') and self.mission_success_label:
                self.mission_success_label.set_text(f'موفق: {success}')
            
            if hasattr(self, 'mission_failed_label') and self.mission_failed_label:
                self.mission_failed_label.set_text(f'ناموفق: {failed}')
            
            # ✅ امتیاز ماموریت (مجموع)
            if hasattr(self, 'mission_score_label') and self.mission_score_label:
                self.mission_score_label.set_text(f'امتیاز: {total_score}')
            
            # ✅ امتیاز کسب شده (فقط موفق‌ها)
            if hasattr(self, 'mission_earned_label') and self.mission_earned_label:
                self.mission_earned_label.set_text(f'امتیاز کسب شده: {earned_score}')
            else:
                # اگر ویجت وجود نداشت، از همان mission_score_label استفاده کن
                if hasattr(self, 'mission_score_label') and self.mission_score_label:
                    self.mission_score_label.set_text(f'امتیاز: {total_score} | کسب شده: {earned_score}')
            
            if hasattr(self, 'mission_target_label') and self.mission_target_label:
                self.mission_target_label.set_text(f'هدف: {total_target:,.0f}')
            
        except Exception as e:
            print(f"خطا در به‌روزرسانی آمار ماموریت: {e}")
            import traceback
            traceback.print_exc()


    def _refresh_mission_report(self, instance):
        """بروزرسانی گزارش ماموریت"""
        try:
            self._show_mission_report_data(None)
            self.show_message('توجه', 'گزارش ماموریت بروزرسانی شد')
        except Exception as e:
            print(f"خطا در بروزرسانی: {e}")
            self.show_message('خطا', f'خطا در بروزرسانی: {str(e)}')


    def _export_mission_excel(self, instance):
        """خروجی اکسل گزارش ماموریت با امتیاز ماموریت و امتیاز کسب شده"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # ========== دریافت داده‌های فیلتر شده ==========
            file_path = os.path.join(get_data_path(), 'do_missions.json')
            if not os.path.exists(file_path):
                self.show_message('خطا', 'هیچ داده‌ای برای خروجی وجود ندارد')
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                all_missions = json.load(f)
            
            # اعمال فیلترها
            selected_user = self.mission_user_combo.text
            if selected_user != 'همه':
                all_missions = [m for m in all_missions if m.get('agent_name') == selected_user]
            
            selected_month = self.mission_month_combo.text
            if selected_month:
                all_missions = [m for m in all_missions if m.get('start_date', '').startswith(selected_month)]
            
            selected_status = self.mission_status_combo.text
            if selected_status != 'همه':
                all_missions = [m for m in all_missions if m.get('status') == selected_status]
            
            if not all_missions:
                self.show_message('خطا', 'هیچ داده‌ای برای خروجی وجود ندارد')
                return
            
            # ========== ایجاد کتاب کار ==========
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش ماموریت"
            ws.right_to_left = True
            
            # استایل‌ها
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ========== اطلاعات هدر ==========
            ws.merge_cells('A1:L1')
            title_cell = ws.cell(row=1, column=1, value=f'گزارش ماموریت‌ها - {selected_month if selected_month else "همه ماه‌ها"}')
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            user_name = self.current_user.get('name', '') if self.current_user else ''
            ws.cell(row=2, column=1, value='کاربر:')
            ws.cell(row=2, column=2, value=user_name)
            ws.cell(row=2, column=1).font = Font(bold=True, size=11)
            ws.cell(row=2, column=4, value='تاریخ تولید:')
            ws.cell(row=2, column=5, value=datetime.now().strftime('%Y/%m/%d %H:%M'))
            ws.cell(row=2, column=4).font = Font(bold=True, size=11)
            
            row_start = 4
            
            # ========== هدر جدول با ستون امتیاز کسب شده ==========
            headers = ['ردیف', 'شناسه', 'نوع', 'روش', 'تاریخ شروع', 'تاریخ پایان', 'مدت', 'امتیاز ماموریت', 'امتیاز کسب شده', 'هدف', 'وضعیت', 'توضیحات']
            col_widths = [6, 12, 12, 10, 14, 14, 8, 14, 14, 16, 14, 30]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_start, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # ========== پر کردن داده‌ها ==========
            for idx, mission in enumerate(all_missions, 1):
                row = row_start + idx
                
                status = mission.get('status', '')
                score = mission.get('score', 0)
                
                # ✅ محاسبه امتیاز کسب شده
                if 'موفق' in status and 'ناموفق' not in status:
                    earned_score = score
                else:
                    earned_score = 0
                
                values = [
                    idx,
                    mission.get('id', ''),
                    mission.get('type', ''),
                    mission.get('method', ''),
                    mission.get('start_date', ''),
                    mission.get('end_date', ''),
                    mission.get('duration', ''),
                    score,  # امتیاز ماموریت
                    earned_score,  # امتیاز کسب شده
                    mission.get('target', 0),
                    status,
                    mission.get('description', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    if idx % 2 == 0:
                        cell.fill = PatternFill(start_color="235347", end_color="235347", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="1A3A2D", end_color="1A3A2D", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
            
            # ========== جمع‌بندی در انتهای جدول ==========
            summary_row = row_start + len(all_missions) + 2
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
            ws.cell(row=summary_row, column=1, value='خلاصه:')
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="FFD700")
            
            # محاسبه جمع کل
            total_score = sum([m.get('score', 0) for m in all_missions])
            total_earned = 0
            for m in all_missions:
                status = m.get('status', '')
                if 'موفق' in status and 'ناموفق' not in status:
                    total_earned += m.get('score', 0)
            total_target = sum([m.get('target', 0) for m in all_missions])
            
            ws.cell(row=summary_row, column=7, value=f'امتیاز ماموریت: {total_score}')
            ws.cell(row=summary_row, column=7).font = Font(bold=True, size=11, color="FFFFFF")
            
            ws.cell(row=summary_row + 1, column=7, value=f'امتیاز کسب شده: {total_earned}')
            ws.cell(row=summary_row + 1, column=7).font = Font(bold=True, size=11, color="00FF00")
            
            ws.cell(row=summary_row + 2, column=7, value=f'هدف کل: {total_target:,.0f}')
            ws.cell(row=summary_row + 2, column=7).font = Font(bold=True, size=11, color="FFD700")
            
            # ========== تنظیم عرض ستون‌ها ==========
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ========== ذخیره فایل ==========
            from utils.storage import get_backup_path
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            today = get_today_jalali().replace('/', '-')
            timestamp = datetime.now().strftime("%H%M%S")
            filename = f'گزارش_ماموریت_{today}_{timestamp}.xlsx'
            filepath = os.path.join(export_dir, filename)
            
            wb.save(filepath)
            self.show_message('موفق', f'فایل اکسل گزارش ماموریت ذخیره شد:\n{filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی اکسل: {str(e)}')
            import traceback
            traceback.print_exc()
    
    # ============================================================
    # توابع کمکی
    # ============================================================
    
    def show_message(self, title, message):
        """نمایش پیام با پشتیبانی از متن طولانی"""
        try:
            from kivy.uix.boxlayout import BoxLayout
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            from kivy.metrics import dp, sp
            from kivy.graphics import Color, Rectangle
            from utils.rtl_widgets import PersianPopup, PersianButton
            
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(message)
                display_text = get_display(reshaped)
            except:
                display_text = message
            
            msg_label = Label(
                text=display_text,
                font_size=sp(15),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                halign='center',
                valign='top',
                text_size=(dp(550), None),
                font_name='fonts/Amiri-Regular.ttf',
                padding=(dp(20), 0, dp(0), 0)
            )
            
            lines = message.count('\n') + 1
            if len(message) > 50 and lines == 1:
                approx_chars_per_line = 35
                lines = (len(message) // approx_chars_per_line) + 1
            
            line_height = sp(15) + dp(6)
            label_height = max(lines * line_height + dp(25), dp(50))
            label_height = min(label_height, dp(490))
            
            msg_label.height = label_height
            msg_label.text_size = (dp(550), label_height)
            
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=None,
                height=label_height + dp(10),
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )
            scroll.add_widget(msg_label)
            content.add_widget(scroll)
            
            btn = PersianButton(
                text='باشه',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(16),
                color=(1, 1, 1, 1),
                background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            
            popup = PersianPopup(
                title=title,
                content=content,
                size_hint=(0.9, None),
                height=label_height + dp(250),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            print(f"خطا در نمایش پیام: {e}")
            import traceback
            traceback.print_exc()
            try:
                from error_handler import ErrorPopup
                ErrorPopup.show_error(str(message))
            except:
                pass
    
    def go_back(self, instance):
        pass
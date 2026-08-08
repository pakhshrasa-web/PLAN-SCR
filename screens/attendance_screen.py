# screens/attendance_screen.py
# ========== صفحه حضور و غیاب ==========
import os
import json
import traceback
from datetime import datetime, timedelta
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.metrics import dp, sp
from kivy.graphics import Color, Rectangle
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.uix.label import Label
from kivy.uix.checkbox import CheckBox

from utils.rtl_widgets import PersianButton, RTLLabel, PersianPopup, RTLTextInput, PersianComboBox
from utils.attendance_manager import AttendanceManager
from utils.jalali_date import get_today_jalali
from utils.file_manager import get_settings
from utils.storage import get_data_path
from error_handler import ErrorPopup

from screens.report_attendance_screen import ReportAttendanceScreen

class AttendanceScreen(Screen):
    """صفحه اصلی حضور و غیاب با ۴ تب"""
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        with self.canvas.before:
            Color(0.08, 0.08, 0.08, 1)
            self.bg_rect = Rectangle(pos=self.pos, size=self.size)
            self.bind(pos=self._update_bg, size=self._update_bg)
        
        self.current_user = None
        self.today = get_today_jalali()
        self.current_tab = 1
        
        # تنظیمات ساعات کاری از settings.json
        settings = get_settings()
        self.work_start_time = settings.get('work_start_time', '08:00')
        self.work_end_time = settings.get('work_end_time', '17:00')
        
        # اطمینان از عددی بودن daily_work_hours
        try:
            self.daily_work_hours = float(settings.get('min_daily_hours', 7))
        except (ValueError, TypeError):
            self.daily_work_hours = 7.0
            print("هشدار: مقدار min_daily_hours نامعتبر است، از مقدار پیش‌فرض ۷ استفاده می‌شود")
        
        # وضعیت‌های دکمه‌ها
        self.is_checked_in = False
        self.is_day_ended = False
        
        # لیست جفت‌های ورود/خروج
        self.entries = []
        
        self.build_ui()
        self.load_today_attendance()
        
        # اضافه کردن پرینت برای دیباگ
        print(f"DEBUG: daily_work_hours = {self.daily_work_hours}")
        print(f"DEBUG: work_start_time = {self.work_start_time}")
        print(f"DEBUG: work_end_time = {self.work_end_time}")
    
    def _update_bg(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size
    
    def set_user(self, user):
        """تنظیم کاربر جاری"""
        self.current_user = user
        self.load_today_attendance()
    
    def build_ui(self):
        """ساخت رابط کاربری"""
        self.clear_widgets()
        
        layout = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(5))
        
        # ========== هدر ==========
        header = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
        header.add_widget(RTLLabel(
            text='حضور و غیاب',
            font_size=sp(20),
            bold=True,
            color=(0.4, 0.8, 1, 1),
            size_hint_x=0.4
        ))
        header.add_widget(RTLLabel(
            text=self.today,
            font_size=sp(14),
            color=(0.6, 0.6, 0.6, 1),
            size_hint_x=0.3,
            halign='center'
        ))
        if self.current_user:
            header.add_widget(RTLLabel(
                text=self.current_user.get('name', ''),
                font_size=sp(14),
                color=(1, 1, 1, 1),
                size_hint_x=0.3,
                halign='left'
            ))
        layout.add_widget(header)
        
        # ========== نوار تب‌ها ==========
        tabs = BoxLayout(size_hint_y=None, height=dp(38), spacing=dp(3))
        
        self.tab1_btn = PersianButton(
            text='ورود و خروج',
            size_hint_x=0.25,
            background_color=(0.2, 0.5, 0.9, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        self.tab1_btn.bind(on_press=lambda x: self.switch_tab(1))
        tabs.add_widget(self.tab1_btn)
        
        self.tab2_btn = PersianButton(
            text='مرخصی',
            size_hint_x=0.25,
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        self.tab2_btn.bind(on_press=lambda x: self.switch_tab(2))
        tabs.add_widget(self.tab2_btn)
        
        self.tab3_btn = PersianButton(
            text='ماموریت',
            size_hint_x=0.25,
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        self.tab3_btn.bind(on_press=lambda x: self.switch_tab(3))
        tabs.add_widget(self.tab3_btn)
        
        self.tab4_btn = PersianButton(
            text='گزارشات',
            size_hint_x=0.25,
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        self.tab4_btn.bind(on_press=lambda x: self.switch_tab(4))
        tabs.add_widget(self.tab4_btn)
        
        layout.add_widget(tabs)
        
        # ========== محتوای تب‌ها ==========
        self.tab_content = BoxLayout(orientation='vertical', size_hint_y=1)
        layout.add_widget(self.tab_content)
        
        # ========== دکمه بازگشت ==========
        back_btn = PersianButton(
            text='بازگشت',
            size_hint_y=None,
            height=dp(40),
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(14)
        )
        back_btn.bind(on_press=self.go_back)
        layout.add_widget(back_btn)
        
        self.add_widget(layout)
        
        # نمایش تب اول
        self.switch_tab(1)
    
    def switch_tab(self, tab_num):
        """تغییر تب"""
        self.current_tab = tab_num
        
        # تغییر رنگ دکمه‌ها
        self.tab1_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 1 else (0.3, 0.3, 0.3, 1)
        self.tab2_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 2 else (0.3, 0.3, 0.3, 1)
        self.tab3_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 3 else (0.3, 0.3, 0.3, 1)
        self.tab4_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 4 else (0.3, 0.3, 0.3, 1)
        
        self.tab_content.clear_widgets()
        
        if tab_num == 1:
            self.show_check_in_out_tab()
        elif tab_num == 2:
            self.show_leave_tab()
        elif tab_num == 3:
            self.show_mission_tab()
        elif tab_num == 4:
            self.show_report_tab()
    
    # ============================================================
    # تب 1: ورود و خروج
    # ============================================================

    def show_check_in_out_tab(self):
        """نمایش تب ورود و خروج"""
        content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(10))
        
        # ========== دکمه‌های ورود/خروج ==========
        btn_row = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(10))
        
        self.check_in_btn = PersianButton(
            text='ثبت ورود',
            size_hint_x=0.5,
            background_color=(0.2, 0.7, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(18),
            bold=True
        )
        self.check_in_btn.bind(on_press=self.do_check_in)
        btn_row.add_widget(self.check_in_btn)
        
        self.check_out_btn = PersianButton(
            text='ثبت خروج',
            size_hint_x=0.5,
            background_color=(0.7, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(18),
            bold=True
        )
        self.check_out_btn.bind(on_press=self.do_check_out)
        btn_row.add_widget(self.check_out_btn)
        
        content.add_widget(btn_row)
        
        # ========== دکمه پایان کار و برچسب وضعیت ==========
        end_day_row = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
        
        self.end_day_btn = PersianButton(
            text='پایان کار',
            size_hint_x=0.5,
            background_color=(0.8, 0.2, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(16),
            bold=True
        )
        self.end_day_btn.bind(on_press=self.do_end_day)
        end_day_row.add_widget(self.end_day_btn)
        
        self.end_day_status_label = RTLLabel(
            text='',
            size_hint_x=0.5,
            font_size=sp(13),
            color=(0.2, 0.9, 0.2, 1),
            halign='right',
            valign='middle'
        )
        end_day_row.add_widget(self.end_day_status_label)
        
        content.add_widget(end_day_row)
        
        # ========== لیست ورود/خروج امروز ==========
        content.add_widget(RTLLabel(
            text='لیست ورود و خروج امروز:',
            size_hint_y=None,
            height=dp(30),
            font_size=sp(14),
            bold=True,
            color=(0.6, 0.8, 1, 1)
        ))
        
        self.entries_list = ScrollView(
            do_scroll_x=False,
            do_scroll_y=True,
            size_hint_y=0.5
        )
        self.entries_container = GridLayout(
            cols=1,
            spacing=dp(3),
            size_hint_y=None,
            padding=dp(5)
        )
        self.entries_container.bind(minimum_height=self.entries_container.setter('height'))
        self.entries_list.add_widget(self.entries_container)
        content.add_widget(self.entries_list)
        
        # ========== خلاصه کارکرد ==========
        self.summary_label = Label(
            text='خلاصه کارکرد: ۰:۰۰',
            size_hint_y=None,
            height=dp(40),
            font_size=sp(16),
            color=(1, 1, 1, 1),
            bold=True,
            halign='center',
            valign='middle',
            font_name='fonts/Amiri-Regular.ttf'
        )
        content.add_widget(self.summary_label)

        # ========== اضافه/کسر کارکرد ==========
        self.overtime_label = Label(
            text='اضافه/کسر کارکرد: ۰:۰۰',
            size_hint_y=None,
            height=dp(35),
            font_size=sp(14),
            color=(1, 0.8, 0.2, 1),
            halign='center',
            valign='middle',
            font_name='fonts/Amiri-Regular.ttf'
        )
        content.add_widget(self.overtime_label)
        
        self.tab_content.add_widget(content)
        
        self.update_buttons_state()
        self.update_entries_list()
        self.update_summary()


    def update_buttons_state(self):
        """به‌روزرسانی وضعیت دکمه‌ها"""
        # بررسی پایان کار
        if self.is_day_ended:
            self.check_in_btn.disabled = True
            self.check_out_btn.disabled = True
            self.end_day_btn.disabled = True
            self.check_in_btn.background_color = (0.3, 0.3, 0.3, 1)
            self.check_out_btn.background_color = (0.3, 0.3, 0.3, 1)
            self.end_day_btn.background_color = (0.3, 0.3, 0.3, 1)
            self.end_day_status_label.text = '✓ پایان کار ثبت شد'
            self.end_day_status_label.color = (0.2, 0.9, 0.2, 1)
            return
        
        # وضعیت عادی
        if self.is_checked_in:
            self.check_in_btn.disabled = True
            self.check_out_btn.disabled = False
            self.check_in_btn.background_color = (0.3, 0.3, 0.3, 1)
            self.check_out_btn.background_color = (0.7, 0.4, 0.2, 1)
        else:
            self.check_in_btn.disabled = False
            self.check_out_btn.disabled = True
            self.check_in_btn.background_color = (0.2, 0.7, 0.2, 1)
            self.check_out_btn.background_color = (0.3, 0.3, 0.3, 1)
        
        # بررسی پایان کار
        if len(self.entries) > 0:
            all_complete = all(e.get('check_out') is not None for e in self.entries)
            if all_complete and not self.is_checked_in:
                self.end_day_btn.disabled = False
                self.end_day_btn.background_color = (0.8, 0.2, 0.2, 1)
                self.end_day_status_label.text = ''
            else:
                self.end_day_btn.disabled = True
                self.end_day_btn.background_color = (0.3, 0.3, 0.3, 1)
                if self.is_checked_in:
                    self.end_day_status_label.text = 'هنوز خروج نزده‌اید'
                    self.end_day_status_label.color = (1, 0.8, 0.2, 1)
                else:
                    self.end_day_status_label.text = ''
        else:
            self.end_day_btn.disabled = True
            self.end_day_btn.background_color = (0.3, 0.3, 0.3, 1)
            self.end_day_status_label.text = 'ورودی ثبت نشده'
            self.end_day_status_label.color = (0.6, 0.6, 0.6, 1)


    def do_check_in(self, instance):
        """ثبت ورود"""
        if not self.current_user:
            ErrorPopup.show_error('کاربری انتخاب نشده است')
            return
        
        user_id = self.current_user.get('id')
        success, message = AttendanceManager.check_in(user_id)
        
        if success:
            now = datetime.now().strftime('%H:%M')
            self.entries.append({'check_in': now, 'check_out': None})
            self.is_checked_in = True
            
            self.update_buttons_state()
            self.update_entries_list()
            self.update_summary()
            self.show_message('موفق', message)
        else:
            ErrorPopup.show_error(message)
    
    def do_check_out(self, instance):
        """ثبت خروج"""
        if not self.current_user:
            ErrorPopup.show_error('کاربری انتخاب نشده است')
            return
        
        user_id = self.current_user.get('id')
        success, message = AttendanceManager.check_out(user_id)
        
        if success:
            now = datetime.now().strftime('%H:%M')
            for entry in reversed(self.entries):
                if entry.get('check_out') is None:
                    entry['check_out'] = now
                    break
            
            self.is_checked_in = False
            
            self.update_buttons_state()
            self.update_entries_list()
            self.update_summary()
            self.show_message('موفق', message)
        else:
            ErrorPopup.show_error(message)
    
    def do_end_day(self, instance):
        """ثبت پایان کار"""
        # بررسی کامل بودن همه جفت‌ها
        for entry in self.entries:
            if entry.get('check_out') is None:
                self.show_message('خطا', 'همه ورودی‌ها خروج ثبت نشده‌اند')
                return
        
        # ثبت پایان کار در دیتابیس
        if self.current_user:
            user_id = self.current_user.get('id')
            records = AttendanceManager.get_daily_report(user_id=user_id)
            attendance = AttendanceManager.load_attendance()
            
            for record in records:
                if record.get('date') == self.today:
                    for att in attendance:
                        if (att.get('user_id') == user_id and 
                            att.get('date') == self.today and
                            att.get('created_at') == record.get('created_at')):
                            att['is_day_ended'] = True
                            break
            
            AttendanceManager.save_attendance(attendance)
        
        self.is_day_ended = True
        
        self.update_buttons_state()
        self.update_entries_list()
        self.update_summary()
        
        self.show_message('موفق', 'پایان کار با موفقیت ثبت شد')
    
    def update_entries_list(self):
        """به‌روزرسانی لیست ورود/خروج"""
        self.entries_container.clear_widgets()
        
        if not self.entries:
            self.entries_container.add_widget(RTLLabel(
                text='هیچ ورودی ثبت نشده است',
                size_hint_y=None,
                height=dp(35),
                font_size=sp(14),
                color=(0.5, 0.5, 0.5, 1)
            ))
            return
        
        for idx, entry in enumerate(self.entries, 1):
            row = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(5))
            with row.canvas.before:
                Color(0.15, 0.15, 0.2, 1)
                rect = Rectangle(pos=row.pos, size=row.size)
                row.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            check_in = entry.get('check_in', '')
            check_out = entry.get('check_out', '')
            
            # محاسبه تفاضل
            diff_text = '-'
            if check_in and check_out:
                in_h, in_m = map(int, check_in.split(':'))
                out_h, out_m = map(int, check_out.split(':'))
                total_min = (out_h - in_h) * 60 + (out_m - in_m)
                diff_h = total_min // 60
                diff_m = total_min % 60
                diff_text = f'{diff_h:02d}:{diff_m:02d}'
            
            row.add_widget(RTLLabel(
                text=f'{idx}',
                size_hint_x=0.1,
                font_size=sp(13),
                color=(0.6, 0.6, 0.6, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'ورود: {check_in or "-"}',
                size_hint_x=0.3,
                font_size=sp(13),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'خروج: {check_out or "-"}',
                size_hint_x=0.3,
                font_size=sp(13),
                color=(1, 1, 1, 1)
            ))
            row.add_widget(RTLLabel(
                text=f'مدت: {diff_text}',
                size_hint_x=0.3,
                font_size=sp(13),
                color=(0.8, 0.8, 0.2, 1)
            ))
            
            self.entries_container.add_widget(row)
    
    def _update_summary_bg(self, instance, value):
        """به‌روزرسانی پس‌زمینه خلاصه"""
        if hasattr(self, 'summary_bg'):
            self.summary_bg.pos = instance.pos
            self.summary_bg.size = instance.size


    def update_summary(self):
        """به‌روزرسانی خلاصه کارکرد - برگشت به روش قبلی"""
        total_minutes = 0
        
        print(f"DEBUG: تعداد entries = {len(self.entries)}")
        
        for entry in self.entries:
            check_in = entry.get('check_in')
            check_out = entry.get('check_out')
            
            print(f"DEBUG: check_in={check_in}, check_out={check_out}")
            
            if check_in and check_out:
                try:
                    in_h, in_m = map(int, check_in.split(':'))
                    out_h, out_m = map(int, check_out.split(':'))
                    total_min = (out_h - in_h) * 60 + (out_m - in_m)
                    if total_min > 0:
                        total_minutes += total_min
                    print(f"DEBUG: مدت این جفت = {total_min} دقیقه")
                except (ValueError, AttributeError) as e:
                    print(f"خطا در محاسبه مدت: {e}")
                    continue
        
        # تبدیل به ساعت و دقیقه
        total_hours = total_minutes // 60
        total_mins = total_minutes % 60
        
        # استفاده از bidi دستی برای Label معمولی
        import arabic_reshaper
        from bidi.algorithm import get_display
        
        # خلاصه کارکرد
        summary_text = f'خلاصه کارکرد: {total_hours:02d}:{total_mins:02d}'
        reshaped_text = arabic_reshaper.reshape(summary_text)
        bidi_text = get_display(reshaped_text)
        self.summary_label.text = bidi_text
        print(f"DEBUG: خلاصه = {summary_text}")
        
        # محاسبه اضافه/کسر کارکرد
        daily_work_minutes = int(self.daily_work_hours * 60)
        diff_minutes = total_minutes - daily_work_minutes
        
        print(f"DEBUG: کارکرد مورد انتظار = {daily_work_minutes} دقیقه")
        print(f"DEBUG: اختلاف = {diff_minutes} دقیقه")
        
        diff_hours = abs(diff_minutes) // 60
        diff_mins = abs(diff_minutes) % 60
        
        if total_minutes == 0:
            overtime_text = 'اضافه/کسر کارکرد: ۰۰:۰۰'
            self.overtime_label.color = (0.6, 0.6, 0.6, 1)
        elif self.is_day_ended:
            if diff_minutes >= 0:
                overtime_text = f'اضافه کارکرد نهایی: {diff_hours:02d}:{diff_mins:02d}'
                self.overtime_label.color = (0.2, 0.9, 0.2, 1)
            else:
                overtime_text = f'کسر کارکرد نهایی: {diff_hours:02d}:{diff_mins:02d}'
                self.overtime_label.color = (0.9, 0.2, 0.2, 1)
        else:
            if diff_minutes >= 0:
                overtime_text = f'اضافه کارکرد: {diff_hours:02d}:{diff_mins:02d}'
                self.overtime_label.color = (0.2, 0.9, 0.2, 1)
            else:
                overtime_text = f'کسر کارکرد: {diff_hours:02d}:{diff_mins:02d}'
                self.overtime_label.color = (0.9, 0.2, 0.2, 1)
        
        reshaped_overtime = arabic_reshaper.reshape(overtime_text)
        bidi_overtime = get_display(reshaped_overtime)
        self.overtime_label.text = bidi_overtime
        print(f"DEBUG: برچسب اضافه/کسر = {overtime_text}")


    # ============================================================
    # تب 2: مرخصی - کامل
    # ============================================================

    def show_leave_tab(self):
        """نمایش تب مرخصی"""
        content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(6))
        
        # ========== ردیف اول: نوع مرخصی و میزان ==========
        row1 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6), size_hint_x=0.94, pos_hint={'center_x': 0.6})
        
        config = AttendanceManager.load_config()
        leave_types = config.get('leave_types', ['ساعتی', 'استحقاقی', 'استعلاجی', 'اضطراری', 'بدون حقوق'])
        
        self.leave_type_combo = PersianComboBox(
            text=leave_types[0] if leave_types else 'ساعتی',
            values=leave_types,
            height=dp(36),
            size_hint_x=0.56
        )
        self.leave_type_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.leave_type_combo.main_btn.color = (1, 1, 1, 1)
        self.leave_type_combo.main_btn.font_size = sp(16)
        self.leave_type_combo.bind(text=self._on_leave_type_change)
        row1.add_widget(self.leave_type_combo)
        
        row1.add_widget(RTLLabel(
            text='نوع مرخصی:',
            size_hint_x=0.06,
            font_size=sp(12),
            color=(1, 1, 1, 1)
        ))
        
        self.leave_duration_input = RTLTextInput(
            text='',
            multiline=False,
            size_hint_x=0.32,
            size_hint_y=None,
            height=dp(36),
            font_size=sp(20),
            hint_text='مدت مرخصی'
        )
        self.leave_duration_input.bg_color = (0.15, 0.15, 0.15, 1)
        self.leave_duration_input.border_color = (0.3, 0.3, 0.3, 1)
        self.leave_duration_input.border_color_focus = (0.2, 0.5, 0.9, 1)
        row1.add_widget(self.leave_duration_input)
        
        row1.add_widget(RTLLabel(
            text='میزان:',
            size_hint_x=0.06,
            font_size=sp(12),
            color=(1, 1, 1, 1)
        ))
        
        content.add_widget(row1)
        
        # ========== ردیف دوم: تاریخ شروع و پایان ==========
        row2 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6), size_hint_x=0.94, pos_hint={'center_x': 0.6})
        
        leave_start_date = RTLTextInput(
            text=get_today_jalali(),
            multiline=False,
            size_hint_x=0.46,
            size_hint_y=None,
            height=dp(36),
            font_size=sp(18),
            hint_text='1404/01/01'
        )
        leave_start_date.bg_color = (0.15, 0.15, 0.15, 1)
        leave_start_date.border_color = (0.3, 0.3, 0.3, 1)
        leave_start_date.border_color_focus = (0.2, 0.5, 0.9, 1)
        self.leave_start_date = leave_start_date
        row2.add_widget(leave_start_date)
        
        row2.add_widget(RTLLabel(
            text='تاریخ شروع:',
            size_hint_x=0.04,
            font_size=sp(12),
            color=(1, 1, 1, 1)
        ))
        
        leave_end_date = RTLTextInput(
            text='',
            multiline=False,
            size_hint_x=0.46,
            size_hint_y=None,
            height=dp(36),
            font_size=sp(18),
            hint_text='1404/01/01'
        )
        leave_end_date.bg_color = (0.15, 0.15, 0.15, 1)
        leave_end_date.border_color = (0.3, 0.3, 0.3, 1)
        leave_end_date.border_color_focus = (0.2, 0.5, 0.9, 1)
        self.leave_end_date = leave_end_date
        row2.add_widget(leave_end_date)
        
        row2.add_widget(RTLLabel(
            text='تاریخ پایان:',
            size_hint_x=0.04,
            font_size=sp(12),
            color=(1, 1, 1, 1)
        ))
        
        content.add_widget(row2)
        
        # ========== ردیف سوم: سقف مرخصی ==========
        row3 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6), padding=[dp(4), dp(4), dp(4), dp(4)])

        self.annual_limit_field = RTLTextInput(
            text='سقف سالانه: ۰ روز',
            multiline=False,
            size_hint_x=0.35,
            size_hint_y=None,
            height=dp(36),
            font_size=sp(15),
            disabled=True
        )
        row3.add_widget(self.annual_limit_field)

        self.dynamic_limit_field = RTLTextInput(
            text='سقف مجاز (تا امروز): ۰ روز',
            multiline=False,
            size_hint_x=0.40,
            size_hint_y=None,
            height=dp(36),
            font_size=sp(15),
            disabled=True
        )
        row3.add_widget(self.dynamic_limit_field)

        refresh_limit_btn = PersianButton(
            text='بروزرسانی',
            size_hint_x=0.20,
            size_hint_y=None,
            height=dp(36),
            background_color=(0.3, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        refresh_limit_btn.bind(on_press=self._refresh_limits)
        row3.add_widget(refresh_limit_btn)

        content.add_widget(row3)

        # ========== ردیف چهارم: دکمه‌ها ==========
        row4 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8))

        self.register_leave_btn = PersianButton(
            text='ثبت/ویرایش',
            size_hint_x=0.5,
            background_color=(0.2, 0.7, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(16),
            bold=True
        )
        self.register_leave_btn.bind(on_press=self._register_leave)
        row4.add_widget(self.register_leave_btn)

        calc_end_btn = PersianButton(
            text='محاسبه تاریخ پایان',
            size_hint_x=0.5,
            background_color=(0.3, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=sp(14)
        )
        calc_end_btn.bind(on_press=self._calculate_end_date)
        row4.add_widget(calc_end_btn)

        content.add_widget(row4)
        
        # ========== ردیف پنجم: آمار مرخصی (نوار آبی) ==========
        from utils.persian_text import PersianLabel

        stats_layout = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8), padding=[dp(8), dp(5), dp(8), dp(5)])
        with stats_layout.canvas.before:
            Color(0.15, 0.2, 0.3, 1)
            rect = Rectangle(pos=stats_layout.pos, size=stats_layout.size)
            stats_layout.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                            size=lambda i, v: setattr(rect, 'size', v))

        self.used_leave_label = PersianLabel(
            text='استفاده شده: ۰ روز',
            size_hint_x=0.33,
            font_size=sp(14),
            color=(255, 255, 255, 255),  # RGB
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.used_leave_label)

        self.remaining_leave_label = PersianLabel(
            text='مانده: ۰ روز',
            size_hint_x=0.33,
            font_size=sp(14),
            color=(51, 230, 51, 255),  # سبز RGB
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.remaining_leave_label)

        self.hourly_used_label = PersianLabel(
            text='ساعتی: ۰۰:۰۰',
            size_hint_x=0.34,
            font_size=sp(14),
            color=(204, 204, 51, 255),  # زرد RGB
            halign='center',
            valign='middle'
        )
        stats_layout.add_widget(self.hourly_used_label)

        content.add_widget(stats_layout)
        
        # ========== لیست درخواست‌های مرخصی ==========
        self.leave_list_scroll = ScrollView(
            do_scroll_x=False,
            do_scroll_y=True,
            size_hint_y=0.35
        )
        self.leave_list_container = GridLayout(
            cols=1,
            spacing=dp(0),
            size_hint_y=None,
            padding=dp(1)
        )
        self.leave_list_container.bind(minimum_height=self.leave_list_container.setter('height'))
        self.leave_list_scroll.add_widget(self.leave_list_container)
        content.add_widget(self.leave_list_scroll)
        
        # ========== دکمه خروجی ==========
        print_btn = PersianButton(
            text='صدور برگ درخواست مرخصی',
            size_hint_y=None,
            height=dp(45),
            background_color=(0.6, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(16),
            bold=True
        )
        print_btn.bind(on_press=self._print_leave_form)
        content.add_widget(print_btn)
        
        self.tab_content.add_widget(content)
        
        # بارگذاری لیست و آمار
        self._load_leave_requests()
        
        # آمار را با تأخیر به روزرسانی کن
        Clock.schedule_once(lambda dt: self._update_leave_stats(), 0.1)
        
        # مقداردهی اولیه برای ویرایش
        self.editing_leave_id = None


    def _refresh_limits(self, instance):
        """بروزرسانی سقف‌های مرخصی"""
        print("🔄 دکمه بروزرسانی سقف زده شد")
        self._update_leave_stats()
        self.show_message('توجه', 'سقف مرخصی بروزرسانی شد')


    def _ensure_fields_visible(self):
        """اطمینان از نمایش فیلدهای سقف"""
        if hasattr(self, 'annual_limit_field'):
            self.annual_limit_field.opacity = 1
        if hasattr(self, 'dynamic_limit_field'):
            self.dynamic_limit_field.opacity = 1


    def _update_leave_stats(self):
        """به روزرسانی آمار مرخصی با نمایش سقف سالانه و داینامیک"""
        try:
            print("="*50)
            print("=== شروع به روزرسانی آمار مرخصی ===")
            
            config = AttendanceManager.load_config()
            annual_limit = config.get('annual_leave_limit', 30)
            print(f"annual_limit: {annual_limit}")
            
            dynamic_limit = self._calculate_dynamic_leave_limit()
            print(f"dynamic_limit: {dynamic_limit}")
            
            # به روزرسانی فیلد سقف سالانه
            if hasattr(self, 'annual_limit_field') and self.annual_limit_field:
                Clock.schedule_once(lambda dt: setattr(self.annual_limit_field, 'text', f'سقف سالانه: {annual_limit} روز'), 0)
            
            # به روزرسانی فیلد سقف داینامیک
            if hasattr(self, 'dynamic_limit_field') and self.dynamic_limit_field:
                Clock.schedule_once(lambda dt: setattr(self.dynamic_limit_field, 'text', f'سقف مجاز (تا امروز): {dynamic_limit} روز'), 0)
            
            # ========== خواندن فایل مرخصی ==========
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            
            if not os.path.exists(file_path):
                if hasattr(self, 'used_leave_label') and self.used_leave_label:
                    Clock.schedule_once(lambda dt: self.used_leave_label.set_text('استفاده شده: ۰ روز'), 0)
                if hasattr(self, 'remaining_leave_label') and self.remaining_leave_label:
                    Clock.schedule_once(lambda dt: self.remaining_leave_label.set_text(f'مانده: {dynamic_limit} روز'), 0)
                    Clock.schedule_once(lambda dt: setattr(self.remaining_leave_label, 'color', (51, 230, 51, 255)), 0)
                if hasattr(self, 'hourly_used_label') and self.hourly_used_label:
                    Clock.schedule_once(lambda dt: self.hourly_used_label.set_text('ساعتی: ۰۰:۰۰'), 0)
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                requests = json.load(f)
            
            if self.current_user:
                user_id = self.current_user.get('id')
                requests = [r for r in requests if r.get('user_id') == user_id]
            
            hourly_to_daily_ratio = config.get('hourly_to_daily_ratio', 5)
            
            total_used_days = 0
            total_hourly_hours = 0
            
            for req in requests:
                leave_type = req.get('leave_type', '')
                status = req.get('status', '')
                duration = req.get('duration', 0)
                
                if status in ['فعال', 'پایان یافته']:
                    if leave_type == 'استحقاقی':
                        total_used_days += duration
                    elif leave_type == 'ساعتی':
                        total_hourly_hours += duration
                        converted_days = duration / hourly_to_daily_ratio
                        total_used_days += converted_days
            
            remaining_days = dynamic_limit - total_used_days
            
            print(f"total_used_days: {total_used_days}")
            print(f"total_hourly_hours: {total_hourly_hours}")
            print(f"remaining_days: {remaining_days}")
            
            # ✅ به روزرسانی UI با PersianLabel.set_text()
            if hasattr(self, 'used_leave_label') and self.used_leave_label:
                Clock.schedule_once(lambda dt: self.used_leave_label.set_text(f'استفاده شده: {total_used_days:.1f} روز'), 0)
            
            if hasattr(self, 'remaining_leave_label') and self.remaining_leave_label:
                Clock.schedule_once(lambda dt: self.remaining_leave_label.set_text(f'مانده: {remaining_days:.1f} روز'), 0)
                # تغییر رنگ
                if remaining_days <= 0:
                    Clock.schedule_once(lambda dt: setattr(self.remaining_leave_label, 'color', (230, 51, 51, 255)), 0)
                elif remaining_days <= 5:
                    Clock.schedule_once(lambda dt: setattr(self.remaining_leave_label, 'color', (230, 204, 51, 255)), 0)
                else:
                    Clock.schedule_once(lambda dt: setattr(self.remaining_leave_label, 'color', (51, 230, 51, 255)), 0)
            
            hours = int(total_hourly_hours)
            minutes = int((total_hourly_hours - hours) * 60)
            
            if hasattr(self, 'hourly_used_label') and self.hourly_used_label:
                Clock.schedule_once(lambda dt: self.hourly_used_label.set_text(f'ساعتی: {hours:02d}:{minutes:02d}'), 0)
            
            print("=== پایان به روزرسانی آمار مرخصی ===")
            print("="*50)
                    
        except Exception as e:
            print(f"خطا در به روزرسانی آمار مرخصی: {e}")
            import traceback
            traceback.print_exc()


    def _force_update_leave_ui(self):
        """اجبار به به روزرسانی UI نوار آبی"""
        try:
            # این تابع صرفاً برای اطمینان از به روزرسانی UI فراخوانی می شود
            if hasattr(self, 'used_leave_label') and self.used_leave_label:
                current_text = self.used_leave_label.text
                self.used_leave_label.text = current_text + ' '
                self.used_leave_label.text = current_text
                print(f"force update used_leave_label: {current_text}")
            
            if hasattr(self, 'remaining_leave_label') and self.remaining_leave_label:
                current_text = self.remaining_leave_label.text
                self.remaining_leave_label.text = current_text + ' '
                self.remaining_leave_label.text = current_text
                print(f"force update remaining_leave_label: {current_text}")
            
            if hasattr(self, 'hourly_used_label') and self.hourly_used_label:
                current_text = self.hourly_used_label.text
                self.hourly_used_label.text = current_text + ' '
                self.hourly_used_label.text = current_text
                print(f"force update hourly_used_label: {current_text}")
                
        except Exception as e:
            print(f"خطا در force_update: {e}")


    def _on_leave_type_change(self, instance, value):
        """تغییر نوع مرخصی - تغییر hint فیلد مدت"""
        if value == 'ساعتی':
            self.leave_duration_input.hint_text = 'ساعت (مثال: 04:30)'
        else:
            self.leave_duration_input.hint_text = 'روز (مثال: 2.5)'


    def _calculate_dynamic_leave_limit(self):
        """محاسبه سقف داینامیک مرخصی - نسخه اصلاح شده"""
        try:
            config = self._load_leave_config()
            annual_limit = config.get('annual_leave_limit', 30)
            
            today = get_today_jalali()
            parts = today.split('/')
            if len(parts) != 3:
                print(f"⚠️ تاریخ نامعتبر: {today}, استفاده از سقف سالانه")
                return annual_limit
            
            current_month = int(parts[1])
            current_day = int(parts[2])
            
            # ماه‌های گذشته (فروردین = ۱)
            months_passed = current_month
            if current_day < 15 and months_passed > 0:
                months_passed -= 0.5
            
            monthly_limit = annual_limit / 12.0
            dynamic_limit = monthly_limit * months_passed
            
            import math
            dynamic_limit = math.floor(dynamic_limit)
            
            print(f"✅ سقف داینامیک: {dynamic_limit} (annual={annual_limit}, months={months_passed})")
            
            return max(0, dynamic_limit)
            
        except Exception as e:
            print(f"❌ خطا در محاسبه سقف داینامیک: {e}")
            return 30


    def _get_used_leave_days(self):
        """محاسبه کل مرخصی استفاده شده (به روز)"""
        try:
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            if not os.path.exists(file_path):
                return 0
            
            with open(file_path, 'r', encoding='utf-8') as f:
                requests = json.load(f)
            
            if self.current_user:
                user_id = self.current_user.get('id')
                requests = [r for r in requests if r.get('user_id') == user_id]
            
            config = AttendanceManager.load_config()
            hourly_to_daily_ratio = config.get('hourly_to_daily_ratio', 5)
            
            total_days = 0
            for req in requests:
                if req.get('status') in ['فعال', 'پایان یافته']:
                    if req.get('leave_type') == 'استحقاقی':
                        total_days += req.get('duration', 0)
                    elif req.get('leave_type') == 'ساعتی':
                        total_days += req.get('duration', 0) / hourly_to_daily_ratio
            
            return total_days
            
        except Exception as e:
            print(f"خطا در محاسبه مرخصی استفاده شده: {e}")
            return 0


    def _calculate_end_date(self, instance):
        """محاسبه خودکار تاریخ پایان بر اساس تاریخ شروع و مدت"""
        try:
            from utils.jalali_date import convert_to_gregorian, to_jalali, validate_jalali_date, get_jalali_month_days
            import datetime as dt
            
            start_date = self.leave_start_date.text.strip()
            duration = self.leave_duration_input.text.strip()
            
            if not start_date or not duration:
                self.show_message('خطا', 'لطفاً تاریخ شروع و مدت را وارد کنید')
                return
            
            # اعتبارسنجی تاریخ شروع
            if not validate_jalali_date(start_date):
                self.show_message('خطا', 'فرمت تاریخ شروع نامعتبر است (مثال: 1404/01/01)')
                return
            
            parts = start_date.split('/')
            year, month, day = map(int, parts)
            
            leave_type = self.leave_type_combo.text
            
            if leave_type == 'ساعتی':
                end_date = start_date
            else:
                try:
                    days = float(duration)
                    if days <= 0:
                        self.show_message('خطا', 'مدت مرخصی باید بیشتر از صفر باشد')
                        return
                    
                    remaining_days = int(days - 1)
                    new_day = day + remaining_days
                    new_month = month
                    new_year = year
                    
                    while True:
                        days_in_month = get_jalali_month_days(new_year, new_month)
                        if new_day <= days_in_month:
                            break
                        new_day -= days_in_month
                        new_month += 1
                        if new_month > 12:
                            new_month = 1
                            new_year += 1
                    
                    end_date = f'{new_year}/{new_month:02d}/{new_day:02d}'
                    
                except ValueError:
                    self.show_message('خطا', 'مدت مرخصی باید عدد باشد')
                    return
            
            self.leave_end_date.text = end_date
            self.show_message('موفق', f'تاریخ پایان محاسبه شد: {end_date}')
            
        except Exception as e:
            self.show_message('خطا', f'خطا در محاسبه تاریخ پایان: {str(e)}')
            import traceback
            traceback.print_exc()


    def _register_leave(self, instance):
        """ثبت یا بروزرسانی درخواست مرخصی"""
        try:
            user_id = self.current_user.get('id') if self.current_user else None
            if not user_id:
                self.show_message('خطا', 'کاربری انتخاب نشده است')
                return
            
            leave_type = self.leave_type_combo.text
            duration = self.leave_duration_input.text.strip()
            start_date = self.leave_start_date.text.strip()
            end_date = self.leave_end_date.text.strip()
            
            if not duration:
                self.show_message('خطا', 'لطفاً مدت مرخصی را وارد کنید')
                return
            
            if not start_date:
                self.show_message('خطا', 'لطفاً تاریخ شروع را وارد کنید')
                return
            
            # تبدیل مدت به عدد
            try:
                if leave_type == 'ساعتی':
                    if ':' in duration:
                        h, m = map(int, duration.split(':'))
                        duration_float = h + (m / 60)
                    else:
                        duration_float = float(duration)
                else:
                    duration_float = float(duration)
            except ValueError:
                self.show_message('خطا', 'فرمت مدت نامعتبر است')
                return
            
            if duration_float <= 0:
                self.show_message('خطا', 'مدت مرخصی باید بیشتر از صفر باشد')
                return
            
            # اگر در حال ویرایش هستیم
            if self.editing_leave_id is not None:
                self._update_leave_request(self.editing_leave_id)
                return
            
            # بررسی سقف مرخصی برای ثبت جدید
            if leave_type in ['استحقاقی', 'ساعتی']:
                used_days = self._get_used_leave_days()
                
                if leave_type == 'ساعتی':
                    config = AttendanceManager.load_config()
                    hourly_to_daily_ratio = config.get('hourly_to_daily_ratio', 5)
                    new_days = duration_float / hourly_to_daily_ratio
                else:
                    new_days = duration_float
                
                dynamic_limit = self._calculate_dynamic_leave_limit()
                
                if used_days + new_days > dynamic_limit:
                    remaining = dynamic_limit - used_days
                    self.show_message('خطا', 
                        f'سقف مرخصی مجاز: {dynamic_limit} روز\n'
                        f'استفاده شده: {used_days:.1f} روز\n'
                        f'مانده قابل استفاده: {remaining:.1f} روز\n'
                        f'مدت درخواستی: {new_days:.1f} روز')
                    return
            
            # ذخیره درخواست جدید
            leave_data = {
                'user_id': user_id,
                'leave_type': leave_type,
                'duration': duration_float,
                'duration_display': duration,
                'start_date': start_date,
                'end_date': end_date if end_date else start_date,
                'status': 'فعال',
                'created_at': get_today_jalali(),
                'is_converted': False
            }
            
            self._save_leave_request(leave_data)
            
            # پاک کردن فرم
            self._clear_leave_form()
            
            # به‌روزرسانی
            self._load_leave_requests()
            self._update_leave_stats()
            
            self.show_message('موفق', 'درخواست مرخصی با موفقیت ثبت شد')
            
        except Exception as e:
            self.show_message('خطا', f'خطا در ثبت مرخصی: {str(e)}')


    def _save_leave_request(self, leave_data):
        """ذخیره درخواست مرخصی در فایل"""
        try:
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    requests = json.load(f)
            else:
                requests = []
            
            max_id = max([r.get('id', 0) for r in requests], default=0)
            leave_data['id'] = max_id + 1
            
            requests.append(leave_data)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(requests, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"خطا در ذخیره مرخصی: {e}")


    def _update_leave_request(self, req_id):
        """بروزرسانی درخواست مرخصی موجود"""
        try:
            leave_type = self.leave_type_combo.text
            duration = self.leave_duration_input.text.strip()
            start_date = self.leave_start_date.text.strip()
            end_date = self.leave_end_date.text.strip()
            
            if not duration or not start_date:
                self.show_message('خطا', 'لطفاً فیلدهای ضروری را پر کنید')
                return
            
            try:
                if leave_type == 'ساعتی':
                    if ':' in duration:
                        h, m = map(int, duration.split(':'))
                        duration_float = h + (m / 60)
                    else:
                        duration_float = float(duration)
                else:
                    duration_float = float(duration)
            except ValueError:
                self.show_message('خطا', 'فرمت مدت نامعتبر است')
                return
            
            # بررسی سقف مرخصی برای ویرایش
            if leave_type in ['استحقاقی', 'ساعتی']:
                file_path = os.path.join(get_data_path(), 'leave_requests.json')
                with open(file_path, 'r', encoding='utf-8') as f:
                    all_requests = json.load(f)
                
                used_days = 0
                for req in all_requests:
                    if req.get('user_id') == self.current_user.get('id') and req.get('id') != req_id:
                        if req.get('status') in ['فعال', 'پایان یافته']:
                            if req.get('leave_type') == 'استحقاقی':
                                used_days += req.get('duration', 0)
                            elif req.get('leave_type') == 'ساعتی':
                                used_days += req.get('duration', 0) / 5
                
                if leave_type == 'ساعتی':
                    config = AttendanceManager.load_config()
                    hourly_to_daily_ratio = config.get('hourly_to_daily_ratio', 5)
                    new_days = duration_float / hourly_to_daily_ratio
                else:
                    new_days = duration_float
                
                dynamic_limit = self._calculate_dynamic_leave_limit()
                
                if used_days + new_days > dynamic_limit:
                    remaining = dynamic_limit - used_days
                    self.show_message('خطا', 
                        f'سقف مرخصی مجاز: {dynamic_limit} روز\n'
                        f'استفاده شده: {used_days:.1f} روز\n'
                        f'مانده قابل استفاده: {remaining:.1f} روز\n'
                        f'مدت درخواستی: {new_days:.1f} روز')
                    return
            
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            with open(file_path, 'r', encoding='utf-8') as f:
                requests = json.load(f)
            
            for req in requests:
                if req.get('id') == req_id:
                    req['leave_type'] = leave_type
                    req['duration'] = duration_float
                    req['duration_display'] = duration
                    req['start_date'] = start_date
                    req['end_date'] = end_date if end_date else start_date
                    break
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(requests, f, ensure_ascii=False, indent=2)
            
            self._clear_leave_form()
            self._load_leave_requests()
            self._update_leave_stats()
            
            self.show_message('موفق', 'درخواست مرخصی با موفقیت بروزرسانی شد')
            
        except Exception as e:
            self.show_message('خطا', f'خطا در بروزرسانی: {str(e)}')


    def _clear_leave_form(self):
        """پاک کردن فرم و برگشت به حالت ثبت"""
        self.leave_duration_input.text = ''
        self.leave_end_date.text = ''
        self.editing_leave_id = None


    def _edit_leave_request(self, instance):
        """آماده‌سازی فرم برای ویرایش"""
        try:
            req_id = instance.req_id
            req_data = instance.req_data
            
            self.leave_type_combo.text = req_data.get('leave_type', '')
            self.leave_duration_input.text = req_data.get('duration_display', '')
            self.leave_start_date.text = req_data.get('start_date', '')
            self.leave_end_date.text = req_data.get('end_date', '')
            
            self.editing_leave_id = req_id
            
            self.show_message('توجه', f'درخواست شماره {req_id} برای ویرایش آماده شد. تغییرات را اعمال کنید')
            
        except Exception as e:
            self.show_message('خطا', f'خطا در ویرایش: {str(e)}')


    def _delete_leave_request(self, instance):
        """حذف درخواست مرخصی"""
        try:
            req_id = instance.req_id
            
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text=f'آیا از حذف درخواست مرخصی شماره {req_id} مطمئن هستید؟',
                size_hint_y=None,
                height=dp(40),
                font_size=sp(16),
                color=(1, 1, 1, 1)
            ))
            
            btn_layout = BoxLayout(spacing=dp(10), size_hint_y=None, height=dp(45))
            yes_btn = PersianButton(
                text='بله، حذف شود',
                size_hint_y=None,
                height=dp(40),
                color=(1, 1, 1, 1),
                background_color=(0.8, 0.2, 0.2, 1),
                font_size=sp(14)
            )
            no_btn = PersianButton(
                text='خیر، انصراف',
                size_hint_y=None,
                height=dp(40),
                color=(1, 1, 1, 1),
                background_color=(0.3, 0.3, 0.3, 1),
                font_size=sp(14)
            )
            btn_layout.add_widget(yes_btn)
            btn_layout.add_widget(no_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='تأیید حذف',
                content=content,
                size_hint=(0.8, 0.35),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            
            def do_delete(inst):
                popup.dismiss()
                self._perform_delete_leave(req_id)
            
            def cancel_delete(inst):
                popup.dismiss()
            
            yes_btn.bind(on_press=do_delete)
            no_btn.bind(on_press=cancel_delete)
            popup.open()
            
        except Exception as e:
            self.show_message('خطا', f'خطا در حذف: {str(e)}')


    def _perform_delete_leave(self, req_id):
        """اجرای حذف درخواست مرخصی"""
        try:
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            if not os.path.exists(file_path):
                self.show_message('خطا', 'فایل مرخصی‌ها یافت نشد')
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                requests = json.load(f)
            
            requests = [r for r in requests if r.get('id') != req_id]
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(requests, f, ensure_ascii=False, indent=2)
            
            self._load_leave_requests()
            self._update_leave_stats()
            
            self.show_message('موفق', 'درخواست مرخصی با موفقیت حذف شد')
            
        except Exception as e:
            self.show_message('خطا', f'خطا در حذف: {str(e)}')


    def _load_leave_requests(self):
        """بارگذاری لیست درخواست‌های مرخصی با غیرفعال کردن تاریخ‌های گذشته"""
        try:
            self.leave_list_container.clear_widgets()
            
            file_path = os.path.join(get_data_path(), 'leave_requests.json')
            if not os.path.exists(file_path):
                self.leave_list_container.add_widget(RTLLabel(
                    text='هیچ درخواست مرخصی ثبت نشده است',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                requests = json.load(f)
            
            if self.current_user:
                user_id = self.current_user.get('id')
                requests = [r for r in requests if r.get('user_id') == user_id]
            
            requests = sorted(requests, key=lambda x: x.get('created_at', ''), reverse=True)
            
            if not requests:
                self.leave_list_container.add_widget(RTLLabel(
                    text='هیچ درخواست مرخصی ثبت نشده است',
                    size_hint_y=None,
                    height=dp(35),
                    font_size=sp(14),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                return
            
            today = get_today_jalali()
            
            # ========== هدر جدول ==========
            header_row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(0))
            
            columns = [
                ('انتخاب', 0.08),
                ('ردیف', 0.05),
                ('تاریخ ثبت', 0.12),
                ('شروع', 0.10),
                ('پایان', 0.10),
                ('مدت', 0.08),
                ('نوع', 0.10),
                ('وضعیت', 0.12),
                ('عملیات', 0.25)
            ]
            
            for text, width in columns:
                header_row.add_widget(RTLLabel(
                    text=text,
                    size_hint_x=width,
                    font_size=sp(10),
                    bold=True,
                    color=(0.4, 0.7, 1, 1),
                    halign='center'
                ))
            self.leave_list_container.add_widget(header_row)
            
            # ========== ردیف‌های داده با چک‌باکس ==========
            self.selected_leave_id = None
            
            for idx, req in enumerate(requests, 1):
                row = BoxLayout(size_hint_y=None, height=dp(34), spacing=dp(0))
                
                with row.canvas.before:
                    Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                    rect = Rectangle(pos=row.pos, size=row.size)
                    row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                            size=lambda i, v, r=rect: setattr(r, 'size', v))
                
                status = req.get('status', 'فعال')
                end_date = req.get('end_date', '')
                
                is_expired = False
                if status == 'فعال' and end_date and end_date < today:
                    is_expired = True
                    status_display = 'منقضی شده'
                    status_color = (0.6, 0.6, 0.6, 1)
                else:
                    status_display = status
                    status_color = (0.2, 0.9, 0.2, 1) if status == 'فعال' else (0.6, 0.6, 0.6, 1)
                
                req_id = req.get('id')
                
                cb = CheckBox(
                    active=False,
                    size_hint_x=0.08,
                    size_hint_y=None,
                    height=dp(30),
                    color=(0.4, 0.7, 1, 1),
                    disabled=(status != 'فعال' or is_expired)
                )
                cb.req_id = req_id
                cb.req_data = req
                cb.bind(active=self._on_select_leave)
                row.add_widget(cb)
                
                row_field = RTLTextInput(
                    text=str(idx),
                    size_hint_x=0.05,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                row_field._hidden_input.disabled = True
                row_field.bg_color = (0, 0, 0, 0)
                row_field.border_color = (0, 0, 0, 0)
                row.add_widget(row_field)
                
                created_field = RTLTextInput(
                    text=req.get('created_at', ''),
                    size_hint_x=0.12,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                created_field._hidden_input.disabled = True
                created_field.bg_color = (0, 0, 0, 0)
                created_field.border_color = (0, 0, 0, 0)
                row.add_widget(created_field)
                
                start_field = RTLTextInput(
                    text=req.get('start_date', ''),
                    size_hint_x=0.10,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                start_field._hidden_input.disabled = True
                start_field.bg_color = (0, 0, 0, 0)
                start_field.border_color = (0, 0, 0, 0)
                row.add_widget(start_field)
                
                end_field = RTLTextInput(
                    text=req.get('end_date', ''),
                    size_hint_x=0.10,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                end_field._hidden_input.disabled = True
                end_field.bg_color = (0, 0, 0, 0)
                end_field.border_color = (0, 0, 0, 0)
                row.add_widget(end_field)
                
                duration_field = RTLTextInput(
                    text=req.get('duration_display', ''),
                    size_hint_x=0.08,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                duration_field._hidden_input.disabled = True
                duration_field.bg_color = (0, 0, 0, 0)
                duration_field.border_color = (0, 0, 0, 0)
                row.add_widget(duration_field)
                
                type_field = RTLTextInput(
                    text=req.get('leave_type', ''),
                    size_hint_x=0.10,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                type_field._hidden_input.disabled = True
                type_field.bg_color = (0, 0, 0, 0)
                type_field.border_color = (0, 0, 0, 0)
                row.add_widget(type_field)
                
                status_field = RTLTextInput(
                    text=status_display,
                    size_hint_x=0.12,
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(10)
                )
                status_field._hidden_input.disabled = True
                status_field.bg_color = (0, 0, 0, 0)
                status_field.border_color = (0, 0, 0, 0)
                status_field._hidden_input.foreground_color = status_color
                row.add_widget(status_field)
                
                action_box = BoxLayout(size_hint_x=0.25, spacing=dp(2))
                
                if status == 'فعال' and not is_expired:
                    edit_btn = PersianButton(
                        text='ویرایش',
                        size_hint_x=0.5,
                        size_hint_y=None,
                        height=dp(28),
                        background_color=(0.3, 0.5, 0.8, 1),
                        color=(1, 1, 1, 1),
                        font_size=sp(9)
                    )
                    edit_btn.req_id = req.get('id')
                    edit_btn.req_data = req
                    edit_btn.bind(on_press=self._edit_leave_request)
                    action_box.add_widget(edit_btn)
                    
                    delete_btn = PersianButton(
                        text='حذف',
                        size_hint_x=0.5,
                        size_hint_y=None,
                        height=dp(28),
                        background_color=(0.8, 0.2, 0.2, 1),
                        color=(1, 1, 1, 1),
                        font_size=sp(9)
                    )
                    delete_btn.req_id = req.get('id')
                    delete_btn.bind(on_press=self._delete_leave_request)
                    action_box.add_widget(delete_btn)
                else:
                    action_box.add_widget(RTLLabel(
                        text='منقضی' if is_expired else '-',
                        size_hint_x=1,
                        font_size=sp(10),
                        color=(0.5, 0.5, 0.5, 1),
                        halign='center'
                    ))
                
                row.add_widget(action_box)
                self.leave_list_container.add_widget(row)
                
        except Exception as e:
            print(f"خطا در بارگذاری مرخصی‌ها: {e}")
            import traceback
            traceback.print_exc()


    def _on_select_leave(self, instance, value):
        """انتخاب مرخصی برای خروجی"""
        if value:
            row = instance.parent
            if row and isinstance(row, BoxLayout):
                for child in self.leave_list_container.children:
                    if isinstance(child, BoxLayout) and child != row:
                        for widget in child.children:
                            if isinstance(widget, CheckBox):
                                widget.active = False
            
            self.selected_leave_id = instance.req_id
            self.selected_leave_data = instance.req_data
        else:
            self.selected_leave_id = None
            self.selected_leave_data = None


    def _print_leave_form(self, instance):
        """خروجی برگ مرخصی - فقط برای مرخصی انتخاب شده"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import arabic_reshaper
            from bidi.algorithm import get_display
            
            if not hasattr(self, 'selected_leave_id') or self.selected_leave_id is None:
                self.show_message('خطا', 'لطفاً ابتدا یک درخواست مرخصی را برای خروجی انتخاب کنید')
                return
            
            req = self.selected_leave_data
            
            font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Amiri-Regular.ttf')
            if not os.path.exists(font_path):
                font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Vazirmatn-Regular.ttf')
            
            try:
                font_title = ImageFont.truetype(font_path, 28)
                font_normal = ImageFont.truetype(font_path, 18)
                font_small = ImageFont.truetype(font_path, 15)
            except:
                font_title = ImageFont.load_default()
                font_normal = ImageFont.load_default()
                font_small = ImageFont.load_default()
            
            def fix_text(text):
                if not text:
                    return ''
                try:
                    reshaped = arabic_reshaper.reshape(str(text))
                    return get_display(reshaped)
                except:
                    return str(text)
            
            width = 900
            height = 700
            img = Image.new('RGB', (width, height), color=(255, 255, 255))
            draw = ImageDraw.Draw(img)
            
            title = 'برگ درخواست مرخصی'
            title_width = draw.textlength(fix_text(title), font=font_title)
            draw.text(((width - title_width) // 2, 25), fix_text(title), fill=(0, 0, 0), font=font_title)
            
            draw.line([(50, 65), (width-50, 65)], fill=(0, 0, 0), width=2)
            
            col1_x = 50
            col2_x = width // 2 + 30
            y = 100
            line_height = 38
            
            user_name = self.current_user.get('name', '') if self.current_user else ''
            user_role = self.current_user.get('role', '') if self.current_user else ''
            
            info_left = [
                f'نام و نام خانوادگی: {user_name}',
                f'سمت/واحد: {user_role}',
                f'نوع مرخصی: {req.get("leave_type", "")}',
                f'مدت مرخصی: {req.get("duration_display", "")}',
            ]
            
            for line in info_left:
                text = fix_text(line)
                text_width = draw.textlength(text, font=font_normal)
                draw.text((col1_x + 350 - text_width, y), text, fill=(0, 0, 0), font=font_normal)
                y += line_height
            
            y = 100
            info_right = [
                f'تاریخ درخواست: {req.get("created_at", "")}',
                f'تاریخ شروع: {req.get("start_date", "")}',
                f'تاریخ پایان: {req.get("end_date", "")}',
            ]
            
            for line in info_right:
                text = fix_text(line)
                text_width = draw.textlength(text, font=font_normal)
                draw.text((col2_x + 350 - text_width, y), text, fill=(0, 0, 0), font=font_normal)
                y += line_height
            
            y = 100 + len(info_left) * line_height + 25
            draw.line([(50, y), (width-50, y)], fill=(200, 200, 200), width=1)
            y += 30
            
            config = AttendanceManager.load_config()
            annual_limit = config.get('annual_leave_limit', 30)
            dynamic_limit = self._calculate_dynamic_leave_limit()
            used_days = self._get_used_leave_days()
            
            history_left = [
                f'سقف مرخصی سالانه: {annual_limit} روز',
                f'سقف مجاز (تا امروز): {dynamic_limit} روز',
            ]
            
            for line in history_left:
                text = fix_text(line)
                text_width = draw.textlength(text, font=font_small)
                draw.text((col1_x + 350 - text_width, y), text, fill=(0, 0, 0), font=font_small)
                y += 30
            
            y = 100 + len(info_left) * line_height + 25 + 30
            history_right = [
                f'مرخصی استفاده شده: {used_days:.1f} روز',
                f'مانده مرخصی: {dynamic_limit - used_days:.1f} روز',
            ]
            
            for line in history_right:
                text = fix_text(line)
                text_width = draw.textlength(text, font=font_small)
                draw.text((col2_x + 350 - text_width, y), text, fill=(0, 0, 0), font=font_small)
                y += 30
            
            y = max(y, 450) + 35
            draw.line([(50, y), (width-50, y)], fill=(200, 200, 200), width=1)
            y += 40
            
            sig1 = 'مسئول واحد: ______________'
            sig2 = 'تایید کننده: ______________'
            
            text1 = fix_text(sig1)
            text2 = fix_text(sig2)
            text1_width = draw.textlength(text1, font=font_normal)
            text2_width = draw.textlength(text2, font=font_normal)
            
            draw.text((col1_x + 350 - text1_width, y), text1, fill=(0, 0, 0), font=font_normal)
            draw.text((col2_x + 350 - text2_width, y), text2, fill=(0, 0, 0), font=font_normal)
            
            y += 45
            
            today = get_today_jalali()
            date_text = f'تاریخ: {today}'
            text_date = fix_text(date_text)
            text_date_width = draw.textlength(text_date, font=font_small)
            draw.text((col1_x + 350 - text_date_width, y), text_date, fill=(0, 0, 0), font=font_small)
            draw.text((col2_x + 350 - text_date_width, y), text_date, fill=(0, 0, 0), font=font_small)
            
            from utils.storage import get_backup_path
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            today_formatted = today.replace('/', '-')
            from datetime import datetime as dt
            timestamp = dt.now().strftime("%H%M%S")
            filename = f'برگ_مرخصی_{today_formatted}_{timestamp}.png'
            filepath = os.path.join(export_dir, filename)
            
            img.save(filepath)
            self.show_message('موفق', f'برگ مرخصی ذخیره شد:\n{filename}')
            
        except ImportError as e:
            self.show_message('خطا', f'کتابخانه مورد نیاز نصب نیست: {str(e)}')
        except Exception as e:
            self.show_message('خطا', f'خطا در ایجاد برگ مرخصی: {str(e)}')
            import traceback
            traceback.print_exc()


    def _load_leave_config(self):
        """بارگذاری تنظیمات مرخصی - نسخه اصلاح شده"""
        try:
            config = AttendanceManager.load_config()
            if config:
                print(f"✅ تنظیمات مرخصی از AttendanceManager: {config}")
                return config
            
            file_path = os.path.join(get_data_path(), 'attendance_config.json')
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    print(f"✅ تنظیمات از فایل مستقیم: {config}")
                    return config
            
            print("⚠️ تنظیمات مرخصی یافت نشد، از مقدار پیش‌فرض استفاده می‌شود")
            return {
                'leave_types': ['ساعتی', 'استحقاقی', 'استعلاجی', 'اضطراری', 'بدون حقوق'],
                'annual_leave_limit': 30,
                'monthly_hourly_leave_limit': '12:30',
                'hourly_to_daily_ratio': 5
            }
        except Exception as e:
            print(f"❌ خطا در بارگذاری تنظیمات مرخصی: {e}")
            return {
                'leave_types': ['ساعتی', 'استحقاقی', 'استعلاجی', 'اضطراری', 'بدون حقوق'],
                'annual_leave_limit': 30,
                'monthly_hourly_leave_limit': '12:30',
                'hourly_to_daily_ratio': 5
            }


    def _update_leave_limits(self):
        """به‌روزرسانی نمایش سقف‌های مرخصی - نسخه اصلاح شده"""
        try:
            config = self._load_leave_config()
            annual_limit = config.get('annual_leave_limit', 30)
            monthly_hourly_limit = config.get('monthly_hourly_leave_limit', '12:30')
            
            dynamic_limit = self._calculate_dynamic_leave_limit()
            
            if hasattr(self, 'annual_limit_field'):
                self.annual_limit_field.text = f'سقف سالانه: {annual_limit} روز'
            if hasattr(self, 'dynamic_limit_field'):
                self.dynamic_limit_field.text = f'سقف مجاز (تا امروز): {dynamic_limit} روز'
            
            print(f"✅ سقف‌ها به‌روز شد: سالانه={annual_limit}, داینامیک={dynamic_limit}")
            
        except Exception as e:
            print(f"❌ خطا در به‌روزرسانی سقف‌ها: {e}")

    # ============================================================
    # تب 3: ماموریت - کامل
    # ============================================================

    def show_mission_tab(self):
        """نمایش تب ماموریت"""
        content = BoxLayout(orientation='vertical', padding=dp(8), spacing=dp(4))
        
        # ========== دکمه فراخوانی ==========
        import_btn = PersianButton(
            text='فراخوانی ماموریت از فایل اکسل',
            size_hint_y=None,
            height=dp(42),
            background_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=sp(14),
            bold=True
        )
        import_btn.bind(on_press=self._import_missions)
        content.add_widget(import_btn)
        
        # ========== لیست ماموریت‌های فعال ==========
        self.mission_list_scroll = ScrollView(
            do_scroll_x=True,
            do_scroll_y=True,
            size_hint_y=0.55,
            scroll_type=['bars', 'content'],
            bar_width=dp(5)
        )
        
        self.mission_list_container = GridLayout(
            cols=1,
            spacing=dp(1),
            size_hint_y=None,
            size_hint_x=None,
            width=dp(780),
            padding=dp(2)
        )
        self.mission_list_container.bind(minimum_height=self.mission_list_container.setter('height'))
        self.mission_list_scroll.add_widget(self.mission_list_container)
        content.add_widget(self.mission_list_scroll)
        
        # ========== فاصله بین لیست و دکمه‌ها ==========
        content.add_widget(BoxLayout(size_hint_y=None, height=dp(3)))
        
        # ========== دکمه‌های پایین ==========
        bottom_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(4))
        
        self.assign_btn = PersianButton(
            text='تعیین تکلیف انتخاب‌ها',
            size_hint_x=0.28,
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12),
            bold=True
        )
        self.assign_btn.bind(on_press=self._assign_selected_missions)
        bottom_row.add_widget(self.assign_btn)
        
        history_btn = PersianButton(
            text='تاریخچه',
            size_hint_x=0.16,
            background_color=(0.4, 0.3, 0.6, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12)
        )
        history_btn.bind(on_press=self._show_history_dialog)
        bottom_row.add_widget(history_btn)
        
        expired_btn = PersianButton(
            text='⏰ ماموریت‌های از دست رفته',
            size_hint_x=0.20,
            background_color=(0.8, 0.3, 0.1, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12),
            bold=True
        )
        expired_btn.bind(on_press=self._show_expired_missions_dialog)
        bottom_row.add_widget(expired_btn)
        
        export_btn = PersianButton(
            text='خروجی اکسل',
            size_hint_x=0.18,
            background_color=(0.6, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12),
            bold=True
        )
        export_btn.bind(on_press=self._export_missions_excel)
        bottom_row.add_widget(export_btn)
        
        print_btn = PersianButton(
            text='چاپ برگه',
            size_hint_x=0.18,
            background_color=(0.2, 0.4, 0.6, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12)
        )
        print_btn.bind(on_press=self._print_mission_form)
        bottom_row.add_widget(print_btn)
        
        content.add_widget(bottom_row)
        
        self.tab_content.add_widget(content)
        
        # بارگذاری ماموریت‌ها
        self._load_missions()


    def _import_missions(self, instance):
        """فراخوانی ماموریت از فایل اکسل"""
        try:
            from utils.file_picker_import import ImportFilePicker
            
            content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(12))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            content.add_widget(RTLLabel(
                text='فراخوانی ماموریت از فایل اکسل',
                font_size=sp(18),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                height=dp(40)
            ))
            
            content.add_widget(RTLLabel(
                text='فایل باید با فرمت .xlsx و شامل ستون‌های مشخص شده باشد',
                font_size=sp(13),
                color=(0.7, 0.7, 0.7, 1),
                size_hint_y=None,
                height=dp(30)
            ))
            
            file_picker = ImportFilePicker(
                on_select=self._process_imported_missions,
                size_hint_y=None,
                height=dp(120)
            )
            content.add_widget(file_picker)
            
            btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
            cancel_btn = PersianButton(
                text='انصراف',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(45),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            cancel_btn.bind(on_press=lambda x: popup.dismiss())
            btn_layout.add_widget(cancel_btn)
            content.add_widget(btn_layout)
            
            popup = PersianPopup(
                title='فراخوانی ماموریت',
                content=content,
                size_hint=(0.9, 0.5),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            
            self._import_popup = popup
            popup.open()
            
        except Exception as e:
            self.show_message('خطا', f'خطا در باز کردن دیالوگ فراخوانی: {str(e)}')


    def _load_missions(self):
        """بارگذاری و نمایش ماموریت‌های فعال"""
        try:
            self.mission_list_container.clear_widgets()
            
            all_missions = self._load_missions_from_file()
            
            current_user_name = ''
            if self.current_user:
                current_user_name = self.current_user.get('name', '')
                if not current_user_name:
                    current_user_name = self.current_user.get('username', '')
            
            active_missions = []
            history_missions = []
            expired_missions = []
            today = get_today_jalali()
            
            for m in all_missions:
                agent_name = m.get('agent_name', '').strip()
                
                if agent_name != current_user_name:
                    continue
                
                status = m.get('status', '')
                end_date = m.get('end_date', '')
                is_active = m.get('active', True)
                
                if status in ['✅ موفق', '❌ ناموفق'] or not is_active:
                    history_missions.append(m)
                elif status in ['⏳ در انتظار', 'در انتظار']:
                    if end_date and end_date < today:
                        expired_missions.append(m)
                    else:
                        active_missions.append(m)
                else:
                    history_missions.append(m)
            
            self._history_missions = history_missions
            self._expired_missions = expired_missions
            
            if not active_missions:
                self.mission_list_container.add_widget(RTLLabel(
                    text='هیچ ماموریت فعالی وجود ندارد',
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(12),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                return
            
            header = BoxLayout(size_hint_y=None, height=dp(26), spacing=dp(1), size_hint_x=None, width=dp(780))
            headers = ['', 'شناسه', 'نوع', 'روش', 'تاریخ شروع', 'مدت', 'امتیاز', 'هدف', 'توضیحات']
            sizes = [0.05, 0.10, 0.12, 0.10, 0.12, 0.07, 0.08, 0.16, 0.20]
            
            for i in range(len(headers)):
                header.add_widget(RTLLabel(
                    text=headers[i],
                    size_hint_x=sizes[i],
                    font_size=sp(10),
                    bold=True,
                    color=(0.4, 0.7, 1, 1)
                ))
            self.mission_list_container.add_widget(header)
            
            for idx, mission in enumerate(active_missions, 1):
                row = BoxLayout(size_hint_y=None, height=dp(32), spacing=dp(1), size_hint_x=None, width=dp(780))
                with row.canvas.before:
                    Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                    rect = Rectangle(pos=row.pos, size=row.size)
                    row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                            size=lambda i, v, r=rect: setattr(r, 'size', v))
                
                # ✅ چک‌باکس با وضعیت selected از فایل
                is_selected = mission.get('selected', False)
                cb = CheckBox(
                    active=is_selected,
                    size_hint_x=0.05,
                    size_hint_y=None,
                    height=dp(26),
                    color=(0.4, 0.7, 1, 1)
                )
                cb.mission_id = mission.get('id')
                cb.mission_data = mission
                cb.bind(active=self._on_mission_select)
                row.add_widget(cb)
                
                row.add_widget(RTLLabel(
                    text=mission.get('id', ''),
                    size_hint_x=0.10,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('type', ''),
                    size_hint_x=0.12,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('method', ''),
                    size_hint_x=0.10,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('start_date', ''),
                    size_hint_x=0.12,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=str(mission.get('duration', '')),
                    size_hint_x=0.07,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=str(mission.get('score', '')),
                    size_hint_x=0.08,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=f"{mission.get('target', 0):,.0f}",
                    size_hint_x=0.16,
                    font_size=sp(10),
                    color=(1, 1, 1, 1)
                ))
                
                description = mission.get('description', '')
                row.add_widget(RTLLabel(
                    text=description,
                    size_hint_x=0.20,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                
                self.mission_list_container.add_widget(row)
                
        except Exception as e:
            print(f"خطا در بارگذاری ماموریت‌ها: {e}")
            import traceback
            traceback.print_exc()


    def _load_missions_from_file(self):
        """بارگذاری ماموریت‌ها از فایل"""
        try:
            file_path = os.path.join(get_data_path(), 'do_missions.json')
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return []
        except Exception as e:
            print(f"خطا در بارگذاری ماموریت‌ها: {e}")
            return []


    def _save_missions_to_file(self, missions):
        """ذخیره ماموریت‌ها در فایل"""
        try:
            file_path = os.path.join(get_data_path(), 'do_missions.json')
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(missions, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"خطا در ذخیره ماموریت‌ها: {e}")


    def _on_mission_select(self, instance, value):
        """انتخاب/لغو انتخاب ماموریت"""
        mission_id = instance.mission_id
        all_missions = self._load_missions_from_file()
        
        for m in all_missions:
            if m.get('id') == mission_id:
                m['selected'] = value
                break
        
        self._save_missions_to_file(all_missions)
        
        instance.active = value


    def _assign_selected_missions(self, instance):
        """تعیین تکلیف ماموریت‌های انتخاب شده با نمایش هدف و توضیحات"""
        all_missions = self._load_missions_from_file()
        selected = [m for m in all_missions if m.get('selected', False)]
        
        if not selected:
            self.show_message('خطا', 'هیچ ماموریتی انتخاب نشده است')
            return
        
        # ========== محاسبه اطلاعات ==========
        total_target = sum([m.get('target', 0) for m in selected])
        total_score = sum([m.get('score', 0) for m in selected])
        
        content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))
        with content.canvas.before:
            Color(0.12, 0.12, 0.12, 1)
            rect = Rectangle(pos=content.pos, size=content.size)
            content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
        
        content.add_widget(RTLLabel(
            text=f'تعداد: {len(selected)} ماموریت انتخاب شده',
            size_hint_y=None,
            height=dp(40),
            font_size=sp(15),
            bold=True,
            color=(1, 1, 1, 1)
        ))
        
        info_row = BoxLayout(size_hint_y=None, height=dp(35), spacing=dp(10))
        info_row.add_widget(RTLLabel(
            text=f'هدف کل: {total_target:,.0f} ریال',
            size_hint_x=0.5,
            font_size=sp(15),
            color=(0.2, 0.9, 0.2, 1)
        ))
        info_row.add_widget(RTLLabel(
            text=f'امتیاز کل: {total_score}',
            size_hint_x=0.5,
            font_size=sp(15),
            color=(0.8, 0.8, 0.2, 1)
        ))
        content.add_widget(info_row)
        
        # ✅ نمایش توضیحات با اسکرول افقی
        desc_text = ""
        for i, m in enumerate(selected[:10], 1):
            desc = m.get('description', 'بدون توضیحات')
            desc_text += f"{i}. {desc}\n"
        
        if len(selected) > 10:
            desc_text += f"... و {len(selected)-10} ماموریت دیگر"
        
        if desc_text:
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            
            desc_container = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(250))
            
            # ✅ اسکرول افقی و عمودی
            scroll = ScrollView(
                do_scroll_x=True,      # ✅ فعال کردن اسکرول افقی
                do_scroll_y=True,      # ✅ فعال کردن اسکرول عمودی
                size_hint=(1, 1),
                bar_width=dp(8),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5),
                scroll_type=['bars', 'content']
            )
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(desc_text)
                display_text = get_display(reshaped)
            except:
                display_text = desc_text
            
            # ✅ محاسبه عرض مورد نیاز برای اسکرول افقی
            max_line_length = 0
            for line in desc_text.split('\n'):
                max_line_length = max(max_line_length, len(line))
            
            # ✅ تنظیم عرض بر اساس طولانی‌ترین خط
            char_width = sp(16) * 0.8  # تقریباً عرض هر کاراکتر
            label_width = max(dp(500), max_line_length * char_width + dp(40))
            
            label = Label(
                text=display_text,
                font_size=sp(16),
                color=(0.9, 0.9, 0.9, 1),
                size_hint=(None, None),  # ✅ None, None برای اسکرول افقی
                halign='right',
                valign='top',
                text_size=(label_width, None),  # ✅ عرض ثابت برای اسکرول افقی
                font_name='fonts/Amiri-Regular.ttf'
            )
            
            lines = desc_text.count('\n') + 1
            line_height = sp(16) + dp(8)
            label.height = max(lines * line_height + dp(20), dp(50))
            label.width = label_width
            
            scroll.add_widget(label)
            desc_container.add_widget(scroll)
            content.add_widget(desc_container)
        
        btn_row = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
        
        success_btn = PersianButton(
            text='موفق',
            size_hint_x=0.5,
            background_color=(0.2, 0.7, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(16),
            bold=True
        )
        success_btn.bind(on_press=lambda x: self._apply_assignment('✅ موفق'))
        btn_row.add_widget(success_btn)
        
        fail_btn = PersianButton(
            text='ناموفق',
            size_hint_x=0.5,
            background_color=(0.8, 0.2, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(16),
            bold=True
        )
        fail_btn.bind(on_press=lambda x: self._apply_assignment('❌ ناموفق'))
        btn_row.add_widget(fail_btn)
        
        content.add_widget(btn_row)
        
        cancel_btn = PersianButton(
            text='انصراف',
            size_hint_y=None,
            height=dp(45),
            background_color=(0.3, 0.3, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(16)
        )
        cancel_btn.bind(on_press=lambda x: popup.dismiss())
        content.add_widget(cancel_btn)
        
        popup = PersianPopup(
            title='تعیین تکلیف ماموریت‌ها',
            content=content,
            size_hint=(0.9, None),
            height=dp(600),
            background_color=(0.08, 0.08, 0.08, 1)
        )
        self._assign_popup = popup
        popup.open()


    def _apply_assignment(self, status):
        """اعمال وضعیت به ماموریت‌های انتخاب شده"""
        if hasattr(self, '_assign_popup'):
            self._assign_popup.dismiss()
        
        all_missions = self._load_missions_from_file()
        updated = 0
        
        for m in all_missions:
            if m.get('selected', False):
                m['status'] = status
                m['active'] = False
                m['selected'] = False
                updated += 1
        
        self._save_missions_to_file(all_missions)
        self._load_missions()
        self.show_message('موفق', f'{updated} ماموریت با موفقیت تعیین تکلیف شد')


    def _show_history_dialog(self, instance):
        """نمایش دیالوگ تاریخچه ماموریت‌ها - ساده مثل ماموریت‌های از دست رفته"""
        try:
            history_missions = getattr(self, '_history_missions', [])
            
            if not history_missions:
                all_missions = self._load_missions_from_file()
                current_user_name = ''
                if self.current_user:
                    current_user_name = self.current_user.get('name', '')
                    if not current_user_name:
                        current_user_name = self.current_user.get('username', '')
                
                history_missions = []
                today = get_today_jalali()
                
                for m in all_missions:
                    agent_name = m.get('agent_name', '').strip()
                    if agent_name != current_user_name:
                        continue
                    
                    status = m.get('status', '')
                    end_date = m.get('end_date', '')
                    is_active = m.get('active', True)
                    
                    if status in ['✅ موفق', '❌ ناموفق'] or not is_active:
                        history_missions.append(m)
                    elif status in ['⏳ در انتظار', 'در انتظار'] and end_date and end_date < today:
                        history_missions.append(m)
            
            # ========== محتوای دیالوگ ==========
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(5))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                            size=lambda i, v: setattr(rect, 'size', v))
            
            if not history_missions:
                content.add_widget(RTLLabel(
                    text='هیچ ماموریتی در تاریخچه وجود ندارد',
                    size_hint_y=None,
                    height=dp(40),
                    font_size=sp(16),
                    color=(0.5, 0.5, 0.5, 1)
                ))
            else:
                # ========== اسکرول ==========
                scroll = ScrollView(
                    do_scroll_x=True,
                    do_scroll_y=True,
                    size_hint_y=0.9,
                    scroll_type=['bars', 'content'],
                    bar_width=dp(5)
                )
                
                container = GridLayout(
                    cols=1,
                    spacing=dp(1),
                    size_hint_y=None,
                    size_hint_x=None,
                    width=dp(650),
                    padding=dp(2)
                )
                container.bind(minimum_height=container.setter('height'))
                
                # ========== هدر ==========
                header = BoxLayout(size_hint_y=None, height=dp(22), spacing=dp(1), size_hint_x=None, width=dp(650))
                headers = ['شناسه', 'نوع', 'وضعیت', 'تاریخ پایان', 'امتیاز', 'امتیاز کسب شده']
                sizes = [0.12, 0.18, 0.18, 0.16, 0.12, 0.14]
                
                for i, (text, size) in enumerate(zip(headers, sizes)):
                    header.add_widget(RTLLabel(
                        text=text,
                        size_hint_x=size,
                        font_size=sp(9),
                        bold=True,
                        color=(0.4, 0.7, 1, 1)
                    ))
                container.add_widget(header)
                
                # ========== ردیف‌ها ==========
                for idx, mission in enumerate(history_missions, 1):
                    row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(1), size_hint_x=None, width=dp(650))
                    
                    status = mission.get('status', '')
                    score = mission.get('score', 0)
                    
                    # ✅ محاسبه امتیاز کسب شده
                    if 'موفق' in status and 'ناموفق' not in status:
                        earned_score = score
                    else:
                        earned_score = 0
                    
                    row.add_widget(RTLLabel(
                        text=mission.get('id', ''),
                        size_hint_x=0.12,
                        font_size=sp(9),
                        color=(1, 1, 1, 1)
                    ))
                    row.add_widget(RTLLabel(
                        text=mission.get('type', ''),
                        size_hint_x=0.18,
                        font_size=sp(9),
                        color=(1, 1, 1, 1)
                    ))
                    row.add_widget(RTLLabel(
                        text=status,
                        size_hint_x=0.18,
                        font_size=sp(9),
                        color=(1, 1, 1, 1)
                    ))
                    row.add_widget(RTLLabel(
                        text=mission.get('end_date', ''),
                        size_hint_x=0.16,
                        font_size=sp(9),
                        color=(1, 1, 1, 1)
                    ))
                    row.add_widget(RTLLabel(
                        text=str(score),
                        size_hint_x=0.12,
                        font_size=sp(9),
                        color=(1, 1, 1, 1)
                    ))
                    row.add_widget(RTLLabel(
                        text=str(earned_score),
                        size_hint_x=0.14,
                        font_size=sp(9),
                        color=(1, 1, 1, 1)
                    ))
                    
                    container.add_widget(row)
                
                scroll.add_widget(container)
                content.add_widget(scroll)
            
            # ========== دکمه بستن ==========
            close_btn = PersianButton(
                text='بستن',
                size_hint_y=None,
                height=dp(40),
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(14)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='تاریخچه ماموریت‌ها',
                content=content,
                size_hint=(0.9, 0.7),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            self.show_message('خطا', f'خطا در نمایش تاریخچه: {str(e)}')
            import traceback
            traceback.print_exc()


    def _show_expired_missions_dialog(self, instance):
        """نمایش دیالوگ ماموریت‌های از دست رفته (منقضی)"""
        try:
            all_missions = self._load_missions_from_file()
            
            current_user_name = ''
            if self.current_user:
                current_user_name = self.current_user.get('name', '')
                if not current_user_name:
                    current_user_name = self.current_user.get('username', '')
            
            today = get_today_jalali()
            
            expired_missions = []
            for m in all_missions:
                agent_name = m.get('agent_name', '').strip()
                if agent_name != current_user_name:
                    continue
                
                end_date = m.get('end_date', '')
                status = m.get('status', '')
                is_active = m.get('active', True)
                
                if is_active and status in ['⏳ در انتظار', 'در انتظار']:
                    if end_date and end_date < today:
                        expired_missions.append(m)
            
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(5))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            if expired_missions:
                content.add_widget(RTLLabel(
                    text=f'تعداد: {len(expired_missions)} ماموریت',
                    size_hint_y=None,
                    height=dp(30),
                    font_size=sp(14),
                    color=(0.6, 0.6, 0.6, 1)
                ))
            else:
                content.add_widget(RTLLabel(
                    text='هیچ ماموریت از دست رفته‌ای وجود ندارد',
                    size_hint_y=None,
                    height=dp(40),
                    font_size=sp(16),
                    color=(0.5, 0.5, 0.5, 1)
                ))
                close_btn = PersianButton(
                    text='بستن',
                    size_hint_y=None,
                    height=dp(40),
                    background_color=(0.3, 0.3, 0.3, 1),
                    color=(1, 1, 1, 1),
                    font_size=sp(14)
                )
                content.add_widget(close_btn)
                
                popup = PersianPopup(
                    title='ماموریت‌های از دست رفته',
                    content=content,
                    size_hint=(0.8, 0.3),
                    background_color=(0.08, 0.08, 0.08, 1)
                )
                close_btn.bind(on_press=popup.dismiss)
                popup.open()
                return
            
            scroll = ScrollView(
                do_scroll_x=True,
                do_scroll_y=True,
                size_hint_y=0.9,
                scroll_type=['bars', 'content'],
                bar_width=dp(5)
            )
            
            container = GridLayout(
                cols=1,
                spacing=dp(1),
                size_hint_y=None,
                size_hint_x=None,
                width=dp(650),
                padding=dp(2)
            )
            container.bind(minimum_height=container.setter('height'))
            
            header = BoxLayout(size_hint_y=None, height=dp(24), spacing=dp(1), size_hint_x=None, width=dp(650))
            headers = ['شناسه', 'نوع', 'تاریخ شروع', 'تاریخ پایان', 'مدت', 'هدف']
            sizes = [0.12, 0.18, 0.16, 0.16, 0.10, 0.28]
            
            for i, (text, size) in enumerate(zip(headers, sizes)):
                header.add_widget(RTLLabel(
                    text=text,
                    size_hint_x=size,
                    font_size=sp(9),
                    bold=True,
                    color=(0.8, 0.4, 0.1, 1)
                ))
            container.add_widget(header)
            
            for idx, mission in enumerate(expired_missions, 1):
                row = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(1), size_hint_x=None, width=dp(650))
                with row.canvas.before:
                    Color(0.15, 0.15, 0.2, 1) if idx % 2 == 0 else Color(0.12, 0.12, 0.16, 1)
                    rect = Rectangle(pos=row.pos, size=row.size)
                    row.bind(pos=lambda i, v, r=rect: setattr(r, 'pos', v),
                            size=lambda i, v, r=rect: setattr(r, 'size', v))
                
                row.add_widget(RTLLabel(
                    text=mission.get('id', ''),
                    size_hint_x=0.12,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('type', ''),
                    size_hint_x=0.18,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('start_date', ''),
                    size_hint_x=0.16,
                    font_size=sp(9),
                    color=(0.8, 0.8, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=mission.get('end_date', ''),
                    size_hint_x=0.16,
                    font_size=sp(9),
                    color=(0.9, 0.2, 0.2, 1)
                ))
                row.add_widget(RTLLabel(
                    text=str(mission.get('duration', '')),
                    size_hint_x=0.10,
                    font_size=sp(9),
                    color=(1, 1, 1, 1)
                ))
                row.add_widget(RTLLabel(
                    text=f"{mission.get('target', 0):,.0f}",
                    size_hint_x=0.28,
                    font_size=sp(9),
                    color=(0.6, 0.9, 0.6, 1)
                ))
                
                container.add_widget(row)
            
            scroll.add_widget(container)
            content.add_widget(scroll)
            
            close_btn = PersianButton(
                text='بستن',
                size_hint_y=None,
                height=dp(40),
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(14)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='لیست ماموریتهای منقضی شده',
                content=content,
                size_hint=(0.9, 0.75),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            self.show_message('خطا', f'خطا در نمایش ماموریت‌های از دست رفته: {str(e)}')
            import traceback
            traceback.print_exc()


    def _export_missions_excel(self, instance):
        """خروجی اکسل ماموریت‌ها"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            all_missions = self._load_missions_from_file()
            
            current_user_name = ''
            if self.current_user:
                current_user_name = self.current_user.get('name', '')
                if not current_user_name:
                    current_user_name = self.current_user.get('username', '')
            
            user_missions = []
            for m in all_missions:
                agent_name = m.get('agent_name', '').strip()
                if agent_name == current_user_name:
                    user_missions.append(m)
            
            if not user_missions:
                self.show_message('خطا', 'هیچ ماموریتی برای خروجی وجود ندارد')
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش ماموریت‌ها"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['شناسه', 'عامل', 'نوع', 'روش', 'تاریخ شروع', 'مدت', 'امتیاز', 'هدف', 'توضیحات', 'تاریخ پایان', 'وضعیت', 'فعال']
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for row_idx, mission in enumerate(user_missions, 2):
                status = mission.get('status', '⏳ در انتظار')
                active = '✅ فعال' if mission.get('active', True) else '🔒 غیرفعال'
                
                values = [
                    mission.get('id', ''),
                    mission.get('agent', ''),
                    mission.get('type', ''),
                    mission.get('method', ''),
                    mission.get('start_date', ''),
                    mission.get('duration', 0),
                    mission.get('score', 0),
                    mission.get('target', 0),
                    mission.get('description', ''),
                    mission.get('end_date', ''),
                    status,
                    active
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            column_widths = [12, 20, 14, 12, 14, 8, 10, 16, 30, 14, 14, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            from utils.storage import get_backup_path
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            today = get_today_jalali().replace('/', '-')
            from datetime import datetime as dt
            timestamp = dt.now().strftime("%H%M%S")
            filename = f'گزارش_ماموریت‌ها_{today}_{timestamp}.xlsx'
            filepath = os.path.join(export_dir, filename)
            
            wb.save(filepath)
            self.show_message('موفق', f'فایل اکسل ماموریت‌ها ذخیره شد:\n{filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی اکسل: {str(e)}')


    def _print_mission_form(self, instance):
        """خروجی برگه ماموریت (تصویر) - فقط برای ماموریت‌های انتخاب شده"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import arabic_reshaper
            from bidi.algorithm import get_display
            
            all_missions = self._load_missions_from_file()
            
            current_user_name = ''
            if self.current_user:
                current_user_name = self.current_user.get('name', '')
                if not current_user_name:
                    current_user_name = self.current_user.get('username', '')
            
            selected_missions = []
            for m in all_missions:
                agent_name = m.get('agent_name', '').strip()
                if agent_name == current_user_name and m.get('selected', False):
                    selected_missions.append(m)
            
            if not selected_missions:
                self.show_message('خطا', 'لطفاً حداقل یک ماموریت را برای چاپ انتخاب کنید')
                return
            
            if len(selected_missions) > 1:
                self.show_message('خطا', 'لطفاً فقط یک ماموریت را برای چاپ انتخاب کنید')
                return
            
            req = selected_missions[0]
            
            font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Amiri-Regular.ttf')
            if not os.path.exists(font_path):
                font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Vazirmatn-Regular.ttf')
            
            try:
                font_title = ImageFont.truetype(font_path, 24)
                font_normal = ImageFont.truetype(font_path, 16)
                font_small = ImageFont.truetype(font_path, 13)
            except:
                font_title = ImageFont.load_default()
                font_normal = ImageFont.load_default()
                font_small = ImageFont.load_default()
            
            def fix_text(text):
                if not text:
                    return ''
                try:
                    reshaped = arabic_reshaper.reshape(str(text))
                    return get_display(reshaped)
                except:
                    return str(text)
            
            width = 800
            height = 500
            img = Image.new('RGB', (width, height), color=(255, 255, 255))
            draw = ImageDraw.Draw(img)
            
            title = 'برگ ماموریت'
            title_width = draw.textlength(fix_text(title), font=font_title)
            draw.text(((width - title_width) // 2, 25), fix_text(title), fill=(0, 0, 0), font=font_title)
            
            draw.line([(50, 60), (width-50, 60)], fill=(0, 0, 0), width=2)
            
            y = 90
            line_height = 35
            
            user_name = self.current_user.get('name', '') if self.current_user else ''
            user_role = self.current_user.get('role', '') if self.current_user else ''
            
            info = [
                f'نام و نام خانوادگی: {user_name}',
                f'سمت/واحد: {user_role}',
                f'نوع ماموریت: {req.get("type", "")}',
                f'روش: {req.get("method", "")}',
                f'تاریخ شروع: {req.get("start_date", "")}',
                f'تاریخ پایان: {req.get("end_date", "")}',
                f'مدت: {req.get("duration", 0)} روز',
                f'امتیاز: {req.get("score", 0)}',
                f'هدف: {req.get("target", 0):,.0f} ریال',
                f'توضیحات: {req.get("description", "")}',
                f'وضعیت: {req.get("status", "⏳ در انتظار")}',
            ]
            
            for line in info:
                text = fix_text(line)
                text_width = draw.textlength(text, font=font_normal)
                draw.text((width - 50 - text_width, y), text, fill=(0, 0, 0), font=font_normal)
                y += line_height
            
            from utils.storage import get_backup_path
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            today = get_today_jalali().replace('/', '-')
            from datetime import datetime as dt
            timestamp = dt.now().strftime("%H%M%S")
            filename = f'برگ_ماموریت_{today}_{timestamp}.png'
            filepath = os.path.join(export_dir, filename)
            
            img.save(filepath)
            self.show_message('موفق', f'برگ ماموریت ذخیره شد:\n{filename}')
            
            for m in all_missions:
                if m.get('selected', False):
                    m['selected'] = False
            self._save_missions_to_file(all_missions)
            self._load_missions()
            
        except ImportError as e:
            self.show_message('خطا', f'کتابخانه مورد نیاز نصب نیست: {str(e)}')
        except Exception as e:
            self.show_message('خطا', f'خطا در ایجاد برگ ماموریت: {str(e)}')
            import traceback
            traceback.print_exc()


    def _process_imported_missions(self, filepath):
        """پردازش فایل اکسل وارد شده"""
        try:
            import openpyxl
            import os
            
            if not filepath or not os.path.exists(filepath):
                self.show_message('خطا', 'فایل انتخاب شده وجود ندارد')
                return
            
            if not filepath.lower().endswith('.xlsx'):
                self.show_message('خطا', 'فایل باید با فرمت .xlsx باشد')
                return
            
            if hasattr(self, '_import_popup'):
                self._import_popup.dismiss()
            
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            current_user_name = ''
            if self.current_user:
                current_user_name = self.current_user.get('name', '')
                if not current_user_name:
                    current_user_name = self.current_user.get('username', '')
            
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f'ستون{col}')
            
            col_map = {}
            for idx, header in enumerate(headers):
                if 'شناسه' in header:
                    col_map['id'] = idx
                elif 'عامل' in header:
                    col_map['agent'] = idx
                elif 'نوع' in header:
                    col_map['type'] = idx
                elif 'روش' in header:
                    col_map['method'] = idx
                elif 'تاریخ شروع' in header:
                    col_map['start_date'] = idx
                elif 'مدت' in header:
                    col_map['duration'] = idx
                elif 'امتیاز' in header:
                    col_map['score'] = idx
                elif 'هدف' in header:
                    col_map['target'] = idx
                elif 'توضیحات' in header:
                    col_map['description'] = idx
                elif 'تاریخ پایان' in header:
                    col_map['end_date'] = idx
                elif 'وضعیت' in header:
                    col_map['status'] = idx
                elif 'فعال' in header:
                    col_map['active'] = idx
            
            imported_count = 0
            missions = self._load_missions_from_file()
            existing_ids = [m.get('id') for m in missions]
            
            for row_idx in range(2, ws.max_row + 1):
                agent_col = col_map.get('agent')
                if agent_col is None:
                    continue
                
                agent_value = ws.cell(row=row_idx, column=agent_col + 1).value
                if not agent_value:
                    continue
                
                agent_full = str(agent_value).strip()
                
                agent_name = agent_full
                if ' - ' in agent_full:
                    agent_name = agent_full.split(' - ')[-1].strip()
                elif '-' in agent_full:
                    agent_name = agent_full.split('-')[-1].strip()
                elif '–' in agent_full:
                    agent_name = agent_full.split('–')[-1].strip()
                
                if agent_name != current_user_name:
                    continue
                
                id_col = col_map.get('id')
                mission_id = None
                if id_col is not None:
                    mission_id = ws.cell(row=row_idx, column=id_col + 1).value
                    if mission_id:
                        mission_id = str(mission_id).strip()
                
                if mission_id and mission_id in existing_ids:
                    continue
                
                mission = {
                    'id': mission_id or f'M{row_idx:04d}',
                    'agent': agent_full,
                    'agent_name': agent_name,
                    'type': ws.cell(row=row_idx, column=col_map.get('type', 0) + 1).value or '',
                    'method': ws.cell(row=row_idx, column=col_map.get('method', 0) + 1).value or '',
                    'start_date': ws.cell(row=row_idx, column=col_map.get('start_date', 0) + 1).value or '',
                    'duration': ws.cell(row=row_idx, column=col_map.get('duration', 0) + 1).value or 0,
                    'score': ws.cell(row=row_idx, column=col_map.get('score', 0) + 1).value or 0,
                    'target': ws.cell(row=row_idx, column=col_map.get('target', 0) + 1).value or 0,
                    'description': ws.cell(row=row_idx, column=col_map.get('description', 0) + 1).value or '',
                    'end_date': ws.cell(row=row_idx, column=col_map.get('end_date', 0) + 1).value or '',
                    'status': '⏳ در انتظار',
                    'active': True,
                    'imported_at': get_today_jalali(),
                    'selected': False
                }
                
                missions.append(mission)
                imported_count += 1
                existing_ids.append(mission_id)
            
            self._save_missions_to_file(missions)
            self._load_missions()
            
            if imported_count > 0:
                self.show_message('موفق', f'{imported_count} ماموریت با موفقیت فراخوانی شد')
            else:
                self.show_message('توجه', 'هیچ ماموریت جدیدی برای کاربر جاری یافت نشد')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در فراخوانی ماموریت‌ها: {str(e)}')
            import traceback
            traceback.print_exc()

    # ============================================================
    # تب 4: گزارشات
    # ============================================================
    
    def show_report_tab(self):
        """نمایش تب گزارشات"""
        self.tab_content.clear_widgets()
        
        # ایجاد صفحه گزارشات
        report_screen = ReportAttendanceScreen(current_user=self.current_user)
        self.tab_content.add_widget(report_screen)
    
    # ============================================================
    # توابع کمکی
    # ============================================================
    
    def load_today_attendance(self):
        """بارگذاری حضور و غیاب امروز از دیتابیس"""
        if not self.current_user:
            return
        
        user_id = self.current_user.get('id')
        records = AttendanceManager.get_daily_report(user_id=user_id)
        
        self.entries = []
        self.is_checked_in = False
        self.is_day_ended = False
        
        for record in records:
            check_in = record.get('check_in')
            check_out = record.get('check_out')
            if check_in:
                self.entries.append({
                    'check_in': check_in,
                    'check_out': check_out
                })
                if check_out is None:
                    self.is_checked_in = True
            
            if record.get('is_day_ended'):
                self.is_day_ended = True
        
        self.update_buttons_state()
        self.update_entries_list()
        self.update_summary()
    
    def show_message(self, title, message):
        """نمایش پیام با پشتیبانی از متن طولانی"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(message)
                display_text = get_display(reshaped)
            except:
                display_text = message
            
            msg_label = Label(
                text=display_text,
                font_size=sp(15),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                halign='center',
                valign='top',
                text_size=(dp(550), None),
                font_name='fonts/Amiri-Regular.ttf',
                padding=(dp(20), 0, dp(0), 0)
            )
            
            lines = message.count('\n') + 1
            if len(message) > 50 and lines == 1:
                approx_chars_per_line = 35
                lines = (len(message) // approx_chars_per_line) + 1
            
            line_height = sp(15) + dp(6)
            label_height = max(lines * line_height + dp(25), dp(50))
            label_height = min(label_height, dp(490))  # ✅ افزایش ارتفاع
            
            msg_label.height = label_height
            msg_label.text_size = (dp(550), label_height)
            
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=None,
                height=label_height + dp(10),
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )
            scroll.add_widget(msg_label)
            content.add_widget(scroll)
            
            btn = PersianButton(
                text='باشه',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(16),
                color=(1, 1, 1, 1),
                background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            
            popup = PersianPopup(
                title=title,
                content=content,
                size_hint=(0.9, None),
                height=label_height + dp(250),  # ✅ افزایش ارتفاع
                background_color=(0.08, 0.08, 0.08, 1)
            )
            btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            print(f"خطا در نمایش پیام: {e}")
    
    def go_back(self, instance):
        """بازگشت به صفحه ورود"""
        self.manager.current = 'login'
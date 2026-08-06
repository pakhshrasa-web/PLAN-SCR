# screens/collection_report_dialog.py
# ========== دیالوگ گزارش وصول مطالبات ==========

import os
import traceback
from datetime import datetime
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.graphics import Color, Rectangle
from kivy.clock import Clock
from kivy.utils import platform

from utils.rtl_widgets import PersianButton, RTLLabel, PersianPopup, RTLTextInput, PersianComboBox
from utils.persian_text import PersianLabel, number_to_words
from utils.file_manager import get_agents, get_routes, get_customers, get_data_path
from utils.bank_manager import get_bank_names
from utils.jalali_date import get_today_jalali
from utils.collection_manager import get_collections
from utils.storage import get_backup_path
from utils.excel_export_helper import ExcelExportHelper
from error_handler import ErrorPopup


class CollectionReportDialog:
    """کلاس مدیریت دیالوگ گزارش وصول مطالبات"""
    
    def __init__(self):
        self.agent_filter = 'همه'
        self.route_filter = 'همه'
        self.status_filter = 'همه'
        self.start_date = ''
        self.end_date = ''
        self.filtered_data = []
        self.agg_filtered_data = []
        
        self.current_tab = 1
        self._all_customer_names = ['همه']
        
        self.show_report_dialog()
    

    def show_report_dialog(self):
        """نمایش دیالوگ اصلی گزارش"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))
            
            # تب بار
            self.tab_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(5))
            
            self.tab1_btn = PersianButton(
                text='گزارش لیستی',
                size_hint_x=0.33,
                background_color=(0.2, 0.5, 0.9, 1),
                color=(1, 1, 1, 1),
                font_size=sp(15),
                bold=True
            )
            self.tab1_btn.bind(on_press=lambda x: self._switch_tab(1))
            self.tab_layout.add_widget(self.tab1_btn)
            
            self.tab2_btn = PersianButton(
                text='خلاصه وصول',
                size_hint_x=0.33,
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(15)
            )
            self.tab2_btn.bind(on_press=lambda x: self._switch_tab(2))
            self.tab_layout.add_widget(self.tab2_btn)
            
            self.tab3_btn = PersianButton(
                text='گزارش تجمیعی',
                size_hint_x=0.34,
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(15)
            )
            self.tab3_btn.bind(on_press=lambda x: self._switch_tab(3))
            self.tab_layout.add_widget(self.tab3_btn)
            
            content.add_widget(self.tab_layout)
            
            # فیلترها
            self.filter_area = BoxLayout(orientation='vertical', size_hint_y=None)
            content.add_widget(self.filter_area)
            
            # محتوای تب
            self.tab_content = BoxLayout(orientation='vertical', size_hint_y=0.75)
            content.add_widget(self.tab_content)
            
            # دکمه بستن
            close_btn = PersianButton(
                text='بستن',
                background_color=(0.3, 0.3, 0.3, 1),
                size_hint_y=None,
                height=dp(48),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            content.add_widget(close_btn)
            
            popup = PersianPopup(
                title='گزارش وصول مطالبات',
                content=content,
                size_hint=(0.95, 0.92),
                background_color=(0.08, 0.08, 0.08, 1),
                auto_dismiss=False
            )
            
            close_btn.bind(on_press=popup.dismiss)
            
            self._switch_tab(1)
            
            popup.open()
            
        except Exception as e:
            error_details = traceback.format_exc()
            ErrorPopup.show_error(f"خطا در نمایش دیالوگ گزارش: {e}", error_details)
    

    def _switch_tab(self, tab_num):
        """تغییر تب"""
        self.current_tab = tab_num
        
        self.tab1_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 1 else (0.3, 0.3, 0.3, 1)
        self.tab2_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 2 else (0.3, 0.3, 0.3, 1)
        self.tab3_btn.background_color = (0.2, 0.5, 0.9, 1) if tab_num == 3 else (0.3, 0.3, 0.3, 1)
        
        if tab_num == 1:
            self._build_list_filters()
            self._apply_filters()
            self._show_report_tab()
        elif tab_num == 2:
            self._build_summary_filters()
            self._apply_filters()
            self._show_summary_tab()
        elif tab_num == 3:
            self._build_aggregated_filters()
            self._apply_aggregated_filters()
            self._show_aggregated_tab()
    

    def _build_list_filters(self):
        """ساخت فیلترهای تب گزارش لیستی"""
        self.filter_area.clear_widgets()
        self.filter_area.height = dp(90)
        
        # ردیف اول
        filter_box1 = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        agents = get_agents()
        agent_names = ['همه'] + [a.get('name', '') for a in agents if a.get('name')]
        self.agent_combo = PersianComboBox(text='همه', values=agent_names, height=dp(36), size_hint_x=0.34)
        self.agent_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.agent_combo.main_btn.color = (1, 1, 1, 1)
        self.agent_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.agent_combo)
        
        routes = get_routes()
        route_names = ['همه'] + [r.get('name', '') for r in routes if r.get('name')]
        self.route_combo = PersianComboBox(text='همه', values=route_names, height=dp(36), size_hint_x=0.33)
        self.route_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.route_combo.main_btn.color = (1, 1, 1, 1)
        self.route_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.route_combo)
        
        self.status_combo = PersianComboBox(
            text='همه', values=['همه', 'موفق', 'ناموفق'], height=dp(36), size_hint_x=0.33
        )
        self.status_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.status_combo.main_btn.color = (1, 1, 1, 1)
        self.status_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.status_combo)
        
        self.filter_area.add_widget(filter_box1)
        
        # ردیف دوم
        filter_box2 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        self.start_date_input = RTLTextInput(
            text='', multiline=False, size_hint_x=0.28, size_hint_y=None, height=dp(34),
            font_size=sp(15), hint_text='از تاریخ'
        )
        self.start_date_input.bg_color = (0.15, 0.15, 0.15, 1)
        self.start_date_input.border_color = (0.3, 0.3, 0.3, 1)
        self.start_date_input._hidden_input.foreground_color = (1, 1, 1, 1)
        filter_box2.add_widget(self.start_date_input)
        
        self.end_date_input = RTLTextInput(
            text='', multiline=False, size_hint_x=0.28, size_hint_y=None, height=dp(34),
            font_size=sp(15), hint_text='تا تاریخ'
        )
        self.end_date_input.bg_color = (0.15, 0.15, 0.15, 1)
        self.end_date_input.border_color = (0.3, 0.3, 0.3, 1)
        self.end_date_input._hidden_input.foreground_color = (1, 1, 1, 1)
        filter_box2.add_widget(self.end_date_input)
        
        apply_btn = PersianButton(
            text='اعمال',
            size_hint_x=0.15, size_hint_y=None, height=dp(34),
            background_color=(0.2, 0.6, 0.2, 1),
            color=(1, 1, 1, 1), font_size=sp(13), bold=True
        )
        apply_btn.bind(on_press=lambda x: self._apply_filters_and_refresh())
        filter_box2.add_widget(apply_btn)
        
        clear_btn = PersianButton(
            text='پاک',
            size_hint_x=0.12, size_hint_y=None, height=dp(34),
            background_color=(0.8, 0.4, 0.2, 1),
            color=(1, 1, 1, 1), font_size=sp(13)
        )
        clear_btn.bind(on_press=lambda x: self._clear_filters())
        filter_box2.add_widget(clear_btn)
        
        self.filter_area.add_widget(filter_box2)
    

    def _build_summary_filters(self):
        """ساخت فیلترهای تب خلاصه وصول"""
        self.filter_area.clear_widgets()
        self.filter_area.height = dp(130)
        
        # ردیف اول
        filter_box1 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        agents = get_agents()
        agent_names = ['همه'] + [a.get('name', '') for a in agents if a.get('name')]
        self.summary_agent_combo = PersianComboBox(text='همه', values=agent_names, height=dp(34), size_hint_x=0.5)
        self.summary_agent_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.summary_agent_combo.main_btn.color = (1, 1, 1, 1)
        self.summary_agent_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.summary_agent_combo)
        
        routes = get_routes()
        route_names = ['همه'] + [r.get('name', '') for r in routes if r.get('name')]
        self.summary_route_combo = PersianComboBox(text='همه', values=route_names, height=dp(34), size_hint_x=0.5)
        self.summary_route_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.summary_route_combo.main_btn.color = (1, 1, 1, 1)
        self.summary_route_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.summary_route_combo)
        
        self.filter_area.add_widget(filter_box1)
        
        # ردیف دوم
        filter_box2 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        customer_container = BoxLayout(size_hint_x=0.5, spacing=dp(3))
        customers = get_customers()
        self._all_customer_names = ['همه'] + [c.get('name', '') for c in customers if c.get('name')]
        
        self.summary_customer_combo = PersianComboBox(
            text='همه',
            values=self._all_customer_names,
            height=dp(34),
            size_hint_x=0.7
        )
        self.summary_customer_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.summary_customer_combo.main_btn.color = (1, 1, 1, 1)
        self.summary_customer_combo.main_btn.font_size = sp(13)
        customer_container.add_widget(self.summary_customer_combo)
        
        search_customer_btn = PersianButton(
            text='جستجو',
            size_hint_x=0.3,
            height=dp(34),
            background_color=(0.2, 0.5, 0.9, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        search_customer_btn.bind(on_press=lambda x: self._show_customer_search_dialog())
        customer_container.add_widget(search_customer_btn)
        filter_box2.add_widget(customer_container)
        
        self.summary_payment_combo = PersianComboBox(
            text='همه',
            values=['همه', 'نقد', 'چک', 'نقد + چک'],
            height=dp(34),
            size_hint_x=0.5
        )
        self.summary_payment_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.summary_payment_combo.main_btn.color = (1, 1, 1, 1)
        self.summary_payment_combo.main_btn.font_size = sp(13)
        filter_box2.add_widget(self.summary_payment_combo)
        
        self.filter_area.add_widget(filter_box2)
        
        # ردیف سوم
        filter_box3 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        bank_names = ['همه'] + get_bank_names()
        self.summary_bank_combo = PersianComboBox(
            text='همه',
            values=bank_names,
            height=dp(34),
            size_hint_x=0.18
        )
        self.summary_bank_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.summary_bank_combo.main_btn.color = (1, 1, 1, 1)
        self.summary_bank_combo.main_btn.font_size = sp(13)
        filter_box3.add_widget(self.summary_bank_combo)
        
        self.summary_start_date = RTLTextInput(
            text='',
            multiline=False,
            size_hint_x=0.20,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(14),
            hint_text='از تاریخ'
        )
        self.summary_start_date.bg_color = (0.15, 0.15, 0.15, 1)
        self.summary_start_date.border_color = (0.3, 0.3, 0.3, 1)
        self.summary_start_date._hidden_input.foreground_color = (1, 1, 1, 1)
        filter_box3.add_widget(self.summary_start_date)
        
        self.summary_end_date = RTLTextInput(
            text='',
            multiline=False,
            size_hint_x=0.20,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(14),
            hint_text='تا تاریخ'
        )
        self.summary_end_date.bg_color = (0.15, 0.15, 0.15, 1)
        self.summary_end_date.border_color = (0.3, 0.3, 0.3, 1)
        self.summary_end_date._hidden_input.foreground_color = (1, 1, 1, 1)
        filter_box3.add_widget(self.summary_end_date)
        
        apply_btn = PersianButton(
            text='اعمال',
            size_hint_x=0.12,
            size_hint_y=None,
            height=dp(32),
            background_color=(0.2, 0.6, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        apply_btn.bind(on_press=lambda x: self._apply_summary_filters())
        filter_box3.add_widget(apply_btn)
        
        clear_btn = PersianButton(
            text='پاک',
            size_hint_x=0.10,
            size_hint_y=None,
            height=dp(32),
            background_color=(0.8, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        clear_btn.bind(on_press=lambda x: self._clear_summary_filters())
        filter_box3.add_widget(clear_btn)
        
        self.filter_area.add_widget(filter_box3)


    def _build_aggregated_filters(self):
        """ساخت فیلترهای تب گزارش تجمیعی"""
        self.filter_area.clear_widgets()
        self.filter_area.height = dp(130)
        
        # ردیف اول
        filter_box1 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        agents = get_agents()
        agent_names = ['همه'] + [a.get('name', '') for a in agents if a.get('name')]
        self.agg_agent_combo = PersianComboBox(text='همه', values=agent_names, height=dp(34), size_hint_x=0.5)
        self.agg_agent_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.agg_agent_combo.main_btn.color = (1, 1, 1, 1)
        self.agg_agent_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.agg_agent_combo)
        
        routes = get_routes()
        route_names = ['همه'] + [r.get('name', '') for r in routes if r.get('name')]
        self.agg_route_combo = PersianComboBox(text='همه', values=route_names, height=dp(34), size_hint_x=0.5)
        self.agg_route_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.agg_route_combo.main_btn.color = (1, 1, 1, 1)
        self.agg_route_combo.main_btn.font_size = sp(13)
        filter_box1.add_widget(self.agg_route_combo)
        
        self.filter_area.add_widget(filter_box1)
        
        # ردیف دوم
        filter_box2 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        bank_names = ['همه'] + get_bank_names()
        self.agg_bank_combo = PersianComboBox(
            text='همه',
            values=bank_names,
            height=dp(34),
            size_hint_x=0.5
        )
        self.agg_bank_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.agg_bank_combo.main_btn.color = (1, 1, 1, 1)
        self.agg_bank_combo.main_btn.font_size = sp(13)
        filter_box2.add_widget(self.agg_bank_combo)
        
        self.agg_payment_combo = PersianComboBox(
            text='همه',
            values=['همه', 'نقد', 'چک', 'نقد + چک'],
            height=dp(34),
            size_hint_x=0.5
        )
        self.agg_payment_combo.main_btn.background_color = (0.2, 0.2, 0.2, 1)
        self.agg_payment_combo.main_btn.color = (1, 1, 1, 1)
        self.agg_payment_combo.main_btn.font_size = sp(13)
        filter_box2.add_widget(self.agg_payment_combo)
        
        self.filter_area.add_widget(filter_box2)
        
        # ردیف سوم
        filter_box3 = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(5), padding=[0, dp(2), 0, dp(2)])
        
        self.agg_start_date = RTLTextInput(
            text='',
            multiline=False,
            size_hint_x=0.22,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(14),
            hint_text='از تاریخ'
        )
        self.agg_start_date.bg_color = (0.15, 0.15, 0.15, 1)
        self.agg_start_date.border_color = (0.3, 0.3, 0.3, 1)
        self.agg_start_date._hidden_input.foreground_color = (1, 1, 1, 1)
        filter_box3.add_widget(self.agg_start_date)
        
        self.agg_end_date = RTLTextInput(
            text='',
            multiline=False,
            size_hint_x=0.22,
            size_hint_y=None,
            height=dp(32),
            font_size=sp(14),
            hint_text='تا تاریخ'
        )
        self.agg_end_date.bg_color = (0.15, 0.15, 0.15, 1)
        self.agg_end_date.border_color = (0.3, 0.3, 0.3, 1)
        self.agg_end_date._hidden_input.foreground_color = (1, 1, 1, 1)
        filter_box3.add_widget(self.agg_end_date)
        
        apply_btn = PersianButton(
            text='اعمال',
            size_hint_x=0.12,
            size_hint_y=None,
            height=dp(32),
            background_color=(0.2, 0.6, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13),
            bold=True
        )
        apply_btn.bind(on_press=lambda x: self._apply_aggregated_filters())
        filter_box3.add_widget(apply_btn)
        
        clear_btn = PersianButton(
            text='پاک',
            size_hint_x=0.10,
            size_hint_y=None,
            height=dp(32),
            background_color=(0.8, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(13)
        )
        clear_btn.bind(on_press=lambda x: self._clear_aggregated_filters())
        filter_box3.add_widget(clear_btn)
        
        self.filter_area.add_widget(filter_box3)


    def _show_customer_search_dialog(self):
        """دیالوگ جستجوی مشتری"""
        try:
            content = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                           size=lambda i, v: setattr(rect, 'size', v))
            
            search_input = RTLTextInput(
                hint_text='جستجوی مشتری...',
                multiline=False,
                size_hint_y=None,
                height=dp(55),
                font_size=sp(20)
            )
            search_input.bg_color = (0.15, 0.15, 0.15, 1)
            search_input.border_color = (0.3, 0.3, 0.3, 1)
            search_input._hidden_input.foreground_color = (1, 1, 1, 1)
            content.add_widget(search_input)
            
            scroll = ScrollView(do_scroll_x=False, do_scroll_y=True, size_hint_y=0.7)
            grid = GridLayout(cols=1, spacing=dp(3), size_hint_y=None, padding=dp(5))
            grid.bind(minimum_height=grid.setter('height'))
            
            def do_search(text):
                grid.clear_widgets()
                search_text = text.strip()
                
                filtered = ['همه']
                for name in self._all_customer_names[1:]:
                    if search_text in name:
                        filtered.append(name)
                
                for name in filtered:
                    btn = PersianButton(
                        text=name,
                        size_hint_y=None,
                        height=dp(48),
                        background_color=(0.2, 0.5, 0.9, 1),
                        color=(1, 1, 1, 1),
                        font_size=sp(16)
                    )
                    btn.bind(on_press=lambda x, n=name: self._select_summary_customer(n))
                    grid.add_widget(btn)
            
            search_input._hidden_input.bind(text=lambda i, v: do_search(v))
            do_search('')
            
            scroll.add_widget(grid)
            content.add_widget(scroll)
            
            close_btn = PersianButton(
                text='بستن',
                size_hint_y=None,
                height=dp(48),
                background_color=(0.3, 0.3, 0.3, 1),
                color=(1, 1, 1, 1),
                font_size=sp(16)
            )
            content.add_widget(close_btn)
            
            self._customer_search_popup = PersianPopup(
                title='انتخاب مشتری',
                content=content,
                size_hint=(0.85, 0.65),
                auto_dismiss=True
            )
            close_btn.bind(on_press=self._customer_search_popup.dismiss)
            self._customer_search_popup.open()
            
        except Exception as e:
            print(f"خطا در دیالوگ جستجوی مشتری: {e}")


    def _select_summary_customer(self, customer_name):
        """انتخاب مشتری از دیالوگ جستجو"""
        self.summary_customer_combo.text = customer_name
        if hasattr(self, '_customer_search_popup'):
            self._customer_search_popup.dismiss()


    def _clear_filters(self):
        """پاک کردن فیلترهای تب گزارش لیستی"""
        self.agent_combo.text = 'همه'
        self.route_combo.text = 'همه'
        self.status_combo.text = 'همه'
        self.start_date_input.text = ''
        self.end_date_input.text = ''
        self._apply_filters_and_refresh()
    

    def _clear_summary_filters(self):
        """پاک کردن فیلترهای تب خلاصه"""
        self.summary_agent_combo.text = 'همه'
        self.summary_route_combo.text = 'همه'
        self.summary_customer_combo.text = 'همه'
        self.summary_payment_combo.text = 'همه'
        self.summary_bank_combo.text = 'همه'
        self.summary_start_date.text = ''
        self.summary_end_date.text = ''
        self._apply_summary_filters()
    

    def _clear_aggregated_filters(self):
        """پاک کردن فیلترهای تب تجمیعی"""
        self.agg_agent_combo.text = 'همه'
        self.agg_route_combo.text = 'همه'
        self.agg_bank_combo.text = 'همه'
        self.agg_payment_combo.text = 'همه'
        self.agg_start_date.text = ''
        self.agg_end_date.text = ''
        self._apply_aggregated_filters()


    def _apply_filters_and_refresh(self):
        """اعمال فیلترها و بروزرسانی جدول"""
        self._apply_filters()
        self._show_report_tab()
    

    def _apply_summary_filters(self):
        """اعمال فیلترهای خلاصه"""
        self._apply_filters()
        self._show_summary_tab()
    

    def _apply_aggregated_filters(self):
        """اعمال فیلترهای تب تجمیعی"""
        self.agg_filtered_data = []
        
        agent = None
        route = None
        bank = None
        payment = None
        
        if hasattr(self, 'agg_agent_combo') and self.agg_agent_combo.text != 'همه':
            agent = self.agg_agent_combo.text
        
        if hasattr(self, 'agg_route_combo') and self.agg_route_combo.text != 'همه':
            route = self.agg_route_combo.text
        
        if hasattr(self, 'agg_bank_combo') and self.agg_bank_combo.text != 'همه':
            bank = self.agg_bank_combo.text
        
        if hasattr(self, 'agg_payment_combo') and self.agg_payment_combo.text != 'همه':
            payment = self.agg_payment_combo.text
        
        data = get_collections(agent_name=agent)
        
        if route:
            data = [c for c in data if c.get('route') == route]
        
        if bank:
            data = [c for c in data if c.get('bank') == bank]
        
        if payment == 'نقد':
            data = [c for c in data if c.get('has_cash') and not c.get('has_check')]
        elif payment == 'چک':
            data = [c for c in data if c.get('has_check') and not c.get('has_cash')]
        elif payment == 'نقد + چک':
            data = [c for c in data if c.get('has_cash') and c.get('has_check')]
        
        start = self.agg_start_date.text.strip() if hasattr(self, 'agg_start_date') else ''
        end = self.agg_end_date.text.strip() if hasattr(self, 'agg_end_date') else ''
        
        if start:
            data = [c for c in data if c.get('date', '') >= start]
        if end:
            data = [c for c in data if c.get('date', '') <= end]
        
        self.agg_filtered_data = data
        self._show_aggregated_tab()


    def _apply_filters(self):
        """اعمال فیلترها بر اساس تب فعلی"""
        if self.current_tab == 1:
            self.agent_filter = self.agent_combo.text if hasattr(self, 'agent_combo') else 'همه'
            self.route_filter = self.route_combo.text if hasattr(self, 'route_combo') else 'همه'
            self.status_filter = self.status_combo.text if hasattr(self, 'status_combo') else 'همه'
            
            agent = self.agent_filter if self.agent_filter != 'همه' else None
            route = self.route_filter if self.route_filter != 'همه' else None
            status = self.status_filter if self.status_filter != 'همه' else None
            
            self.filtered_data = get_collections(agent_name=agent, status=status)
            
            if route:
                self.filtered_data = [c for c in self.filtered_data if c.get('route') == route]
            
            start = self.start_date_input.text.strip() if hasattr(self, 'start_date_input') else ''
            end = self.end_date_input.text.strip() if hasattr(self, 'end_date_input') else ''
            
            if start:
                self.filtered_data = [c for c in self.filtered_data if c.get('date', '') >= start]
            if end:
                self.filtered_data = [c for c in self.filtered_data if c.get('date', '') <= end]
        
        elif self.current_tab == 2:
            self.agent_filter = self.summary_agent_combo.text if hasattr(self, 'summary_agent_combo') else 'همه'
            self.route_filter = self.summary_route_combo.text if hasattr(self, 'summary_route_combo') else 'همه'
            
            agent = self.agent_filter if self.agent_filter != 'همه' else None
            route = self.route_filter if self.route_filter != 'همه' else None
            
            self.filtered_data = get_collections(agent_name=agent)
            
            if route:
                self.filtered_data = [c for c in self.filtered_data if c.get('route') == route]
            
            customer_filter = self.summary_customer_combo.text if hasattr(self, 'summary_customer_combo') else 'همه'
            if customer_filter != 'همه':
                self.filtered_data = [c for c in self.filtered_data if c.get('customer') == customer_filter]
            
            payment_filter = self.summary_payment_combo.text if hasattr(self, 'summary_payment_combo') else 'همه'
            if payment_filter == 'نقد':
                self.filtered_data = [c for c in self.filtered_data if c.get('has_cash') and not c.get('has_check')]
            elif payment_filter == 'چک':
                self.filtered_data = [c for c in self.filtered_data if c.get('has_check') and not c.get('has_cash')]
            elif payment_filter == 'نقد + چک':
                self.filtered_data = [c for c in self.filtered_data if c.get('has_cash') and c.get('has_check')]
            
            bank_filter = self.summary_bank_combo.text if hasattr(self, 'summary_bank_combo') else 'همه'
            if bank_filter != 'همه':
                self.filtered_data = [c for c in self.filtered_data if c.get('bank') == bank_filter]
            
            start = self.summary_start_date.text.strip() if hasattr(self, 'summary_start_date') else ''
            end = self.summary_end_date.text.strip() if hasattr(self, 'summary_end_date') else ''
            
            if start:
                self.filtered_data = [c for c in self.filtered_data if c.get('date', '') >= start]
            if end:
                self.filtered_data = [c for c in self.filtered_data if c.get('date', '') <= end]


    def _show_report_tab(self):
        """نمایش تب گزارش لیستی وصول"""
        self.tab_content.clear_widgets()
        
        if not self.filtered_data:
            self.tab_content.add_widget(RTLLabel(
                text='هیچ وصولی با این فیلترها یافت نشد',
                size_hint_y=None, height=dp(50),
                font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
            ))
            return
        
        toolbar = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8), padding=[dp(5), dp(3), dp(5), dp(3)])
        toolbar.add_widget(RTLLabel(
            text=f'تعداد وصول: {len(self.filtered_data)}',
            size_hint_x=0.5,
            font_size=sp(14),
            color=(0.6, 0.8, 1, 1),
            halign='right'
        ))
        
        btn_layout = BoxLayout(size_hint_x=0.5, spacing=dp(5))
        excel_btn = PersianButton(
            text='خروجی اکسل',
            size_hint_x=0.5,
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12)
        )
        excel_btn.bind(on_press=lambda x: self._export_excel_only())
        btn_layout.add_widget(excel_btn)
        
        image_btn = PersianButton(
            text='خروجی تصویری',
            size_hint_x=0.5,
            background_color=(0.6, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12)
        )
        image_btn.bind(on_press=lambda x: self._export_image_only())
        btn_layout.add_widget(image_btn)
        
        toolbar.add_widget(btn_layout)
        self.tab_content.add_widget(toolbar)
        
        # هدر جدول
        header_box = BoxLayout(size_hint_y=None, height=dp(38), spacing=dp(2))
        headers = [
            ('ردیف', 0.09), ('تاریخ', 0.11), ('مشتری', 0.18), ('وضعیت', 0.09),
            ('نوع', 0.12), ('نقد', 0.13), ('چک', 0.12), ('جمع', 0.13)
        ]
        for text, size in headers:
            header_box.add_widget(RTLLabel(
                text=text, size_hint_x=size, size_hint_y=None, height=dp(36),
                font_size=sp(14), bold=True, color=(0.4, 0.7, 1, 1), halign='center'
            ))
        self.tab_content.add_widget(header_box)
        
        separator = BoxLayout(size_hint_y=None, height=dp(1))
        with separator.canvas.before:
            Color(0.3, 0.3, 0.3, 1)
            Rectangle(pos=separator.pos, size=separator.size)
        self.tab_content.add_widget(separator)
        
        table_scroll = ScrollView(
            do_scroll_x=False, do_scroll_y=True, size_hint_y=1,
            scroll_type=['bars', 'content'], bar_width=dp(5)
        )
        
        table_grid = GridLayout(cols=1, spacing=dp(1), size_hint_y=None, padding=[0, dp(2), 0, dp(2)])
        table_grid.bind(minimum_height=table_grid.setter('height'))
        
        for idx, col in enumerate(self.filtered_data):
            date = col.get('date', '')
            customer = col.get('customer', '')
            status = col.get('status', '')
            
            if col.get('has_cash') and col.get('has_check'):
                payment_type = 'نقد + چک'
            elif col.get('has_cash'):
                payment_type = 'نقد'
            elif col.get('has_check'):
                payment_type = 'چک'
            else:
                payment_type = '-'
            
            cash_amount = col.get('net_cash', 0)
            check_amount = col.get('total_check_amount', 0)
            total = col.get('total_collection', 0)
            
            is_even = idx % 2 == 0
            row_color = (0.18, 0.18, 0.22, 1) if is_even else (0.14, 0.14, 0.18, 1)
            
            row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(2))
            with row.canvas.before:
                Color(*row_color)
                Rectangle(pos=row.pos, size=row.size)
            
            status_color = (0.2, 0.8, 0.2, 1) if status == 'موفق' else (0.8, 0.2, 0.2, 1)
            
            row_data = [
                (str(idx + 1), 0.09, (0.7, 0.7, 0.7, 1)),
                (date, 0.11, (1, 1, 1, 1)),
                (customer, 0.18, (1, 1, 1, 1)),
                (status, 0.09, status_color),
                (payment_type, 0.12, (1, 1, 1, 1)),
                (f"{cash_amount:,.0f}" if cash_amount else '-', 0.13, (1, 1, 1, 1)),
                (f"{check_amount:,.0f}" if check_amount else '-', 0.12, (1, 1, 1, 1)),
                (f"{total:,.0f}", 0.13, (1, 1, 1, 1)),
            ]
            
            for text, size, color in row_data:
                row.add_widget(RTLLabel(
                    text=str(text), size_hint_x=size, size_hint_y=None, height=dp(40),
                    font_size=sp(13), color=color, halign='center'
                ))
            
            table_grid.add_widget(row)
        
        table_scroll.add_widget(table_grid)
        self.tab_content.add_widget(table_scroll)


    def _show_summary_tab(self):
        """نمایش تب خلاصه وصول با کارت های عامل، مسیر و مشتری"""
        self.tab_content.clear_widgets()
        
        if not self.filtered_data:
            self.tab_content.add_widget(RTLLabel(
                text='داده ای برای نمایش وجود ندارد',
                size_hint_y=None, height=dp(50),
                font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
            ))
            return
        
        toolbar = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(4), padding=[dp(3), dp(3), dp(3), dp(3)])
        
        toolbar.add_widget(RTLLabel(
            text=f'تعداد وصول: {len(self.filtered_data)}',
            size_hint_x=0.30,
            font_size=sp(12),
            color=(0.6, 0.8, 1, 1),
            halign='right'
        ))
        
        btn_layout = BoxLayout(size_hint_x=0.70, spacing=dp(4))
        
        # دکمه خروجی اکسل
        excel_btn = PersianButton(
            text='اکسل',
            size_hint_x=0.20,
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(10)
        )
        excel_btn.bind(on_press=lambda x: self._export_excel_only())
        btn_layout.add_widget(excel_btn)
        
        # دکمه خروجی تصویری
        image_btn = PersianButton(
            text='تصویر',
            size_hint_x=0.20,
            background_color=(0.6, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(10)
        )
        image_btn.bind(on_press=lambda x: self._export_image_only())
        btn_layout.add_widget(image_btn)
        
        # دکمه خروجی چک اکسل
        check_excel_btn = PersianButton(
            text='چک اکسل',
            size_hint_x=0.20,
            background_color=(0.4, 0.2, 0.6, 1),
            color=(1, 1, 1, 1),
            font_size=sp(10)
        )
        check_excel_btn.bind(on_press=lambda x: self._export_checks_excel())
        btn_layout.add_widget(check_excel_btn)
        
        # دکمه خروجی چک تصویری
        check_image_btn = PersianButton(
            text='چک تصویر',
            size_hint_x=0.20,
            background_color=(0.6, 0.2, 0.4, 1),
            color=(1, 1, 1, 1),
            font_size=sp(10)
        )
        check_image_btn.bind(on_press=lambda x: self._export_checks_image())
        btn_layout.add_widget(check_image_btn)
        
        toolbar.add_widget(btn_layout)
        self.tab_content.add_widget(toolbar)
        
        scroll = ScrollView(
            do_scroll_x=False, do_scroll_y=True, size_hint_y=1,
            scroll_type=['bars', 'content'], bar_width=dp(5)
        )
        
        content = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None, padding=dp(5))
        content.bind(minimum_height=content.setter('height'))
        
        # کارت عامل
        agent_stats = {}
        for c in self.filtered_data:
            agent = c.get('agent_name', 'نامشخص')
            if agent not in agent_stats:
                agent_stats[agent] = {'count': 0, 'cash': 0, 'check': 0, 'success': 0, 'fail': 0}
            agent_stats[agent]['count'] += 1
            if c.get('status') == 'موفق':
                agent_stats[agent]['success'] += 1
                agent_stats[agent]['cash'] += c.get('net_cash', 0)
                agent_stats[agent]['check'] += c.get('total_check_amount', 0)
            else:
                agent_stats[agent]['fail'] += 1
        
        for agent, data in agent_stats.items():
            total = data['cash'] + data['check']
            
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(130), padding=dp(10), spacing=dp(5))
            with card.canvas.before:
                Color(0.15, 0.2, 0.3, 1)
                card.rect = Rectangle(pos=card.pos, size=card.size)
            card.bind(pos=self._update_card_rect, size=self._update_card_rect)
            
            card.add_widget(RTLLabel(
                text=f'عامل: {agent}', size_hint_y=None, height=dp(30),
                font_size=sp(16), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            
            info_layout = BoxLayout(size_hint_y=None, height=dp(25), spacing=dp(10))
            info_layout.add_widget(RTLLabel(text=f'کل: {data["count"]}', size_hint_x=0.25, font_size=sp(14), color=(1, 1, 1, 1)))
            info_layout.add_widget(RTLLabel(text=f'موفق: {data["success"]}', size_hint_x=0.25, font_size=sp(14), color=(0.2, 0.8, 0.2, 1)))
            info_layout.add_widget(RTLLabel(text=f'ناموفق: {data["fail"]}', size_hint_x=0.25, font_size=sp(14), color=(0.8, 0.2, 0.2, 1)))
            info_layout.add_widget(RTLLabel(text=f'جمع: {total:,.0f}', size_hint_x=0.25, font_size=sp(14), color=(0.8, 0.6, 0.2, 1)))
            card.add_widget(info_layout)
            
            detail_layout = BoxLayout(size_hint_y=None, height=dp(25), spacing=dp(10))
            detail_layout.add_widget(RTLLabel(text=f'نقد: {data["cash"]:,.0f}', size_hint_x=0.33, font_size=sp(13), color=(0.6, 0.8, 0.6, 1)))
            detail_layout.add_widget(RTLLabel(text=f'چک: {data["check"]:,.0f}', size_hint_x=0.33, font_size=sp(13), color=(0.8, 0.6, 0.8, 1)))
            avg = (total/data["count"]) if data["count"] > 0 else 0
            detail_layout.add_widget(RTLLabel(text=f'متوسط: {avg:,.0f}', size_hint_x=0.34, font_size=sp(13), color=(0.6, 0.6, 0.6, 1)))
            card.add_widget(detail_layout)
            
            content.add_widget(card)
        
        content.add_widget(Label(size_hint_y=None, height=dp(10)))
        
        # کارت مسیر
        route_stats = {}
        for c in self.filtered_data:
            route = c.get('route', 'نامشخص')
            if route not in route_stats:
                route_stats[route] = {'count': 0, 'cash': 0, 'check': 0, 'success': 0, 'fail': 0}
            route_stats[route]['count'] += 1
            if c.get('status') == 'موفق':
                route_stats[route]['success'] += 1
                route_stats[route]['cash'] += c.get('net_cash', 0)
                route_stats[route]['check'] += c.get('total_check_amount', 0)
            else:
                route_stats[route]['fail'] += 1
        
        for route, data in route_stats.items():
            total = data['cash'] + data['check']
            
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(130), padding=dp(10), spacing=dp(5))
            with card.canvas.before:
                Color(0.2, 0.15, 0.25, 1)
                card.rect = Rectangle(pos=card.pos, size=card.size)
            card.bind(pos=self._update_card_rect, size=self._update_card_rect)
            
            card.add_widget(RTLLabel(
                text=f'مسیر: {route}', size_hint_y=None, height=dp(30),
                font_size=sp(16), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            
            info_layout = BoxLayout(size_hint_y=None, height=dp(25), spacing=dp(10))
            info_layout.add_widget(RTLLabel(text=f'کل: {data["count"]}', size_hint_x=0.25, font_size=sp(14), color=(1, 1, 1, 1)))
            info_layout.add_widget(RTLLabel(text=f'موفق: {data["success"]}', size_hint_x=0.25, font_size=sp(14), color=(0.2, 0.8, 0.2, 1)))
            info_layout.add_widget(RTLLabel(text=f'ناموفق: {data["fail"]}', size_hint_x=0.25, font_size=sp(14), color=(0.8, 0.2, 0.2, 1)))
            info_layout.add_widget(RTLLabel(text=f'جمع: {total:,.0f}', size_hint_x=0.25, font_size=sp(14), color=(0.8, 0.6, 0.2, 1)))
            card.add_widget(info_layout)
            
            detail_layout = BoxLayout(size_hint_y=None, height=dp(25), spacing=dp(10))
            detail_layout.add_widget(RTLLabel(text=f'نقد: {data["cash"]:,.0f}', size_hint_x=0.33, font_size=sp(13), color=(0.6, 0.8, 0.6, 1)))
            detail_layout.add_widget(RTLLabel(text=f'چک: {data["check"]:,.0f}', size_hint_x=0.33, font_size=sp(13), color=(0.8, 0.6, 0.8, 1)))
            avg = (total/data["count"]) if data["count"] > 0 else 0
            detail_layout.add_widget(RTLLabel(text=f'متوسط: {avg:,.0f}', size_hint_x=0.34, font_size=sp(13), color=(0.6, 0.6, 0.6, 1)))
            card.add_widget(detail_layout)
            
            content.add_widget(card)
        
        content.add_widget(Label(size_hint_y=None, height=dp(10)))
        
        # کارت مشتری
        customer_stats = {}
        for c in self.filtered_data:
            customer = c.get('customer', 'نامشخص')
            if customer not in customer_stats:
                customer_stats[customer] = {
                    'count': 0, 'cash': 0, 'check': 0, 'success': 0, 'fail': 0,
                    'agent': c.get('agent_name', ''), 'route': c.get('route', '')
                }
            customer_stats[customer]['count'] += 1
            if c.get('status') == 'موفق':
                customer_stats[customer]['success'] += 1
                customer_stats[customer]['cash'] += c.get('net_cash', 0)
                customer_stats[customer]['check'] += c.get('total_check_amount', 0)
            else:
                customer_stats[customer]['fail'] += 1
        
        for customer, data in customer_stats.items():
            total = data['cash'] + data['check']
            
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(150), padding=dp(10), spacing=dp(5))
            with card.canvas.before:
                Color(0.15, 0.22, 0.18, 1)
                card.rect = Rectangle(pos=card.pos, size=card.size)
            card.bind(pos=self._update_card_rect, size=self._update_card_rect)
            
            card.add_widget(RTLLabel(
                text=f'مشتری: {customer}', size_hint_y=None, height=dp(30),
                font_size=sp(16), bold=True, color=(0.4, 0.7, 1, 1)
            ))
            
            sub_info = BoxLayout(size_hint_y=None, height=dp(22), spacing=dp(10))
            sub_info.add_widget(RTLLabel(text=f'عامل: {data["agent"]}', size_hint_x=0.5, font_size=sp(12), color=(0.6, 0.6, 0.6, 1)))
            sub_info.add_widget(RTLLabel(text=f'مسیر: {data["route"]}', size_hint_x=0.5, font_size=sp(12), color=(0.6, 0.6, 0.6, 1)))
            card.add_widget(sub_info)
            
            info_layout = BoxLayout(size_hint_y=None, height=dp(25), spacing=dp(10))
            info_layout.add_widget(RTLLabel(text=f'کل: {data["count"]}', size_hint_x=0.25, font_size=sp(14), color=(1, 1, 1, 1)))
            info_layout.add_widget(RTLLabel(text=f'موفق: {data["success"]}', size_hint_x=0.25, font_size=sp(14), color=(0.2, 0.8, 0.2, 1)))
            info_layout.add_widget(RTLLabel(text=f'ناموفق: {data["fail"]}', size_hint_x=0.25, font_size=sp(14), color=(0.8, 0.2, 0.2, 1)))
            info_layout.add_widget(RTLLabel(text=f'جمع: {total:,.0f}', size_hint_x=0.25, font_size=sp(14), color=(0.8, 0.6, 0.2, 1)))
            card.add_widget(info_layout)
            
            detail_layout = BoxLayout(size_hint_y=None, height=dp(25), spacing=dp(10))
            detail_layout.add_widget(RTLLabel(text=f'نقد: {data["cash"]:,.0f}', size_hint_x=0.33, font_size=sp(13), color=(0.6, 0.8, 0.6, 1)))
            detail_layout.add_widget(RTLLabel(text=f'چک: {data["check"]:,.0f}', size_hint_x=0.33, font_size=sp(13), color=(0.8, 0.6, 0.8, 1)))
            avg = (total/data["count"]) if data["count"] > 0 else 0
            detail_layout.add_widget(RTLLabel(text=f'متوسط: {avg:,.0f}', size_hint_x=0.34, font_size=sp(13), color=(0.6, 0.6, 0.6, 1)))
            card.add_widget(detail_layout)
            
            content.add_widget(card)
        
        scroll.add_widget(content)
        self.tab_content.add_widget(scroll)


    def _show_aggregated_tab(self):
        """نمایش تب گزارش تجمیعی - کارت هر عامل"""
        self.tab_content.clear_widgets()
        
        if not hasattr(self, 'agg_filtered_data') or not self.agg_filtered_data:
            self.tab_content.add_widget(RTLLabel(
                text='داده ای برای نمایش وجود ندارد',
                size_hint_y=None, height=dp(50),
                font_size=sp(16), color=(0.5, 0.5, 0.5, 1)
            ))
            return
        
        toolbar = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8), padding=[dp(5), dp(3), dp(5), dp(3)])
        toolbar.add_widget(RTLLabel(
            text=f'تعداد عوامل: {len(self.agg_filtered_data)}',
            size_hint_x=0.5,
            font_size=sp(14),
            color=(0.6, 0.8, 1, 1),
            halign='right'
        ))
        
        btn_layout = BoxLayout(size_hint_x=0.5, spacing=dp(5))
        excel_btn = PersianButton(
            text='خروجی اکسل',
            size_hint_x=0.5,
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12)
        )
        excel_btn.bind(on_press=lambda x: self._export_aggregated_excel())
        btn_layout.add_widget(excel_btn)
        
        image_btn = PersianButton(
            text='خروجی تصویری',
            size_hint_x=0.5,
            background_color=(0.6, 0.4, 0.2, 1),
            color=(1, 1, 1, 1),
            font_size=sp(12)
        )
        image_btn.bind(on_press=lambda x: self._export_aggregated_image())
        btn_layout.add_widget(image_btn)
        
        toolbar.add_widget(btn_layout)
        self.tab_content.add_widget(toolbar)
        
        scroll = ScrollView(
            do_scroll_x=False, do_scroll_y=True, size_hint_y=1,
            scroll_type=['bars', 'content'], bar_width=dp(5)
        )
        
        content = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None, padding=dp(5))
        content.bind(minimum_height=content.setter('height'))
        
        agent_data = {}
        for item in self.agg_filtered_data:
            agent_name = item.get('agent_name', 'نامشخص')
            if agent_name not in agent_data:
                agent_data[agent_name] = {
                    'count': 0,
                    'success_count': 0,
                    'fail_count': 0,
                    'total_cash': 0,
                    'total_check': 0,
                    'total_amount': 0,
                    'customers': set(),
                    'routes': set()
                }
            
            agent_data[agent_name]['count'] += 1
            if item.get('status') == 'موفق':
                agent_data[agent_name]['success_count'] += 1
                agent_data[agent_name]['total_cash'] += item.get('net_cash', 0)
                agent_data[agent_name]['total_check'] += item.get('total_check_amount', 0)
            else:
                agent_data[agent_name]['fail_count'] += 1
            
            agent_data[agent_name]['total_amount'] += item.get('total_collection', 0)
            agent_data[agent_name]['customers'].add(item.get('customer', ''))
            agent_data[agent_name]['routes'].add(item.get('route', ''))
        
        for idx, (agent_name, data) in enumerate(agent_data.items()):
            total = data['total_amount']
            cash = data['total_cash']
            check = data['total_check']
            count = data['count']
            success = data['success_count']
            fail = data['fail_count']
            customers_count = len(data['customers'])
            routes_count = len(data['routes'])
            
            colors = [
                (0.15, 0.2, 0.3, 1),
                (0.2, 0.15, 0.3, 1),
                (0.15, 0.25, 0.2, 1),
                (0.25, 0.15, 0.15, 1),
                (0.15, 0.2, 0.25, 1)
            ]
            color = colors[idx % len(colors)]
            
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(160), padding=dp(12), spacing=dp(5))
            with card.canvas.before:
                Color(*color)
                card.rect = Rectangle(pos=card.pos, size=card.size)
            card.bind(pos=self._update_card_rect, size=self._update_card_rect)
            
            card.add_widget(RTLLabel(
                text=f'عامل: {agent_name}',
                size_hint_y=None,
                height=dp(32),
                font_size=sp(17),
                bold=True,
                color=(0.4, 0.8, 1, 1)
            ))
            
            row1 = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(10))
            row1.add_widget(RTLLabel(
                text=f'کل وصول: {count}',
                size_hint_x=0.2,
                font_size=sp(14),
                color=(1, 1, 1, 1)
            ))
            row1.add_widget(RTLLabel(
                text=f'موفق: {success}',
                size_hint_x=0.2,
                font_size=sp(14),
                color=(0.2, 0.9, 0.2, 1)
            ))
            row1.add_widget(RTLLabel(
                text=f'ناموفق: {fail}',
                size_hint_x=0.2,
                font_size=sp(14),
                color=(0.9, 0.2, 0.2, 1)
            ))
            row1.add_widget(RTLLabel(
                text=f'مشتریان: {customers_count}',
                size_hint_x=0.2,
                font_size=sp(14),
                color=(0.8, 0.8, 0.2, 1)
            ))
            row1.add_widget(RTLLabel(
                text=f'مسیرها: {routes_count}',
                size_hint_x=0.2,
                font_size=sp(14),
                color=(0.2, 0.8, 0.8, 1)
            ))
            card.add_widget(row1)
            
            row2 = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(10))
            row2.add_widget(RTLLabel(
                text=f'جمع نقد: {cash:,.0f}',
                size_hint_x=0.33,
                font_size=sp(14),
                color=(0.6, 0.9, 0.6, 1)
            ))
            row2.add_widget(RTLLabel(
                text=f'جمع چک: {check:,.0f}',
                size_hint_x=0.33,
                font_size=sp(14),
                color=(0.9, 0.6, 0.9, 1)
            ))
            row2.add_widget(RTLLabel(
                text=f'جمع کل: {total:,.0f}',
                size_hint_x=0.34,
                font_size=sp(14),
                bold=True,
                color=(0.9, 0.8, 0.2, 1)
            ))
            card.add_widget(row2)
            
            avg = total / count if count > 0 else 0
            success_rate = (success / count * 100) if count > 0 else 0
            
            row3 = BoxLayout(size_hint_y=None, height=dp(28), spacing=dp(10))
            row3.add_widget(RTLLabel(
                text=f'میانگین هر وصول: {avg:,.0f}',
                size_hint_x=0.5,
                font_size=sp(13),
                color=(0.7, 0.7, 0.7, 1)
            ))
            row3.add_widget(RTLLabel(
                text=f'نرخ موفقیت: {success_rate:.1f}%',
                size_hint_x=0.5,
                font_size=sp(13),
                color=(0.4, 0.8, 0.4, 1)
            ))
            card.add_widget(row3)
            
            content.add_widget(card)
        
        scroll.add_widget(content)
        self.tab_content.add_widget(scroll)


    def _update_card_rect(self, instance, value):
        """بروزرسانی مستطیل پس زمینه کارت"""
        if hasattr(instance, 'rect'):
            instance.rect.pos = instance.pos
            instance.rect.size = instance.size


    def _export_excel_only(self):
        """خروجی اکسل"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            if not self.filtered_data:
                self.show_message('خطا', 'داده ای برای خروجی وجود ندارد')
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش وصول"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            success_fill = PatternFill(start_color="235347", end_color="235347", fill_type="solid")
            fail_fill = PatternFill(start_color="78281F", end_color="78281F", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = [
                'ردیف', 'شناسه', 'تاریخ', 'عامل', 'مشتری', 'مسیر', 'وضعیت',
                'نوع پرداخت', 'نوع وصول نقدی', 'مبلغ نقد (ریال)', 'کسورات (ریال)',
                'خالص نقد (ریال)', 'بانک', 'شماره پیگیری', 'تعداد چک',
                'جمع چک ها (ریال)', 'جمع کل (ریال)', 'علت عدم وصول',
                'تاریخ پیگیری بعدی', 'توضیحات'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for row_idx, col in enumerate(self.filtered_data, 2):
                is_success = col.get('status') == 'موفق'
                row_fill = success_fill if is_success else fail_fill
                
                if col.get('has_cash') and col.get('has_check'):
                    payment_type = 'نقد + چک'
                elif col.get('has_cash'):
                    payment_type = 'نقد'
                elif col.get('has_check'):
                    payment_type = 'چک'
                else:
                    payment_type = '-'
                
                values = [
                    row_idx - 1,
                    col.get('id', ''),
                    col.get('date', ''),
                    col.get('agent_name', ''),
                    col.get('customer', ''),
                    col.get('route', ''),
                    col.get('status', ''),
                    payment_type,
                    col.get('cash_type', ''),
                    col.get('cash_amount', 0),
                    col.get('cash_deductions', 0),
                    col.get('net_cash', 0),
                    col.get('bank', ''),
                    col.get('tracking_number', ''),
                    len(col.get('checks', [])),
                    col.get('total_check_amount', 0),
                    col.get('total_collection', 0),
                    col.get('fail_reason', ''),
                    col.get('next_follow_up_date', ''),
                    col.get('description', '')
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.font = Font(color="FFFFFF")
            
            column_widths = [6, 16, 12, 18, 20, 14, 10, 12, 14, 16, 14, 16, 14, 16, 10, 16, 16, 18, 14, 22]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            today = get_today_jalali().replace('/', '-')
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            excel_filename = f'گزارش_وصول_{today}_{datetime.now().strftime("%H%M%S")}.xlsx'
            excel_path = os.path.join(export_dir, excel_filename)
            wb.save(excel_path)
            
            self.show_message('موفق', f'فایل اکسل ذخیره شد:\n{excel_filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی اکسل: {str(e)}')


    def _export_image_only(self):
        """خروجی تصویری با bidi"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import arabic_reshaper
            from bidi.algorithm import get_display
            
            if not self.filtered_data:
                self.show_message('خطا', 'داده ای برای خروجی وجود ندارد')
                return
            
            font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Vazir.ttf')
            if not os.path.exists(font_path):
                font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Amiri-Regular.ttf')
            
            try:
                font_title = ImageFont.truetype(font_path, 24)
                font_header = ImageFont.truetype(font_path, 14)
                font_row = ImageFont.truetype(font_path, 12)
            except:
                font_title = ImageFont.load_default()
                font_header = ImageFont.load_default()
                font_row = ImageFont.load_default()
            
            def fix_text(text):
                if not text:
                    return ''
                try:
                    reshaped = arabic_reshaper.reshape(str(text))
                    return get_display(reshaped)
                except:
                    return str(text)
            
            col_widths = [35, 80, 100, 100, 50, 70, 80, 80, 80]
            headers = ['ردیف', 'تاریخ', 'عامل', 'مشتری', 'وضعیت', 'نوع', 'نقد', 'چک', 'جمع']
            
            row_height = 32
            header_height = 40
            title_height = 50
            padding = 12
            
            table_width = sum(col_widths) + (len(col_widths) + 1) * 2
            table_height = title_height + header_height + len(self.filtered_data) * row_height + padding * 3
            
            img = Image.new('RGB', (table_width + 30, table_height + 30), color=(18, 18, 24))
            draw = ImageDraw.Draw(img)
            
            title_text = fix_text(f'گزارش وصول مطالبات - {get_today_jalali()}')
            draw.text((table_width // 2, padding), title_text, fill=(255, 215, 0), font=font_title, anchor='ma')
            
            y = title_height + padding
            x_start = 15
            
            for col_idx, (header, width) in enumerate(zip(headers, col_widths)):
                x = x_start + sum(col_widths[:col_idx]) + col_idx * 2
                draw.rectangle([x, y, x + width, y + header_height], fill=(30, 60, 90))
                draw.rectangle([x, y, x + width, y + header_height], outline=(50, 50, 60))
                draw.text((x + width // 2, y + header_height // 2), fix_text(header), fill=(255, 255, 255), font=font_header, anchor='mm')
            
            for row_idx, col in enumerate(self.filtered_data):
                y = title_height + header_height + padding * 2 + row_idx * row_height
                
                is_success = col.get('status') == 'موفق'
                bg_color = (25, 45, 35) if is_success else (50, 30, 25)
                
                if col.get('has_cash') and col.get('has_check'):
                    payment_type = 'نقد + چک'
                elif col.get('has_cash'):
                    payment_type = 'نقد'
                elif col.get('has_check'):
                    payment_type = 'چک'
                else:
                    payment_type = '-'
                
                status_text = 'موفق' if is_success else 'ناموفق'
                status_color = (100, 255, 100) if is_success else (255, 100, 100)
                
                row_values = [
                    str(row_idx + 1),
                    col.get('date', ''),
                    col.get('agent_name', ''),
                    col.get('customer', ''),
                    status_text,
                    payment_type,
                    f"{col.get('net_cash', 0):,.0f}",
                    f"{col.get('total_check_amount', 0):,.0f}",
                    f"{col.get('total_collection', 0):,.0f}",
                ]
                
                colors = [
                    (200, 200, 200), (255, 255, 255), (200, 200, 200), (255, 255, 255),
                    status_color, (255, 255, 255), (255, 255, 255), (255, 255, 255), (255, 255, 255),
                ]
                
                for col_idx, (value, width) in enumerate(zip(row_values, col_widths)):
                    x = x_start + sum(col_widths[:col_idx]) + col_idx * 2
                    draw.rectangle([x, y, x + width, y + row_height], fill=bg_color)
                    draw.rectangle([x, y, x + width, y + row_height], outline=(40, 40, 50))
                    draw.text((x + width // 2, y + row_height // 2), fix_text(value), fill=colors[col_idx], font=font_row, anchor='mm')
            
            today = get_today_jalali().replace('/', '-')
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            image_filename = f'گزارش_وصول_{today}_{datetime.now().strftime("%H%M%S")}.png'
            image_path = os.path.join(export_dir, image_filename)
            img.save(image_path, quality=95)
            
            self.show_message('موفق', f'تصویر گزارش ذخیره شد:\n{image_filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول Pillow یا bidi نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی تصویری: {str(e)}')


    def _export_aggregated_excel(self):
        """خروجی اکسل گزارش تجمیعی"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            data_source = self.agg_filtered_data if hasattr(self, 'agg_filtered_data') and self.agg_filtered_data else self.filtered_data
            
            if not data_source:
                self.show_message('خطا', 'داده ای برای خروجی وجود ندارد')
                return
            
            agent_data = {}
            for item in data_source:
                agent_name = item.get('agent_name', 'نامشخص')
                if agent_name not in agent_data:
                    agent_data[agent_name] = {
                        'count': 0,
                        'success_count': 0,
                        'fail_count': 0,
                        'total_cash': 0,
                        'total_check': 0,
                        'total_amount': 0,
                        'customers': set(),
                        'routes': set()
                    }
                
                agent_data[agent_name]['count'] += 1
                if item.get('status') == 'موفق':
                    agent_data[agent_name]['success_count'] += 1
                    agent_data[agent_name]['total_cash'] += item.get('net_cash', 0)
                    agent_data[agent_name]['total_check'] += item.get('total_check_amount', 0)
                else:
                    agent_data[agent_name]['fail_count'] += 1
                
                agent_data[agent_name]['total_amount'] += item.get('total_collection', 0)
                agent_data[agent_name]['customers'].add(item.get('customer', ''))
                agent_data[agent_name]['routes'].add(item.get('route', ''))
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "گزارش تجمیعی"
            ws.right_to_left = True
            
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = [
                'ردیف', 'عامل', 'تعداد وصول', 'موفق', 'ناموفق',
                'تعداد مشتریان', 'تعداد مسیرها', 'جمع نقد (ریال)',
                'جمع چک (ریال)', 'جمع کل (ریال)', 'میانگین', 'نرخ موفقیت'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            row_idx = 2
            for agent_name, data in agent_data.items():
                total = data['total_amount']
                count = data['count']
                success = data['success_count']
                fail = data['fail_count']
                success_rate = (success / count * 100) if count > 0 else 0
                avg = total / count if count > 0 else 0
                
                values = [
                    row_idx - 1,
                    agent_name,
                    count,
                    success,
                    fail,
                    len(data['customers']),
                    len(data['routes']),
                    data['total_cash'],
                    data['total_check'],
                    total,
                    round(avg, 0),
                    f"{success_rate:.1f}%"
                ]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    if col_idx == 10:
                        cell.font = Font(bold=True, color="FFD700")
                
                row_idx += 1
            
            column_widths = [6, 20, 12, 10, 10, 14, 14, 16, 16, 16, 14, 14]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            today = get_today_jalali().replace('/', '-')
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            excel_filename = f'گزارش_تجمیعی_{today}_{datetime.now().strftime("%H%M%S")}.xlsx'
            excel_path = os.path.join(export_dir, excel_filename)
            wb.save(excel_path)
            
            self.show_message('موفق', f'فایل اکسل ذخیره شد:\n{excel_filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول openpyxl نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی اکسل: {str(e)}')


    def _export_aggregated_image(self):
        """خروجی تصویری گزارش تجمیعی"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import arabic_reshaper
            from bidi.algorithm import get_display
            
            data_source = self.agg_filtered_data if hasattr(self, 'agg_filtered_data') and self.agg_filtered_data else self.filtered_data
            
            if not data_source:
                self.show_message('خطا', 'داده ای برای خروجی وجود ندارد')
                return
            
            agent_data = {}
            for item in data_source:
                agent_name = item.get('agent_name', 'نامشخص')
                if agent_name not in agent_data:
                    agent_data[agent_name] = {
                        'count': 0,
                        'success_count': 0,
                        'fail_count': 0,
                        'total_cash': 0,
                        'total_check': 0,
                        'total_amount': 0,
                        'customers': set(),
                        'routes': set()
                    }
                
                agent_data[agent_name]['count'] += 1
                if item.get('status') == 'موفق':
                    agent_data[agent_name]['success_count'] += 1
                    agent_data[agent_name]['total_cash'] += item.get('net_cash', 0)
                    agent_data[agent_name]['total_check'] += item.get('total_check_amount', 0)
                else:
                    agent_data[agent_name]['fail_count'] += 1
                
                agent_data[agent_name]['total_amount'] += item.get('total_collection', 0)
                agent_data[agent_name]['customers'].add(item.get('customer', ''))
                agent_data[agent_name]['routes'].add(item.get('route', ''))
            
            font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Vazir.ttf')
            if not os.path.exists(font_path):
                font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'Amiri-Regular.ttf')
            
            try:
                font_title = ImageFont.truetype(font_path, 22)
                font_header = ImageFont.truetype(font_path, 13)
                font_row = ImageFont.truetype(font_path, 11)
            except:
                font_title = ImageFont.load_default()
                font_header = ImageFont.load_default()
                font_row = ImageFont.load_default()
            
            def fix_text(text):
                if not text:
                    return ''
                try:
                    reshaped = arabic_reshaper.reshape(str(text))
                    return get_display(reshaped)
                except:
                    return str(text)
            
            headers = ['ردیف', 'عامل', 'تعداد', 'موفق', 'ناموفق', 'مشتریان', 'مسیرها', 'نقد', 'چک', 'جمع کل']
            col_widths = [35, 130, 55, 50, 50, 55, 50, 70, 70, 80]
            
            row_height = 30
            header_height = 38
            title_height = 50
            padding = 12
            
            table_width = sum(col_widths) + (len(col_widths) + 1) * 2
            table_height = title_height + header_height + len(agent_data) * row_height + padding * 3
            
            img = Image.new('RGB', (table_width + 30, table_height + 30), color=(18, 18, 24))
            draw = ImageDraw.Draw(img)
            
            title_text = fix_text(f'گزارش تجمیعی وصول - {get_today_jalali()}')
            draw.text((table_width // 2, padding), title_text, fill=(255, 215, 0), font=font_title, anchor='ma')
            
            y = title_height + padding
            x_start = 15
            
            for col_idx, (header, width) in enumerate(zip(headers, col_widths)):
                x = x_start + sum(col_widths[:col_idx]) + col_idx * 2
                draw.rectangle([x, y, x + width, y + header_height], fill=(30, 60, 90))
                draw.rectangle([x, y, x + width, y + header_height], outline=(50, 50, 60))
                draw.text((x + width // 2, y + header_height // 2), fix_text(header), fill=(255, 255, 255), font=font_header, anchor='mm')
            
            row_idx = 0
            for agent_name, data in agent_data.items():
                y = title_height + header_height + padding * 2 + row_idx * row_height
                
                total = data['total_amount']
                count = data['count']
                success = data['success_count']
                fail = data['fail_count']
                
                bg_color = (25, 35, 45) if row_idx % 2 == 0 else (20, 28, 38)
                
                row_values = [
                    str(row_idx + 1),
                    agent_name,
                    str(count),
                    str(success),
                    str(fail),
                    str(len(data['customers'])),
                    str(len(data['routes'])),
                    f"{data['total_cash']:,.0f}",
                    f"{data['total_check']:,.0f}",
                    f"{total:,.0f}",
                ]
                
                colors = [
                    (200, 200, 200), (255, 255, 255), (255, 255, 255),
                    (100, 255, 100), (255, 100, 100), (255, 255, 100),
                    (100, 255, 255), (150, 255, 150), (255, 150, 255),
                    (255, 215, 0)
                ]
                
                for col_idx, (value, width) in enumerate(zip(row_values, col_widths)):
                    x = x_start + sum(col_widths[:col_idx]) + col_idx * 2
                    draw.rectangle([x, y, x + width, y + row_height], fill=bg_color)
                    draw.rectangle([x, y, x + width, y + row_height], outline=(40, 40, 50))
                    draw.text((x + width // 2, y + row_height // 2), fix_text(value), fill=colors[col_idx], font=font_row, anchor='mm')
                
                row_idx += 1
            
            today = get_today_jalali().replace('/', '-')
            export_dir = get_backup_path()
            os.makedirs(export_dir, exist_ok=True)
            
            image_filename = f'گزارش_تجمیعی_{today}_{datetime.now().strftime("%H%M%S")}.png'
            image_path = os.path.join(export_dir, image_filename)
            img.save(image_path, quality=95)
            
            self.show_message('موفق', f'تصویر گزارش ذخیره شد:\n{image_filename}')
            
        except ImportError:
            self.show_message('خطا', 'ماژول Pillow یا bidi نصب نیست')
        except Exception as e:
            self.show_message('خطا', f'خطا در خروجی تصویری: {str(e)}')


    def _export_checks_excel(self):
        """خروجی اکسل گزارش چک - استفاده از کلاس کمکی"""
        ExcelExportHelper.export_checks_excel(self.filtered_data, self.show_message)


    def _export_checks_image(self):
        """خروجی تصویری گزارش چک - استفاده از کلاس کمکی"""
        ExcelExportHelper.export_checks_image(self.filtered_data, self.show_message)


    def show_message(self, title, message):
        """نمایش پیام با پشتیبانی از متن طولانی"""
        try:
            from kivy.uix.boxlayout import BoxLayout
            from kivy.uix.scrollview import ScrollView
            from kivy.uix.label import Label
            from kivy.metrics import dp, sp
            from kivy.graphics import Color, Rectangle
            from utils.rtl_widgets import PersianPopup, PersianButton
            
            content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
            with content.canvas.before:
                Color(0.12, 0.12, 0.12, 1)
                rect = Rectangle(pos=content.pos, size=content.size)
                content.bind(pos=lambda i, v: setattr(rect, 'pos', v),
                        size=lambda i, v: setattr(rect, 'size', v))
            
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                reshaped = arabic_reshaper.reshape(message)
                display_text = get_display(reshaped)
            except:
                display_text = message
            
            msg_label = Label(
                text=display_text,
                font_size=sp(15),
                color=(1, 1, 1, 1),
                size_hint_y=None,
                halign='center',
                valign='top',
                text_size=(dp(550), None),
                font_name='fonts/Amiri-Regular.ttf',
                padding=(dp(20), 0, dp(0), 0)
            )
            
            lines = message.count('\n') + 1
            if len(message) > 50 and lines == 1:
                approx_chars_per_line = 35
                lines = (len(message) // approx_chars_per_line) + 1
            
            line_height = sp(15) + dp(6)
            label_height = max(lines * line_height + dp(25), dp(50))
            label_height = min(label_height, dp(490))
            
            msg_label.height = label_height
            msg_label.text_size = (dp(550), label_height)
            
            scroll = ScrollView(
                do_scroll_x=False,
                do_scroll_y=True,
                size_hint_y=None,
                height=label_height + dp(10),
                bar_width=dp(6),
                bar_color=(0.3, 0.5, 0.8, 0.8),
                bar_inactive_color=(0.2, 0.2, 0.2, 0.5)
            )
            scroll.add_widget(msg_label)
            content.add_widget(scroll)
            
            btn = PersianButton(
                text='باشه',
                size_hint_y=None,
                height=dp(45),
                font_size=sp(16),
                color=(1, 1, 1, 1),
                background_color=(0.2, 0.6, 1, 1)
            )
            content.add_widget(btn)
            
            popup = PersianPopup(
                title=title,
                content=content,
                size_hint=(0.9, None),
                height=label_height + dp(250),
                background_color=(0.08, 0.08, 0.08, 1)
            )
            btn.bind(on_press=popup.dismiss)
            popup.open()
            
        except Exception as e:
            print(f"خطا در نمایش پیام: {e}")
            import traceback
            traceback.print_exc()
            try:
                from error_handler import ErrorPopup
                ErrorPopup.show_error(str(message))
            except:
                pass